from collections import defaultdict
from typing import Optional
import asyncio

import csv
import json
import os
import re
import smtplib
import subprocess
import tempfile
from email.message import EmailMessage

import io
from pathlib import Path
from urllib import request as urlrequest
from urllib.parse import parse_qs, quote_plus, urlencode, urlparse, urlunparse
from dotenv import dotenv_values
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from jose import JWTError, jwt
from sqlalchemy import and_, create_engine, func, or_
from sqlalchemy.orm import Session, aliased

from ..config import (
    get_active_company_key,
    get_company_profiles,
    set_active_company,
    settings,
    upsert_company_profile,
)
from ..core.init_db import init_db
from ..core.deps import get_db, require_admin
from ..core.security import (
    ALGORITHM,
    SECRET_KEY,
    create_access_token,
    hash_password,
    verify_password,
)
from ..core.utils import local_now, local_now_naive, local_today
from ..database import get_current_database_url, refresh_engine
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from decimal import InvalidOperation

from ..models.inventory import (
    Bodega,
    EgresoInventario,
    EgresoItem,
    EgresoTipo,
    ExchangeRate,
    IngresoInventario,
    IngresoItem,
    IngresoTipo,
    Linea,
    Marca,
    Producto,
    ProductoCombo,
    Proveedor,
    SaldoProducto,
    Segmento,
)
from ..models.sales import (
    AccountingPolicySetting,
    AccountingEntry,
    AccountingEntryLine,
    AccountingVoucherType,
    Banco,
    CajaDiaria,
    CierreCaja,
    Cliente,
    CompanyProfileSetting,
    CobranzaAbono,
    CuentaBancaria,
    CuentaContable,
    DepositoCliente,
    EmailConfig,
    FormaPago,
    NotificationRecipient,
    PosPrintSetting,
    ReciboCaja,
    ReciboMotivo,
    ReciboRubro,
    ReversionToken,
    SalesInterfaceSetting,
    ProductoComision,
    Preventa,
    PreventaItem,
    VentaComisionAsignacion,
    VentaComisionFinal,
    Vendedor,
    VendedorBodega,
    VentaFactura,
    VentaItem,
    VentaPago,
)
from ..models.user import Branch, Permission, Role, User

router = APIRouter()

SALES_INTERFACE_OPTIONS = [
    {"code": "ropa", "label": "Interfaz Ventas Ropa"},
    {"code": "ferreteria", "label": "Interfaz Ferreteria"},
    {"code": "farmacia", "label": "Interfaz Farmacia"},
    {"code": "comestibles", "label": "Interfaz Tienda de Comestibles"},
]


def to_decimal(value: Optional[float]) -> Decimal:
    return Decimal(str(value or 0))


def _get_user_from_cookie(request: Request, db: Session) -> Optional[User]:
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        return None
    email = payload.get("sub")
    if not email:
        return None
    return db.query(User).filter(func.lower(User.email) == email.lower()).first()


def _permission_names(user: User) -> set[str]:
    return {perm.name for perm in (user.permissions or [])}


def _has_permission(user: User, perm: str) -> bool:
    if any(role.name == "administrador" for role in user.roles or []):
        return True
    return perm in _permission_names(user)


def _set_request_permissions(request: Request, user: User) -> None:
    if any(role.name == "administrador" for role in user.roles or []):
        perm_names = _permission_catalog_names()
    else:
        perm_names = _permission_names(user)
    request.state.permission_names = perm_names
    request.state.has_permissions = True


def _is_api_request(request: Request) -> bool:
    accept = (request.headers.get("accept") or "").lower()
    if "application/json" in accept:
        return True
    if request.headers.get("x-requested-with") == "fetch":
        return True
    if request.headers.get("hx-request"):
        return True
    return False


def _permission_redirect_url(request: Request, message: str) -> str:
    referer = request.headers.get("referer") or "/home"
    base_url = str(request.base_url).rstrip("/")
    if referer.startswith("http") and not referer.startswith(base_url):
        referer = "/home"
    parsed = urlparse(referer)
    query = parse_qs(parsed.query)
    query["perm_error"] = [message]
    new_query = urlencode(query, doseq=True)
    if parsed.scheme:
        return urlunparse(
            (parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment)
        )
    return f"{parsed.path}?{new_query}" if new_query else parsed.path


def _permission_sms_recipients() -> list[str]:
    raw = (settings.SMS_ALERT_RECIPIENTS or "").strip()
    if not raw:
        return []
    return [item.strip() for item in raw.split(",") if item.strip()]


def _send_permission_sms_alert(user: Optional[User], request: Request, perm: str) -> None:
    sms_url = (settings.SMS_WEBHOOK_URL or "").strip()
    recipients = _permission_sms_recipients()
    if not sms_url or not recipients:
        return
    user_label = user.full_name if user and user.full_name else (user.email if user else "desconocido")
    message = (
        "Acceso denegado. "
        f"Usuario: {user_label}. "
        f"Permiso: {perm}. "
        f"Ruta: {request.url.path}. "
        f"Hora: {local_now().strftime('%Y-%m-%d %H:%M:%S')}."
    )
    payload = {
        "to": recipients,
        "message": message,
        "user": user.email if user else None,
        "perm": perm,
        "path": request.url.path,
        "ip": request.client.host if request.client else None,
    }
    headers = {"Content-Type": "application/json"}
    token = (settings.SMS_WEBHOOK_TOKEN or "").strip()
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = json.dumps(payload).encode("utf-8")
    req = urlrequest.Request(sms_url, data=data, headers=headers, method="POST")
    with urlrequest.urlopen(req, timeout=3) as response:
        response.read()


def _enforce_permission(request: Request, user: User, perm: str) -> None:
    if not _has_permission(user, perm):
        message = "No tienes permisos para acceder a esta opcion del sistema."
        try:
            _send_permission_sms_alert(user, request, perm)
        except Exception:
            pass
        if _is_api_request(request):
            raise HTTPException(status_code=403, detail=message)
        raise HTTPException(
            status_code=303, headers={"Location": _permission_redirect_url(request, message)}
        )


PERMISSION_GROUPS = [
    {
        "title": "Menus principales (visibilidad)",
        "items": [
            {"name": "menu.home", "label": "Panel principal"},
            {"name": "menu.sales", "label": "Ventas"},
            {"name": "menu.sales.caliente", "label": "Ventas en caliente"},
            {"name": "menu.inventory", "label": "Inventarios"},
            {"name": "menu.inventory.caliente", "label": "Inventario en caliente"},
            {"name": "menu.finance", "label": "Finanzas"},
            {"name": "menu.accounting", "label": "Contabilidad"},
            {"name": "menu.reports", "label": "Informes"},
            {"name": "menu.data", "label": "Datos / catalogos"},
        ],
    },
    {
        "title": "Ventas (visibilidad sub-menu)",
        "items": [
            {"name": "menu.sales.utilitario", "label": "Utilitario de ventas"},
            {"name": "menu.sales.etiquetas", "label": "Impresion de etiquetas"},
            {"name": "menu.sales.cobranza", "label": "Gestion de cobranza"},
            {"name": "menu.sales.roc", "label": "Recibos de caja"},
            {"name": "menu.sales.depositos", "label": "Registro de depositos"},
            {"name": "menu.sales.cierre", "label": "Cierre de caja"},
            {"name": "menu.sales.comisiones", "label": "Registro de comisiones"},
            {"name": "menu.sales.preventas", "label": "Panel de preventas"},
            {"name": "menu.sales.preventas.mobile", "label": "Nueva preventa movil"},
        ],
    },
    {
        "title": "Inventarios (visibilidad sub-menu)",
        "items": [
            {"name": "menu.inventory.ingresos", "label": "Ingresos de inventario"},
            {"name": "menu.inventory.egresos", "label": "Egresos de inventario"},
        ],
    },
    {
        "title": "Accesos (acciones)",
        "items": [
            {"name": "access.sales", "label": "Acceso a ventas"},
            {"name": "access.sales.caliente", "label": "Ventas en caliente"},
            {"name": "access.sales.registrar", "label": "Registrar facturas"},
            {"name": "access.sales.pagos", "label": "Aplicar pagos"},
            {"name": "access.sales.utilitario", "label": "Utilitario de ventas"},
            {"name": "access.sales.etiquetas", "label": "Impresion de etiquetas"},
            {"name": "access.sales.cobranza", "label": "Cobranza / abonos"},
            {"name": "access.sales.roc", "label": "Recibos de caja"},
            {"name": "access.sales.depositos", "label": "Depositos bancarios"},
            {"name": "access.sales.cierre", "label": "Cierre de caja"},
            {"name": "access.sales.reversion", "label": "Reversion de facturas"},
            {"name": "access.sales.comisiones", "label": "Registro de comisiones"},
            {"name": "access.sales.preventas", "label": "Gestion de preventas"},
            {"name": "access.sales.preventas.mobile", "label": "Crear preventas (movil)"},
            {"name": "access.inventory", "label": "Acceso a inventarios"},
            {"name": "access.inventory.caliente", "label": "Inventario en caliente"},
            {"name": "access.inventory.ingresos", "label": "Ingresos de inventario"},
            {"name": "access.inventory.egresos", "label": "Egresos de inventario"},
            {"name": "access.inventory.productos", "label": "Crear/editar productos"},
            {"name": "access.finance", "label": "Acceso a finanzas"},
            {"name": "access.finance.rates", "label": "Configurar tasas"},
            {"name": "access.accounting", "label": "Acceso a contabilidad"},
            {"name": "access.accounting.financial_data", "label": "Datos financieros"},
            {"name": "access.accounting.entries", "label": "Comprobantes contables"},
            {"name": "access.accounting.voucher_types", "label": "Tipos de comprobantes"},
            {"name": "access.reports", "label": "Acceso a informes"},
            {"name": "access.data", "label": "Acceso a datos"},
            {"name": "access.data.permissions", "label": "Gestion de permisos"},
            {"name": "access.data.users", "label": "Usuarios"},
            {"name": "access.data.roles", "label": "Roles"},
            {"name": "access.data.catalogs", "label": "Catalogos"},
        ],
    },
]


def _permission_catalog_names() -> set[str]:
    names: set[str] = set()
    for group in PERMISSION_GROUPS:
        for item in group["items"]:
            names.add(item["name"])
    return names


def _ensure_permission_catalog_in_db(db: Session) -> None:
    catalog_names = _permission_catalog_names()
    existing_names = {
        name
        for (name,) in db.query(Permission.name).filter(Permission.name.in_(catalog_names)).all()
    }
    missing_names = sorted(catalog_names - existing_names)
    if not missing_names:
        return
    db.add_all([Permission(name=name) for name in missing_names])
    db.commit()


def _require_admin_web(
    request: Request, db: Session = Depends(get_db)
) -> User:
    user = _get_user_from_cookie(request, db)
    if not user:
        raise HTTPException(status_code=302, headers={"Location": "/login"})
    if user.is_active is False:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    _set_request_permissions(request, user)
    return user


def _require_user_web(
    request: Request, db: Session = Depends(get_db)
) -> User:
    user = _get_user_from_cookie(request, db)
    if not user:
        raise HTTPException(status_code=302, headers={"Location": "/login"})
    if user.is_active is False:
        raise HTTPException(status_code=403, detail="Acceso denegado")
    _set_request_permissions(request, user)
    return user


def _resolve_branch_bodega(db: Session, user: User) -> tuple[Optional[Branch], Optional[Bodega]]:
    allowed_codes = _allowed_branch_codes(db)
    user_branches = [b for b in (user.branches or []) if (b.code or "").lower() in allowed_codes]
    allowed_branch_ids = {b.id for b in user_branches}
    bodega = None
    if user.default_bodega_id:
        bodega = (
            db.query(Bodega)
            .join(Branch, Branch.id == Bodega.branch_id)
            .filter(Bodega.id == user.default_bodega_id, Bodega.activo.is_(True))
            .filter(func.lower(Branch.code).in_(allowed_codes))
            .first()
        )
        if bodega and allowed_branch_ids and bodega.branch_id not in allowed_branch_ids:
            bodega = None

    branch = None
    if bodega:
        branch = db.query(Branch).filter(Branch.id == bodega.branch_id).first()
    if not branch and user.default_branch_id:
        if not allowed_branch_ids or user.default_branch_id in allowed_branch_ids:
            branch = (
                db.query(Branch)
                .filter(Branch.id == user.default_branch_id)
                .filter(func.lower(Branch.code).in_(allowed_codes))
                .first()
            )
    if not branch and user_branches:
        branch = user_branches[0]
    if not branch:
        branch = (
            db.query(Branch)
            .filter(func.lower(Branch.code).in_(allowed_codes))
            .order_by(Branch.id)
            .first()
        )

    if branch and (not bodega or bodega.branch_id != branch.id):
        bodega = (
            db.query(Bodega)
            .filter(Bodega.branch_id == branch.id, Bodega.activo.is_(True))
            .order_by(Bodega.id)
            .first()
        )
    return branch, bodega


def _vendedores_for_bodega(db: Session, bodega: Optional[Bodega]) -> list[Vendedor]:
    base = db.query(Vendedor).filter(Vendedor.activo.is_(True))
    if not bodega:
        return base.order_by(Vendedor.nombre).all()
    assigned = (
        base.join(VendedorBodega, VendedorBodega.vendedor_id == Vendedor.id)
        .filter(VendedorBodega.bodega_id == bodega.id)
        .order_by(Vendedor.nombre)
        .all()
    )
    if assigned:
        return assigned
    return base.order_by(Vendedor.nombre).all()


def _vendedores_for_branch(db: Session, branch_id: Optional[int]) -> list[Vendedor]:
    base = db.query(Vendedor).filter(Vendedor.activo.is_(True))
    if not branch_id:
        return base.order_by(Vendedor.nombre).all()
    assigned = (
        base.join(VendedorBodega, VendedorBodega.vendedor_id == Vendedor.id)
        .join(Bodega, Bodega.id == VendedorBodega.bodega_id)
        .filter(Bodega.branch_id == branch_id)
        .distinct()
        .order_by(Vendedor.nombre)
        .all()
    )
    if assigned:
        return assigned
    return base.order_by(Vendedor.nombre).all()


def _default_vendedor_id(db: Session, bodega: Optional[Bodega]) -> Optional[int]:
    if not bodega:
        return None
    row = (
        db.query(VendedorBodega)
        .filter(VendedorBodega.bodega_id == bodega.id, VendedorBodega.is_default.is_(True))
        .first()
    )
    return row.vendedor_id if row else None


def _vendedor_id_for_user(db: Session, user: User, bodega: Optional[Bodega]) -> Optional[int]:
    user_name = (user.full_name or "").strip().lower()
    if not user_name:
        return None
    query = db.query(Vendedor).filter(Vendedor.activo.is_(True))
    if bodega:
        query = (
            query.join(VendedorBodega, VendedorBodega.vendedor_id == Vendedor.id)
            .filter(VendedorBodega.bodega_id == bodega.id)
            .distinct()
        )
    vendedor = query.filter(func.lower(Vendedor.nombre) == user_name).first()
    return vendedor.id if vendedor else None


def _is_vendedor_role(user: User) -> bool:
    return any((role.name or "").lower() == "vendedor" for role in (user.roles or []))


def _enforce_preventas_mobile_access(request: Request, user: User) -> None:
    if _has_permission(user, "access.sales.preventas.mobile") or _has_permission(
        user, "access.sales.preventas"
    ):
        return
    _enforce_permission(request, user, "access.sales.preventas.mobile")


def _preventa_estado_badge(estado: str) -> dict[str, str]:
    mapping = {
        "PENDIENTE": {"label": "Pendiente", "class": "bg-pink-100 text-pink-800"},
        "REVISION": {"label": "En revision", "class": "bg-amber-100 text-amber-800"},
        "FACTURADA": {"label": "Facturada", "class": "bg-emerald-100 text-emerald-800"},
        "ANULADA": {"label": "Anulada", "class": "bg-violet-100 text-violet-800"},
    }
    return mapping.get((estado or "").upper(), {"label": estado or "-", "class": "bg-slate-100 text-slate-700"})


def _get_or_create_consumidor_final(db: Session) -> Cliente:
    cliente = (
        db.query(Cliente)
        .filter(func.lower(Cliente.nombre) == "consumidor final")
        .first()
    )
    if cliente:
        return cliente
    cliente = Cliente(nombre="Consumidor final", activo=True)
    db.add(cliente)
    db.flush()
    return cliente


def _next_preventa_number(db: Session, branch: Branch) -> tuple[int, str]:
    last = (
        db.query(Preventa)
        .filter(Preventa.branch_id == branch.id)
        .order_by(Preventa.secuencia.desc())
        .first()
    )
    seq = (last.secuencia if last else 0) + 1
    branch_code = (branch.code or "").lower()
    prefix = "C" if branch_code == "central" else "E" if branch_code == "esteli" else branch_code[:1].upper()
    return seq, f"PV{prefix}-{seq:06d}"


def _preventa_active_conflict(
    db: Session,
    *,
    bodega_id: int,
    producto_id: int,
    vendedor_id: int,
) -> Optional[tuple[Preventa, Optional[Vendedor]]]:
    row = (
        db.query(Preventa, Vendedor)
        .join(PreventaItem, PreventaItem.preventa_id == Preventa.id)
        .join(Vendedor, Vendedor.id == Preventa.vendedor_id, isouter=True)
        .filter(
            Preventa.bodega_id == bodega_id,
            Preventa.estado.in_(["PENDIENTE", "REVISION"]),
            PreventaItem.producto_id == producto_id,
            Preventa.vendedor_id != vendedor_id,
        )
        .order_by(Preventa.id.desc())
        .first()
    )
    return row if row else None


def _preventa_reserved_by_others(
    db: Session,
    *,
    bodega_id: int,
    producto_id: int,
    vendedor_id: int,
) -> tuple[Decimal, list[tuple[str, str, Decimal]]]:
    rows = (
        db.query(
            Preventa.numero,
            Vendedor.nombre,
            func.sum(PreventaItem.cantidad).label("qty"),
        )
        .join(PreventaItem, PreventaItem.preventa_id == Preventa.id)
        .join(Vendedor, Vendedor.id == Preventa.vendedor_id, isouter=True)
        .filter(
            Preventa.bodega_id == bodega_id,
            Preventa.estado.in_(["PENDIENTE", "REVISION"]),
            PreventaItem.producto_id == producto_id,
            Preventa.vendedor_id != vendedor_id,
        )
        .group_by(Preventa.numero, Vendedor.nombre)
        .order_by(Preventa.numero.asc())
        .all()
    )
    details: list[tuple[str, str, Decimal]] = []
    total = Decimal("0")
    for numero, vendedor_nombre, qty in rows:
        qty_dec = Decimal(str(qty or 0))
        details.append((str(numero or "-"), str(vendedor_nombre or "Vendedor"), qty_dec))
        total += qty_dec
    return total, details


def _preventa_reserved_bulk_by_others(
    db: Session,
    *,
    bodega_id: int,
    producto_ids: list[int],
    vendedor_id: Optional[int] = None,
    include_same_vendedor: bool = False,
) -> tuple[dict[int, Decimal], dict[int, list[dict[str, object]]]]:
    if not producto_ids:
        return {}, {}
    base = (
        db.query(
            PreventaItem.producto_id.label("producto_id"),
            Preventa.id.label("preventa_id"),
            Preventa.numero.label("numero"),
            Preventa.vendedor_id.label("vendedor_id"),
            Vendedor.nombre.label("vendedor_nombre"),
            func.sum(PreventaItem.cantidad).label("qty"),
        )
        .join(Preventa, Preventa.id == PreventaItem.preventa_id)
        .join(Vendedor, Vendedor.id == Preventa.vendedor_id, isouter=True)
        .filter(
            Preventa.bodega_id == bodega_id,
            Preventa.estado.in_(["PENDIENTE", "REVISION"]),
            PreventaItem.producto_id.in_(producto_ids),
        )
    )
    if vendedor_id and not include_same_vendedor:
        base = base.filter(Preventa.vendedor_id != vendedor_id)
    rows = (
        base.group_by(
            PreventaItem.producto_id,
            Preventa.id,
            Preventa.numero,
            Preventa.vendedor_id,
            Vendedor.nombre,
        )
        .order_by(PreventaItem.producto_id.asc(), Preventa.id.asc())
        .all()
    )
    totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    details: dict[int, list[dict[str, object]]] = defaultdict(list)
    for producto_id, preventa_id, numero, row_vendedor_id, vendedor_nombre, qty in rows:
        qty_dec = Decimal(str(qty or 0))
        pid = int(producto_id)
        totals[pid] += qty_dec
        vend_id = int(row_vendedor_id) if row_vendedor_id else 0
        details[pid].append(
            {
                "preventa_id": int(preventa_id),
                "numero": str(numero or "-"),
                "vendedor_id": vend_id,
                "vendedor": str(vendedor_nombre or "Vendedor"),
                "same_vendedor": bool(vendedor_id and vend_id and int(vendedor_id) == vend_id),
                "qty": float(qty_dec),
            }
        )
    return dict(totals), dict(details)


def _preventa_required_qty_map(
    item_rows: list[tuple["PreventaItem", "Producto"]],
) -> dict[int, Decimal]:
    required: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    for row, _producto in item_rows:
        required[int(row.producto_id)] += Decimal(str(row.cantidad or 0))
    return required


def _repair_preventa_currency_if_needed(db: Session, preventa: Optional[Preventa]) -> bool:
    if not preventa or preventa.estado not in {"PENDIENTE", "REVISION"}:
        return False
    rows = (
        db.query(PreventaItem, Producto)
        .join(Producto, Producto.id == PreventaItem.producto_id)
        .filter(PreventaItem.preventa_id == preventa.id)
        .all()
    )
    touched = False
    total_usd = Decimal("0")
    total_cs = Decimal("0")
    total_items = Decimal("0")
    for item, producto in rows:
        qty = Decimal(str(item.cantidad or 0))
        role = (item.combo_role or "").strip().lower() if getattr(item, "combo_role", None) else ""
        if role == "":
            item_cs = Decimal(str(item.precio_unitario_cs or 0))
            prod_usd = Decimal(str(producto.precio_venta1_usd or 0))
            prod_cs = Decimal(str(producto.precio_venta1 or 0))
            bug_pattern = prod_usd > 0 and prod_cs > 0 and abs(item_cs - prod_usd) <= Decimal("0.01")
            if bug_pattern:
                item.precio_unitario_usd = prod_usd
                item.precio_unitario_cs = prod_cs
                item.subtotal_usd = (prod_usd * qty).quantize(Decimal("0.01"))
                item.subtotal_cs = (prod_cs * qty).quantize(Decimal("0.01"))
                touched = True
        total_usd += Decimal(str(item.subtotal_usd or 0))
        total_cs += Decimal(str(item.subtotal_cs or 0))
        total_items += qty
    if touched:
        preventa.total_usd = total_usd.quantize(Decimal("0.01"))
        preventa.total_cs = total_cs.quantize(Decimal("0.01"))
        preventa.total_items = total_items.quantize(Decimal("0.01"))
    return touched


def _get_sumatra_path(config_path: Optional[str] = None) -> Optional[Path]:
    env_path = os.getenv("SUMATRA_PATH")
    candidates = [
        config_path,
        env_path,
        r"C:\Program Files\SumatraPDF\SumatraPDF.exe",
        r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe",
        r"C:\Users\USER\AppData\Local\SumatraPDF\SumatraPDF.exe",
    ]
    for candidate in candidates:
        if not candidate:
            continue
        path = Path(candidate)
        if path.exists():
            return path
    return None


def _balances_by_bodega(
    db: Session,
    bodega_ids: list[int],
    product_ids: list[int],
) -> dict[tuple[int, int], Decimal]:
    if not bodega_ids or not product_ids:
        return {}
    ingreso_rows = (
        db.query(IngresoItem.producto_id, IngresoInventario.bodega_id, func.sum(IngresoItem.cantidad))
        .join(IngresoInventario, IngresoInventario.id == IngresoItem.ingreso_id)
        .filter(IngresoInventario.bodega_id.in_(bodega_ids))
        .filter(IngresoItem.producto_id.in_(product_ids))
        .group_by(IngresoItem.producto_id, IngresoInventario.bodega_id)
        .all()
    )
    egreso_rows = (
        db.query(EgresoItem.producto_id, EgresoInventario.bodega_id, func.sum(EgresoItem.cantidad))
        .join(EgresoInventario, EgresoInventario.id == EgresoItem.egreso_id)
        .filter(EgresoInventario.bodega_id.in_(bodega_ids))
        .filter(EgresoItem.producto_id.in_(product_ids))
        .group_by(EgresoItem.producto_id, EgresoInventario.bodega_id)
        .all()
    )
    venta_rows = (
        db.query(VentaItem.producto_id, VentaFactura.bodega_id, func.sum(VentaItem.cantidad))
        .join(VentaFactura, VentaFactura.id == VentaItem.factura_id)
        .filter(VentaFactura.bodega_id.in_(bodega_ids))
        .filter(VentaItem.producto_id.in_(product_ids))
        .filter(VentaFactura.estado != "ANULADA")
        .group_by(VentaItem.producto_id, VentaFactura.bodega_id)
        .all()
    )
    balances: dict[tuple[int, int], Decimal] = {}
    for producto_id, bodega_id, qty in ingreso_rows:
        balances[(producto_id, bodega_id)] = Decimal(str(qty or 0))
    for producto_id, bodega_id, qty in egreso_rows:
        balances[(producto_id, bodega_id)] = balances.get((producto_id, bodega_id), Decimal("0")) - Decimal(str(qty or 0))
    for producto_id, bodega_id, qty in venta_rows:
        balances[(producto_id, bodega_id)] = balances.get((producto_id, bodega_id), Decimal("0")) - Decimal(str(qty or 0))
    return balances


def _default_company_profile_payload() -> dict[str, str]:
    multi_branch_enabled = get_active_company_key() != "comestibles"
    return {
        "legal_name": "Hollywood Pacas",
        "trade_name": "Hollywood Pacas",
        "app_title": "ERP Hollywood Pacas",
        "sidebar_subtitle": "ERP Central",
        "website": "http://hollywoodpacas.com.ni",
        "ruc": "",
        "phone": "8900-0300",
        "address": "",
        "email": "admin@hollywoodpacas.com",
        "logo_url": "/static/logo_hollywood.png",
        "pos_logo_url": "/static/logo_hollywood.png",
        "favicon_url": "/static/favicon.ico",
        "inventory_cs_only": False,
        "multi_branch_enabled": multi_branch_enabled,
        "price_auto_from_cost_enabled": False,
        "price_margin_percent": 0,
    }


def _company_profile_payload(db: Session) -> dict[str, str]:
    payload = _default_company_profile_payload()
    row = db.query(CompanyProfileSetting).order_by(CompanyProfileSetting.id.asc()).first()
    if not row:
        return payload
    payload.update(
        {
            "legal_name": row.legal_name or payload["legal_name"],
            "trade_name": row.trade_name or payload["trade_name"],
            "app_title": row.app_title or payload["app_title"],
            "sidebar_subtitle": row.sidebar_subtitle or payload["sidebar_subtitle"],
            "website": row.website or payload["website"],
            "ruc": row.ruc or payload["ruc"],
            "phone": row.phone or payload["phone"],
            "address": row.address or payload["address"],
            "email": row.email or payload["email"],
            "logo_url": row.logo_url or payload["logo_url"],
            "pos_logo_url": row.pos_logo_url or payload["pos_logo_url"],
            "favicon_url": row.favicon_url or payload["favicon_url"],
            "inventory_cs_only": bool(row.inventory_cs_only),
            "multi_branch_enabled": bool(row.multi_branch_enabled),
            "price_auto_from_cost_enabled": bool(row.price_auto_from_cost_enabled),
            "price_margin_percent": int(row.price_margin_percent or 0),
        }
    )
    return payload


def _inventory_cs_only_mode(db: Session) -> bool:
    profile = _company_profile_payload(db)
    return bool(profile.get("inventory_cs_only"))


def _multi_branch_enabled_mode(db: Session) -> bool:
    if get_active_company_key() == "comestibles":
        return False
    profile = _company_profile_payload(db)
    return bool(profile.get("multi_branch_enabled", True))


def _price_margin_mode(db: Session) -> tuple[bool, int]:
    profile = _company_profile_payload(db)
    enabled = bool(profile.get("price_auto_from_cost_enabled", False))
    try:
        percent = int(profile.get("price_margin_percent") or 0)
    except (TypeError, ValueError):
        percent = 0
    return enabled, max(0, percent)


def _allowed_branch_codes(db: Session) -> set[str]:
    return {"central", "esteli"} if _multi_branch_enabled_mode(db) else {"central"}


def _allowed_branch_ids(db: Session) -> list[int]:
    return [row.id for row in _scoped_branches_query(db).all()]


def _scoped_branches_query(db: Session):
    codes = _allowed_branch_codes(db)
    return db.query(Branch).filter(func.lower(Branch.code).in_(codes))


def _scoped_bodegas_query(db: Session):
    codes = _allowed_branch_codes(db)
    return (
        db.query(Bodega)
        .join(Branch, Branch.id == Bodega.branch_id)
        .filter(Bodega.activo.is_(True))
        .filter(func.lower(Branch.code).in_(codes))
    )


def _user_scoped_branch_ids(db: Session, user: User) -> set[int]:
    allowed_codes = _allowed_branch_codes(db)
    user_ids = {
        int(branch.id)
        for branch in (user.branches or [])
        if (branch.code or "").lower() in allowed_codes
    }
    if user_ids:
        return user_ids
    if user.default_branch_id:
        branch = (
            db.query(Branch)
            .filter(Branch.id == user.default_branch_id)
            .filter(func.lower(Branch.code).in_(allowed_codes))
            .first()
        )
        if branch:
            return {int(branch.id)}
    return {int(row.id) for row in _scoped_branches_query(db).all()}


def _resolve_logo_path(logo_url: str, *, prefer_pos: bool = False, pos_logo_url: Optional[str] = None) -> Path:
    static_dir = Path(__file__).resolve().parents[1] / "static"
    if prefer_pos:
        normalized_pos = (pos_logo_url or "").strip()
        if normalized_pos:
            if normalized_pos.startswith("/static/"):
                pos_candidate = static_dir / normalized_pos.replace("/static/", "", 1)
                if pos_candidate.exists():
                    return pos_candidate
            absolute_pos_candidate = Path(normalized_pos)
            if absolute_pos_candidate.exists():
                return absolute_pos_candidate
        # Compatibilidad con instalaciones antiguas.
        pos_logo = static_dir / "logopos.png"
        if pos_logo.exists():
            return pos_logo

    normalized = (logo_url or "").strip()
    if normalized:
        if normalized.startswith("/static/"):
            candidate = static_dir / normalized.replace("/static/", "", 1)
            if candidate.exists():
                return candidate
        absolute_candidate = Path(normalized)
        if absolute_candidate.exists():
            return absolute_candidate

    fallback = static_dir / "logo_hollywood.png"
    return fallback


def _company_identity(branch: Optional[Branch], profile: dict[str, str]) -> dict[str, str]:
    company_name = (
        profile.get("trade_name", "")
        or (branch.company_name if branch and branch.company_name else "")
        or "Empresa"
    ).strip()
    ruc = (
        (profile.get("ruc", "") or "").strip()
        or (branch.ruc if branch and branch.ruc else "")
        or "-"
    )
    telefono = (
        (profile.get("phone", "") or "").strip()
        or (branch.telefono if branch and branch.telefono else "")
        or "-"
    ).strip() or "-"
    direccion = (
        (profile.get("address", "") or "").strip()
        or (branch.direccion if branch and branch.direccion else "")
        or "-"
    ).strip() or "-"
    sucursal = branch.name if branch and branch.name else "-"
    return {
        "company_name": company_name or "Empresa",
        "ruc": ruc,
        "telefono": telefono,
        "direccion": direccion,
        "sucursal": sucursal,
    }


def _build_pos_ticket_pdf_bytes(factura: VentaFactura, profile: Optional[dict[str, str]] = None) -> bytes:
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    def wrap_text(text: str, max_chars: int) -> list[str]:
        if not text:
            return [""]
        words = text.split()
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if len(candidate) > max_chars:
                if current:
                    lines.append(current)
                current = word
            else:
                current = candidate
        if current:
            lines.append(current)
        return lines or [text]

    def format_qty(value: float) -> str:
        return f"{value:.2f}".rstrip("0").rstrip(".")

    def format_amount(value: float) -> str:
        return f"{value:,.2f}"

    def extract_weight_lbs(text: str) -> float:
        if not text:
            return 0.0
        match = re.search(r"\b(\d+(?:\.\d+)?)\s*(lbs)\b", text.lower())
        if not match:
            return 0.0
        try:
            return float(match.group(1))
        except ValueError:
            return 0.0

    branch = factura.bodega.branch if factura.bodega else None
    company_profile = profile or _default_company_profile_payload()
    identity = _company_identity(branch, company_profile)
    company_name = identity["company_name"]
    ruc = identity["ruc"]
    telefono = identity["telefono"]
    direccion = identity["direccion"]
    sucursal = identity["sucursal"]

    cliente = factura.cliente.nombre if factura.cliente else "Consumidor final"
    cliente_id = factura.cliente.identificacion if factura.cliente and factura.cliente.identificacion else "-"
    vendedor = factura.vendedor.nombre if factura.vendedor else "-"

    fecha_base = factura.created_at or factura.fecha
    fecha_str = ""
    hora_str = ""
    if fecha_base:
        try:
            fecha_str = fecha_base.strftime("%d/%m/%Y")
            hora_str = fecha_base.strftime("%H:%M")
        except AttributeError:
            fecha_str = str(fecha_base)

    moneda = factura.moneda or "CS"
    currency_label = "C$" if moneda == "CS" else "$"
    total_amount = float(factura.total_cs or 0) if moneda == "CS" else float(factura.total_usd or 0)
    subtotal_amount = total_amount

    pagos = factura.pagos or []
    total_paid = sum(
        float(pago.monto_cs or 0) if moneda == "CS" else float(pago.monto_usd or 0)
        for pago in pagos
    )
    saldo = total_paid - total_amount

    lines: list[tuple[str, str, bool, int]] = []

    def add_line(text: str, align: str = "left", bold: bool = False, size: int = 8):
        lines.append((text, align, bold, size))

    add_line(company_name.upper(), "center", True, 10)
    add_line(f"RUC: {ruc}", "center")
    add_line(f"Tel: {telefono}", "center")
    direccion_lines = wrap_text(direccion, 32)[:2]
    for line in direccion_lines:
        add_line(line, "center")
    add_line(f"Sucursal: {sucursal}", "center", True, 9)
    add_line("-" * 32, "center")
    add_line(f"Factura: {factura.numero}", "left", True)
    add_line(f"Fecha: {fecha_str} {hora_str}".strip())
    add_line(f"Cliente: {cliente}")
    add_line(f"Identificacion R/C: {cliente_id}")
    add_line(f"Vendedor: {vendedor}")
    add_line("-" * 32, "center")

    max_desc = 32
    total_bultos = 0.0
    total_lbs = 0.0
    for item in factura.items:
        codigo = item.producto.cod_producto if item.producto else "-"
        descripcion = item.producto.descripcion if item.producto else "-"
        combo_label = ""
        if item.combo_role == "gift":
            combo_label = " [REGALO]"
        elif item.combo_role == "parent":
            combo_label = " [OFERTA]"
        qty = float(item.cantidad or 0)
        price = (
            float(item.precio_unitario_cs or 0)
            if moneda == "CS"
            else float(item.precio_unitario_usd or 0)
        )
        subtotal = (
            float(item.subtotal_cs or 0)
            if moneda == "CS"
            else float(item.subtotal_usd or 0)
        )
        total_bultos += qty
        lbs_per_unit = extract_weight_lbs(descripcion)
        if lbs_per_unit:
            total_lbs += lbs_per_unit * qty
        add_line(f"Codigo: {codigo}{combo_label}", "left", True, 9)
        for part in wrap_text(descripcion, max_desc):
            add_line(part, "left", False, 9)
        add_line(
            f"Cant: {format_qty(qty)}  Precio: {currency_label} {format_amount(price)}",
            "left",
            False,
            9,
        )
        add_line(
            f"Desc: {currency_label} 0.00  Subtotal: {currency_label} {format_amount(subtotal)}",
            "left",
            False,
            9,
        )
        add_line("-" * 32, "center")

    add_line(f"Total bultos: {format_qty(total_bultos)}", "left", True, 9)
    add_line(f"Total libras: {format_qty(total_lbs)}", "left", True, 9)
    add_line(f"Subtotal: {currency_label} {format_amount(subtotal_amount)}", "right", True, 9)
    add_line(f"Descuentos: {currency_label} 0.00", "right", False, 9)
    add_line(f"Total: {currency_label} {format_amount(total_amount)}", "right", True, 10)

    if pagos:
        add_line("-" * 32, "center")
        add_line("Pagos aplicados", "left", True, 9)
        for pago in pagos:
            forma = pago.forma_pago.nombre if pago.forma_pago else "Pago"
            banco = pago.banco.nombre if pago.banco else ""
            label = f"{forma} {banco}".strip()
            monto = (
                float(pago.monto_cs or 0)
                if moneda == "CS"
                else float(pago.monto_usd or 0)
            )
            add_line(f"{label}: {currency_label} {format_amount(monto)}", "left", False, 9)

    if saldo >= 0:
        add_line(f"Vuelto: {currency_label} {format_amount(saldo)}", "left", True, 9)
    else:
        add_line(f"Saldo: {currency_label} {format_amount(abs(saldo))}", "left", True, 9)

    add_line("")
    add_line("Gracias por su compra", "center", True, 9)
    add_line("Revise su mercaderia antes de salir.", "center", False, 9)
    add_line("No se aceptan cambios ni devoluciones", "center", False, 9)
    add_line("de mercaderia ni de dinero.", "center", False, 9)

    content_width = 70 * mm
    width = 80 * mm
    margin = (width - content_width) / 2
    top_margin = 6 * mm
    bottom_margin = 6 * mm
    logo_path = _resolve_logo_path(
        company_profile.get("logo_url", ""),
        prefer_pos=True,
        pos_logo_url=company_profile.get("pos_logo_url", ""),
    )
    logo_height = 60 * mm if logo_path.exists() else 0
    logo_spacing = 3 * mm if logo_height else 0

    def line_gap(size: int) -> float:
        return size + 4

    total_height = top_margin + bottom_margin + logo_height + logo_spacing
    total_height += sum(line_gap(size) for _, _, _, size in lines)
    total_height = max(total_height, 120 * mm)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(width, total_height))
    y = total_height - top_margin

    if logo_height:
        logo_width = 78 * mm
        pdf.drawImage(
            str(logo_path),
            (width - logo_width) / 2,
            y - logo_height,
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask="auto",
        )
        y -= logo_height + logo_spacing

    for text, align, bold, size in lines:
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        if align == "center":
            pdf.drawCentredString(width / 2, y, text)
        elif align == "right":
            pdf.drawRightString(width - margin, y, text)
        else:
            pdf.drawString(margin, y, text)
        y -= line_gap(size)

    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


def _print_pos_ticket(
    factura: VentaFactura,
    printer_name: str,
    copies: int,
    profile: Optional[dict[str, str]] = None,
    sumatra_override: Optional[str] = None,
) -> None:
    sumatra_path = _get_sumatra_path(sumatra_override)
    if not sumatra_path:
        return
    pdf_bytes = _build_pos_ticket_pdf_bytes(factura, profile)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(pdf_bytes)
        tmp_path = tmp_file.name
    try:
        for _ in range(max(copies, 1)):
            subprocess.run(
                [
                    str(sumatra_path),
                    "-print-to",
                    printer_name,
                    "-print-settings",
                    "noscale",
                    "-silent",
                    tmp_path,
                ],
                check=False,
            )
    finally:
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except OSError:
            pass


def _build_roc_ticket_pdf_bytes(recibo: ReciboCaja, profile: Optional[dict[str, str]] = None) -> bytes:
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    def wrap_text(text: str, max_chars: int) -> list[str]:
        if not text:
            return [""]
        words = text.split()
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if len(candidate) > max_chars:
                if current:
                    lines.append(current)
                current = word
            else:
                current = candidate
        if current:
            lines.append(current)
        return lines or [text]

    def format_amount(value: float) -> str:
        return f"{value:,.2f}"

    branch = recibo.branch
    company_profile = profile or _default_company_profile_payload()
    identity = _company_identity(branch, company_profile)
    company_name = identity["company_name"]
    ruc = identity["ruc"]
    telefono = identity["telefono"]
    direccion = identity["direccion"]
    sucursal = identity["sucursal"]
    bodega_name = recibo.bodega.name if recibo.bodega else "-"

    fecha_base = recibo.created_at or recibo.fecha
    fecha_str = ""
    hora_str = ""
    if fecha_base:
        try:
            fecha_str = fecha_base.strftime("%d/%m/%Y")
            hora_str = fecha_base.strftime("%H:%M")
        except AttributeError:
            fecha_str = str(fecha_base)

    moneda = recibo.moneda or "CS"
    currency_label = "C$" if moneda == "CS" else "$"
    monto_total = float(recibo.monto_cs or 0) if moneda == "CS" else float(recibo.monto_usd or 0)
    monto_cs = float(recibo.monto_cs or 0)

    rubro = recibo.rubro.nombre if recibo.rubro else "-"
    motivo = recibo.motivo.nombre if recibo.motivo else "-"
    descripcion = recibo.descripcion or "-"
    rubro_codigo = recibo.rubro.cuenta.codigo if recibo.rubro and recibo.rubro.cuenta else "9999"
    rubro_cuenta = f"{rubro_codigo} {rubro}"
    caja_cuenta = "1101 Caja"

    lines: list[tuple[str, str, bool, int]] = []

    def add_line(text: str, align: str = "left", bold: bool = False, size: int = 8):
        lines.append((text, align, bold, size))

    add_line(company_name.upper(), "center", True, 10)
    add_line(f"RUC: {ruc}", "center")
    add_line(f"Tel: {telefono}", "center")
    direccion_lines = wrap_text(direccion, 32)[:2]
    for line in direccion_lines:
        add_line(line, "center")
    add_line(f"Sucursal: {sucursal}", "center", True, 9)
    add_line("-" * 32, "center")
    add_line("RECIBO OFICIAL DE CAJA", "center", True, 10)
    add_line("-" * 32, "center")
    add_line(f"No. Recibo: {recibo.numero}", "left", True, 9)
    add_line(f"Fecha: {fecha_str} {hora_str}".strip())
    add_line(f"Bodega: {bodega_name}")
    add_line(f"Tipo: {recibo.tipo}")
    add_line(f"Rubro: {rubro}")
    add_line(f"Motivo: {motivo}")
    add_line(f"Monto: {currency_label} {format_amount(monto_total)}", "left", True, 9)
    add_line(f"Equivalente C$: {format_amount(monto_cs)}")
    add_line("-" * 32, "center")
    add_line("Detalle", "left", True, 9)
    for part in wrap_text(descripcion, 32):
        add_line(part, "left", False, 9)

    add_line("-" * 32, "center")
    add_line("Mini asiento contable", "left", True, 9)
    if recibo.tipo == "EGRESO":
        add_line(f"Debe: {rubro_cuenta}", "left", False, 9)
        add_line(f"{currency_label} {format_amount(monto_total)}", "right", True, 9)
        add_line(f"Haber: {caja_cuenta}", "left", False, 9)
        add_line(f"{currency_label} {format_amount(monto_total)}", "right", True, 9)
    else:
        add_line(f"Debe: {caja_cuenta}", "left", False, 9)
        add_line(f"{currency_label} {format_amount(monto_total)}", "right", True, 9)
        add_line(f"Haber: {rubro_cuenta}", "left", False, 9)
        add_line(f"{currency_label} {format_amount(monto_total)}", "right", True, 9)

    add_line("-" * 32, "center")
    add_line("Realizado por: __________________", "left", False, 9)
    add_line("Recibido por: ___________________", "left", False, 9)
    add_line("Autorizado: _____________________", "left", False, 9)

    content_width = 70 * mm
    width = 80 * mm
    margin = (width - content_width) / 2
    top_margin = 6 * mm
    bottom_margin = 6 * mm
    logo_path = _resolve_logo_path(
        company_profile.get("logo_url", ""),
        prefer_pos=True,
        pos_logo_url=company_profile.get("pos_logo_url", ""),
    )
    logo_height = 42 * mm if logo_path.exists() else 0
    logo_spacing = 2 * mm if logo_height else 0

    def line_gap(size: int) -> float:
        return size + 4

    total_height = top_margin + bottom_margin + logo_height + logo_spacing
    total_height += sum(line_gap(size) for _, _, _, size in lines)
    total_height = max(total_height, 140 * mm)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(width, total_height))
    y = total_height - top_margin

    if logo_height:
        logo_width = 65 * mm
        pdf.drawImage(
            str(logo_path),
            (width - logo_width) / 2,
            y - logo_height,
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask="auto",
        )
        y -= logo_height + logo_spacing

    for text, align, bold, size in lines:
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        if align == "center":
            pdf.drawCentredString(width / 2, y, text)
        elif align == "right":
            pdf.drawRightString(width - margin, y, text)
        else:
            pdf.drawString(margin, y, text)
        y -= line_gap(size)

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer.read()


def _print_roc_ticket(
    recibo: ReciboCaja,
    printer_name: str,
    copies: int,
    profile: Optional[dict[str, str]] = None,
    sumatra_override: Optional[str] = None,
) -> None:
    sumatra_path = _get_sumatra_path(sumatra_override)
    if not sumatra_path:
        return
    pdf_bytes = _build_roc_ticket_pdf_bytes(recibo, profile)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(pdf_bytes)
        tmp_path = tmp_file.name
    try:
        for _ in range(max(copies, 1)):
            subprocess.run(
                [
                    str(sumatra_path),
                    "-print-to",
                    printer_name,
                    "-print-settings",
                    "noscale",
                    "-silent",
                    tmp_path,
                ],
                check=False,
            )
    finally:
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except OSError:
            pass


def _build_cierre_ticket_pdf_bytes(
    cierre: CierreCaja,
    tasa: Decimal,
    resumen: dict,
    total_bultos: Decimal,
    profile: Optional[dict[str, str]] = None,
) -> bytes:
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    def format_amount(value: Decimal) -> str:
        return f"{Decimal(str(value or 0)):,.2f}"

    def wrap_text(text: str, max_chars: int) -> list[str]:
        if not text:
            return [""]
        words = text.split()
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if len(candidate) > max_chars:
                if current:
                    lines.append(current)
                current = word
            else:
                current = candidate
        if current:
            lines.append(current)
        return lines or [text]

    branch = cierre.branch
    company_profile = profile or _default_company_profile_payload()
    identity = _company_identity(branch, company_profile)
    company_name = identity["company_name"]
    ruc = identity["ruc"]
    telefono = identity["telefono"]
    direccion = identity["direccion"]
    sucursal = identity["sucursal"]
    bodega_name = cierre.bodega.name if cierre.bodega else "-"

    fecha_str = cierre.fecha.strftime("%d/%m/%Y") if cierre.fecha else ""

    lines: list[tuple[str, str, bool, int]] = []

    def add_line(text: str, align: str = "left", bold: bool = False, size: int = 9):
        lines.append((text, align, bold, size))

    add_line(company_name.upper(), "center", True, 12)
    add_line(f"RUC: {ruc}", "center")
    add_line(f"Tel: {telefono}", "center")
    direccion_lines = wrap_text(direccion, 32)[:2]
    for line in direccion_lines:
        add_line(line, "center")
    add_line(f"Sucursal: {sucursal}", "center", True, 10)
    add_line("-" * 32, "center")
    add_line("CIERRE OFICIAL DE CAJA", "center", True, 12)
    add_line("-" * 32, "center")
    add_line(f"Fecha: {fecha_str}", "left", True, 10)
    add_line(f"Bodega: {bodega_name}", "left", False, 10)
    add_line("-" * 32, "center")

    add_line("Resumen arqueo (USD)", "left", True, 10)
    add_line(f"Ventas: $ {format_amount(resumen['ventas_usd'])}", "left", False, 10)
    add_line(f"Ingresos: + $ {format_amount(resumen['ingresos_usd'])}", "left", False, 10)
    add_line(f"Egresos: - $ {format_amount(resumen['egresos_usd'])}", "left", False, 10)
    add_line(f"Depositos: - $ {format_amount(resumen['depositos_usd'])}", "left", False, 10)
    add_line(f"Creditos: - $ {format_amount(resumen['creditos_usd'])}", "left", False, 10)
    add_line(f"Total esperado: $ {format_amount(resumen['total_calculado_usd'])}", "left", True, 10)
    add_line("")

    add_line("Efectivo contado", "left", True, 10)
    add_line(f"Total C$: {format_amount(cierre.total_efectivo_cs)}", "left", False, 10)
    add_line(f"Total USD: {format_amount(cierre.total_efectivo_usd)}", "left", False, 10)
    add_line(f"Total USD equiv: {format_amount(cierre.total_efectivo_usd_equiv)}", "left", True, 10)
    add_line("-" * 32, "center")

    add_line(f"Faltante/Sobrante: $ {format_amount(cierre.diferencia_usd)}", "left", True, 10)
    add_line(f"Total bultos vendidos: {format_amount(total_bultos)}", "left", False, 10)
    add_line("")

    try:
        detalle_cs = json.loads(cierre.detalle_cs or "{}")
    except Exception:
        detalle_cs = {}
    try:
        detalle_usd = json.loads(cierre.detalle_usd or "{}")
    except Exception:
        detalle_usd = {}

    add_line("Desglose USD (cantidades)", "left", True, 10)
    usd_items = []
    for denom, qty in detalle_usd.items():
        try:
            usd_items.append((Decimal(str(denom)), qty))
        except Exception:
            usd_items.append((Decimal("0"), qty))
    usd_items = sorted(usd_items, key=lambda item: item[0], reverse=True)
    subtotal_usd_breakdown = Decimal("0")
    for denom, qty in usd_items:
        qty_dec = Decimal(str(qty or 0))
        total = denom * qty_dec
        subtotal_usd_breakdown += total
        add_line(f"$ {denom} x {qty} = $ {format_amount(total)}", "left", False, 9)
    add_line("")
    add_line(f"Total desglose USD: {format_amount(subtotal_usd_breakdown)}", "left", True, 10)
    add_line("")

    add_line("Desglose C$ (cantidades)", "left", True, 10)
    cs_items = []
    for denom, qty in detalle_cs.items():
        try:
            cs_items.append((Decimal(str(denom)), qty))
        except Exception:
            cs_items.append((Decimal("0"), qty))
    cs_items = sorted(cs_items, key=lambda item: item[0], reverse=True)
    subtotal_cs_breakdown = Decimal("0")
    for denom, qty in cs_items:
        qty_dec = Decimal(str(qty or 0))
        total = denom * qty_dec
        subtotal_cs_breakdown += total
        add_line(f"C$ {denom} x {qty} = C$ {format_amount(total)}", "left", False, 9)
    add_line("")
    add_line(f"Total desglose C$: {format_amount(subtotal_cs_breakdown)}", "left", True, 10)
    add_line("")
    add_line(f"Total C$: {format_amount(cierre.total_efectivo_cs)}", "left", True, 10)
    add_line(f"Total USD: {format_amount(cierre.total_efectivo_usd)}", "left", True, 10)
    add_line(f"Total USD equiv: {format_amount(cierre.total_efectivo_usd_equiv)}", "left", True, 10)

    add_line("-" * 32, "center")
    add_line("Realizado por: __________________", "left", False, 9)
    add_line("Recibido por: ___________________", "left", False, 9)
    add_line("Autorizado: _____________________", "left", False, 9)

    content_width = 70 * mm
    width = 80 * mm
    margin = (width - content_width) / 2
    top_margin = 6 * mm
    bottom_margin = 6 * mm
    logo_path = _resolve_logo_path(
        company_profile.get("logo_url", ""),
        prefer_pos=True,
        pos_logo_url=company_profile.get("pos_logo_url", ""),
    )
    logo_height = 45 * mm if logo_path.exists() else 0
    logo_spacing = 2 * mm if logo_height else 0

    def line_gap(size: int) -> float:
        return size + 5

    total_height = top_margin + bottom_margin + logo_height + logo_spacing
    total_height += sum(line_gap(size) for _, _, _, size in lines)
    total_height = total_height

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(width, total_height))
    y = total_height - top_margin

    if logo_height:
        logo_width = 65 * mm
        pdf.drawImage(
            str(logo_path),
            (width - logo_width) / 2,
            y - logo_height,
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask="auto",
        )
        y -= logo_height + logo_spacing

    for text, align, bold, size in lines:
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        if "||" in text:
            left_text, right_text = text.split("||", 1)
            pdf.drawString(margin, y, left_text.strip())
            if right_text.strip():
                pdf.drawString(width / 2, y, right_text.strip())
        elif align == "center":
            pdf.drawCentredString(width / 2, y, text)
        elif align == "right":
            pdf.drawRightString(width - margin, y, text)
        else:
            pdf.drawString(margin, y, text)
        y -= line_gap(size)

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer.read()


def _print_cierre_ticket(
    cierre: CierreCaja,
    tasa: Decimal,
    resumen: dict,
    total_bultos: Decimal,
    printer_name: str,
    copies: int,
    profile: Optional[dict[str, str]] = None,
    sumatra_override: Optional[str] = None,
) -> None:
    sumatra_path = _get_sumatra_path(sumatra_override)
    if not sumatra_path:
        return
    pdf_bytes = _build_cierre_ticket_pdf_bytes(cierre, tasa, resumen, total_bultos, profile)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(pdf_bytes)
        tmp_path = tmp_file.name
    try:
        for _ in range(max(copies, 1)):
            subprocess.run(
                [
                    str(sumatra_path),
                    "-print-to",
                    printer_name,
                    "-print-settings",
                    "noscale",
                    "-silent",
                    tmp_path,
                ],
                check=False,
            )
    finally:
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except OSError:
            pass


def _generate_token(length: int = 6) -> str:
    import random
    import string

    return "".join(random.choice(string.digits) for _ in range(length))


def _normalize_company_key(value: str) -> str:
    return re.sub(r"[^a-z0-9_]+", "_", (value or "").strip().lower()).strip("_")


def _validate_database_url(database_url: str) -> Optional[str]:
    try:
        test_engine = create_engine(database_url, pool_pre_ping=True)
        with test_engine.connect():
            pass
        test_engine.dispose()
    except Exception as exc:
        return f"No se pudo conectar: {exc.__class__.__name__}"
    return None


def _allowed_sales_interface_codes() -> set[str]:
    return {item["code"] for item in SALES_INTERFACE_OPTIONS}


def _get_sales_interface_setting(db: Session) -> SalesInterfaceSetting:
    setting = db.query(SalesInterfaceSetting).order_by(SalesInterfaceSetting.id.asc()).first()
    if setting:
        if (setting.interface_code or "").strip().lower() not in _allowed_sales_interface_codes():
            setting.interface_code = "ropa"
            db.commit()
        return setting

    setting = SalesInterfaceSetting(interface_code="ropa")
    db.add(setting)
    db.commit()
    db.refresh(setting)
    return setting


def _get_company_profile_setting(db: Session) -> CompanyProfileSetting:
    row = db.query(CompanyProfileSetting).order_by(CompanyProfileSetting.id.asc()).first()
    if row:
        return row
    row = CompanyProfileSetting()
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def _send_reversion_email(
    subject: str,
    html_body: str,
    recipients: list[str],
    sender_email: Optional[str] = None,
    sender_name: Optional[str] = None,
    ) -> Optional[str]:
    if not recipients:
        return "No hay destinatarios activos"
    smtp_user = settings.SMTP_USER
    smtp_password = settings.SMTP_PASSWORD
    smtp_host = settings.SMTP_HOST
    smtp_port = settings.SMTP_PORT
    if not smtp_user or not smtp_password:
        env_path = Path(__file__).resolve().parents[2] / ".env"
        env_values = dotenv_values(env_path)
        def _env_get(key: str) -> str:
            return (env_values.get(key) or env_values.get(f"\ufeff{key}") or "").strip()
        smtp_user = smtp_user or _env_get("SMTP_USER")
        smtp_password = smtp_password or _env_get("SMTP_PASSWORD")
        smtp_host = smtp_host or _env_get("SMTP_HOST")
        smtp_port = smtp_port or int(_env_get("SMTP_PORT") or 0)
    if not smtp_user or not smtp_password:
        env_exists = env_path.exists()
        user_flag = "si" if smtp_user else "no"
        pass_flag = "si" if smtp_password else "no"
        return f"SMTP sin configurar (env={env_path}, existe={env_exists}, user={user_flag}, pass={pass_flag})"

    config = {
        "host": smtp_host or settings.SMTP_HOST,
        "port": smtp_port or settings.SMTP_PORT,
        "user": smtp_user,
        "password": smtp_password,
    }
    message = EmailMessage()
    message["Subject"] = subject
    from_email = sender_email or smtp_user
    if sender_name:
        message["From"] = f"{sender_name} <{from_email}>"
    else:
        message["From"] = from_email
    message["To"] = ", ".join(recipients)
    message.set_content("Se requiere un cliente de correo compatible con HTML.")
    message.add_alternative(html_body, subtype="html")

    try:
        with smtplib.SMTP(config["host"], config["port"]) as smtp:
            smtp.starttls()
            smtp.login(config["user"], config["password"])
            smtp.send_message(message)
    except Exception as exc:
        return f"Error SMTP: {exc.__class__.__name__}"
    return None

@router.get("/login")
def login_page(request: Request):
    return request.app.state.templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/login")
def login_action(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    remember: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    identifier = username.strip().lower()
    user = db.query(User).filter(
        (func.lower(User.email) == identifier)
        | (func.lower(User.full_name) == identifier)
    ).first()
    if not user or not verify_password(password, user.hashed_password):
        return request.app.state.templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Credenciales incorrectas",
                "version": settings.UI_VERSION,
            },
            status_code=401,
        )
    if not user.is_active:
        return request.app.state.templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Usuario inactivo",
                "version": settings.UI_VERSION,
            },
            status_code=403,
        )

    # Keep sessions alive by default for POS usage (1 year).
    expires = timedelta(days=365)
    token = create_access_token({"sub": user.email}, expires_delta=expires)
    response = RedirectResponse("/home", status_code=302)
    max_age = int(expires.total_seconds())
    expires_at = datetime.now(timezone.utc) + expires
    host = request.url.hostname or ""
    cookie_domain = None
    if host and host not in {"localhost", "127.0.0.1"} and host.count(".") >= 1:
        cookie_domain = f".{host.split('.', 1)[1]}"
    response.set_cookie(
        "access_token",
        token,
        httponly=True,
        samesite="lax",
        max_age=max_age,
        expires=expires_at,
        path="/",
        domain=cookie_domain,
    )
    return response


@router.get("/logout")
def logout(request: Request):
    response = RedirectResponse("/login", status_code=302)
    host = request.url.hostname or ""
    cookie_domain = None
    if host and host not in {"localhost", "127.0.0.1"} and host.count(".") >= 1:
        cookie_domain = f".{host.split('.', 1)[1]}"
    response.delete_cookie("access_token", path="/", domain=cookie_domain)
    return response



@router.get("/")
def root():
    return RedirectResponse("/home", status_code=302)


@router.get("/home")
def home(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    def _fmt_money(value: Optional[Decimal], symbol: str) -> str:
        amount = Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{symbol}{amount:,.2f}"

    branch, bodega = _resolve_branch_bodega(db, user)
    vendedor_id = _vendedor_id_for_user(db, user, bodega)
    sales_interface = _get_sales_interface_setting(db)
    sales_interface_code = (sales_interface.interface_code or "ropa").strip().lower()
    active_company_key = (get_active_company_key() or "").strip().lower()
    comestibles_theme_enabled = sales_interface_code == "comestibles" or active_company_key == "comestibles"
    home_preventas: list[dict] = []

    if vendedor_id:
        start_dt = datetime.combine(local_today() - timedelta(days=1), datetime.min.time())
        end_dt = datetime.combine(local_today() + timedelta(days=1), datetime.min.time())

        query = (
            db.query(Preventa, Cliente)
            .outerjoin(Cliente, Cliente.id == Preventa.cliente_id)
            .filter(
                Preventa.vendedor_id == vendedor_id,
                Preventa.fecha >= start_dt,
                Preventa.fecha < end_dt,
            )
        )
        if bodega:
            query = query.filter(Preventa.bodega_id == bodega.id)
        elif branch:
            query = query.filter(Preventa.branch_id == branch.id)
        query = query.order_by(Preventa.fecha.desc(), Preventa.id.desc()).limit(20)

        estado_bootstrap = {
            "PENDIENTE": "text-bg-danger",
            "REVISION": "text-bg-warning",
            "FACTURADA": "text-bg-success",
            "ANULADA": "text-bg-secondary",
        }

        for preventa, cliente in query.all():
            total_usd = Decimal(str(preventa.total_usd or 0))
            total_cs = Decimal(str(preventa.total_cs or 0))
            if total_usd > 0:
                monto_label = _fmt_money(total_usd, "$")
            else:
                monto_label = _fmt_money(total_cs, "C$")
            estado = (preventa.estado or "").upper()
            badge = _preventa_estado_badge(estado)
            home_preventas.append(
                {
                    "numero": preventa.numero,
                    "fecha": preventa.fecha,
                    "fecha_label": preventa.fecha.strftime("%Y-%m-%d %H:%M") if preventa.fecha else "-",
                    "cliente": cliente.nombre if cliente else "Consumidor final",
                    "monto_label": monto_label,
                    "estado_label": badge["label"],
                    "estado_class": estado_bootstrap.get(estado, "text-bg-light"),
                }
            )

    return request.app.state.templates.TemplateResponse(
        "home.html",
        {
            "request": request,
            "user": user,
            "version": settings.UI_VERSION,
            "home_preventas": home_preventas,
            "sales_interface_code": sales_interface_code,
            "comestibles_theme_enabled": comestibles_theme_enabled,
        },
    )


@router.get("/finance")
def finance_home(
    request: Request,
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.finance")
    return request.app.state.templates.TemplateResponse(
        "finance.html",
        {
            "request": request,
            "user": user,
            "version": settings.UI_VERSION,
        },
    )


def _accounting_period(entry_date: date) -> str:
    return entry_date.strftime("%Y-%m")


def _next_accounting_sequence(
    db: Session,
    tipo_id: int,
    branch_id: Optional[int],
    period: str,
) -> int:
    query = db.query(func.max(AccountingEntry.secuencia)).filter(
        AccountingEntry.tipo_id == tipo_id,
        AccountingEntry.periodo == period,
    )
    if branch_id is None:
        query = query.filter(AccountingEntry.branch_id.is_(None))
    else:
        query = query.filter(AccountingEntry.branch_id == branch_id)
    last_seq = query.scalar() or 0
    return int(last_seq) + 1


def _build_accounting_entry_number(voucher_type: AccountingVoucherType, period: str, seq: int) -> str:
    compact_period = period.replace("-", "")
    prefix = (voucher_type.prefijo or voucher_type.code or "CPB").strip().upper()
    return f"{prefix}-{compact_period}-{seq:05d}"


def _build_accounting_entry_pdf_bytes(entry: AccountingEntry) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 42
    y = height - 50

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(margin, y, "Comprobante contable")
    y -= 20
    pdf.setFont("Helvetica", 10)
    pdf.drawString(margin, y, f"Numero: {entry.numero}")
    y -= 14
    tipo_nombre = entry.tipo.nombre if entry.tipo else "-"
    pdf.drawString(margin, y, f"Tipo: {tipo_nombre}")
    y -= 14
    branch_name = entry.branch.name if entry.branch else "General"
    pdf.drawString(margin, y, f"Sucursal: {branch_name}")
    y -= 14
    pdf.drawString(margin, y, f"Fecha: {entry.fecha.strftime('%Y-%m-%d') if entry.fecha else '-'}")
    y -= 14
    pdf.drawString(margin, y, f"Estado: {entry.estado}")
    y -= 14
    if entry.referencia:
        pdf.drawString(margin, y, f"Referencia: {entry.referencia}")
        y -= 14
    pdf.drawString(margin, y, f"Concepto: {entry.descripcion or '-'}")

    y -= 24
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(margin, y, "Codigo")
    pdf.drawString(margin + 70, y, "Cuenta")
    pdf.drawString(margin + 280, y, "Detalle")
    pdf.drawRightString(width - 120, y, "Debe")
    pdf.drawRightString(width - margin, y, "Haber")
    y -= 12
    pdf.line(margin, y, width - margin, y)
    y -= 14

    pdf.setFont("Helvetica", 8)
    for line in entry.lines:
        if y < 70:
            pdf.showPage()
            y = height - 50
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(margin, y, "Codigo")
            pdf.drawString(margin + 70, y, "Cuenta")
            pdf.drawString(margin + 280, y, "Detalle")
            pdf.drawRightString(width - 120, y, "Debe")
            pdf.drawRightString(width - margin, y, "Haber")
            y -= 12
            pdf.line(margin, y, width - margin, y)
            y -= 14
            pdf.setFont("Helvetica", 8)
        code = line.cuenta.codigo if line.cuenta else "-"
        name = line.cuenta.nombre if line.cuenta else "-"
        detail = line.descripcion or ""
        pdf.drawString(margin, y, code[:14])
        pdf.drawString(margin + 70, y, name[:38])
        pdf.drawString(margin + 280, y, detail[:22])
        pdf.drawRightString(width - 120, y, f"{float(line.debe or 0):,.2f}")
        pdf.drawRightString(width - margin, y, f"{float(line.haber or 0):,.2f}")
        y -= 12

    y -= 8
    pdf.line(margin, y, width - margin, y)
    y -= 14
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(width - 120, y, f"{float(entry.total_debe or 0):,.2f}")
    pdf.drawRightString(width - margin, y, f"{float(entry.total_haber or 0):,.2f}")
    y -= 18
    pdf.drawString(margin, y, "Cuadre contable (debe = haber):")
    pdf.drawString(margin + 175, y, "SI" if to_decimal(entry.total_debe) == to_decimal(entry.total_haber) else "NO")

    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


def _suggest_counter_account_id(
    db: Session,
    selected_account: CuentaContable,
    active_accounts: list[CuentaContable],
) -> Optional[int]:
    if not selected_account:
        return None

    line_base = aliased(AccountingEntryLine)
    line_other = aliased(AccountingEntryLine)
    historical = (
        db.query(line_other.cuenta_id, func.count(line_other.id).label("hits"))
        .select_from(line_base)
        .join(AccountingEntry, AccountingEntry.id == line_base.entry_id)
        .join(line_other, line_other.entry_id == line_base.entry_id)
        .filter(line_base.cuenta_id == selected_account.id)
        .filter(line_other.cuenta_id != selected_account.id)
        .filter(AccountingEntry.estado != "ANULADO")
    )
    if (selected_account.naturaleza or "").upper() == "DEBE":
        historical = historical.filter(line_base.debe > 0, line_other.haber > 0)
    else:
        historical = historical.filter(line_base.haber > 0, line_other.debe > 0)
    best = (
        historical.group_by(line_other.cuenta_id)
        .order_by(func.count(line_other.id).desc(), line_other.cuenta_id.asc())
        .first()
    )
    if best and best[0]:
        return int(best[0])

    target_nature = "HABER" if (selected_account.naturaleza or "").upper() == "DEBE" else "DEBE"
    candidates = [
        acc
        for acc in active_accounts
        if acc.id != selected_account.id and (acc.naturaleza or "").upper() == target_nature
    ]
    if not candidates:
        return None

    haystack = f"{(selected_account.codigo or '').lower()} {(selected_account.nombre or '').lower()}"
    priority_terms = []
    if any(term in haystack for term in ["venta", "ingreso"]):
        priority_terms = ["caja", "banco", "cliente", "cobrar"]
    elif any(term in haystack for term in ["gasto", "costo"]):
        priority_terms = ["caja", "banco", "proveedor", "pagar"]
    elif any(term in haystack for term in ["inventario", "mercaderia"]):
        priority_terms = ["proveedor", "pagar", "caja", "banco"]
    elif any(term in haystack for term in ["caja", "banco"]):
        priority_terms = ["venta", "ingreso", "gasto", "costo", "inventario"]

    if priority_terms:
        for term in priority_terms:
            match = next(
                (
                    acc
                    for acc in candidates
                    if term in (acc.nombre or "").lower() or term in (acc.codigo or "").lower()
                ),
                None,
            )
            if match:
                return int(match.id)

    candidates.sort(key=lambda acc: ((acc.codigo or ""), (acc.nombre or "")))
    return int(candidates[0].id) if candidates else None


def _find_account_by_terms(
    active_accounts: list[CuentaContable],
    terms: list[str],
    naturaleza: Optional[str] = None,
) -> Optional[int]:
    target_nat = (naturaleza or "").upper().strip()
    for term in terms:
        term_lower = term.lower()
        for account in active_accounts:
            if target_nat and (account.naturaleza or "").upper() != target_nat:
                continue
            name = (account.nombre or "").lower()
            code = (account.codigo or "").lower()
            if term_lower in name or term_lower in code:
                return int(account.id)
    return None


def _terms_from_csv(value: Optional[str], fallback: list[str]) -> list[str]:
    raw = (value or "").strip()
    if not raw:
        return fallback
    items = [item.strip().lower() for item in raw.split(",") if item.strip()]
    return items or fallback


def _get_accounting_policy(db: Session) -> dict:
    row = db.query(AccountingPolicySetting).order_by(AccountingPolicySetting.id.asc()).first()
    defaults = {
        "strict_mode": True,
        "auto_entry_enabled": False,
        "ingreso_debe_terms": ["caja", "banco", "cliente", "cobrar"],
        "ingreso_haber_terms": ["venta", "ingreso"],
        "egreso_debe_terms": ["gasto", "costo", "compra", "inventario"],
        "egreso_haber_terms": ["caja", "banco", "proveedor", "pagar"],
    }
    if not row:
        return defaults
    return {
        "strict_mode": bool(row.strict_mode),
        "auto_entry_enabled": bool(row.auto_entry_enabled),
        "ingreso_debe_terms": _terms_from_csv(row.ingreso_debe_terms, defaults["ingreso_debe_terms"]),
        "ingreso_haber_terms": _terms_from_csv(row.ingreso_haber_terms, defaults["ingreso_haber_terms"]),
        "egreso_debe_terms": _terms_from_csv(row.egreso_debe_terms, defaults["egreso_debe_terms"]),
        "egreso_haber_terms": _terms_from_csv(row.egreso_haber_terms, defaults["egreso_haber_terms"]),
    }


def _build_voucher_template(
    voucher_type: AccountingVoucherType,
    active_accounts: list[CuentaContable],
    policy: dict,
) -> dict:
    code = (voucher_type.code or "").upper().strip()
    debit_terms = ["caja", "banco", "cliente", "cobrar"]
    credit_terms = ["venta", "ingreso"]
    if code == "EGRESO":
        debit_terms = ["gasto", "costo", "compra", "inventario"]
        credit_terms = ["caja", "banco", "proveedor", "pagar"]
    elif code == "AJUSTE":
        debit_terms = ["inventario", "costo", "gasto", "ajuste"]
        credit_terms = ["inventario", "costo", "ingreso", "ajuste"]
    elif code == "DIARIO":
        debit_terms = ["gasto", "inventario", "cliente", "caja"]
        credit_terms = ["ingreso", "venta", "proveedor", "banco"]
    if code == "INGRESO":
        debit_terms = list(policy.get("ingreso_debe_terms") or debit_terms)
        credit_terms = list(policy.get("ingreso_haber_terms") or credit_terms)
    elif code == "EGRESO":
        debit_terms = list(policy.get("egreso_debe_terms") or debit_terms)
        credit_terms = list(policy.get("egreso_haber_terms") or credit_terms)

    debit_id = _find_account_by_terms(active_accounts, debit_terms, "DEBE")
    credit_id = _find_account_by_terms(active_accounts, credit_terms, "HABER")
    if not debit_id:
        debit_id = _find_account_by_terms(active_accounts, debit_terms)
    if not credit_id:
        credit_id = _find_account_by_terms(active_accounts, credit_terms)
    return {
        "debit_account_id": debit_id,
        "credit_account_id": credit_id,
        "debit_hint": " / ".join(debit_terms),
        "credit_hint": " / ".join(credit_terms),
        "concept": f"Asiento sugerido para {voucher_type.nombre}",
    }


def _validate_accounting_entry_policy(
    voucher_type: AccountingVoucherType,
    line_payloads: list[dict],
    policy: dict,
) -> Optional[str]:
    if not policy.get("strict_mode", True):
        return None
    code = (voucher_type.code or "").upper().strip()
    if code not in {"INGRESO", "EGRESO"}:
        return None

    if not line_payloads:
        return "El comprobante no tiene lineas validas."

    has_cross_same_account = False
    by_account: dict[int, dict[str, Decimal]] = {}
    for item in line_payloads:
        account_id = int(item["account"].id)
        if account_id not in by_account:
            by_account[account_id] = {"debe": Decimal("0"), "haber": Decimal("0")}
        by_account[account_id]["debe"] += item["debe"]
        by_account[account_id]["haber"] += item["haber"]
    for totals in by_account.values():
        if totals["debe"] > 0 and totals["haber"] > 0:
            has_cross_same_account = True
            break
    if has_cross_same_account:
        return "Politica contable: una misma cuenta no debe ir al Debe y Haber en el mismo comprobante."

    debit_lines = [item for item in line_payloads if item["debe"] > 0]
    credit_lines = [item for item in line_payloads if item["haber"] > 0]
    if not debit_lines or not credit_lines:
        return "Politica contable: debe existir afectacion en ambos lados (Debe y Haber)."

    def _has_terms(lines: list[dict], terms: list[str]) -> bool:
        for item in lines:
            text = f"{(item['account'].codigo or '').lower()} {(item['account'].nombre or '').lower()}"
            if any(term in text for term in terms):
                return True
        return False

    if code == "INGRESO":
        ingreso_debe_terms = list(policy.get("ingreso_debe_terms") or ["caja", "banco", "cliente", "cobrar"])
        ingreso_haber_terms = list(policy.get("ingreso_haber_terms") or ["venta", "ingreso"])
        if not _has_terms(debit_lines, ingreso_debe_terms):
            return "Politica INGRESO: en Debe debe participar caja/banco/cliente por cobrar."
        if not _has_terms(credit_lines, ingreso_haber_terms):
            return "Politica INGRESO: en Haber debe participar una cuenta de ventas/ingresos."
    if code == "EGRESO":
        egreso_debe_terms = list(policy.get("egreso_debe_terms") or ["gasto", "costo", "compra", "inventario"])
        egreso_haber_terms = list(policy.get("egreso_haber_terms") or ["caja", "banco", "proveedor", "pagar"])
        if not _has_terms(debit_lines, egreso_debe_terms):
            return "Politica EGRESO: en Debe debe participar gasto/costo/compra/inventario."
        if not _has_terms(credit_lines, egreso_haber_terms):
            return "Politica EGRESO: en Haber debe participar caja/banco/proveedor por pagar."
    return None


def _find_voucher_type_for_code(db: Session, code: str) -> Optional[AccountingVoucherType]:
    normalized = (code or "").strip().upper()
    if not normalized:
        return None
    vt = (
        db.query(AccountingVoucherType)
        .filter(func.upper(AccountingVoucherType.code) == normalized, AccountingVoucherType.activo.is_(True))
        .first()
    )
    if vt:
        return vt
    return (
        db.query(AccountingVoucherType)
        .filter(func.upper(AccountingVoucherType.code) == "DIARIO", AccountingVoucherType.activo.is_(True))
        .first()
    )


def _build_auto_accounting_entry(
    db: Session,
    *,
    event_code: str,
    branch_id: Optional[int],
    entry_date: date,
    amount: Decimal,
    reference: str,
    description: str,
) -> Optional[AccountingEntry]:
    policy = _get_accounting_policy(db)
    if not policy.get("auto_entry_enabled", False):
        return None
    if amount <= 0:
        return None

    voucher_type = _find_voucher_type_for_code(db, "INGRESO" if event_code == "SALE" else "EGRESO")
    if not voucher_type:
        return None

    active_accounts = (
        db.query(CuentaContable)
        .filter(CuentaContable.activo.is_(True))
        .order_by(CuentaContable.codigo)
        .all()
    )
    template = _build_voucher_template(voucher_type, active_accounts, policy)
    debit_id = int(template["debit_account_id"]) if template.get("debit_account_id") else None
    credit_id = int(template["credit_account_id"]) if template.get("credit_account_id") else None

    if event_code == "INV_IN":
        debit_id = _find_account_by_terms(active_accounts, ["inventario"], "DEBE") or debit_id
        credit_id = (
            _find_account_by_terms(active_accounts, ["proveedor", "pagar", "caja", "banco"], "HABER")
            or credit_id
        )
    elif event_code == "INV_OUT":
        debit_id = _find_account_by_terms(active_accounts, ["costo", "gasto", "merma"], "DEBE") or debit_id
        credit_id = _find_account_by_terms(active_accounts, ["inventario"], "HABER") or credit_id

    if not debit_id or not credit_id or debit_id == credit_id:
        return None
    debit_account = db.query(CuentaContable).filter(CuentaContable.id == debit_id).first()
    credit_account = db.query(CuentaContable).filter(CuentaContable.id == credit_id).first()
    if not debit_account or not credit_account:
        return None

    period = _accounting_period(entry_date)
    seq = _next_accounting_sequence(db, voucher_type.id, branch_id, period)
    number = _build_accounting_entry_number(voucher_type, period, seq)
    line_amount = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    entry = AccountingEntry(
        tipo_id=voucher_type.id,
        branch_id=branch_id,
        fecha=entry_date,
        periodo=period,
        secuencia=seq,
        numero=number,
        referencia=reference[:160],
        descripcion=description[:400],
        estado="POSTEADO",
        total_debe=line_amount,
        total_haber=line_amount,
        creado_por="auto-system",
        lines=[
            AccountingEntryLine(
                cuenta_id=debit_account.id,
                descripcion=f"Auto {event_code} - Debe",
                debe=line_amount,
                haber=Decimal("0.00"),
            ),
            AccountingEntryLine(
                cuenta_id=credit_account.id,
                descripcion=f"Auto {event_code} - Haber",
                debe=Decimal("0.00"),
                haber=line_amount,
            ),
        ],
    )
    return entry


@router.get("/accounting")
def accounting_home(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting")
    return request.app.state.templates.TemplateResponse(
        "accounting.html",
        {
            "request": request,
            "user": user,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/accounting/financial-data")
def accounting_financial_data(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.financial_data")
    cuentas_total = db.query(func.count(CuentaContable.id)).scalar() or 0
    cuentas_activas = (
        db.query(func.count(CuentaContable.id)).filter(CuentaContable.activo.is_(True)).scalar() or 0
    )
    voucher_types = db.query(AccountingVoucherType).order_by(AccountingVoucherType.code).all()
    policy = _get_accounting_policy(db)
    return request.app.state.templates.TemplateResponse(
        "accounting_financial_data.html",
        {
            "request": request,
            "user": user,
            "voucher_types": voucher_types,
            "policy": policy,
            "cuentas_total": cuentas_total,
            "cuentas_activas": cuentas_activas,
            "version": settings.UI_VERSION,
            "error": request.query_params.get("error"),
            "success": request.query_params.get("success"),
        },
    )


@router.post("/accounting/voucher-types")
def accounting_voucher_type_create(
    request: Request,
    code: str = Form(...),
    nombre: str = Form(...),
    prefijo: str = Form(...),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.voucher_types")
    code = (code or "").strip().upper()
    nombre = (nombre or "").strip()
    prefijo = (prefijo or "").strip().upper()
    if not code or not nombre or not prefijo:
        return RedirectResponse("/accounting/financial-data?error=Datos+incompletos", status_code=303)
    exists = (
        db.query(AccountingVoucherType)
        .filter(func.lower(AccountingVoucherType.code) == code.lower())
        .first()
    )
    if exists:
        return RedirectResponse("/accounting/financial-data?error=Codigo+ya+existe", status_code=303)
    db.add(
        AccountingVoucherType(
            code=code,
            nombre=nombre,
            prefijo=prefijo[:10],
            activo=activo == "on",
        )
    )
    db.commit()
    return RedirectResponse("/accounting/financial-data?success=Tipo+creado", status_code=303)


@router.post("/accounting/voucher-types/{item_id}/update")
def accounting_voucher_type_update(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    prefijo: str = Form(...),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.voucher_types")
    row = db.query(AccountingVoucherType).filter(AccountingVoucherType.id == item_id).first()
    if not row:
        return RedirectResponse("/accounting/financial-data?error=Tipo+no+existe", status_code=303)
    row.nombre = (nombre or "").strip()
    row.prefijo = (prefijo or "").strip().upper()[:10]
    row.activo = activo == "on"
    db.commit()
    return RedirectResponse("/accounting/financial-data?success=Tipo+actualizado", status_code=303)


@router.post("/accounting/policies")
def accounting_policy_update(
    request: Request,
    strict_mode: Optional[str] = Form(None),
    auto_entry_enabled: Optional[str] = Form(None),
    ingreso_debe_terms: str = Form(""),
    ingreso_haber_terms: str = Form(""),
    egreso_debe_terms: str = Form(""),
    egreso_haber_terms: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.financial_data")
    row = db.query(AccountingPolicySetting).order_by(AccountingPolicySetting.id.asc()).first()
    if not row:
        row = AccountingPolicySetting()
        db.add(row)
    row.strict_mode = strict_mode == "on"
    row.auto_entry_enabled = auto_entry_enabled == "on"
    row.ingreso_debe_terms = (ingreso_debe_terms or "").strip()
    row.ingreso_haber_terms = (ingreso_haber_terms or "").strip()
    row.egreso_debe_terms = (egreso_debe_terms or "").strip()
    row.egreso_haber_terms = (egreso_haber_terms or "").strip()
    row.updated_by = user.email
    db.commit()
    return RedirectResponse("/accounting/financial-data?success=Politicas+contables+actualizadas", status_code=303)


@router.get("/accounting/entries")
def accounting_entries_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.entries")
    branch_ids = _user_scoped_branch_ids(db, user)
    entries_query = (
        db.query(AccountingEntry)
        .outerjoin(Branch, Branch.id == AccountingEntry.branch_id)
        .filter(or_(AccountingEntry.branch_id.is_(None), AccountingEntry.branch_id.in_(branch_ids)))
    )
    start = request.query_params.get("start")
    end = request.query_params.get("end")
    if start:
        try:
            start_date = datetime.strptime(start, "%Y-%m-%d").date()
            entries_query = entries_query.filter(AccountingEntry.fecha >= start_date)
        except ValueError:
            pass
    if end:
        try:
            end_date = datetime.strptime(end, "%Y-%m-%d").date()
            entries_query = entries_query.filter(AccountingEntry.fecha <= end_date)
        except ValueError:
            pass
    entries = (
        entries_query.order_by(AccountingEntry.fecha.desc(), AccountingEntry.id.desc())
        .limit(120)
        .all()
    )
    voucher_types = (
        db.query(AccountingVoucherType)
        .filter(AccountingVoucherType.activo.is_(True))
        .order_by(AccountingVoucherType.code)
        .all()
    )
    policy = _get_accounting_policy(db)
    cuentas = (
        db.query(CuentaContable)
        .filter(CuentaContable.activo.is_(True))
        .order_by(CuentaContable.codigo)
        .all()
    )
    counter_suggestions: dict[int, int] = {}
    for account in cuentas:
        suggested = _suggest_counter_account_id(db, account, cuentas)
        if suggested:
            counter_suggestions[int(account.id)] = int(suggested)
    voucher_templates: dict[int, dict] = {}
    for vt in voucher_types:
        voucher_templates[int(vt.id)] = _build_voucher_template(vt, cuentas, policy)
    branches = (
        db.query(Branch)
        .filter(Branch.id.in_(branch_ids))
        .order_by(Branch.name)
        .all()
    )
    return request.app.state.templates.TemplateResponse(
        "accounting_entries.html",
        {
            "request": request,
            "user": user,
            "entries": entries,
            "voucher_types": voucher_types,
            "cuentas": cuentas,
            "cuentas_json": [
                {
                    "id": int(c.id),
                    "codigo": c.codigo,
                    "nombre": c.nombre,
                    "naturaleza": c.naturaleza,
                    "tipo": c.tipo,
                }
                for c in cuentas
            ],
            "counter_suggestions": counter_suggestions,
            "voucher_templates": voucher_templates,
            "policy": policy,
            "branches": branches,
            "version": settings.UI_VERSION,
            "error": request.query_params.get("error"),
            "success": request.query_params.get("success"),
            "today": local_today().isoformat(),
        },
    )


@router.get("/accounting/accounts/search")
def accounting_accounts_search(
    request: Request,
    q: str = "",
    side: str = "",
    limit: int = 20,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.entries")
    limit = max(5, min(int(limit or 20), 50))
    query = (q or "").strip().lower()
    side_norm = (side or "").strip().lower()

    rows_q = db.query(CuentaContable).filter(CuentaContable.activo.is_(True))
    if side_norm == "debe":
        rows_q = rows_q.filter(func.upper(CuentaContable.naturaleza) == "DEBE")
    elif side_norm == "haber":
        rows_q = rows_q.filter(func.upper(CuentaContable.naturaleza) == "HABER")

    if query:
        like = f"%{query}%"
        rows_q = rows_q.filter(
            or_(
                func.lower(CuentaContable.codigo).like(like),
                func.lower(CuentaContable.nombre).like(like),
                func.lower(CuentaContable.tipo).like(like),
                func.lower(CuentaContable.naturaleza).like(like),
            )
        )

    rows = rows_q.order_by(CuentaContable.codigo).limit(limit).all()
    return JSONResponse(
        {
            "items": [
                {
                    "id": int(r.id),
                    "codigo": r.codigo,
                    "nombre": r.nombre,
                    "naturaleza": r.naturaleza,
                    "tipo": r.tipo,
                }
                for r in rows
            ]
        }
    )


@router.post("/accounting/entries")
def accounting_entries_create(
    request: Request,
    fecha: str = Form(...),
    tipo_id: int = Form(...),
    branch_id: Optional[str] = Form(None),
    referencia: str = Form(""),
    descripcion: str = Form(...),
    line_cuenta_id: list[str] = Form(...),
    line_descripcion: list[str] = Form([]),
    line_debe: list[str] = Form([]),
    line_haber: list[str] = Form([]),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.entries")
    try:
        entry_date = datetime.strptime(fecha, "%Y-%m-%d").date()
    except ValueError:
        return RedirectResponse("/accounting/entries?error=Fecha+invalida", status_code=303)

    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    selected_branch_id: Optional[int] = None
    if branch_id and branch_id.strip():
        try:
            selected_branch_id = int(branch_id)
        except ValueError:
            return RedirectResponse("/accounting/entries?error=Sucursal+invalida", status_code=303)
    if selected_branch_id and selected_branch_id not in scoped_branch_ids:
        return RedirectResponse("/accounting/entries?error=Sucursal+no+permitida", status_code=303)

    voucher_type = (
        db.query(AccountingVoucherType)
        .filter(AccountingVoucherType.id == tipo_id, AccountingVoucherType.activo.is_(True))
        .first()
    )
    if not voucher_type:
        return RedirectResponse("/accounting/entries?error=Tipo+de+comprobante+invalido", status_code=303)

    if len(line_cuenta_id) < 2:
        return RedirectResponse("/accounting/entries?error=Debes+registrar+al+menos+2+lineas", status_code=303)

    lines: list[AccountingEntryLine] = []
    line_payloads: list[dict] = []
    total_debe = Decimal("0")
    total_haber = Decimal("0")
    for index, cuenta_raw in enumerate(line_cuenta_id):
        if not cuenta_raw:
            return RedirectResponse("/accounting/entries?error=Cuenta+contable+requerida", status_code=303)
        try:
            cuenta_id = int(cuenta_raw)
        except ValueError:
            return RedirectResponse("/accounting/entries?error=Cuenta+contable+invalida", status_code=303)
        cuenta = db.query(CuentaContable).filter(CuentaContable.id == cuenta_id).first()
        if not cuenta:
            return RedirectResponse("/accounting/entries?error=Cuenta+contable+invalida", status_code=303)
        debe_raw = (line_debe[index] if index < len(line_debe) else "0").replace(",", ".")
        haber_raw = (line_haber[index] if index < len(line_haber) else "0").replace(",", ".")
        try:
            debe_value = to_decimal(debe_raw).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            haber_value = to_decimal(haber_raw).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
        except (InvalidOperation, ValueError):
            return RedirectResponse("/accounting/entries?error=Monto+invalido+en+lineas", status_code=303)
        if debe_value < 0 or haber_value < 0:
            return RedirectResponse("/accounting/entries?error=No+se+permiten+valores+negativos", status_code=303)
        if debe_value == 0 and haber_value == 0:
            continue
        if debe_value > 0 and haber_value > 0:
            return RedirectResponse("/accounting/entries?error=Cada+linea+solo+puede+tener+Debe+o+Haber", status_code=303)
        total_debe += debe_value
        total_haber += haber_value
        line_detail = line_descripcion[index].strip() if index < len(line_descripcion) else ""
        lines.append(
            AccountingEntryLine(
                cuenta_id=cuenta_id,
                descripcion=line_detail[:200] if line_detail else None,
                debe=debe_value,
                haber=haber_value,
            )
        )
        line_payloads.append(
            {
                "account": cuenta,
                "debe": debe_value,
                "haber": haber_value,
            }
        )

    if len(lines) < 2:
        return RedirectResponse("/accounting/entries?error=Debes+registrar+lineas+con+monto", status_code=303)
    if total_debe <= 0 or total_haber <= 0:
        return RedirectResponse("/accounting/entries?error=Debe+y+Haber+deben+ser+mayores+a+0", status_code=303)
    if total_debe != total_haber:
        return RedirectResponse("/accounting/entries?error=El+comprobante+no+cuadra+(Debe+!=+Haber)", status_code=303)
    policy = _get_accounting_policy(db)
    policy_error = _validate_accounting_entry_policy(voucher_type, line_payloads, policy)
    if policy_error:
        return RedirectResponse(f"/accounting/entries?error={quote_plus(policy_error)}", status_code=303)

    period = _accounting_period(entry_date)
    seq = _next_accounting_sequence(db, tipo_id, selected_branch_id, period)
    number = _build_accounting_entry_number(voucher_type, period, seq)
    entry = AccountingEntry(
        tipo_id=tipo_id,
        branch_id=selected_branch_id,
        fecha=entry_date,
        periodo=period,
        secuencia=seq,
        numero=number,
        referencia=(referencia or "").strip()[:160] or None,
        descripcion=(descripcion or "").strip(),
        estado="POSTEADO",
        total_debe=total_debe,
        total_haber=total_haber,
        creado_por=user.email,
        lines=lines,
    )
    db.add(entry)
    db.commit()
    return RedirectResponse("/accounting/entries?success=Comprobante+registrado", status_code=303)


@router.post("/accounting/entries/{entry_id}/annul")
def accounting_entry_annul(
    request: Request,
    entry_id: int,
    motivo: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.entries")
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    entry = (
        db.query(AccountingEntry)
        .filter(AccountingEntry.id == entry_id)
        .filter(or_(AccountingEntry.branch_id.is_(None), AccountingEntry.branch_id.in_(scoped_branch_ids)))
        .first()
    )
    if not entry:
        return RedirectResponse("/accounting/entries?error=Comprobante+no+encontrado", status_code=303)
    if entry.estado == "ANULADO":
        return RedirectResponse("/accounting/entries?error=El+comprobante+ya+esta+anulado", status_code=303)
    entry.estado = "ANULADO"
    entry.anulado_motivo = (motivo or "").strip()[:260] or "Anulacion manual"
    entry.anulado_por = user.email
    entry.anulado_at = local_now_naive()
    db.commit()
    return RedirectResponse("/accounting/entries?success=Comprobante+anulado", status_code=303)


@router.get("/accounting/entries/{entry_id}/pdf")
def accounting_entry_pdf(
    request: Request,
    entry_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.accounting.entries")
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    entry = (
        db.query(AccountingEntry)
        .filter(AccountingEntry.id == entry_id)
        .filter(or_(AccountingEntry.branch_id.is_(None), AccountingEntry.branch_id.in_(scoped_branch_ids)))
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")
    pdf_bytes = _build_accounting_entry_pdf_bytes(entry)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={entry.numero}.pdf"},
    )


@router.get("/inventory")
def inventory_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory")
    edit_id = request.query_params.get("edit")
    show_inactive = request.query_params.get("show_inactive") in {"1", "true", "True"}
    edit_product = None
    if edit_id and edit_id.isdigit():
        edit_product = db.query(Producto).filter(Producto.id == int(edit_id)).first()
    productos_query = db.query(Producto).order_by(Producto.descripcion)
    if not show_inactive:
        productos_query = productos_query.filter(Producto.activo.is_(True))
    productos = productos_query.all()
    bodegas = _scoped_bodegas_query(db).order_by(Bodega.id).all()
    lineas = db.query(Linea).order_by(Linea.linea).all()
    segmentos = db.query(Segmento).order_by(Segmento.segmento).all()
    marcas = db.query(Marca).filter(Marca.activo.is_(True)).order_by(Marca.nombre).all()
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    product_ids = [p.id for p in productos]
    bodega_ids = [b.id for b in bodegas]
    balances = _balances_by_bodega(db, bodega_ids, product_ids)
    bodega_central = next((b for b in bodegas if (b.code or "").lower() == "central"), None)
    if not bodega_central:
        bodega_central = next((b for b in bodegas if "central" in (b.name or "").lower()), None)
    bodega_esteli = next((b for b in bodegas if (b.code or "").lower() == "esteli"), None)
    if not bodega_esteli:
        bodega_esteli = next((b for b in bodegas if "esteli" in (b.name or "").lower()), None)

    def _sum_for_bodega(bodega: Optional[Bodega]) -> tuple[Decimal, int]:
        if not bodega:
            return Decimal("0"), 0
        total_qty = Decimal("0")
        count_items = 0
        for producto in productos:
            qty = balances.get((producto.id, bodega.id), Decimal("0"))
            total_qty += qty
            if qty > 0:
                count_items += 1
        return total_qty, count_items

    central_qty, central_items = _sum_for_bodega(bodega_central)
    esteli_qty, esteli_items = _sum_for_bodega(bodega_esteli)
    global_qty = central_qty + esteli_qty
    global_items = len({p.id for p in productos if (balances.get((p.id, bodega_central.id), Decimal("0")) if bodega_central else Decimal("0")) > 0 or (balances.get((p.id, bodega_esteli.id), Decimal("0")) if bodega_esteli else Decimal("0")) > 0})
    inventory_cs_only = _inventory_cs_only_mode(db)
    auto_price_margin_enabled, auto_price_margin_pct = _price_margin_mode(db)
    return request.app.state.templates.TemplateResponse(
        "inventory.html",
        {
            "request": request,
            "user": user,
            "productos": productos,
            "bodegas": bodegas,
            "central_qty": float(central_qty),
            "esteli_qty": float(esteli_qty),
            "global_qty": float(global_qty),
            "central_items": central_items,
            "esteli_items": esteli_items,
            "global_items": global_items,
            "lineas": lineas,
            "segmentos": segmentos,
            "edit_product": edit_product,
            "rate_today": rate_today,
            "error": error,
            "success": success,
            "show_inactive": show_inactive,
            "inventory_cs_only": inventory_cs_only,
            "auto_price_margin_enabled": auto_price_margin_enabled,
            "auto_price_margin_pct": auto_price_margin_pct,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/inventory/caliente")
def inventory_caliente(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.inventory.caliente")
    q = (request.query_params.get("q") or "").strip()
    scope_param = (request.query_params.get("scope") or "central").strip().lower()
    scope = scope_param
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    rate_value = Decimal(str(rate_today.rate)) if rate_today and rate_today.rate else Decimal("0")

    productos_query = (
        db.query(Producto)
        .outerjoin(SaldoProducto)
        .filter(Producto.activo.is_(True))
    )
    if q:
        q_like = f"%{q.lower()}%"
        productos_query = productos_query.filter(
            or_(
                func.lower(Producto.cod_producto).like(q_like),
                func.lower(Producto.descripcion).like(q_like),
            )
        )
    productos = productos_query.order_by(Producto.descripcion).all()

    branches = _scoped_branches_query(db).all()
    branch_map = {b.code.lower(): b for b in branches if b.code}
    central_branch = branch_map.get("central")
    esteli_branch = branch_map.get("esteli")
    bodegas_query = db.query(Bodega).filter(Bodega.activo.is_(True))
    if branches:
        bodegas_query = bodegas_query.filter(Bodega.branch_id.in_([b.id for b in branches]))
    bodegas = bodegas_query.all()
    bodega_map = {b.branch_id: b for b in bodegas}

    branch, user_bodega = _resolve_branch_bodega(db, user)
    user_branches = list(user.branches or [])
    allowed_scopes: list[str] = []
    if user_branches:
        allowed_scopes = [b.code.lower() for b in user_branches if b.code and b.code.lower() in _allowed_branch_codes(db)]
    if not allowed_scopes and branch and branch.code:
        if branch.code.lower() in _allowed_branch_codes(db):
            allowed_scopes = [branch.code.lower()]
    if not allowed_scopes:
        allowed_scopes = ["central"]

    if len(allowed_scopes) == 1:
        scope = allowed_scopes[0]
    else:
        allowed_scope_values = set(allowed_scopes + ["ambas"])
        if scope not in allowed_scope_values:
            scope = "central"
        if scope != "ambas" and scope not in allowed_scopes:
            scope = allowed_scopes[0]
        if scope == "ambas" and not all(code in allowed_scopes for code in _allowed_branch_codes(db)):
            scope = allowed_scopes[0]

    def _balances_by_bodega(bodega_ids: list[int], product_ids: list[int]) -> dict[tuple[int, int], Decimal]:
        if not bodega_ids or not product_ids:
            return {}
        ingreso_rows = (
            db.query(IngresoItem.producto_id, IngresoInventario.bodega_id, func.sum(IngresoItem.cantidad))
            .join(IngresoInventario, IngresoInventario.id == IngresoItem.ingreso_id)
            .filter(IngresoInventario.bodega_id.in_(bodega_ids))
            .filter(IngresoItem.producto_id.in_(product_ids))
            .group_by(IngresoItem.producto_id, IngresoInventario.bodega_id)
            .all()
        )
        egreso_rows = (
            db.query(EgresoItem.producto_id, EgresoInventario.bodega_id, func.sum(EgresoItem.cantidad))
            .join(EgresoInventario, EgresoInventario.id == EgresoItem.egreso_id)
            .filter(EgresoInventario.bodega_id.in_(bodega_ids))
            .filter(EgresoItem.producto_id.in_(product_ids))
            .group_by(EgresoItem.producto_id, EgresoInventario.bodega_id)
            .all()
        )
        venta_rows = (
            db.query(VentaItem.producto_id, VentaFactura.bodega_id, func.sum(VentaItem.cantidad))
            .join(VentaFactura, VentaFactura.id == VentaItem.factura_id)
            .filter(VentaFactura.bodega_id.in_(bodega_ids))
            .filter(VentaItem.producto_id.in_(product_ids))
            .filter(VentaFactura.estado != "ANULADA")
            .group_by(VentaItem.producto_id, VentaFactura.bodega_id)
            .all()
        )
        balances: dict[tuple[int, int], Decimal] = {}
        for producto_id, bodega_id, qty in ingreso_rows:
            balances[(producto_id, bodega_id)] = Decimal(str(qty or 0))
        for producto_id, bodega_id, qty in egreso_rows:
            balances[(producto_id, bodega_id)] = balances.get((producto_id, bodega_id), Decimal("0")) - Decimal(str(qty or 0))
        for producto_id, bodega_id, qty in venta_rows:
            balances[(producto_id, bodega_id)] = balances.get((producto_id, bodega_id), Decimal("0")) - Decimal(str(qty or 0))
        return balances

    product_ids = [p.id for p in productos]
    if scope == "ambas":
        bodega_ids = [b.id for b in bodegas]
    else:
        selected_branch = central_branch if scope == "central" else esteli_branch if scope == "esteli" else branch
        selected_bodega = None
        if selected_branch:
            selected_bodega = bodega_map.get(selected_branch.id)
        if scope not in {"central", "esteli"} and user_bodega:
            selected_bodega = user_bodega
        bodega_ids = [selected_bodega.id] if selected_bodega else []
    balances = _balances_by_bodega(bodega_ids, product_ids)

    productos_view = []
    for producto in productos:
        price_usd = Decimal(str(producto.precio_venta1_usd or 0))
        price_cs = Decimal(str(producto.precio_venta1 or 0))
        if price_usd == 0 and price_cs and rate_value:
            price_usd = (price_cs / rate_value).quantize(Decimal("0.01"))
        if price_cs == 0 and price_usd and rate_value:
            price_cs = (price_usd * rate_value).quantize(Decimal("0.01"))
        item = {
            "id": producto.id,
            "codigo": producto.cod_producto,
            "descripcion": producto.descripcion,
            "precio_usd": float(price_usd or 0),
            "precio_cs": float(price_cs or 0),
        }
        if scope == "ambas":
            rows = []
            if central_branch:
                central_bodega = bodega_map.get(central_branch.id)
                qty = balances.get((producto.id, central_bodega.id), Decimal("0")) if central_bodega else Decimal("0")
                rows.append({"label": central_branch.name, "existencia": float(qty or 0)})
            if esteli_branch:
                esteli_bodega = bodega_map.get(esteli_branch.id)
                qty = balances.get((producto.id, esteli_bodega.id), Decimal("0")) if esteli_bodega else Decimal("0")
                rows.append({"label": esteli_branch.name, "existencia": float(qty or 0)})
            total_qty = sum(Decimal(str(row["existencia"])) for row in rows)
            item["existencias"] = rows
            item["existencia_total"] = float(total_qty or 0)
        else:
            selected_branch = central_branch if scope == "central" else esteli_branch if scope == "esteli" else branch
            selected_bodega = None
            if selected_branch:
                selected_bodega = bodega_map.get(selected_branch.id)
            if scope not in {"central", "esteli"} and user_bodega:
                selected_bodega = user_bodega
            qty = balances.get((producto.id, selected_bodega.id), Decimal("0")) if selected_bodega else Decimal("0")
            item["existencia"] = float(qty or 0)
            item["scope_label"] = selected_branch.name if selected_branch else "Bodega"
        productos_view.append(item)

    template_name = "inventory_caliente_grid.html" if request.headers.get("HX-Request") else "inventory_caliente.html"
    return request.app.state.templates.TemplateResponse(
        template_name,
        {
            "request": request,
            "user": user,
            "q": q,
            "productos": productos_view,
            "rate_today": rate_today,
            "scope": scope,
            "scope_param": scope_param,
            "allowed_scopes": allowed_scopes,
            "current_branch": branch,
            "central_branch": central_branch,
            "esteli_branch": esteli_branch,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/inventory/ingresos")
def inventory_ingresos_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.ingresos")
    def _parse_date(value: Optional[str]) -> Optional[date]:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    start_date = _parse_date(request.query_params.get("start_date"))
    end_date = _parse_date(request.query_params.get("end_date"))
    if not start_date and not end_date:
        end_date = local_today()
        start_date = end_date - timedelta(days=30)
    ingresos_query = db.query(IngresoInventario)
    if start_date:
        ingresos_query = ingresos_query.filter(IngresoInventario.fecha >= start_date)
    if end_date:
        ingresos_query = ingresos_query.filter(IngresoInventario.fecha <= end_date)
    ingresos = (
        ingresos_query.order_by(IngresoInventario.fecha.desc(), IngresoInventario.id.desc())
        .all()
    )
    # "Traslado entre bodegas" se genera automaticamente desde egresos.
    tipos = (
        db.query(IngresoTipo)
        .filter(func.lower(IngresoTipo.nombre) != "traslado entre bodegas")
        .order_by(IngresoTipo.nombre)
        .all()
    )
    bodegas = _scoped_bodegas_query(db).order_by(Bodega.name).all()
    proveedores = db.query(Proveedor).order_by(Proveedor.nombre).all()
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .order_by(Producto.descripcion)
        .all()
    )
    product_ids = [p.id for p in productos]
    bodega_ids = [b.id for b in bodegas]
    balances = _balances_by_bodega(db, bodega_ids, product_ids)
    saldos_por_bodega: dict[int, dict[int, float]] = {}
    for producto in productos:
        per_bodega: dict[int, float] = {}
        for bodega in bodegas:
            qty = balances.get((producto.id, bodega.id), Decimal("0"))
            per_bodega[bodega.id] = float(qty or 0)
        saldos_por_bodega[producto.id] = per_bodega
    lineas = db.query(Linea).order_by(Linea.linea).all()
    segmentos = db.query(Segmento).order_by(Segmento.segmento).all()
    marcas = db.query(Marca).filter(Marca.activo.is_(True)).order_by(Marca.nombre).all()
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    print_id = request.query_params.get("print_id")
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    inventory_cs_only = _inventory_cs_only_mode(db)
    auto_price_margin_enabled, auto_price_margin_pct = _price_margin_mode(db)
    return request.app.state.templates.TemplateResponse(
        "inventory_ingresos.html",
        {
            "request": request,
            "user": user,
            "ingresos": ingresos,
            "tipos": tipos,
            "bodegas": bodegas,
            "proveedores": proveedores,
            "productos": productos,
            "saldos_por_bodega": saldos_por_bodega,
            "lineas": lineas,
            "segmentos": segmentos,
            "marcas": marcas,
            "rate_today": rate_today,
            "error": error,
            "start_date": start_date.isoformat() if start_date else "",
            "end_date": end_date.isoformat() if end_date else "",
            "success": success,
            "print_id": print_id,
            "inventory_cs_only": inventory_cs_only,
            "auto_price_margin_enabled": auto_price_margin_enabled,
            "auto_price_margin_pct": auto_price_margin_pct,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/inventory/egresos")
def inventory_egresos_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.egresos")
    def _parse_date(value: Optional[str]) -> Optional[date]:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    start_date = _parse_date(request.query_params.get("start_date"))
    end_date = _parse_date(request.query_params.get("end_date"))
    if not start_date and not end_date:
        end_date = local_today()
        start_date = end_date - timedelta(days=30)
    egresos_query = db.query(EgresoInventario)
    if start_date:
        egresos_query = egresos_query.filter(EgresoInventario.fecha >= start_date)
    if end_date:
        egresos_query = egresos_query.filter(EgresoInventario.fecha <= end_date)
    egresos = (
        egresos_query.order_by(EgresoInventario.fecha.desc(), EgresoInventario.id.desc())
        .all()
    )
    tipos = db.query(EgresoTipo).order_by(EgresoTipo.nombre).all()
    bodegas = _scoped_bodegas_query(db).order_by(Bodega.name).all()
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .order_by(Producto.descripcion)
        .all()
    )
    product_ids = [p.id for p in productos]
    bodega_ids = [b.id for b in bodegas]
    balances = _balances_by_bodega(db, bodega_ids, product_ids)
    saldos_por_bodega: dict[int, dict[int, float]] = {}
    for producto in productos:
        per_bodega: dict[int, float] = {}
        for bodega in bodegas:
            qty = balances.get((producto.id, bodega.id), Decimal("0"))
            per_bodega[bodega.id] = float(qty or 0)
        saldos_por_bodega[producto.id] = per_bodega
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    print_id = request.query_params.get("print_id")
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    inventory_cs_only = _inventory_cs_only_mode(db)
    return request.app.state.templates.TemplateResponse(
        "inventory_egresos.html",
        {
            "request": request,
            "user": user,
            "egresos": egresos,
            "tipos": tipos,
            "bodegas": bodegas,
            "productos": productos,
            "saldos_por_bodega": saldos_por_bodega,
            "rate_today": rate_today,
            "error": error,
            "start_date": start_date.isoformat() if start_date else "",
            "end_date": end_date.isoformat() if end_date else "",
            "success": success,
            "print_id": print_id,
            "inventory_cs_only": inventory_cs_only,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/sales")
def sales_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales")
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .order_by(Producto.descripcion)
        .all()
    )
    clientes_preview = [
        {"id": c.id, "nombre": c.nombre}
        for c in db.query(Cliente).order_by(Cliente.nombre).limit(1000).all()
    ]
    formas_pago = db.query(FormaPago).order_by(FormaPago.nombre).all()
    bancos = db.query(Banco).order_by(Banco.nombre).all()
    cuentas = db.query(CuentaBancaria).order_by(CuentaBancaria.banco_id).all()
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    print_id = request.query_params.get("print_id")
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    branch, bodega = _resolve_branch_bodega(db, user)
    vendedores = _vendedores_for_bodega(db, bodega)
    next_invoice = None
    if branch and bodega:
        last_factura = (
            db.query(VentaFactura)
            .filter(VentaFactura.bodega_id == bodega.id)
            .order_by(VentaFactura.secuencia.desc())
            .first()
        )
        next_seq = (last_factura.secuencia if last_factura else 0) + 1
        branch_code = (branch.code or "").lower()
        prefix = "C" if branch_code == "central" else "E" if branch_code == "esteli" else branch_code[:1].upper()
        width = 6
        next_invoice = f"{prefix}-{next_seq:0{width}d}"
    pos_print = (
        db.query(PosPrintSetting)
        .filter(PosPrintSetting.branch_id == branch.id)
        .first()
        if branch
        else None
    )
    initial_preventa = None
    preventa_id_raw = (request.query_params.get("preventa_id") or "").strip()
    if preventa_id_raw and preventa_id_raw.isdigit():
        preventa = (
            db.query(Preventa)
            .filter(Preventa.id == int(preventa_id_raw))
            .first()
        )
        if preventa and preventa.estado in {"PENDIENTE", "REVISION"}:
            item_rows = (
                db.query(PreventaItem, Producto)
                .join(Producto, Producto.id == PreventaItem.producto_id)
                .filter(PreventaItem.preventa_id == preventa.id)
                .all()
            )
            product_ids = [producto.id for _, producto in item_rows]
            balances = _balances_by_bodega(db, [preventa.bodega_id], product_ids) if product_ids else {}
            required_by_product = _preventa_required_qty_map(item_rows)
            items = []
            for row, producto in item_rows:
                existencia = float(balances.get((producto.id, preventa.bodega_id), Decimal("0")) or 0)
                qty = float(row.cantidad or 0)
                required_qty = float(required_by_product.get(int(producto.id), Decimal("0")) or 0)
                if existencia < required_qty:
                    return RedirectResponse(
                        f"/sales/preventas?error=Sin+saldo+actual+para+{producto.cod_producto}+en+preventa",
                        status_code=303,
                    )
                items.append(
                    {
                        "product_id": producto.id,
                        "cod_producto": producto.cod_producto,
                        "descripcion": producto.descripcion,
                        "cantidad": qty,
                        "existencia": existencia,
                        "precio_usd": float(row.precio_unitario_usd or 0),
                        "precio_cs": float(row.precio_unitario_cs or 0),
                        "combo_role": row.combo_role,
                        "combo_group": row.combo_group,
                    }
                )
            initial_preventa = {
                "id": preventa.id,
                "numero": preventa.numero,
                "cliente_id": preventa.cliente_id,
                "cliente_nombre": preventa.cliente.nombre if preventa.cliente else "Consumidor final",
                "vendedor_id": preventa.vendedor_id,
                "fecha": preventa.fecha.date().isoformat() if preventa.fecha else local_today().isoformat(),
                "moneda": "USD",
                "items": items,
            }
            if preventa.estado == "PENDIENTE":
                preventa.estado = "REVISION"
                preventa.reviewed_at = local_now_naive()
                db.commit()

    sales_interface = _get_sales_interface_setting(db)
    interface_code = (sales_interface.interface_code or "ropa").strip().lower()
    template_name = "sales_comestibles.html" if interface_code == "comestibles" else "sales.html"

    return request.app.state.templates.TemplateResponse(
        template_name,
        {
            "request": request,
            "user": user,
            "productos": productos,
            "clientes_preview": clientes_preview,
            "vendedores": vendedores,
            "formas_pago": formas_pago,
            "bancos": bancos,
            "cuentas": cuentas,
            "rate_today": rate_today,
            "error": error,
            "success": success,
            "print_id": print_id,
            "next_invoice": next_invoice,
            "pos_print": pos_print,
            "default_vendedor_id": _default_vendedor_id(db, bodega),
            "initial_preventa": initial_preventa,
            "sales_interface_code": interface_code,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/sales/preventas/notifications/stream")
async def sales_preventas_notifications_stream(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales")
    branch, bodega = _resolve_branch_bodega(db, user)
    last_id_raw = (request.query_params.get("last_id") or "0").strip()
    last_id = int(last_id_raw) if last_id_raw.isdigit() else 0

    async def event_stream():
        nonlocal last_id
        headers_payload = {"ok": True}
        yield f"event: ready\ndata: {json.dumps(headers_payload, ensure_ascii=False)}\n\n"
        while True:
            if await request.is_disconnected():
                break
            try:
                query = (
                    db.query(
                        Preventa.id,
                        Preventa.numero,
                        Preventa.fecha,
                        Vendedor.nombre.label("vendedor_nombre"),
                        Cliente.nombre.label("cliente_nombre"),
                    )
                    .join(Vendedor, Vendedor.id == Preventa.vendedor_id)
                    .outerjoin(Cliente, Cliente.id == Preventa.cliente_id)
                    .filter(Preventa.estado == "PENDIENTE", Preventa.id > last_id)
                    .order_by(Preventa.id.asc())
                    .limit(10)
                )
                if bodega:
                    query = query.filter(Preventa.bodega_id == bodega.id)
                elif branch:
                    query = query.filter(Preventa.branch_id == branch.id)

                rows = query.all()
                for row in rows:
                    last_id = max(last_id, int(row.id or 0))
                    payload = {
                        "id": int(row.id),
                        "numero": row.numero,
                        "fecha": row.fecha.strftime("%Y-%m-%d %H:%M") if row.fecha else "-",
                        "vendedor": row.vendedor_nombre or "Sin vendedor",
                        "cliente": row.cliente_nombre or "Consumidor final",
                        "url": "/sales/preventas",
                    }
                    yield f"event: preventa\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"

                yield "event: ping\ndata: {}\n\n"
            except Exception:
                yield "event: error\ndata: {\"message\":\"stream_error\"}\n\n"
            await asyncio.sleep(8)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/sales/preventas/notifications/poll")
def sales_preventas_notifications_poll(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales")
    branch, bodega = _resolve_branch_bodega(db, user)
    last_id_raw = (request.query_params.get("last_id") or "0").strip()
    last_id = int(last_id_raw) if last_id_raw.isdigit() else 0

    query = (
        db.query(
            Preventa.id,
            Preventa.numero,
            Preventa.fecha,
            Vendedor.nombre.label("vendedor_nombre"),
            Cliente.nombre.label("cliente_nombre"),
        )
        .join(Vendedor, Vendedor.id == Preventa.vendedor_id)
        .outerjoin(Cliente, Cliente.id == Preventa.cliente_id)
        .filter(
            Preventa.id > last_id,
            Preventa.estado.in_(["PENDIENTE", "REVISION"]),
        )
    )
    if bodega:
        query = query.filter(Preventa.bodega_id == bodega.id)
    elif branch:
        query = query.filter(Preventa.branch_id == branch.id)
    query = query.order_by(Preventa.id.asc()).limit(20)

    items = []
    for row in query.all():
        items.append(
            {
                "id": int(row.id),
                "numero": row.numero,
                "fecha": row.fecha.strftime("%Y-%m-%d %H:%M") if row.fecha else "-",
                "vendedor": row.vendedor_nombre or "Sin vendedor",
                "cliente": row.cliente_nombre or "Consumidor final",
                "url": "/sales/preventas",
            }
        )
    return {"ok": True, "items": items}


@router.get("/m/preventas")
def mobile_preventas_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_preventas_mobile_access(request, user)
    branch, bodega = _resolve_branch_bodega(db, user)
    if not branch or not bodega:
        raise HTTPException(status_code=400, detail="Usuario sin sucursal/bodega asignada")
    vendedores = _vendedores_for_bodega(db, bodega)
    default_vendedor_id = _default_vendedor_id(db, bodega)
    vendedor_user_id = _vendedor_id_for_user(db, user, bodega)
    if vendedor_user_id:
        default_vendedor_id = vendedor_user_id
    if not default_vendedor_id and vendedores:
        default_vendedor_id = vendedores[0].id
    consumidor = _get_or_create_consumidor_final(db)
    db.commit()
    success = request.query_params.get("success")
    error = request.query_params.get("error")
    return request.app.state.templates.TemplateResponse(
        "sales_preventas_mobile.html",
        {
            "request": request,
            "user": user,
            "branch": branch,
            "bodega": bodega,
            "vendedores": vendedores,
            "default_vendedor_id": default_vendedor_id,
            "vendedor_user_id": vendedor_user_id,
            "is_vendedor_role": _is_vendedor_role(user),
            "consumidor_final_id": consumidor.id,
            "consumidor_final_nombre": consumidor.nombre,
            "success": success,
            "error": error,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/m/preventas/productos/search")
def mobile_preventas_products_search(
    request: Request,
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_preventas_mobile_access(request, user)
    query = q.strip()
    if len(query) < 2:
        return JSONResponse({"ok": True, "items": []})
    branch, bodega = _resolve_branch_bodega(db, user)
    if not bodega:
        return JSONResponse({"ok": False, "message": "Usuario sin bodega asignada"}, status_code=400)
    like = f"%{query.lower()}%"
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .filter(
            or_(
                func.lower(Producto.cod_producto).like(like),
                func.lower(Producto.descripcion).like(like),
            )
        )
        .order_by(Producto.descripcion)
        .limit(100)
        .all()
    )
    balances: dict[tuple[int, int], Decimal] = {}
    if productos:
        balances = _balances_by_bodega(db, [bodega.id], [p.id for p in productos])
    items = []
    for producto in productos:
        existencia = float(balances.get((producto.id, bodega.id), Decimal("0")) or 0)
        items.append(
            {
                "id": producto.id,
                "cod_producto": producto.cod_producto,
                "descripcion": producto.descripcion,
                "precio_venta1_usd": float(producto.precio_venta1_usd or 0),
                "precio_venta1": float(producto.precio_venta1 or 0),
                "existencia": existencia,
                "combo_count": len(producto.combo_children or []),
            }
        )
    return JSONResponse(
        {
            "ok": True,
            "items": items,
            "branch": branch.name if branch else "-",
            "bodega": bodega.name,
        }
    )


@router.get("/m/preventas/combos/search")
def mobile_preventas_combos_search(
    request: Request,
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_preventas_mobile_access(request, user)
    query = (q or "").strip()
    if len(query) < 2:
        return JSONResponse({"ok": True, "items": []})
    branch, bodega = _resolve_branch_bodega(db, user)
    if not bodega:
        return JSONResponse({"ok": False, "message": "Usuario sin bodega asignada"}, status_code=400)

    like = f"%{query.lower()}%"
    rows = (
        db.query(
            Producto,
            func.count(ProductoCombo.id).label("combo_count"),
        )
        .join(
            ProductoCombo,
            ProductoCombo.parent_producto_id == Producto.id,
        )
        .filter(Producto.activo.is_(True))
        .filter(
            or_(
                func.lower(Producto.cod_producto).like(like),
                func.lower(Producto.descripcion).like(like),
            )
        )
        .group_by(Producto.id)
        .order_by(Producto.descripcion)
        .limit(100)
        .all()
    )

    productos = [producto for producto, _ in rows]
    balances: dict[tuple[int, int], Decimal] = {}
    if productos:
        balances = _balances_by_bodega(db, [bodega.id], [p.id for p in productos])

    items = []
    for producto, combo_count in rows:
        existencia = float(balances.get((producto.id, bodega.id), Decimal("0")) or 0)
        items.append(
            {
                "id": producto.id,
                "cod_producto": producto.cod_producto,
                "descripcion": producto.descripcion,
                "precio_venta1_usd": float(producto.precio_venta1_usd or 0),
                "precio_venta1": float(producto.precio_venta1 or 0),
                "existencia": existencia,
                "combo_count": int(combo_count or 0),
            }
        )
    return JSONResponse(
        {
            "ok": True,
            "items": items,
            "branch": branch.name if branch else "-",
            "bodega": bodega.name,
        }
    )


@router.get("/m/preventas/product/{product_id}/combo")
def mobile_preventas_product_combo(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_preventas_mobile_access(request, user)
    _, bodega = _resolve_branch_bodega(db, user)
    if not bodega:
        return JSONResponse({"ok": False, "message": "Usuario sin bodega asignada"}, status_code=400)
    producto = db.query(Producto).filter(Producto.id == product_id, Producto.activo.is_(True)).first()
    if not producto:
        return JSONResponse({"ok": False, "message": "Producto no encontrado"}, status_code=404)
    combos = (
        db.query(ProductoCombo)
        .filter(
            ProductoCombo.parent_producto_id == product_id,
        )
        .order_by(ProductoCombo.id)
        .all()
    )
    child_ids = [c.child_producto_id for c in combos if c.child_producto_id]
    balances = _balances_by_bodega(db, [bodega.id], child_ids) if child_ids else {}
    items = []
    for combo in combos:
        child = combo.child
        if not child:
            continue
        existencia = float(balances.get((child.id, bodega.id), Decimal("0")) or 0)
        items.append(
            {
                "id": combo.id,
                "child_id": child.id,
                "cod_producto": child.cod_producto,
                "descripcion": child.descripcion,
                "cantidad": float(combo.cantidad or 0),
                "precio_venta1_usd": float(child.precio_venta1_usd or 0),
                "precio_venta1": float(child.precio_venta1 or 0),
                "existencia": existencia,
            }
        )
    return JSONResponse({"ok": True, "items": items})


@router.get("/m/preventas/clientes/search")
def mobile_preventas_clientes_search(
    request: Request,
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_preventas_mobile_access(request, user)
    query = (q or "").strip().lower()
    if len(query) < 2:
        return JSONResponse({"ok": True, "items": []})
    like = f"%{query}%"
    items = (
        db.query(Cliente)
        .filter(
            or_(
                func.lower(Cliente.nombre).like(like),
                func.lower(Cliente.identificacion).like(like),
                func.lower(Cliente.telefono).like(like),
            )
        )
        .order_by(Cliente.nombre)
        .limit(60)
        .all()
    )
    return JSONResponse(
        {
            "ok": True,
            "items": [{"id": c.id, "nombre": c.nombre, "telefono": c.telefono or ""} for c in items],
        }
    )


@router.post("/m/preventas/cliente")
async def mobile_preventas_create_cliente(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_preventas_mobile_access(request, user)
    form = await request.form()
    nombre = str(form.get("nombre") or "").strip()
    telefono = str(form.get("telefono") or "").strip() or None
    identificacion = str(form.get("identificacion") or "").strip() or None
    direccion = str(form.get("direccion") or "").strip() or None
    if not nombre:
        return JSONResponse({"ok": False, "message": "Nombre requerido"}, status_code=400)
    exists = db.query(Cliente).filter(func.lower(Cliente.nombre) == nombre.lower()).first()
    if exists:
        return JSONResponse({"ok": True, "id": exists.id, "nombre": exists.nombre, "existing": True})
    cliente = Cliente(
        nombre=nombre,
        telefono=telefono,
        identificacion=identificacion,
        direccion=direccion,
        activo=True,
    )
    db.add(cliente)
    db.commit()
    return JSONResponse({"ok": True, "id": cliente.id, "nombre": cliente.nombre, "existing": False})


@router.post("/m/preventas")
async def mobile_preventas_create(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_preventas_mobile_access(request, user)
    form = await request.form()
    cliente_id = form.get("cliente_id") or None
    vendedor_id = form.get("vendedor_id") or None
    fecha = form.get("fecha")
    moneda = (form.get("moneda") or "USD").upper()
    if moneda not in {"USD", "CS"}:
        moneda = "USD"
    observacion = (form.get("observacion") or "").strip() or None
    item_ids = form.getlist("item_producto_id")
    item_qtys = form.getlist("item_cantidad")
    item_prices = form.getlist("item_precio")
    item_price_usds = form.getlist("item_precio_usd")
    item_price_css = form.getlist("item_precio_cs")
    item_roles = form.getlist("item_role")
    item_combo_groups = form.getlist("item_combo_group")
    preventa_id_raw = str(form.get("preventa_id") or "").strip()

    branch, bodega = _resolve_branch_bodega(db, user)
    if not branch or not bodega:
        return RedirectResponse("/m/preventas?error=Usuario+sin+sucursal+bodega", status_code=303)
    if _is_vendedor_role(user):
        vendedor_user_id = _vendedor_id_for_user(db, user, bodega)
        if vendedor_user_id:
            vendedor_id = str(vendedor_user_id)

    if not vendedor_id:
        return RedirectResponse("/m/preventas?error=Selecciona+vendedor", status_code=303)
    if not item_ids:
        return RedirectResponse("/m/preventas?error=Agrega+items+a+la+preventa", status_code=303)
    vendedor = db.query(Vendedor).filter(Vendedor.id == int(vendedor_id), Vendedor.activo.is_(True)).first()
    if not vendedor:
        return RedirectResponse("/m/preventas?error=Vendedor+invalido", status_code=303)

    if not cliente_id:
        cliente_id = _get_or_create_consumidor_final(db).id
        db.flush()
    try:
        fecha_value = date.fromisoformat(str(fecha).split("T")[0]) if fecha else local_today()
    except (TypeError, ValueError):
        fecha_value = local_today()

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    tasa = Decimal(str(rate_today.rate if rate_today else 0))
    if moneda == "USD" and (not rate_today or tasa <= 0):
        return RedirectResponse("/m/preventas?error=Tasa+de+cambio+no+configurada", status_code=303)

    product_ids = [int(pid) for pid in item_ids if str(pid).isdigit()]
    balances = _balances_by_bodega(db, [bodega.id], list(set(product_ids))) if product_ids else {}

    def _to_dec(value: object, default: str = "0") -> Decimal:
        try:
            return Decimal(str(value if value is not None else default))
        except Exception:
            return Decimal(default)

    parsed_items: list[dict] = []
    for idx, raw_product_id in enumerate(item_ids):
        if not str(raw_product_id).isdigit():
            continue
        producto_id = int(raw_product_id)
        qty = _to_dec(item_qtys[idx] if idx < len(item_qtys) else "0").quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        if qty <= 0:
            continue
        producto = db.query(Producto).filter(Producto.id == producto_id, Producto.activo.is_(True)).first()
        if not producto:
            return RedirectResponse("/m/preventas?error=Producto+no+encontrado", status_code=303)
        existencia = Decimal(str(balances.get((producto.id, bodega.id), Decimal("0")) or 0))
        reserved_qty, reserved_details = _preventa_reserved_by_others(
            db,
            bodega_id=bodega.id,
            producto_id=producto.id,
            vendedor_id=vendedor.id,
        )
        libre = max(Decimal("0"), existencia - reserved_qty)
        if qty > libre:
            if reserved_details:
                detail_parts = [
                    f"{vend} ({numero}): {int(det_qty) if det_qty == det_qty.to_integral() else det_qty}"
                    for numero, vend, det_qty in reserved_details
                ]
                detail_txt = "; ".join(detail_parts)
                msg = (
                    f"Saldo libre insuficiente para {producto.cod_producto}. "
                    f"Solicitas {int(qty) if qty == qty.to_integral() else qty}, "
                    f"libre {int(libre) if libre == libre.to_integral() else libre} "
                    f"(existencia {int(existencia) if existencia == existencia.to_integral() else existencia}, "
                    f"reservado en preventas activas: {detail_txt})."
                )
            else:
                msg = f"Sin saldo para {producto.cod_producto}. Disponible {int(existencia) if existencia == existencia.to_integral() else existencia}."
            return RedirectResponse(f"/m/preventas?{urlencode({'error': msg})}", status_code=303)
        price_input = _to_dec(item_prices[idx] if idx < len(item_prices) else "0")
        price_usd_input = _to_dec(item_price_usds[idx] if idx < len(item_price_usds) else "0")
        price_cs_input = _to_dec(item_price_css[idx] if idx < len(item_price_css) else "0")
        combo_role = item_roles[idx] if idx < len(item_roles) else None
        combo_group = item_combo_groups[idx] if idx < len(item_combo_groups) else None
        combo_role = str(combo_role or "").strip() or None
        combo_group = str(combo_group or "").strip() or None
        if price_usd_input > 0 or price_cs_input > 0:
            precio_usd = price_usd_input
            precio_cs = price_cs_input
            if precio_usd <= 0 and tasa > 0:
                precio_usd = (precio_cs / tasa).quantize(Decimal("0.01"))
            if precio_cs <= 0:
                precio_cs = (precio_usd * tasa).quantize(Decimal("0.01")) if tasa > 0 else Decimal("0")
        else:
            if price_input <= 0:
                price_input = _to_dec(producto.precio_venta1 if moneda == "CS" else producto.precio_venta1_usd)
            if moneda == "USD":
                precio_usd = price_input
                precio_cs = (price_input * tasa).quantize(Decimal("0.01"))
            else:
                precio_cs = price_input
                precio_usd = (price_input / tasa).quantize(Decimal("0.01")) if tasa > 0 else Decimal("0")
        parsed_items.append(
            {
                "producto": producto,
                "cantidad": qty,
                "precio_usd": precio_usd,
                "precio_cs": precio_cs,
                "subtotal_usd": (precio_usd * qty).quantize(Decimal("0.01")),
                "subtotal_cs": (precio_cs * qty).quantize(Decimal("0.01")),
                "combo_role": combo_role,
                "combo_group": combo_group,
            }
        )

    if not parsed_items:
        return RedirectResponse("/m/preventas?error=No+hay+items+validos", status_code=303)

    seq, numero = _next_preventa_number(db, branch)
    now_local = local_now()
    fecha_dt = datetime.combine(fecha_value, now_local.time()).replace(tzinfo=None)
    preventa = Preventa(
        secuencia=seq,
        numero=numero,
        branch_id=branch.id,
        bodega_id=bodega.id,
        cliente_id=int(cliente_id) if cliente_id else None,
        vendedor_id=vendedor.id,
        fecha=fecha_dt,
        estado="PENDIENTE",
        observacion=observacion,
        total_usd=sum((x["subtotal_usd"] for x in parsed_items), Decimal("0")),
        total_cs=sum((x["subtotal_cs"] for x in parsed_items), Decimal("0")),
        total_items=sum((x["cantidad"] for x in parsed_items), Decimal("0")),
        usuario_registro=user.full_name,
        created_at=local_now_naive(),
    )
    db.add(preventa)
    db.flush()
    for item in parsed_items:
        db.add(
            PreventaItem(
                preventa_id=preventa.id,
                producto_id=item["producto"].id,
                cantidad=item["cantidad"],
                precio_unitario_usd=item["precio_usd"],
                precio_unitario_cs=item["precio_cs"],
                subtotal_usd=item["subtotal_usd"],
                subtotal_cs=item["subtotal_cs"],
                combo_role=item["combo_role"],
                combo_group=item["combo_group"],
            )
        )
    db.commit()
    return RedirectResponse(
        f"/m/preventas?success=Preventa+{preventa.numero}+registrada",
        status_code=303,
    )


@router.get("/sales/preventas")
def sales_preventas_panel(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.preventas")
    today = local_today()
    fecha = (request.query_params.get("fecha") or today.isoformat()).strip()
    estado = (request.query_params.get("estado") or "").strip().upper()
    vendedor_id = (request.query_params.get("vendedor_id") or "").strip()
    branch_id = (request.query_params.get("branch_id") or "all").strip()
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    try:
        fecha_value = date.fromisoformat(fecha)
    except ValueError:
        fecha_value = today

    query = (
        db.query(Preventa)
        .join(Cliente, Cliente.id == Preventa.cliente_id, isouter=True)
        .join(Vendedor, Vendedor.id == Preventa.vendedor_id, isouter=True)
        .join(Branch, Branch.id == Preventa.branch_id, isouter=True)
        .filter(func.date(Preventa.fecha) == fecha_value)
    )
    if branch_id and branch_id != "all":
        if branch_id.isdigit():
            query = query.filter(Preventa.branch_id == int(branch_id))
    if vendedor_id and vendedor_id.isdigit():
        query = query.filter(Preventa.vendedor_id == int(vendedor_id))
    if estado:
        query = query.filter(Preventa.estado == estado)
    preventas = query.order_by(Preventa.id.desc()).all()
    repaired_any = False
    for p in preventas:
        if _repair_preventa_currency_if_needed(db, p):
            repaired_any = True
    if repaired_any:
        db.commit()
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    branches = (
        _scoped_branches_query(db)
        .filter(Branch.id.in_(scoped_branch_ids))
        .order_by(Branch.name)
        .all()
    )
    vendedores = db.query(Vendedor).filter(Vendedor.activo.is_(True)).order_by(Vendedor.nombre).all()
    rows = [
        {
            "id": p.id,
            "numero": p.numero,
            "fecha": p.fecha,
            "cliente": p.cliente.nombre if p.cliente else "Consumidor final",
            "vendedor": p.vendedor.nombre if p.vendedor else "-",
            "sucursal": p.branch.name if p.branch else "-",
            "total_usd": float(p.total_usd or 0),
            "total_cs": float(p.total_cs or 0),
            "estado": p.estado,
            "badge": _preventa_estado_badge(p.estado),
        }
        for p in preventas
    ]
    return request.app.state.templates.TemplateResponse(
        "sales_preventas_panel.html",
        {
            "request": request,
            "user": user,
            "rows": rows,
            "fecha": fecha_value.isoformat(),
            "estado": estado,
            "vendedor_id": vendedor_id,
            "branch_id": branch_id,
            "branches": branches,
            "vendedores": vendedores,
            "is_vendedor_role": _is_vendedor_role(user),
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/sales/preventas/{preventa_id}/detail")
def sales_preventas_detail(
    request: Request,
    preventa_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.preventas")
    preventa = db.query(Preventa).filter(Preventa.id == preventa_id).first()
    if not preventa:
        return JSONResponse({"ok": False, "message": "Preventa no encontrada"}, status_code=404)
    if _repair_preventa_currency_if_needed(db, preventa):
        db.commit()
    rows = (
        db.query(PreventaItem, Producto)
        .join(Producto, Producto.id == PreventaItem.producto_id)
        .filter(PreventaItem.preventa_id == preventa.id)
        .order_by(PreventaItem.id)
        .all()
    )
    balances = _balances_by_bodega(db, [preventa.bodega_id], [p.id for _, p in rows]) if rows else {}
    items = []
    total_usd_items = Decimal("0")
    total_cs_items = Decimal("0")
    for item, producto in rows:
        existencia = float(balances.get((producto.id, preventa.bodega_id), Decimal("0")) or 0)
        qty = float(item.cantidad or 0)
        combo_role = (item.combo_role or "").strip().lower() if getattr(item, "combo_role", None) else ""
        combo_group = (item.combo_group or "").strip() if getattr(item, "combo_group", None) else ""
        items.append(
            {
                "codigo": producto.cod_producto,
                "descripcion": producto.descripcion,
                "cantidad": qty,
                "precio_usd": float(item.precio_unitario_usd or 0),
                "precio_cs": float(item.precio_unitario_cs or 0),
                "subtotal_usd": float(item.subtotal_usd or 0),
                "subtotal_cs": float(item.subtotal_cs or 0),
                "existencia": existencia,
                "ok_stock": existencia >= qty,
                "combo_role": combo_role,
                "combo_group": combo_group,
            }
        )
        total_usd_items += Decimal(str(item.subtotal_usd or 0))
        total_cs_items += Decimal(str(item.subtotal_cs or 0))
    if preventa.estado == "PENDIENTE":
        preventa.estado = "REVISION"
        preventa.reviewed_at = local_now_naive()
        db.commit()
    return JSONResponse(
        {
            "ok": True,
            "preventa": {
                "id": preventa.id,
                "numero": preventa.numero,
                "estado": preventa.estado,
                "cliente": preventa.cliente.nombre if preventa.cliente else "Consumidor final",
                "vendedor": preventa.vendedor.nombre if preventa.vendedor else "-",
                "fecha": preventa.fecha.isoformat() if preventa.fecha else "",
                "total_usd": float(total_usd_items),
                "total_cs": float(total_cs_items),
            },
            "items": items,
        }
    )


@router.post("/sales/preventas/{preventa_id}/anular")
async def sales_preventas_anular(
    request: Request,
    preventa_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.preventas")
    if _is_vendedor_role(user):
        return RedirectResponse("/sales/preventas?error=Rol+vendedor+no+puede+anular+preventas", status_code=303)
    preventa = db.query(Preventa).filter(Preventa.id == preventa_id).first()
    if not preventa:
        return RedirectResponse("/sales/preventas?error=Preventa+no+encontrada", status_code=303)
    if preventa.estado == "FACTURADA":
        return RedirectResponse("/sales/preventas?error=No+se+puede+anular+una+preventa+facturada", status_code=303)
    preventa.estado = "ANULADA"
    preventa.anulada_at = local_now_naive()
    preventa.anulada_por = user.full_name
    db.commit()
    return RedirectResponse("/sales/preventas?success=Preventa+anulada", status_code=303)


@router.post("/sales/preventas/{preventa_id}/usar")
async def sales_preventas_usar_en_factura(
    request: Request,
    preventa_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.preventas")
    _enforce_permission(request, user, "access.sales.registrar")
    preventa = db.query(Preventa).filter(Preventa.id == preventa_id).first()
    if not preventa:
        return RedirectResponse("/sales/preventas?error=Preventa+no+encontrada", status_code=303)
    if preventa.estado not in {"PENDIENTE", "REVISION"}:
        return RedirectResponse("/sales/preventas?error=Preventa+no+disponible+para+facturar", status_code=303)
    item_rows = (
        db.query(PreventaItem, Producto)
        .join(Producto, Producto.id == PreventaItem.producto_id)
        .filter(PreventaItem.preventa_id == preventa.id)
        .all()
    )
    balances = _balances_by_bodega(db, [preventa.bodega_id], [p.id for _, p in item_rows]) if item_rows else {}
    required_by_product = _preventa_required_qty_map(item_rows)
    for producto_id, required_qty in required_by_product.items():
        producto = next((p for _item, p in item_rows if int(p.id) == int(producto_id)), None)
        if not producto:
            continue
        existencia = Decimal(str(balances.get((producto.id, preventa.bodega_id), Decimal("0")) or 0))
        if existencia < required_qty:
            return RedirectResponse(
                f"/sales/preventas?error=Sin+saldo+actual+para+{producto.cod_producto}",
                status_code=303,
            )
    if preventa.estado == "PENDIENTE":
        preventa.estado = "REVISION"
        preventa.reviewed_at = local_now_naive()
    db.commit()
    return RedirectResponse(f"/sales?preventa_id={preventa.id}", status_code=303)


@router.post("/sales/preventas/{preventa_id}/release")
async def sales_preventas_release_from_sales(
    request: Request,
    preventa_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.registrar")
    preventa = db.query(Preventa).filter(Preventa.id == preventa_id).first()
    if not preventa:
        return JSONResponse({"ok": False, "message": "Preventa no encontrada"}, status_code=404)
    if preventa.estado == "FACTURADA":
        return JSONResponse(
            {"ok": False, "message": "No se puede liberar una preventa facturada"},
            status_code=400,
        )
    if preventa.estado == "ANULADA":
        return JSONResponse({"ok": True, "message": "La preventa ya estaba anulada"})
    preventa.estado = "ANULADA"
    preventa.anulada_at = local_now_naive()
    preventa.anulada_por = user.full_name or user.email
    db.commit()
    return JSONResponse({"ok": True, "message": f"Preventa {preventa.numero} liberada"})


@router.get("/sales/utilitario")
def sales_utilitario(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.utilitario")
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    def _parse_date(value: Optional[str]) -> Optional[date]:
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    start_date = _parse_date(request.query_params.get("start_date"))
    end_date = _parse_date(request.query_params.get("end_date"))
    branch_id = request.query_params.get("branch_id") or "all"
    cliente_q = (request.query_params.get("cliente") or "").strip()
    vendedor_q = (request.query_params.get("vendedor") or "").strip()
    producto_q = (request.query_params.get("producto") or "").strip()
    if not start_date and not end_date:
        start_date = local_today()
        end_date = local_today()

    ventas_query = (
        db.query(VentaFactura)
        .join(Bodega, Bodega.id == VentaFactura.bodega_id, isouter=True)
        .join(Branch, Branch.id == Bodega.branch_id, isouter=True)
        .filter(Branch.id.in_(scoped_branch_ids))
    )
    if branch_id and branch_id != "all":
        try:
            branch_id_int = int(branch_id)
            if branch_id_int not in scoped_branch_ids:
                ventas_query = ventas_query.filter(Branch.id == -1)
            else:
                ventas_query = ventas_query.filter(Branch.id == branch_id_int)
        except ValueError:
            pass
    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time())
        ventas_query = ventas_query.filter(VentaFactura.fecha >= start_dt)
    if end_date:
        end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())
        ventas_query = ventas_query.filter(VentaFactura.fecha < end_dt)
    if cliente_q:
        ventas_query = ventas_query.join(Cliente, isouter=True).filter(
            func.lower(Cliente.nombre).like(f"%{cliente_q.lower()}%")
        )
    if vendedor_q:
        ventas_query = ventas_query.join(Vendedor, isouter=True).filter(
            func.lower(Vendedor.nombre).like(f"%{vendedor_q.lower()}%")
        )
    if producto_q:
        ventas_query = (
            ventas_query.join(VentaItem)
            .join(Producto)
            .filter(
                (func.lower(Producto.descripcion).like(f"%{producto_q.lower()}%"))
                | (func.lower(Producto.cod_producto).like(f"%{producto_q.lower()}%"))
            )
            .distinct()
        )
    ventas = (
        ventas_query.order_by(VentaFactura.fecha.desc(), VentaFactura.id.desc()).all()
    )
    _, bodega = _resolve_branch_bodega(db, user)
    vendedores = _vendedores_for_bodega(db, bodega)
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    branches = (
        _scoped_branches_query(db)
        .filter(Branch.id.in_(scoped_branch_ids))
        .order_by(Branch.name)
        .all()
    )

    return request.app.state.templates.TemplateResponse(
        "sales_utilitario.html",
        {
            "request": request,
            "user": user,
            "ventas": ventas,
            "vendedores": vendedores,
            "branches": branches,
            "start_date": start_date.isoformat() if start_date else "",
            "end_date": end_date.isoformat() if end_date else "",
            "branch_id": branch_id,
            "cliente_q": cliente_q,
            "vendedor_q": vendedor_q,
            "producto_q": producto_q,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/sales/etiquetas")
def sales_etiquetas(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.etiquetas")
    return request.app.state.templates.TemplateResponse(
        "sales_etiquetas.html",
        {
            "request": request,
            "user": user,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/sales/etiquetas/background/upload")
async def sales_etiquetas_upload_background(
    request: Request,
    target_format: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.etiquetas")
    format_map = {
        "pacas_1": "etiquetaPacasuna",
        "pacas_2": "etiquetaPacas2",
        "bolsas_2": "etiquetaBolsas2",
        "bolsa_50": "etiquetaBolsa50",
    }
    key = (target_format or "").strip().lower()
    if key not in format_map:
        return JSONResponse({"ok": False, "message": "Formato no valido"}, status_code=400)
    if not file or not file.filename:
        return JSONResponse({"ok": False, "message": "Archivo requerido"}, status_code=400)

    ext = Path(file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".webp"}:
        return JSONResponse({"ok": False, "message": "Formato de imagen no permitido"}, status_code=400)

    try:
        payload = await file.read()
    except Exception:
        return JSONResponse({"ok": False, "message": "No se pudo leer el archivo"}, status_code=400)
    if not payload:
        return JSONResponse({"ok": False, "message": "Archivo vacio"}, status_code=400)
    if len(payload) > 15 * 1024 * 1024:
        return JSONResponse({"ok": False, "message": "Archivo excede 15MB"}, status_code=400)

    labels_dir = Path(__file__).resolve().parents[1] / "static" / "labels"
    labels_dir.mkdir(parents=True, exist_ok=True)
    base_name = format_map[key]
    out_path = labels_dir / f"{base_name}{ext}"
    # Evitar fondos duplicados con distintas extensiones para el mismo formato.
    for candidate in labels_dir.glob(f"{base_name}.*"):
        if candidate != out_path:
            try:
                candidate.unlink()
            except OSError:
                pass

    with out_path.open("wb") as fh:
        fh.write(payload)

    return JSONResponse({"ok": True, "message": "Background actualizado", "url": f"/static/labels/{out_path.name}"})


def _sales_commissions_filters(request: Request):
    fecha_raw = request.query_params.get("fecha")
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    branch_id = request.query_params.get("branch_id") or "all"
    vendedor_facturacion_id = (request.query_params.get("vendedor_facturacion_id") or "").strip()
    vendedor_asignado_id = (request.query_params.get("vendedor_asignado_id") or "").strip()
    producto_asig_q = (request.query_params.get("producto_asig_q") or "").strip()
    active_tab = (request.query_params.get("tab") or "precios").strip().lower()
    if active_tab not in {"precios", "asignacion", "reportes"}:
        active_tab = "precios"
    today_value = local_today()
    start_date = today_value
    end_date = today_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
            if start_raw and not end_raw:
                end_date = start_date
            if end_raw and not start_raw:
                start_date = end_date
        except ValueError:
            start_date = today_value
            end_date = today_value
    elif fecha_raw:
        # Compatibilidad con URLs antiguas que enviaban solo "fecha".
        try:
            only_date = date.fromisoformat(fecha_raw)
            start_date = only_date
            end_date = only_date
        except ValueError:
            start_date = today_value
            end_date = today_value
    if end_date < start_date:
        end_date = start_date
    fecha_value = start_date
    return (
        fecha_value,
        start_date,
        end_date,
        branch_id,
        vendedor_facturacion_id,
        vendedor_asignado_id,
        producto_asig_q,
        active_tab,
    )


def _sales_commissions_report_filters(request: Request):
    start_raw = request.query_params.get("rep_start_date")
    end_raw = request.query_params.get("rep_end_date")
    branch_id = request.query_params.get("rep_branch_id") or "all"
    vendedor_id = (request.query_params.get("rep_vendedor_id") or "").strip()

    today = local_today()
    start_date = today
    end_date = today
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today
            end_date = today
    if end_date < start_date:
        end_date = start_date
    return start_date, end_date, branch_id, vendedor_id


def _commission_sales_rows_query(
    db: Session,
    fecha_value: date,
    branch_id: str | None,
    vendedor_id: str | None,
    producto_asig_q: str,
):
    return _commission_sales_rows_query_range(
        db,
        fecha_value,
        fecha_value,
        branch_id,
        vendedor_id,
        producto_asig_q,
    )


def _commission_sales_rows_query_range(
    db: Session,
    start_date: date,
    end_date: date,
    branch_id: str | None,
    vendedor_id: str | None,
    producto_asig_q: str,
):
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    query = (
        db.query(VentaFactura, VentaItem, Producto, Cliente, Vendedor, Branch, Bodega)
        .join(VentaItem, VentaItem.factura_id == VentaFactura.id)
        .join(Producto, Producto.id == VentaItem.producto_id)
        .join(Bodega, Bodega.id == VentaFactura.bodega_id, isouter=True)
        .join(Branch, Branch.id == Bodega.branch_id, isouter=True)
        .join(Cliente, Cliente.id == VentaFactura.cliente_id, isouter=True)
        .join(Vendedor, Vendedor.id == VentaFactura.vendedor_id, isouter=True)
        .filter(VentaFactura.fecha >= start_dt, VentaFactura.fecha < end_dt)
        .filter(VentaFactura.estado != "ANULADA")
    )
    if branch_id and branch_id != "all":
        try:
            query = query.filter(Branch.id == int(branch_id))
        except ValueError:
            pass
    if vendedor_id:
        try:
            query = query.filter(VentaFactura.vendedor_id == int(vendedor_id))
        except ValueError:
            pass
    if producto_asig_q:
        like = f"%{producto_asig_q.lower()}%"
        query = query.filter(
            or_(
                func.lower(Producto.cod_producto).like(like),
                func.lower(Producto.descripcion).like(like),
            )
        )
    return query.order_by(VentaFactura.id.asc(), VentaItem.id.asc())


def _commission_dates_in_range(start_date: date, end_date: date) -> list[date]:
    if end_date < start_date:
        return [start_date]
    days = (end_date - start_date).days
    return [start_date + timedelta(days=i) for i in range(days + 1)]


def _commission_branch_scope(branch_id: str | None) -> Optional[int]:
    if not branch_id or branch_id == "all":
        return None
    try:
        return int(branch_id)
    except ValueError:
        return None


def _ensure_commission_temp_snapshot(
    db: Session,
    fecha_value: date,
    branch_id: str | None,
) -> tuple[int, int]:
    scope_branch_id = _commission_branch_scope(branch_id)
    source_rows = _commission_sales_rows_query(
        db, fecha_value, str(scope_branch_id) if scope_branch_id else "all", None, ""
    ).all()

    source_map: dict[int, tuple] = {
        item.id: (factura, item, producto, cliente, vendedor, branch, bodega)
        for factura, item, producto, cliente, vendedor, branch, bodega in source_rows
    }
    source_item_ids = set(source_map.keys())

    temp_query = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha == fecha_value
    )
    if scope_branch_id:
        temp_query = temp_query.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
    temp_rows = temp_query.all()
    temp_by_item: dict[int, list[VentaComisionAsignacion]] = {}
    for row in temp_rows:
        temp_by_item.setdefault(row.venta_item_id, []).append(row)
    fallback_vendor = (
        db.query(Vendedor.id)
        .filter(Vendedor.activo.is_(True))
        .order_by(Vendedor.nombre)
        .first()
    )
    fallback_vendor_id = fallback_vendor[0] if fallback_vendor else None

    created = 0
    updated = 0
    for item_id, source in source_map.items():
        factura, item, producto, cliente, vendedor, branch, bodega = source
        assigned_vendor_id = (vendedor.id if vendedor else None) or fallback_vendor_id
        if not assigned_vendor_id:
            continue
        sold_qty = int(
            Decimal(str(item.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        )
        if sold_qty <= 0:
            continue
        existing_rows = temp_by_item.get(item_id, [])
        if existing_rows:
            price_usd = Decimal(str(item.precio_unitario_usd or 0))
            price_cs = Decimal(str(item.precio_unitario_cs or 0))
            for row in existing_rows:
                row.factura_id = factura.id
                row.branch_id = branch.id if branch else None
                row.bodega_id = bodega.id if bodega else None
                row.cliente_id = cliente.id if cliente else None
                row.producto_id = producto.id
                row.vendedor_origen_id = vendedor.id if vendedor else None
                if not row.vendedor_asignado_id:
                    row.vendedor_asignado_id = assigned_vendor_id

            rows_sorted = sorted(existing_rows, key=lambda r: r.id)
            primary = next(
                (r for r in rows_sorted if r.vendedor_asignado_id == (vendedor.id if vendedor else None)),
                rows_sorted[0],
            )
            secondary_rows = [r for r in rows_sorted if r.id != primary.id]
            sum_secondary = 0
            for row in secondary_rows:
                q = int(Decimal(str(row.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
                if q < 0:
                    q = 0
                row.cantidad = q
                sum_secondary += q

            primary_qty = sold_qty - sum_secondary
            if primary_qty < 0:
                primary_qty = 0
            primary.cantidad = primary_qty

            for row in existing_rows:
                q = Decimal(str(row.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                row.subtotal_usd = price_usd * q
                row.subtotal_cs = price_cs * q
                row.precio_unitario_usd = price_usd
                row.precio_unitario_cs = price_cs
            updated += len(existing_rows)
            continue

        db.add(
            VentaComisionAsignacion(
                venta_item_id=item.id,
                factura_id=factura.id,
                branch_id=branch.id if branch else None,
                bodega_id=bodega.id if bodega else None,
                cliente_id=cliente.id if cliente else None,
                producto_id=producto.id,
                fecha=fecha_value,
                vendedor_origen_id=vendedor.id if vendedor else None,
                vendedor_asignado_id=assigned_vendor_id,
                cantidad=sold_qty,
                precio_unitario_usd=Decimal(str(item.precio_unitario_usd or 0)),
                precio_unitario_cs=Decimal(str(item.precio_unitario_cs or 0)),
                subtotal_usd=Decimal(str(item.precio_unitario_usd or 0)) * sold_qty,
                subtotal_cs=Decimal(str(item.precio_unitario_cs or 0)) * sold_qty,
                usuario_registro="snapshot",
            )
        )
        created += 1

    removed = 0
    for row in temp_rows:
        if row.venta_item_id not in source_item_ids:
            db.delete(row)
            removed += 1

    if created or removed or updated:
        db.commit()
    return created, removed


def _build_commission_assignment_rows(
    db: Session,
    start_date: date,
    end_date: date,
    branch_id: str | None,
    vendedor_facturacion_id: str | None,
    vendedor_asignado_id: str | None,
    producto_asig_q: str,
):
    scope_branch_id = _commission_branch_scope(branch_id)
    temp_query = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha >= start_date,
        VentaComisionAsignacion.fecha <= end_date,
    )
    if scope_branch_id:
        temp_query = temp_query.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
    if vendedor_asignado_id:
        try:
            temp_query = temp_query.filter(
                VentaComisionAsignacion.vendedor_asignado_id == int(vendedor_asignado_id)
            )
        except ValueError:
            pass
    source_rows = _commission_sales_rows_query_range(
        db,
        start_date,
        end_date,
        branch_id if branch_id else "all",
        None,
        "",
    ).all()
    source_qty_map: dict[int, int] = {}
    source_meta_map: dict[int, tuple] = {}
    for factura, item, producto, cliente, vendedor, branch, bodega in source_rows:
        qty = int(
            Decimal(str(item.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        )
        source_qty_map[item.id] = qty
        source_meta_map[item.id] = (factura, item, producto, cliente, vendedor, branch, bodega)

    temp_rows = temp_query.order_by(VentaComisionAsignacion.id.asc()).all()

    # Blindaje: si la tabla temporal quedo incompleta, recrea filas faltantes
    # desde ventas reales del dia para evitar facturas "perdidas" en la grilla.
    source_item_ids = set(source_qty_map.keys())
    temp_item_ids = {row.venta_item_id for row in temp_rows}
    missing_item_ids = source_item_ids - temp_item_ids
    if missing_item_ids:
        fallback_vendor = (
            db.query(Vendedor.id)
            .filter(Vendedor.activo.is_(True))
            .order_by(Vendedor.nombre)
            .first()
        )
        fallback_vendor_id = fallback_vendor[0] if fallback_vendor else None
        for item_id in missing_item_ids:
            source = source_meta_map.get(item_id)
            if not source:
                continue
            factura, item, producto, cliente, vendedor, branch, bodega = source
            qty = int(
                Decimal(str(item.cantidad or 0)).quantize(
                    Decimal("1"), rounding=ROUND_HALF_UP
                )
            )
            if qty <= 0:
                continue
            assigned_vendor_id = (vendedor.id if vendedor else None) or fallback_vendor_id
            if not assigned_vendor_id:
                continue
            db.add(
                VentaComisionAsignacion(
                    venta_item_id=item.id,
                    factura_id=factura.id,
                    branch_id=branch.id if branch else None,
                    bodega_id=bodega.id if bodega else None,
                    cliente_id=cliente.id if cliente else None,
                    producto_id=producto.id,
                    fecha=factura.fecha.date() if factura and factura.fecha else start_date,
                    vendedor_origen_id=vendedor.id if vendedor else None,
                    vendedor_asignado_id=assigned_vendor_id,
                    cantidad=qty,
                    precio_unitario_usd=Decimal(str(item.precio_unitario_usd or 0)),
                    precio_unitario_cs=Decimal(str(item.precio_unitario_cs or 0)),
                    subtotal_usd=Decimal(str(item.precio_unitario_usd or 0)) * qty,
                    subtotal_cs=Decimal(str(item.precio_unitario_cs or 0)) * qty,
                    usuario_registro="autobackfill",
                )
            )
        db.commit()
        temp_rows = temp_query.order_by(VentaComisionAsignacion.id.asc()).all()

    if not temp_rows:
        return [], 0, 0, 0, {}, 0.0, 0.0, 0, []

    product_ids = list({row.producto_id for row in temp_rows})
    factura_ids = list({row.factura_id for row in temp_rows})
    cliente_ids = list({row.cliente_id for row in temp_rows if row.cliente_id})
    vendor_ids = list(
        {
            row.vendedor_origen_id
            for row in temp_rows
            if row.vendedor_origen_id
        }
        | {
            row.vendedor_asignado_id
            for row in temp_rows
            if row.vendedor_asignado_id
        }
    )

    products = (
        db.query(Producto).filter(Producto.id.in_(product_ids)).all() if product_ids else []
    )
    product_map = {p.id: p for p in products}
    factura_rows = (
        db.query(VentaFactura).filter(VentaFactura.id.in_(factura_ids)).all()
        if factura_ids
        else []
    )
    factura_map = {f.id: f for f in factura_rows}
    cliente_rows = (
        db.query(Cliente).filter(Cliente.id.in_(cliente_ids)).all()
        if cliente_ids
        else []
    )
    cliente_map = {c.id: c for c in cliente_rows}
    vendor_rows = (
        db.query(Vendedor).filter(Vendedor.id.in_(vendor_ids)).all()
        if vendor_ids
        else []
    )
    vendor_map = {v.id: v.nombre for v in vendor_rows}

    if producto_asig_q:
        like = producto_asig_q.lower()
        filtered = []
        for row in temp_rows:
            p = product_map.get(row.producto_id)
            if not p:
                continue
            haystack = f"{p.cod_producto or ''} {p.descripcion or ''}".lower()
            if like in haystack:
                filtered.append(row)
        temp_rows = filtered

    if vendedor_facturacion_id:
        try:
            vendedor_facturacion_id_int = int(vendedor_facturacion_id)
            temp_rows = [
                row
                for row in temp_rows
                if int(row.vendedor_origen_id or 0) == vendedor_facturacion_id_int
            ]
        except ValueError:
            pass

    if not temp_rows:
        return [], 0, 0, 0, {}, 0.0, 0.0, 0, []

    commission_rows = (
        db.query(ProductoComision)
        .filter(ProductoComision.producto_id.in_(product_ids))
        .all()
        if product_ids
        else []
    )
    commission_map: dict[int, Decimal] = {
        row.producto_id: Decimal(str(row.comision_usd or 0))
        for row in commission_rows
    }

    def _row_qty_int(r: VentaComisionAsignacion) -> int:
        return int(
            Decimal(str(r.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        )

    primary_ids: set[int] = set()
    grouped_rows: dict[int, list[VentaComisionAsignacion]] = {}
    for row in temp_rows:
        grouped_rows.setdefault(row.venta_item_id, []).append(row)
    for rows in grouped_rows.values():
        rows_sorted = sorted(rows, key=lambda r: r.id)
        positive_rows = [r for r in rows_sorted if _row_qty_int(r) > 0]
        preferred = next(
            (
                r
                for r in rows_sorted
                if r.vendedor_asignado_id == r.vendedor_origen_id and _row_qty_int(r) > 0
            ),
            (positive_rows[0] if positive_rows else rows_sorted[0]),
        )
        primary_ids.add(preferred.id)

    output_rows: list[dict] = []
    for row in temp_rows:
        producto = product_map.get(row.producto_id)
        factura = factura_map.get(row.factura_id)
        cliente = cliente_map.get(row.cliente_id) if row.cliente_id else None
        if not producto or not factura:
            continue
        precio = (
            Decimal(str(row.precio_unitario_usd or 0))
            if (factura.moneda or "CS") == "USD"
            else Decimal(str(row.precio_unitario_cs or 0))
        )
        comision_unit = commission_map.get(producto.id, Decimal("0"))
        precio_label = "$" if (factura.moneda or "CS") == "USD" else "C$"
        qty_int = int(Decimal(str(row.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        total_qty_int = source_qty_map.get(row.venta_item_id, qty_int)
        vendedor_origen_nombre = vendor_map.get(row.vendedor_origen_id, "-")
        vendedor_asignado_nombre = vendor_map.get(
            row.vendedor_asignado_id, vendedor_origen_nombre
        )
        output_rows.append(
            {
                "temp_id": row.id,
                "venta_item_id": row.venta_item_id,
                "factura_id": row.factura_id,
                "fecha": row.fecha,
                "fecha_label": row.fecha.strftime("%d/%m/%Y") if row.fecha else "-",
                "producto_id": row.producto_id,
                "factura_numero": factura.numero,
                "cliente": cliente.nombre if cliente else "Consumidor final",
                "descripcion": f"{producto.cod_producto} - {producto.descripcion}",
                "cantidad": qty_int,
                "precio": float(precio),
                "precio_usd_unit": float(row.precio_unitario_usd or 0),
                "precio_label": precio_label,
                "comision_unit_usd": float(comision_unit),
                "comision_total_usd": float(comision_unit * Decimal(str(qty_int))),
                "subtotal_usd": float(row.subtotal_usd or 0),
                "vendedor_origen_id": row.vendedor_origen_id,
                "vendedor_origen": vendedor_origen_nombre,
                "vendedor_id": row.vendedor_asignado_id,
                "vendedor_nombre": vendedor_asignado_nombre,
                "is_primary": row.id in primary_ids,
                "cantidad_total_item": total_qty_int,
            }
        )

    output_rows.sort(
        key=lambda row: (
            (row.get("vendedor_nombre") or "-").lower(),
            str(row.get("factura_numero") or ""),
            str(row.get("descripcion") or ""),
        )
    )
    visible_item_ids = {int(row.venta_item_id) for row in temp_rows}
    total_bultos_vendidos = int(
        sum(int(source_qty_map.get(item_id, 0)) for item_id in visible_item_ids)
    )
    total_bultos = int(sum(int(row.get("cantidad") or 0) for row in output_rows))
    total_comision = Decimal("0")
    total_ventas_usd = Decimal("0")
    total_facturas = len({int(row.get("factura_id") or 0) for row in output_rows if row.get("factura_id")})
    period_summary_by_date: dict[date, dict] = {}
    by_vendor: dict[str, dict[str, float]] = {}
    for row in output_rows:
        vendor_name = row.get("vendedor_nombre") or "-"
        qty = int(row.get("cantidad") or 0)
        comision_total = Decimal(str(row.get("comision_total_usd") or 0))
        subtotal_usd = Decimal(str(row.get("subtotal_usd") or 0))
        fecha_val = row.get("fecha")
        factura_id = int(row.get("factura_id") or 0)
        total_comision += comision_total
        total_ventas_usd += subtotal_usd
        if isinstance(fecha_val, date):
            if fecha_val not in period_summary_by_date:
                period_summary_by_date[fecha_val] = {
                    "fecha_label": fecha_val.strftime("%d/%m/%Y"),
                    "bultos": 0,
                    "ventas_usd": Decimal("0"),
                    "facturas": set(),
                }
            period_summary_by_date[fecha_val]["bultos"] += qty
            period_summary_by_date[fecha_val]["ventas_usd"] += subtotal_usd
            if factura_id:
                period_summary_by_date[fecha_val]["facturas"].add(factura_id)
        if vendor_name not in by_vendor:
            by_vendor[vendor_name] = {
                "items_vendidos": 0.0,
                "bultos": 0.0,
                "ventas_usd": 0.0,
                "comision_usd": 0.0,
            }
        by_vendor[vendor_name]["items_vendidos"] += qty
        by_vendor[vendor_name]["bultos"] += qty
        by_vendor[vendor_name]["ventas_usd"] += float(subtotal_usd)
        by_vendor[vendor_name]["comision_usd"] += float(comision_total)
    period_summary_rows = []
    for fecha_val in sorted(period_summary_by_date.keys()):
        item = period_summary_by_date[fecha_val]
        period_summary_rows.append(
            {
                "fecha_label": item["fecha_label"],
                "facturas": len(item["facturas"]),
                "bultos": int(item["bultos"]),
                "ventas_usd": float(item["ventas_usd"]),
            }
        )
    return (
        output_rows,
        total_bultos,
        len(output_rows),
        total_bultos_vendidos,
        by_vendor,
        float(total_comision),
        float(total_ventas_usd),
        total_facturas,
        period_summary_rows,
    )


def _commission_missing_prices(
    db: Session,
    start_date: date,
    end_date: date,
    branch_id: str | None,
) -> list[dict]:
    scope_branch_id = _commission_branch_scope(branch_id)
    temp_query = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha >= start_date,
        VentaComisionAsignacion.fecha <= end_date,
    )
    if scope_branch_id:
        temp_query = temp_query.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
    product_ids = {
        row[0]
        for row in temp_query.with_entities(VentaComisionAsignacion.producto_id)
        .distinct()
        .all()
    }
    if not product_ids:
        return []

    commission_map = {
        row.producto_id: Decimal(str(row.comision_usd or 0))
        for row in db.query(ProductoComision)
        .filter(ProductoComision.producto_id.in_(list(product_ids)))
        .all()
    }
    missing_ids = [pid for pid in product_ids if commission_map.get(pid, Decimal("0")) <= 0]
    if not missing_ids:
        return []

    products = (
        db.query(Producto)
        .filter(Producto.id.in_(missing_ids))
        .order_by(Producto.cod_producto, Producto.descripcion)
        .all()
    )
    return [{"id": p.id, "codigo": p.cod_producto or "-", "descripcion": p.descripcion or "-"} for p in products]


def _build_commission_reports_data(
    db: Session,
    start_date: date,
    end_date: date,
    branch_id: str | None,
    vendedor_id: str | None,
) -> dict:
    query = (
        db.query(
            VentaComisionAsignacion,
            VentaFactura,
            Producto,
            Cliente,
            Vendedor,
            Branch,
            ProductoComision,
        )
        .join(VentaFactura, VentaFactura.id == VentaComisionAsignacion.factura_id, isouter=True)
        .join(Producto, Producto.id == VentaComisionAsignacion.producto_id, isouter=True)
        .join(Cliente, Cliente.id == VentaComisionAsignacion.cliente_id, isouter=True)
        .join(Vendedor, Vendedor.id == VentaComisionAsignacion.vendedor_asignado_id, isouter=True)
        .join(Branch, Branch.id == VentaComisionAsignacion.branch_id, isouter=True)
        .join(ProductoComision, ProductoComision.producto_id == VentaComisionAsignacion.producto_id, isouter=True)
        .filter(
            VentaComisionAsignacion.fecha >= start_date,
            VentaComisionAsignacion.fecha <= end_date,
        )
    )
    if branch_id and branch_id != "all":
        try:
            query = query.filter(VentaComisionAsignacion.branch_id == int(branch_id))
        except ValueError:
            pass
    if vendedor_id:
        try:
            query = query.filter(VentaComisionAsignacion.vendedor_asignado_id == int(vendedor_id))
        except ValueError:
            pass

    rows = query.order_by(
        VentaComisionAsignacion.fecha.asc(),
        Vendedor.nombre.asc(),
        VentaFactura.numero.asc(),
        VentaComisionAsignacion.id.asc(),
    ).all()

    detail_rows: list[dict] = []
    summary_map: dict[str, dict[str, Decimal]] = {}
    pivot_map: dict[date, dict[str, dict[str, Decimal]]] = {}
    pivot_vendors_set: set[str] = set()
    total_bultos = Decimal("0")
    total_comision = Decimal("0")
    total_ventas_usd = Decimal("0")

    for temp_row, factura, producto, cliente, vendedor, branch, producto_comision in rows:
        qty = Decimal(str(temp_row.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        comision_unit = Decimal(str(producto_comision.comision_usd or 0)) if producto_comision else Decimal("0")
        comision_total = comision_unit * qty
        subtotal_usd = Decimal(str(temp_row.subtotal_usd or 0))
        vendor_name = vendedor.nombre if vendedor else "Sin asignar"
        fecha_value = temp_row.fecha

        detail_rows.append(
            {
                "fecha": fecha_value,
                "fecha_label": fecha_value.strftime("%d/%m/%Y") if fecha_value else "-",
                "sucursal": branch.name if branch else "-",
                "factura": factura.numero if factura else "-",
                "cliente": cliente.nombre if cliente else "Consumidor final",
                "producto": f"{(producto.cod_producto if producto else '-') or '-'} - {(producto.descripcion if producto else '-') or '-'}",
                "vendedor": vendor_name,
                "cantidad": int(qty),
                "subtotal_usd": float(subtotal_usd),
                "comision_unit_usd": float(comision_unit),
                "comision_total_usd": float(comision_total),
            }
        )

        total_bultos += qty
        total_comision += comision_total
        total_ventas_usd += subtotal_usd

        if vendor_name not in summary_map:
            summary_map[vendor_name] = {
                "bultos": Decimal("0"),
                "ventas_usd": Decimal("0"),
                "comision_usd": Decimal("0"),
            }
        summary_map[vendor_name]["bultos"] += qty
        summary_map[vendor_name]["ventas_usd"] += subtotal_usd
        summary_map[vendor_name]["comision_usd"] += comision_total

        pivot_vendors_set.add(vendor_name)
        if fecha_value not in pivot_map:
            pivot_map[fecha_value] = {}
        if vendor_name not in pivot_map[fecha_value]:
            pivot_map[fecha_value][vendor_name] = {
                "bultos": Decimal("0"),
                "comision_usd": Decimal("0"),
            }
        pivot_map[fecha_value][vendor_name]["bultos"] += qty
        pivot_map[fecha_value][vendor_name]["comision_usd"] += comision_total

    summary_rows = [
        {
            "vendedor": vendor_name,
            "bultos": int(values["bultos"]),
            "ventas_usd": float(values["ventas_usd"]),
            "comision_usd": float(values["comision_usd"]),
        }
        for vendor_name, values in sorted(
            summary_map.items(),
            key=lambda pair: (pair[0] or "").lower(),
        )
    ]

    pivot_vendors = sorted(list(pivot_vendors_set), key=lambda name: (name or "").lower())
    pivot_rows = []
    pivot_vendor_totals = {
        vendor_name: {"bultos": Decimal("0"), "comision_usd": Decimal("0")}
        for vendor_name in pivot_vendors
    }
    for fecha_value in sorted(pivot_map.keys()):
        by_vendor = pivot_map.get(fecha_value, {})
        cells = []
        day_bultos = Decimal("0")
        day_comision = Decimal("0")
        for vendor_name in pivot_vendors:
            values = by_vendor.get(
                vendor_name, {"bultos": Decimal("0"), "comision_usd": Decimal("0")}
            )
            bultos = Decimal(str(values["bultos"]))
            comision_usd = Decimal(str(values["comision_usd"]))
            cells.append(
                {
                    "vendor": vendor_name,
                    "bultos": int(bultos),
                    "comision_usd": float(comision_usd),
                }
            )
            day_bultos += bultos
            day_comision += comision_usd
            pivot_vendor_totals[vendor_name]["bultos"] += bultos
            pivot_vendor_totals[vendor_name]["comision_usd"] += comision_usd
        pivot_rows.append(
            {
                "fecha": fecha_value,
                "fecha_label": fecha_value.strftime("%d/%m/%Y"),
                "cells": cells,
                "day_bultos": int(day_bultos),
                "day_comision_usd": float(day_comision),
            }
        )

    return {
        "detail_rows": detail_rows,
        "summary_rows": summary_rows,
        "pivot_vendors": pivot_vendors,
        "pivot_rows": pivot_rows,
        "pivot_vendor_totals": {
            vendor_name: {
                "bultos": int(values["bultos"]),
                "comision_usd": float(values["comision_usd"]),
            }
            for vendor_name, values in pivot_vendor_totals.items()
        },
        "total_bultos": int(total_bultos),
        "total_ventas_usd": float(total_ventas_usd),
        "total_comision_usd": float(total_comision),
    }


def _commission_day_status(
    db: Session,
    fecha_value: date,
    branch_id: str | None,
) -> dict:
    scope_branch_id = _commission_branch_scope(branch_id)

    temp_q = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha == fecha_value
    )
    final_q = db.query(VentaComisionFinal).filter(VentaComisionFinal.fecha == fecha_value)
    if scope_branch_id:
        temp_q = temp_q.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
        final_q = final_q.filter(VentaComisionFinal.branch_id == scope_branch_id)

    temp_rows = temp_q.all()
    final_rows = final_q.all()

    if not final_rows:
        return {
            "code": "abierto",
            "label": "Abierto",
            "final_count": 0,
            "temp_count": len(temp_rows),
        }

    def pack_temp(row: VentaComisionAsignacion) -> tuple:
        return (
            int(row.venta_item_id or 0),
            int(row.vendedor_asignado_id or 0),
            int(Decimal(str(row.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
        )

    def pack_final(row: VentaComisionFinal) -> tuple:
        return (
            int(row.venta_item_id or 0),
            int(row.vendedor_asignado_id or 0),
            int(Decimal(str(row.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
        )

    temp_set = sorted(pack_temp(row) for row in temp_rows)
    final_set = sorted(pack_final(row) for row in final_rows)
    in_editing = temp_set != final_set

    if in_editing:
        return {
            "code": "reabierto",
            "label": "Reabierto en edicion",
            "final_count": len(final_rows),
            "temp_count": len(temp_rows),
        }
    return {
        "code": "cerrado",
        "label": "Cerrado",
        "final_count": len(final_rows),
        "temp_count": len(temp_rows),
    }


@router.get("/sales/comisiones")
def sales_comisiones(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    (
        fecha_value,
        start_date,
        end_date,
        branch_id,
        vendedor_facturacion_id,
        vendedor_asignado_id,
        producto_asig_q,
        active_tab,
    ) = _sales_commissions_filters(request)
    rep_start_date, rep_end_date, rep_branch_id, rep_vendedor_id = (
        _sales_commissions_report_filters(request)
    )
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    selected_branch_id: Optional[int] = None
    if branch_id and branch_id != "all":
        try:
            selected_branch_id = int(branch_id)
        except ValueError:
            selected_branch_id = None

    vendedores = (
        db.query(Vendedor)
        .filter(Vendedor.activo.is_(True))
        .order_by(Vendedor.nombre)
        .all()
    )
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .order_by(Producto.descripcion)
        .all()
    )
    product_ids = [p.id for p in productos]
    commission_map = {
        row.producto_id: float(row.comision_usd or 0)
        for row in db.query(ProductoComision)
        .filter(ProductoComision.producto_id.in_(product_ids))
        .all()
    } if product_ids else {}
    product_rows = [
        {
            "id": producto.id,
            "codigo": producto.cod_producto,
            "descripcion": producto.descripcion,
            "costo_producto": float(producto.costo_producto or 0),
            "precio_venta_usd": float(producto.precio_venta1_usd or 0),
            "comision": float(commission_map.get(producto.id, 0) or 0),
        }
        for producto in productos
    ]

    for day_value in _commission_dates_in_range(start_date, end_date):
        _ensure_commission_temp_snapshot(db, day_value, branch_id)
    (
        assignment_rows,
        total_bultos,
        total_rows,
        total_bultos_vendidos,
        by_vendor,
        total_comision_usd,
        total_ventas_usd,
        total_facturas,
        period_summary_rows,
    ) = _build_commission_assignment_rows(
        db,
        start_date,
        end_date,
        branch_id,
        vendedor_facturacion_id,
        vendedor_asignado_id,
        producto_asig_q,
    )
    if start_date == end_date:
        day_status = _commission_day_status(db, start_date, branch_id)
    else:
        day_status = {
            "code": "abierto",
            "label": f"Periodo {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            "final_count": 0,
            "temp_count": total_rows,
        }
    for day_value in _commission_dates_in_range(rep_start_date, rep_end_date):
        _ensure_commission_temp_snapshot(db, day_value, rep_branch_id)
    reports_data = _build_commission_reports_data(
        db,
        rep_start_date,
        rep_end_date,
        rep_branch_id,
        rep_vendedor_id,
    )
    success = request.query_params.get("success")
    error = request.query_params.get("error")

    return request.app.state.templates.TemplateResponse(
        "sales_comisiones.html",
        {
            "request": request,
            "user": user,
            "branches": branches,
            "vendedores": vendedores,
            "product_rows": product_rows,
            "assignment_rows": assignment_rows,
            "total_bultos": total_bultos,
            "total_rows": total_rows,
            "total_bultos_vendidos": total_bultos_vendidos,
            "total_ventas_usd": total_ventas_usd,
            "total_facturas": total_facturas,
            "total_comision_usd": total_comision_usd,
            "by_vendor": by_vendor,
            "period_summary_rows": period_summary_rows,
            "day_status": day_status,
            "fecha": fecha_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "selected_branch": branch_id or "all",
            "selected_vendedor_facturacion": vendedor_facturacion_id or "",
            "selected_vendedor_asignado": vendedor_asignado_id or "",
            "producto_asig_q": producto_asig_q,
            "rep_start_date": rep_start_date.isoformat(),
            "rep_end_date": rep_end_date.isoformat(),
            "rep_selected_branch": rep_branch_id or "all",
            "rep_selected_vendedor": rep_vendedor_id or "",
            "rep_detail_rows": reports_data["detail_rows"],
            "rep_summary_rows": reports_data["summary_rows"],
            "rep_pivot_vendors": reports_data["pivot_vendors"],
            "rep_pivot_rows": reports_data["pivot_rows"],
            "rep_pivot_vendor_totals": reports_data["pivot_vendor_totals"],
            "rep_total_bultos": reports_data["total_bultos"],
            "rep_total_ventas_usd": reports_data["total_ventas_usd"],
            "rep_total_comision_usd": reports_data["total_comision_usd"],
            "active_tab": active_tab,
            "success": success,
            "error": error,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/sales/comisiones/precios")
async def sales_comisiones_save_prices(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    form = await request.form()
    fecha_raw = str(form.get("fecha") or "")
    start_raw = str(form.get("start_date") or "")
    end_raw = str(form.get("end_date") or "")
    branch_id = str(form.get("branch_id") or "all")
    vendedor_facturacion_id = str(form.get("vendedor_facturacion_id") or "").strip()
    vendedor_asignado_id = str(form.get("vendedor_asignado_id") or "").strip()
    producto_asig_q = str(form.get("producto_asig_q") or "").strip()
    updates = 0

    def parse_amount(raw: Optional[str]) -> Decimal:
        val = str(raw or "").strip()
        if not val:
            return Decimal("0")
        val = val.replace(",", "")
        try:
            return Decimal(val)
        except Exception:
            return Decimal("0")

    for key, value in form.items():
        if not key.startswith("comision_"):
            continue
        try:
            product_id = int(key.replace("comision_", ""))
        except ValueError:
            continue
        comision = parse_amount(value)
        row = (
            db.query(ProductoComision)
            .filter(ProductoComision.producto_id == product_id)
            .first()
        )
        if row:
            row.comision_usd = comision
            row.usuario_registro = user.full_name
        else:
            db.add(
                ProductoComision(
                    producto_id=product_id,
                    comision_usd=comision,
                    usuario_registro=user.full_name,
                )
            )
        updates += 1

    db.commit()
    msg = (
        f"Comisiones registradas ({updates})"
        if updates
        else "No se detectaron cambios para guardar"
    )
    params = {
        "tab": "precios",
        "success": msg,
        "fecha": fecha_raw,
        "start_date": start_raw,
        "end_date": end_raw,
        "branch_id": branch_id or "all",
        "vendedor_facturacion_id": vendedor_facturacion_id,
        "vendedor_asignado_id": vendedor_asignado_id,
        "producto_asig_q": producto_asig_q,
    }
    return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)


@router.post("/sales/comisiones/asignaciones")
async def sales_comisiones_save_assignments(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    form = await request.form()
    fecha_raw = str(form.get("fecha") or "")
    start_raw = str(form.get("start_date") or "")
    end_raw = str(form.get("end_date") or "")
    branch_id = str(form.get("branch_id") or "all")
    vendedor_facturacion_id = str(form.get("vendedor_facturacion_id") or "").strip()
    vendedor_asignado_id = str(form.get("vendedor_asignado_id") or "").strip()
    producto_asig_q = str(form.get("producto_asig_q") or "").strip()
    payload_raw = str(form.get("rows_payload") or "[]")

    try:
        fecha_value = date.fromisoformat(fecha_raw)
    except ValueError:
        fecha_value = local_today()
    start_date = fecha_value
    end_date = fecha_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
            if start_raw and not end_raw:
                end_date = start_date
            if end_raw and not start_raw:
                start_date = end_date
        except ValueError:
            start_date = fecha_value
            end_date = fecha_value
    if end_date < start_date:
        end_date = start_date
    try:
        payload = json.loads(payload_raw)
    except json.JSONDecodeError:
        payload = []

    for day_value in _commission_dates_in_range(start_date, end_date):
        _ensure_commission_temp_snapshot(db, day_value, branch_id)
    scope_branch_id = _commission_branch_scope(branch_id)
    source_rows = _commission_sales_rows_query_range(
        db,
        start_date,
        end_date,
        branch_id if branch_id else "all",
        None,
        "",
    ).all()
    source_map = {
        item.id: (factura, item, producto, cliente, vendedor, branch, bodega)
        for factura, item, producto, cliente, vendedor, branch, bodega in source_rows
    }
    source_qty_map = {
        item.id: int(
            Decimal(str(item.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        )
        for _, item, *_ in source_rows
    }

    payload_temp_ids: set[int] = set()
    payload_item_ids: set[int] = set()
    for row in payload if isinstance(payload, list) else []:
        try:
            temp_id = int(row.get("temp_id"))
            venta_item_id = int(row.get("venta_item_id"))
            if temp_id > 0:
                payload_temp_ids.add(temp_id)
            payload_item_ids.add(venta_item_id)
        except Exception:
            continue

    saved = 0
    current_query = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha >= start_date,
        VentaComisionAsignacion.fecha <= end_date,
    )
    if payload_item_ids:
        current_query = current_query.filter(
            VentaComisionAsignacion.venta_item_id.in_(list(payload_item_ids))
        )
    else:
        current_query = current_query.filter(VentaComisionAsignacion.id == -1)
    if scope_branch_id:
        current_query = current_query.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
    current_rows = current_query.all()
    current_by_temp_id: dict[int, VentaComisionAsignacion] = {int(r.id): r for r in current_rows}

    # Solo trabajamos sobre items tocados en payload para no afectar otros filtros/registros.
    updates_by_item: dict[int, list[dict]] = {}
    payload_rows = payload if isinstance(payload, list) else []
    for row in payload_rows:
        try:
            temp_id = int(row.get("temp_id"))
            venta_item_id = int(row.get("venta_item_id"))
            row_vendedor_asignado_id = int(row.get("vendedor_id"))
            cantidad = int(Decimal(str(row.get("cantidad") or "0")))
        except Exception:
            continue
        if temp_id > 0:
            current_row = current_by_temp_id.get(temp_id)
            if not current_row:
                continue
            if int(current_row.venta_item_id or 0) != venta_item_id:
                continue
        if venta_item_id not in source_map:
            continue
        if cantidad < 0:
            continue
        updates_by_item.setdefault(venta_item_id, []).append(
            {
                "temp_id": temp_id,
                "vendedor_id": row_vendedor_asignado_id,
                "cantidad": cantidad,
            }
        )

    if not updates_by_item:
        params = {
            "tab": "asignacion",
            "fecha": fecha_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "branch_id": branch_id or "all",
            "vendedor_facturacion_id": vendedor_facturacion_id,
            "vendedor_asignado_id": vendedor_asignado_id,
            "producto_asig_q": producto_asig_q,
            "error": "No hay filas validas para guardar asignaciones.",
        }
        return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)

    invalid_items: list[int] = []
    touched_item_ids = list(updates_by_item.keys())

    # Incluye filas existentes no enviadas (por ejemplo, ocultas por paginado/filtro) para
    # validar contra el total vendido sin romper el item completo.
    current_by_item: dict[int, list[VentaComisionAsignacion]] = {}
    for row in current_rows:
        if int(row.venta_item_id or 0) in touched_item_ids:
            current_by_item.setdefault(int(row.venta_item_id), []).append(row)

    for venta_item_id, rows in updates_by_item.items():
        sold_qty = source_qty_map.get(venta_item_id)
        if sold_qty is None:
            invalid_items.append(venta_item_id)
            continue

        merged_by_temp_id: dict[int, dict] = {}
        merged_new_rows: list[dict] = []
        for existing in current_by_item.get(venta_item_id, []):
            merged_by_temp_id[int(existing.id)] = {
                "temp_id": int(existing.id),
                "vendedor_id": int(existing.vendedor_asignado_id or 0),
                "cantidad": int(
                    Decimal(str(existing.cantidad or 0)).quantize(
                        Decimal("1"), rounding=ROUND_HALF_UP
                    )
                ),
            }
        for incoming in rows:
            incoming_temp_id = int(incoming["temp_id"])
            if incoming_temp_id > 0:
                merged_by_temp_id[incoming_temp_id] = incoming
            else:
                merged_new_rows.append(incoming)

        merged_rows = list(merged_by_temp_id.values()) + merged_new_rows
        total_payload_qty = sum(int(r["cantidad"]) for r in merged_rows)
        positive_rows = [r for r in merged_rows if int(r["cantidad"]) > 0]
        if total_payload_qty != sold_qty or not positive_rows:
            invalid_items.append(venta_item_id)
            continue

        for r in rows:
            exists = (
                db.query(Vendedor.id)
                .filter(Vendedor.id == int(r["vendedor_id"]), Vendedor.activo.is_(True))
                .first()
            )
            if not exists:
                invalid_items.append(venta_item_id)
                break
    if invalid_items:
        params = {
            "tab": "asignacion",
            "fecha": fecha_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "branch_id": branch_id or "all",
            "vendedor_facturacion_id": vendedor_facturacion_id,
            "vendedor_asignado_id": vendedor_asignado_id,
            "producto_asig_q": producto_asig_q,
            "error": "Hay items con cantidades invalidas. Verifica que la suma por item sea igual a la cantidad vendida.",
        }
        return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)

    for venta_item_id, incoming_rows in updates_by_item.items():
        source = source_map.get(venta_item_id)
        if not source:
            continue
        _, item, _, _, _, _, _ = source
        precio_usd = Decimal(str(item.precio_unitario_usd or 0))
        precio_cs = Decimal(str(item.precio_unitario_cs or 0))

        existing_rows = current_by_item.get(venta_item_id, [])
        incoming_by_temp = {
            int(r["temp_id"]): r for r in incoming_rows if int(r.get("temp_id") or 0) > 0
        }
        incoming_new_rows = [
            r for r in incoming_rows if int(r.get("temp_id") or 0) <= 0 and int(r.get("cantidad") or 0) > 0
        ]

        for existing in existing_rows:
            incoming = incoming_by_temp.get(int(existing.id))
            if not incoming:
                continue
            cantidad_dec = Decimal(str(incoming["cantidad"])).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
            existing.vendedor_asignado_id = int(incoming["vendedor_id"])
            existing.cantidad = cantidad_dec
            existing.precio_unitario_usd = precio_usd
            existing.precio_unitario_cs = precio_cs
            existing.subtotal_usd = precio_usd * cantidad_dec
            existing.subtotal_cs = precio_cs * cantidad_dec
            existing.usuario_registro = user.full_name
            saved += 1

        source_factura, _source_item, source_producto, source_cliente, source_vendedor, source_branch, source_bodega = source
        for incoming in incoming_new_rows:
            cantidad_dec = Decimal(str(incoming["cantidad"])).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
            if cantidad_dec <= 0:
                continue
            db.add(
                VentaComisionAsignacion(
                    venta_item_id=venta_item_id,
                    factura_id=source_factura.id if source_factura else None,
                    branch_id=source_branch.id if source_branch else None,
                    bodega_id=source_bodega.id if source_bodega else None,
                    cliente_id=source_cliente.id if source_cliente else None,
                    producto_id=source_producto.id if source_producto else None,
                    fecha=source_factura.fecha.date() if source_factura and source_factura.fecha else start_date,
                    vendedor_origen_id=source_vendedor.id if source_vendedor else None,
                    vendedor_asignado_id=int(incoming["vendedor_id"]),
                    cantidad=cantidad_dec,
                    precio_unitario_usd=precio_usd,
                    precio_unitario_cs=precio_cs,
                    subtotal_usd=precio_usd * cantidad_dec,
                    subtotal_cs=precio_cs * cantidad_dec,
                    usuario_registro=user.full_name,
                )
            )
            saved += 1

        # Limpieza: elimina filas en cero cuando existe al menos una fila positiva para el item.
        # Evita dejar filas "primarias" bloqueadas en 0 despues de repartir a otros vendedores.
        existing_rows_refreshed = (
            db.query(VentaComisionAsignacion)
            .filter(
                VentaComisionAsignacion.venta_item_id == venta_item_id,
            )
            .all()
        )
        positive_count = 0
        for row in existing_rows_refreshed:
            qty_int = int(
                Decimal(str(row.cantidad or 0)).quantize(
                    Decimal("1"), rounding=ROUND_HALF_UP
                )
            )
            if qty_int > 0:
                positive_count += 1
        if positive_count > 0:
            for row in existing_rows_refreshed:
                qty_int = int(
                    Decimal(str(row.cantidad or 0)).quantize(
                        Decimal("1"), rounding=ROUND_HALF_UP
                    )
                )
                if qty_int <= 0:
                    db.delete(row)

        # Si por alguna razon no llego una fila existente (payload incompleto), no la tocamos.
        # Evita perder datos por filtros parciales del frontend.

    db.commit()
    params = {
        "tab": "asignacion",
        "fecha": fecha_value.isoformat(),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "branch_id": branch_id or "all",
        "vendedor_facturacion_id": vendedor_facturacion_id,
        "vendedor_asignado_id": vendedor_asignado_id,
        "producto_asig_q": producto_asig_q,
        "success": f"Asignaciones guardadas ({saved})",
    }
    return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)


@router.post("/sales/comisiones/asignaciones/validar-precios")
async def sales_comisiones_validate_prices(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    form = await request.form()
    fecha_raw = str(form.get("fecha") or "")
    start_raw = str(form.get("start_date") or "")
    end_raw = str(form.get("end_date") or "")
    branch_id = str(form.get("branch_id") or "all")
    vendedor_facturacion_id = str(form.get("vendedor_facturacion_id") or "").strip()
    vendedor_asignado_id = str(form.get("vendedor_asignado_id") or "").strip()
    producto_asig_q = str(form.get("producto_asig_q") or "").strip()

    try:
        fecha_value = date.fromisoformat(fecha_raw)
    except ValueError:
        fecha_value = local_today()
    start_date = fecha_value
    end_date = fecha_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
            if start_raw and not end_raw:
                end_date = start_date
            if end_raw and not start_raw:
                start_date = end_date
        except ValueError:
            start_date = fecha_value
            end_date = fecha_value
    if end_date < start_date:
        end_date = start_date

    for day_value in _commission_dates_in_range(start_date, end_date):
        _ensure_commission_temp_snapshot(db, day_value, branch_id)
    missing_products = _commission_missing_prices(db, start_date, end_date, branch_id)

    if missing_products:
        preview = ", ".join(
            f"{row['codigo']} ({row['descripcion']})" for row in missing_products[:8]
        )
        if len(missing_products) > 8:
            preview += f" ... +{len(missing_products) - 8} mas"
        params = {
            "tab": "asignacion",
            "fecha": fecha_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "branch_id": branch_id or "all",
            "vendedor_facturacion_id": vendedor_facturacion_id,
            "vendedor_asignado_id": vendedor_asignado_id,
            "producto_asig_q": producto_asig_q,
            "error": f"Productos sin precio de comision: {preview}",
        }
        return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)

    params = {
        "tab": "asignacion",
        "fecha": fecha_value.isoformat(),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "branch_id": branch_id or "all",
        "vendedor_facturacion_id": vendedor_facturacion_id,
        "vendedor_asignado_id": vendedor_asignado_id,
        "producto_asig_q": producto_asig_q,
        "success": "Validacion completada. Todos los productos tienen precio de comision.",
    }
    return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)


@router.post("/sales/comisiones/asignaciones/regenerar")
async def sales_comisiones_regenerate_assignments(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    form = await request.form()
    fecha_raw = str(form.get("fecha") or "")
    start_raw = str(form.get("start_date") or "")
    end_raw = str(form.get("end_date") or "")
    branch_id = str(form.get("branch_id") or "all")
    vendedor_facturacion_id = str(form.get("vendedor_facturacion_id") or "").strip()
    vendedor_asignado_id = str(form.get("vendedor_asignado_id") or "").strip()
    producto_asig_q = str(form.get("producto_asig_q") or "").strip()

    try:
        fecha_value = date.fromisoformat(fecha_raw)
    except ValueError:
        fecha_value = local_today()
    start_date = fecha_value
    end_date = fecha_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
            if start_raw and not end_raw:
                end_date = start_date
            if end_raw and not start_raw:
                start_date = end_date
        except ValueError:
            start_date = fecha_value
            end_date = fecha_value
    if end_date < start_date:
        end_date = start_date

    scope_branch_id = _commission_branch_scope(branch_id)
    temp_query = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha == fecha_value
    )
    if scope_branch_id:
        temp_query = temp_query.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
    deleted = 0
    deleted = temp_query.delete(synchronize_session=False)
    db.commit()
    created, _removed = _ensure_commission_temp_snapshot(db, fecha_value, branch_id)

    params = {
        "tab": "asignacion",
        "fecha": fecha_value.isoformat(),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "branch_id": branch_id or "all",
        "vendedor_facturacion_id": vendedor_facturacion_id,
        "vendedor_asignado_id": vendedor_asignado_id,
        "producto_asig_q": producto_asig_q,
        "success": f"Regenerado completado. Temporal reiniciado: {deleted} eliminadas, {created} recreadas.",
    }
    return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)


@router.post("/sales/comisiones/asignaciones/finalizar")
async def sales_comisiones_finalize_day(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    form = await request.form()
    fecha_raw = str(form.get("fecha") or "")
    start_raw = str(form.get("start_date") or "")
    end_raw = str(form.get("end_date") or "")
    branch_id = str(form.get("branch_id") or "all")
    vendedor_facturacion_id = str(form.get("vendedor_facturacion_id") or "").strip()
    vendedor_asignado_id = str(form.get("vendedor_asignado_id") or "").strip()
    producto_asig_q = str(form.get("producto_asig_q") or "").strip()

    try:
        fecha_value = date.fromisoformat(fecha_raw)
    except ValueError:
        fecha_value = local_today()
    start_date = fecha_value
    end_date = fecha_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
            if start_raw and not end_raw:
                end_date = start_date
            if end_raw and not start_raw:
                start_date = end_date
        except ValueError:
            start_date = fecha_value
            end_date = fecha_value
    if end_date < start_date:
        end_date = start_date

    _ensure_commission_temp_snapshot(db, fecha_value, branch_id)
    scope_branch_id = _commission_branch_scope(branch_id)
    temp_query = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha == fecha_value
    )
    if scope_branch_id:
        temp_query = temp_query.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
    temp_rows = temp_query.all()
    if not temp_rows:
        params = {
            "tab": "asignacion",
            "fecha": fecha_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "branch_id": branch_id or "all",
            "vendedor_facturacion_id": vendedor_facturacion_id,
            "vendedor_asignado_id": vendedor_asignado_id,
            "producto_asig_q": producto_asig_q,
            "error": "No hay datos temporales para finalizar.",
        }
        return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)

    product_ids = list({row.producto_id for row in temp_rows})
    commission_map = {
        row.producto_id: Decimal(str(row.comision_usd or 0))
        for row in db.query(ProductoComision)
        .filter(ProductoComision.producto_id.in_(product_ids))
        .all()
    } if product_ids else {}

    final_query = db.query(VentaComisionFinal).filter(VentaComisionFinal.fecha == fecha_value)
    if scope_branch_id:
        final_query = final_query.filter(VentaComisionFinal.branch_id == scope_branch_id)
    replaced = final_query.delete(synchronize_session=False)

    inserted = 0
    for row in temp_rows:
        qty = Decimal(str(row.cantidad or 0)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        comision_unit = commission_map.get(row.producto_id, Decimal("0"))
        comision_total = comision_unit * qty
        db.add(
            VentaComisionFinal(
                fecha=row.fecha,
                branch_id=row.branch_id,
                bodega_id=row.bodega_id,
                factura_id=row.factura_id,
                venta_item_id=row.venta_item_id,
                cliente_id=row.cliente_id,
                producto_id=row.producto_id,
                vendedor_origen_id=row.vendedor_origen_id,
                vendedor_asignado_id=row.vendedor_asignado_id,
                cantidad=qty,
                precio_unitario_usd=row.precio_unitario_usd,
                precio_unitario_cs=row.precio_unitario_cs,
                subtotal_usd=row.subtotal_usd,
                subtotal_cs=row.subtotal_cs,
                comision_unit_usd=comision_unit,
                comision_total_usd=comision_total,
                usuario_registro=user.full_name,
            )
        )
        inserted += 1
    db.commit()

    params = {
        "tab": "asignacion",
        "fecha": fecha_value.isoformat(),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "branch_id": branch_id or "all",
        "vendedor_facturacion_id": vendedor_facturacion_id,
        "vendedor_asignado_id": vendedor_asignado_id,
        "producto_asig_q": producto_asig_q,
        "success": f"Comisiones finales actualizadas: {inserted} filas (reemplazadas {replaced}).",
    }
    return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)


@router.post("/sales/comisiones/asignaciones/reabrir")
async def sales_comisiones_reopen_day(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    form = await request.form()
    fecha_raw = str(form.get("fecha") or "")
    start_raw = str(form.get("start_date") or "")
    end_raw = str(form.get("end_date") or "")
    branch_id = str(form.get("branch_id") or "all")
    vendedor_facturacion_id = str(form.get("vendedor_facturacion_id") or "").strip()
    vendedor_asignado_id = str(form.get("vendedor_asignado_id") or "").strip()
    producto_asig_q = str(form.get("producto_asig_q") or "").strip()

    try:
        fecha_value = date.fromisoformat(fecha_raw)
    except ValueError:
        fecha_value = local_today()
    start_date = fecha_value
    end_date = fecha_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
            if start_raw and not end_raw:
                end_date = start_date
            if end_raw and not start_raw:
                start_date = end_date
        except ValueError:
            start_date = fecha_value
            end_date = fecha_value
    if end_date < start_date:
        end_date = start_date

    scope_branch_id = _commission_branch_scope(branch_id)
    final_query = db.query(VentaComisionFinal).filter(VentaComisionFinal.fecha == fecha_value)
    if scope_branch_id:
        final_query = final_query.filter(VentaComisionFinal.branch_id == scope_branch_id)
    final_rows = final_query.all()
    if not final_rows:
        params = {
            "tab": "asignacion",
            "fecha": fecha_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "branch_id": branch_id or "all",
            "vendedor_facturacion_id": vendedor_facturacion_id,
            "vendedor_asignado_id": vendedor_asignado_id,
            "producto_asig_q": producto_asig_q,
            "error": "No hay cierre final para reabrir en ese dia/sucursal.",
        }
        return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)

    temp_query = db.query(VentaComisionAsignacion).filter(
        VentaComisionAsignacion.fecha == fecha_value
    )
    if scope_branch_id:
        temp_query = temp_query.filter(VentaComisionAsignacion.branch_id == scope_branch_id)
    temp_query.delete(synchronize_session=False)

    recreated = 0
    for row in final_rows:
        db.add(
            VentaComisionAsignacion(
                venta_item_id=row.venta_item_id,
                factura_id=row.factura_id,
                branch_id=row.branch_id,
                bodega_id=row.bodega_id,
                cliente_id=row.cliente_id,
                producto_id=row.producto_id,
                fecha=row.fecha,
                vendedor_origen_id=row.vendedor_origen_id,
                vendedor_asignado_id=row.vendedor_asignado_id,
                cantidad=row.cantidad,
                precio_unitario_usd=row.precio_unitario_usd,
                precio_unitario_cs=row.precio_unitario_cs,
                subtotal_usd=row.subtotal_usd,
                subtotal_cs=row.subtotal_cs,
                usuario_registro=user.full_name,
            )
        )
        recreated += 1
    db.commit()

    params = {
        "tab": "asignacion",
        "fecha": fecha_value.isoformat(),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "branch_id": branch_id or "all",
        "vendedor_facturacion_id": vendedor_facturacion_id,
        "vendedor_asignado_id": vendedor_asignado_id,
        "producto_asig_q": producto_asig_q,
        "success": f"Cierre reabierto en temporal: {recreated} filas restauradas.",
    }
    return RedirectResponse("/sales/comisiones?" + urlencode(params), status_code=303)


@router.get("/sales/comisiones/reportes/pdf")
def sales_comisiones_reports_pdf(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    rep_start_date, rep_end_date, rep_branch_id, rep_vendedor_id = (
        _sales_commissions_report_filters(request)
    )
    for day_value in _commission_dates_in_range(rep_start_date, rep_end_date):
        _ensure_commission_temp_snapshot(db, day_value, rep_branch_id)
    reports_data = _build_commission_reports_data(
        db, rep_start_date, rep_end_date, rep_branch_id, rep_vendedor_id
    )
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    selected_branch = None
    if rep_branch_id and rep_branch_id != "all":
        try:
            selected_branch = next((b for b in branches if b.id == int(rep_branch_id)), None)
        except ValueError:
            selected_branch = None

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape

    buffer = io.BytesIO()
    page_size = landscape(A4)
    page_w, page_h = page_size
    margin = 24
    content_w = page_w - (margin * 2)
    c = canvas.Canvas(buffer, pagesize=page_size)
    y = page_h - 24

    def draw_header() -> float:
        header_y = page_h - 24
        c.setFont("Helvetica-Bold", 14)
        c.drawString(margin, header_y, "Sabana de comisiones por vendedor")
        c.setFont("Helvetica", 9)
        branch_label = selected_branch.name if selected_branch else "Todas las sucursales"
        c.drawString(
            margin,
            header_y - 15,
            f"Rango: {rep_start_date.strftime('%d/%m/%Y')} - {rep_end_date.strftime('%d/%m/%Y')} | Sucursal: {branch_label}",
        )
        c.drawString(
            margin,
            header_y - 28,
            f"Vendedor filtro: {rep_vendedor_id or 'Todos'} | Total bultos: {reports_data['total_bultos']} | Total comision USD: ${reports_data['total_comision_usd']:,.2f}",
        )
        c.line(margin, header_y - 34, page_w - margin, header_y - 34)
        return header_y - 48

    def short_vendor(name: str) -> str:
        clean = (name or "-").strip()
        if len(clean) <= 14:
            return clean
        return clean[:12] + ".."

    y = draw_header()
    pivot_vendors = reports_data["pivot_vendors"]
    pivot_rows = reports_data["pivot_rows"]
    pivot_vendor_totals = reports_data["pivot_vendor_totals"]

    if not pivot_rows or not pivot_vendors:
        c.setFont("Helvetica", 11)
        c.drawString(margin, y, "Sin datos de comisiones para el rango seleccionado.")
        c.showPage()
        c.save()
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "inline; filename=comisiones_sabana.pdf"},
        )

    date_col_w = 64
    pair_col_w = 76
    total_cols_w = 110
    available_for_pairs = max(1, content_w - date_col_w - total_cols_w)
    max_vendors_per_page = max(1, int(available_for_pairs // pair_col_w))
    vendor_chunks = [
        pivot_vendors[i : i + max_vendors_per_page]
        for i in range(0, len(pivot_vendors), max_vendors_per_page)
    ]

    for chunk_idx, vendors_chunk in enumerate(vendor_chunks):
        if chunk_idx > 0:
            c.showPage()
            y = draw_header()

        col_x = [margin]
        col_x.append(col_x[-1] + date_col_w)
        for _ in vendors_chunk:
            col_x.append(col_x[-1] + pair_col_w / 2)
            col_x.append(col_x[-1] + pair_col_w / 2)
        col_x.append(col_x[-1] + total_cols_w / 2)
        col_x.append(col_x[-1] + total_cols_w / 2)

        c.setFont("Helvetica-Bold", 8)
        c.drawString(col_x[0] + 2, y, "Fecha")
        header_cursor = 1
        for vendor_name in vendors_chunk:
            title = short_vendor(vendor_name)
            c.drawString(col_x[header_cursor] + 2, y, title)
            c.drawRightString(col_x[header_cursor + 1] - 2, y - 10, "Bul")
            c.drawRightString(col_x[header_cursor + 2] - 2, y - 10, "USD")
            header_cursor += 2
        c.drawRightString(col_x[-2] - 2, y, "Tot Bul")
        c.drawRightString(col_x[-1] - 2, y, "Tot USD")
        y -= 18
        c.line(margin, y + 4, page_w - margin, y + 4)

        c.setFont("Helvetica", 8)
        for row in pivot_rows:
            if y < 50:
                c.showPage()
                y = draw_header()
                c.setFont("Helvetica-Bold", 8)
                c.drawString(col_x[0] + 2, y, "Fecha")
                header_cursor = 1
                for vendor_name in vendors_chunk:
                    title = short_vendor(vendor_name)
                    c.drawString(col_x[header_cursor] + 2, y, title)
                    c.drawRightString(col_x[header_cursor + 1] - 2, y - 10, "Bul")
                    c.drawRightString(col_x[header_cursor + 2] - 2, y - 10, "USD")
                    header_cursor += 2
                c.drawRightString(col_x[-2] - 2, y, "Tot Bul")
                c.drawRightString(col_x[-1] - 2, y, "Tot USD")
                y -= 18
                c.line(margin, y + 4, page_w - margin, y + 4)
                c.setFont("Helvetica", 8)

            c.drawString(col_x[0] + 2, y, row["fecha_label"])
            header_cursor = 1
            row_by_vendor = {
                cell["vendor"]: cell for cell in row.get("cells", [])
            }
            for vendor_name in vendors_chunk:
                cell = row_by_vendor.get(
                    vendor_name, {"bultos": 0, "comision_usd": 0.0}
                )
                c.drawRightString(col_x[header_cursor + 1] - 2, y, f"{int(cell['bultos'])}")
                c.drawRightString(
                    col_x[header_cursor + 2] - 2,
                    y,
                    f"{float(cell['comision_usd']):,.2f}",
                )
                header_cursor += 2
            c.drawRightString(col_x[-2] - 2, y, f"{int(row['day_bultos'])}")
            c.drawRightString(col_x[-1] - 2, y, f"{float(row['day_comision_usd']):,.2f}")
            y -= 12

        y -= 4
        c.line(margin, y + 4, page_w - margin, y + 4)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(col_x[0] + 2, y, "TOTAL")
        header_cursor = 1
        for vendor_name in vendors_chunk:
            total_cell = pivot_vendor_totals.get(vendor_name, {"bultos": 0, "comision_usd": 0})
            c.drawRightString(col_x[header_cursor + 1] - 2, y, f"{int(total_cell['bultos'])}")
            c.drawRightString(
                col_x[header_cursor + 2] - 2,
                y,
                f"{float(total_cell['comision_usd']):,.2f}",
            )
            header_cursor += 2
        c.drawRightString(col_x[-2] - 2, y, f"{int(reports_data['total_bultos'])}")
        c.drawRightString(
            col_x[-1] - 2,
            y,
            f"{float(reports_data['total_comision_usd']):,.2f}",
        )

    c.showPage()
    c.save()
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=comisiones_sabana.pdf"},
    )


@router.get("/sales/comisiones/reportes/xlsx")
def sales_comisiones_reports_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.comisiones")
    rep_start_date, rep_end_date, rep_branch_id, rep_vendedor_id = (
        _sales_commissions_report_filters(request)
    )
    for day_value in _commission_dates_in_range(rep_start_date, rep_end_date):
        _ensure_commission_temp_snapshot(db, day_value, rep_branch_id)
    reports_data = _build_commission_reports_data(
        db, rep_start_date, rep_end_date, rep_branch_id, rep_vendedor_id
    )

    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    selected_branch = None
    if rep_branch_id and rep_branch_id != "all":
        try:
            selected_branch = next((b for b in branches if b.id == int(rep_branch_id)), None)
        except ValueError:
            selected_branch = None
    branch_label = selected_branch.name if selected_branch else "Todas las sucursales"

    wb = Workbook()

    ws = wb.active
    ws.title = "Sabana"
    ws["A1"] = "Sabana de comisiones por vendedor"
    ws["A2"] = (
        f"Rango: {rep_start_date.strftime('%d/%m/%Y')} - {rep_end_date.strftime('%d/%m/%Y')} | "
        f"Sucursal: {branch_label} | Vendedor: {rep_vendedor_id or 'Todos'}"
    )
    ws["A3"] = (
        f"Total bultos: {int(reports_data['total_bultos'])} | "
        f"Total comision USD: {float(reports_data['total_comision_usd']):,.2f}"
    )
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].font = Font(size=10)
    ws["A3"].font = Font(size=10)

    pivot_vendors = reports_data.get("pivot_vendors", [])
    headers = ["Fecha"]
    for vendor_name in pivot_vendors:
        headers.append(f"{vendor_name} - Bultos")
        headers.append(f"{vendor_name} - Comision USD")
    headers.extend(["Total bultos dia", "Total comision dia USD"])
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in reports_data.get("pivot_rows", []):
        row_values = [row.get("fecha_label", "-")]
        by_vendor = {cell.get("vendor"): cell for cell in row.get("cells", [])}
        for vendor_name in pivot_vendors:
            cell = by_vendor.get(vendor_name, {"bultos": 0, "comision_usd": 0})
            row_values.append(int(cell.get("bultos", 0) or 0))
            row_values.append(float(cell.get("comision_usd", 0) or 0))
        row_values.append(int(row.get("day_bultos", 0) or 0))
        row_values.append(float(row.get("day_comision_usd", 0) or 0))
        ws.append(row_values)

    total_values = ["TOTAL"]
    vendor_totals = reports_data.get("pivot_vendor_totals", {})
    for vendor_name in pivot_vendors:
        totals = vendor_totals.get(vendor_name, {"bultos": 0, "comision_usd": 0})
        total_values.append(int(totals.get("bultos", 0) or 0))
        total_values.append(float(totals.get("comision_usd", 0) or 0))
    total_values.append(int(reports_data.get("total_bultos", 0) or 0))
    total_values.append(float(reports_data.get("total_comision_usd", 0) or 0))
    ws.append(total_values)
    total_row = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    ws_detail = wb.create_sheet("Detalle")
    detail_headers = [
        "Fecha",
        "Sucursal",
        "Vendedor",
        "Factura",
        "Cliente",
        "Producto",
        "Bultos",
        "Venta USD",
        "Comision unit USD",
        "Comision total USD",
    ]
    ws_detail.append(detail_headers)
    for col in range(1, len(detail_headers) + 1):
        cell = ws_detail.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in reports_data.get("detail_rows", []):
        ws_detail.append(
            [
                row.get("fecha_label", "-"),
                row.get("sucursal", "-"),
                row.get("vendedor", "-"),
                row.get("factura", "-"),
                row.get("cliente", "-"),
                row.get("producto", "-"),
                int(row.get("cantidad", 0) or 0),
                float(row.get("subtotal_usd", 0) or 0),
                float(row.get("comision_unit_usd", 0) or 0),
                float(row.get("comision_total_usd", 0) or 0),
            ]
        )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=comisiones_sabana.xlsx"},
    )


@router.get("/sales/roc")
def sales_roc(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.roc")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    print_id = request.query_params.get("print_id")
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    today_value = local_today()
    start_date = today_value
    end_date = today_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today_value
            end_date = today_value
    branch, bodega = _resolve_branch_bodega(db, user)
    rubros = db.query(ReciboRubro).filter(ReciboRubro.activo.is_(True)).order_by(ReciboRubro.nombre).all()
    motivos = db.query(ReciboMotivo).filter(ReciboMotivo.activo.is_(True)).order_by(ReciboMotivo.tipo, ReciboMotivo.nombre).all()
    recibos_query = db.query(ReciboCaja)
    if bodega:
        recibos_query = recibos_query.filter(ReciboCaja.bodega_id == bodega.id)
    recibos_query = recibos_query.filter(ReciboCaja.fecha.between(start_date, end_date))
    recibos = recibos_query.order_by(ReciboCaja.fecha.desc(), ReciboCaja.id.desc()).limit(200).all()
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    return request.app.state.templates.TemplateResponse(
        "sales_roc.html",
        {
            "request": request,
            "user": user,
            "rubros": rubros,
            "motivos": motivos,
            "recibos": recibos,
            "rate_today": rate_today,
            "branch": branch,
            "bodega": bodega,
            "today": today_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "error": error,
            "success": success,
            "print_id": print_id,
            "version": settings.UI_VERSION,
        },
    )

@router.get("/sales/cierre")
def sales_cierre(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.cierre")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    print_id = request.query_params.get("print_id")
    fecha_raw = request.query_params.get("fecha")
    fecha_value = local_today()
    if fecha_raw:
        try:
            fecha_value = date.fromisoformat(str(fecha_raw))
        except ValueError:
            fecha_value = local_today()

    branch, bodega = _resolve_branch_bodega(db, user)
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= fecha_value)
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

    def to_usd(moneda: str, monto_cs: Decimal, monto_usd: Decimal) -> Decimal:
        if moneda == "USD":
            return Decimal(str(monto_usd or 0))
        return (Decimal(str(monto_cs or 0)) / tasa) if tasa else Decimal("0")

    ventas_query = db.query(VentaFactura).filter(func.date(VentaFactura.fecha) == fecha_value)
    if bodega:
        ventas_query = ventas_query.filter(VentaFactura.bodega_id == bodega.id)
    ventas_query = ventas_query.filter(VentaFactura.estado != "ANULADA")
    ventas = ventas_query.all()
    total_ventas_usd = sum(
        to_usd(f.moneda or "CS", f.total_cs or 0, f.total_usd or 0) for f in ventas
    )

    recibos_query = db.query(ReciboCaja).filter(func.date(ReciboCaja.fecha) == fecha_value)
    if bodega:
        recibos_query = recibos_query.filter(ReciboCaja.bodega_id == bodega.id)
    recibos_query = recibos_query.filter(ReciboCaja.afecta_caja.is_(True))
    recibos = recibos_query.all()
    total_ingresos_usd = sum(
        to_usd(r.moneda or "CS", r.monto_cs or 0, r.monto_usd or 0)
        for r in recibos
        if r.tipo == "INGRESO"
    )
    total_egresos_usd = sum(
        to_usd(r.moneda or "CS", r.monto_cs or 0, r.monto_usd or 0)
        for r in recibos
        if r.tipo == "EGRESO"
    )

    depositos_query = db.query(DepositoCliente).filter(func.date(DepositoCliente.fecha) == fecha_value)
    if bodega:
        depositos_query = depositos_query.filter(DepositoCliente.bodega_id == bodega.id)
    depositos = depositos_query.all()
    total_depositos_usd = sum(
        to_usd(d.moneda or "CS", d.monto_cs or 0, d.monto_usd or 0) for d in depositos
    )

    creditos_query = db.query(VentaFactura).filter(
        func.date(VentaFactura.fecha) == fecha_value,
        VentaFactura.estado != "ANULADA",
        VentaFactura.estado_cobranza == "PENDIENTE",
    )
    if bodega:
        creditos_query = creditos_query.filter(VentaFactura.bodega_id == bodega.id)
    creditos = creditos_query.all()
    total_creditos_usd = Decimal("0")
    for factura in creditos:
        if (factura.moneda or "CS") == "USD":
            paid_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos)
            due_usd = Decimal(str(factura.total_usd or 0))
            saldo_usd = max(due_usd - paid_usd, Decimal("0"))
            if saldo_usd > 0:
                total_creditos_usd += saldo_usd
        else:
            paid_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos)
            due_cs = Decimal(str(factura.total_cs or 0))
            saldo_cs = max(due_cs - paid_cs, Decimal("0"))
            if saldo_cs > 0:
                total_creditos_usd += (saldo_cs / tasa) if tasa else Decimal("0")

    total_calculado_usd = (
        Decimal(str(total_ventas_usd))
        - Decimal(str(total_egresos_usd))
        + Decimal(str(total_ingresos_usd))
        - Decimal(str(total_depositos_usd))
        - Decimal(str(total_creditos_usd))
    )

    denominaciones_cs = [0.5, 1, 5, 10, 20, 50, 100, 200, 500, 1000]
    denominaciones_usd = [1, 2, 5, 10, 20, 50, 100]

    return request.app.state.templates.TemplateResponse(
        "sales_cierre.html",
        {
            "request": request,
            "user": user,
            "branch": branch,
            "bodega": bodega,
            "fecha": fecha_value.isoformat(),
            "rate_today": rate_today,
            "tasa": float(tasa) if tasa else 0,
            "total_ventas_usd": float(total_ventas_usd),
            "total_ingresos_usd": float(total_ingresos_usd),
            "total_egresos_usd": float(total_egresos_usd),
            "total_depositos_usd": float(total_depositos_usd),
            "total_creditos_usd": float(total_creditos_usd),
            "total_calculado_usd": float(total_calculado_usd),
            "denominaciones_cs": denominaciones_cs,
            "denominaciones_usd": denominaciones_usd,
            "error": error,
            "success": success,
            "print_id": print_id,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/sales/cierre")
async def sales_cierre_create(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.cierre")
    form = await request.form()
    fecha_raw = form.get("fecha")
    fecha_value = local_today()
    if fecha_raw:
        try:
            fecha_value = date.fromisoformat(str(fecha_raw))
        except ValueError:
            fecha_value = local_today()

    branch, bodega = _resolve_branch_bodega(db, user)
    if not branch or not bodega:
        return RedirectResponse("/sales/cierre?error=Usuario+sin+sucursal+o+bodega", status_code=303)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= fecha_value)
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

    detalle_cs = {}
    detalle_usd = {}
    total_cs = Decimal("0")
    total_usd = Decimal("0")
    def parse_qty(val: str) -> Decimal:
        raw = re.sub(r"[^0-9.,-]", "", str(val or "0"))
        if "," in raw and "." in raw:
            if raw.rfind(",") > raw.rfind("."):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "," in raw and "." not in raw:
            raw = raw.replace(",", "")
        try:
            return Decimal(raw or "0")
        except Exception:
            return Decimal("0")

    for key, value in form.items():
        if key.startswith("cs_"):
            denom = Decimal(key.replace("cs_", "").replace("_", "."))
            qty = parse_qty(str(value))
            detalle_cs[str(denom)] = float(qty)
            total_cs += denom * qty
        if key.startswith("usd_"):
            denom = Decimal(key.replace("usd_", "").replace("_", "."))
            qty = parse_qty(str(value))
            detalle_usd[str(denom)] = float(qty)
            total_usd += denom * qty

    total_usd_equiv = total_usd + (total_cs / tasa if tasa else Decimal("0"))

    def to_usd(moneda: str, monto_cs: Decimal, monto_usd: Decimal) -> Decimal:
        if moneda == "USD":
            return Decimal(str(monto_usd or 0))
        return (Decimal(str(monto_cs or 0)) / tasa) if tasa else Decimal("0")

    ventas_query = db.query(VentaFactura).filter(func.date(VentaFactura.fecha) == fecha_value)
    ventas_query = ventas_query.filter(VentaFactura.estado != "ANULADA")
    if bodega:
        ventas_query = ventas_query.filter(VentaFactura.bodega_id == bodega.id)
    ventas = ventas_query.all()
    total_ventas_usd = sum(
        to_usd(f.moneda or "CS", f.total_cs or 0, f.total_usd or 0) for f in ventas
    )

    recibos_query = db.query(ReciboCaja).filter(func.date(ReciboCaja.fecha) == fecha_value)
    if bodega:
        recibos_query = recibos_query.filter(ReciboCaja.bodega_id == bodega.id)
    recibos_query = recibos_query.filter(ReciboCaja.afecta_caja.is_(True))
    recibos = recibos_query.all()
    total_ingresos_usd = sum(
        to_usd(r.moneda or "CS", r.monto_cs or 0, r.monto_usd or 0)
        for r in recibos
        if r.tipo == "INGRESO"
    )
    total_egresos_usd = sum(
        to_usd(r.moneda or "CS", r.monto_cs or 0, r.monto_usd or 0)
        for r in recibos
        if r.tipo == "EGRESO"
    )

    depositos_query = db.query(DepositoCliente).filter(func.date(DepositoCliente.fecha) == fecha_value)
    if bodega:
        depositos_query = depositos_query.filter(DepositoCliente.bodega_id == bodega.id)
    depositos = depositos_query.all()
    total_depositos_usd = sum(
        to_usd(d.moneda or "CS", d.monto_cs or 0, d.monto_usd or 0) for d in depositos
    )

    creditos_query = db.query(VentaFactura).filter(
        func.date(VentaFactura.fecha) == fecha_value,
        VentaFactura.estado != "ANULADA",
        VentaFactura.estado_cobranza == "PENDIENTE",
    )
    if bodega:
        creditos_query = creditos_query.filter(VentaFactura.bodega_id == bodega.id)
    creditos = creditos_query.all()
    total_creditos_usd = Decimal("0")
    for factura in creditos:
        if (factura.moneda or "CS") == "USD":
            paid_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos)
            due_usd = Decimal(str(factura.total_usd or 0))
            saldo_usd = max(due_usd - paid_usd, Decimal("0"))
            if saldo_usd > 0:
                total_creditos_usd += saldo_usd
        else:
            paid_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos)
            due_cs = Decimal(str(factura.total_cs or 0))
            saldo_cs = max(due_cs - paid_cs, Decimal("0"))
            if saldo_cs > 0:
                total_creditos_usd += (saldo_cs / tasa) if tasa else Decimal("0")

    total_calculado_usd = (
        Decimal(str(total_ventas_usd))
        - Decimal(str(total_egresos_usd))
        + Decimal(str(total_ingresos_usd))
        - Decimal(str(total_depositos_usd))
        - Decimal(str(total_creditos_usd))
    )
    diferencia = total_usd_equiv - total_calculado_usd

    cierre = CierreCaja(
        branch_id=branch.id,
        bodega_id=bodega.id,
        fecha=fecha_value,
        detalle_cs=json.dumps(detalle_cs),
        detalle_usd=json.dumps(detalle_usd),
        total_efectivo_cs=total_cs,
        total_efectivo_usd=total_usd,
        total_efectivo_usd_equiv=total_usd_equiv,
        total_ventas_usd=total_ventas_usd,
        total_ingresos_usd=total_ingresos_usd,
        total_egresos_usd=total_egresos_usd,
        total_depositos_usd=total_depositos_usd,
        total_creditos_usd=total_creditos_usd,
        total_calculado_usd=total_calculado_usd,
        diferencia_usd=diferencia,
        usuario_registro=user.full_name,
    )
    db.add(cierre)
    db.commit()
    pos_print = (
        db.query(PosPrintSetting)
        .filter(PosPrintSetting.branch_id == branch.id)
        .first()
    )
    if pos_print and pos_print.cierre_auto_print:
        try:
            company_profile = _company_profile_payload(db)
            resumen = {
                "ventas_usd": total_ventas_usd,
                "ingresos_usd": total_ingresos_usd,
                "egresos_usd": total_egresos_usd,
                "depositos_usd": total_depositos_usd,
                "creditos_usd": total_creditos_usd,
                "total_calculado_usd": total_calculado_usd,
            }
            total_bultos = (
                db.query(func.coalesce(func.sum(VentaItem.cantidad), 0))
                .join(VentaFactura)
                .filter(
                    func.date(VentaFactura.fecha) == fecha_value,
                    VentaFactura.estado != "ANULADA",
                    VentaFactura.bodega_id == bodega.id,
                )
                .scalar()
            )
            _print_cierre_ticket(
                cierre,
                tasa,
                resumen,
                Decimal(str(total_bultos or 0)),
                pos_print.cierre_printer_name or pos_print.printer_name,
                pos_print.cierre_copies or 1,
                company_profile,
                pos_print.sumatra_path,
            )
        except Exception:
            pass
    return RedirectResponse(f"/sales/cierre?success=Cierre+registrado&print_id={cierre.id}", status_code=303)


@router.get("/sales/cierre/{cierre_id}/pdf")
def sales_cierre_pdf(
    request: Request,
    cierre_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.cierre")
    cierre = db.query(CierreCaja).filter(CierreCaja.id == cierre_id).first()
    if not cierre:
        return JSONResponse({"ok": False, "message": "Cierre no encontrado"}, status_code=404)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and cierre.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Cierre fuera de tu bodega"}, status_code=403)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= cierre.fecha)
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
    resumen = {
        "ventas_usd": cierre.total_ventas_usd,
        "ingresos_usd": cierre.total_ingresos_usd,
        "egresos_usd": cierre.total_egresos_usd,
        "depositos_usd": cierre.total_depositos_usd,
        "creditos_usd": cierre.total_creditos_usd,
        "total_calculado_usd": cierre.total_calculado_usd,
    }
    total_bultos = (
        db.query(func.coalesce(func.sum(VentaItem.cantidad), 0))
        .join(VentaFactura)
        .filter(
            func.date(VentaFactura.fecha) == cierre.fecha,
            VentaFactura.estado != "ANULADA",
            VentaFactura.bodega_id == cierre.bodega_id,
        )
        .scalar()
    )
    pdf_bytes = _build_cierre_ticket_pdf_bytes(
        cierre,
        tasa,
        resumen,
        Decimal(str(total_bultos or 0)),
        _company_profile_payload(db),
    )
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=cierre_{cierre.fecha}.pdf"},
    )

@router.post("/sales/roc")
async def sales_roc_create(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.roc")
    form = await request.form()
    tipo = (form.get("tipo") or "EGRESO").upper()
    rubro_id = form.get("rubro_id")
    motivo_id = form.get("motivo_id")
    descripcion = (form.get("descripcion") or "").strip()
    fecha_raw = form.get("fecha")
    moneda = (form.get("moneda") or "CS").upper()
    monto_raw = form.get("monto")
    afecta_caja = form.get("afecta_caja") == "on"

    if tipo not in {"INGRESO", "EGRESO"}:
        return RedirectResponse("/sales/roc?error=Tipo+no+valido", status_code=303)
    if not rubro_id or not motivo_id or not monto_raw:
        return RedirectResponse("/sales/roc?error=Datos+incompletos", status_code=303)
    if moneda not in {"CS", "USD"}:
        return RedirectResponse("/sales/roc?error=Moneda+no+valida", status_code=303)

    try:
        monto = float(monto_raw)
    except ValueError:
        monto = 0.0
    if monto <= 0:
        return RedirectResponse("/sales/roc?error=Monto+no+valido", status_code=303)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if moneda == "USD" and not rate_today:
        return RedirectResponse("/sales/roc?error=Tasa+de+cambio+no+configurada", status_code=303)
    tasa = float(rate_today.rate) if rate_today else 0

    branch, bodega = _resolve_branch_bodega(db, user)
    if not branch:
        return RedirectResponse("/sales/roc?error=Usuario+sin+sucursal+asignada", status_code=303)
    if not bodega:
        return RedirectResponse("/sales/roc?error=Bodega+no+configurada+para+la+sucursal", status_code=303)

    rubro = db.query(ReciboRubro).filter(ReciboRubro.id == int(rubro_id), ReciboRubro.activo.is_(True)).first()
    motivo = db.query(ReciboMotivo).filter(ReciboMotivo.id == int(motivo_id), ReciboMotivo.tipo == tipo, ReciboMotivo.activo.is_(True)).first()
    if not rubro or not motivo:
        return RedirectResponse("/sales/roc?error=Rubro+o+motivo+no+valido", status_code=303)

    if fecha_raw:
        try:
            fecha_value = date.fromisoformat(str(fecha_raw).split("T")[0])
        except ValueError:
            fecha_value = local_today()
    else:
        fecha_value = local_today()

    last_recibo = (
        db.query(ReciboCaja)
        .filter(ReciboCaja.bodega_id == bodega.id)
        .order_by(ReciboCaja.secuencia.desc())
        .first()
    )
    next_seq = (last_recibo.secuencia if last_recibo else 0) + 1
    branch_code = (branch.code or "").lower()
    prefix = "ROC-C" if branch_code == "central" else "ROC-E" if branch_code == "esteli" else f"ROC-{branch_code[:1].upper()}"
    numero = f"{prefix}-{next_seq:05d}"

    monto_usd = Decimal("0")
    monto_cs = Decimal("0")
    if moneda == "USD":
        monto_usd = Decimal(str(monto))
        monto_cs = Decimal(str(monto * tasa))
    else:
        monto_cs = Decimal(str(monto))
        monto_usd = Decimal(str(monto / tasa)) if tasa else Decimal("0")

    recibo = ReciboCaja(
        secuencia=next_seq,
        numero=numero,
        branch_id=branch.id,
        bodega_id=bodega.id,
        tipo=tipo,
        rubro_id=rubro.id,
        motivo_id=motivo.id,
        descripcion=descripcion,
        fecha=fecha_value,
        moneda=moneda,
        tasa_cambio=tasa if moneda == "USD" else None,
        monto_usd=monto_usd,
        monto_cs=monto_cs,
        afecta_caja=afecta_caja,
        usuario_registro=user.full_name,
    )
    db.add(recibo)

    if afecta_caja:
        caja = (
            db.query(CajaDiaria)
            .filter(
                CajaDiaria.branch_id == branch.id,
                CajaDiaria.bodega_id == bodega.id,
                CajaDiaria.fecha == fecha_value,
            )
            .first()
        )
        if not caja:
            caja = CajaDiaria(
                branch_id=branch.id,
                bodega_id=bodega.id,
                fecha=fecha_value,
                saldo_usd=Decimal("0"),
                saldo_cs=Decimal("0"),
            )
            db.add(caja)
        if tipo == "INGRESO":
            caja.saldo_usd = Decimal(str(caja.saldo_usd or 0)) + monto_usd
            caja.saldo_cs = Decimal(str(caja.saldo_cs or 0)) + monto_cs
        else:
            caja.saldo_usd = Decimal(str(caja.saldo_usd or 0)) - monto_usd
            caja.saldo_cs = Decimal(str(caja.saldo_cs or 0)) - monto_cs

    db.commit()
    print_id = recibo.id
    pos_print = (
        db.query(PosPrintSetting)
        .filter(PosPrintSetting.branch_id == branch.id)
        .first()
    )
    if pos_print and pos_print.roc_auto_print:
        try:
            company_profile = _company_profile_payload(db)
            _print_roc_ticket(
                recibo,
                pos_print.roc_printer_name or pos_print.printer_name,
                pos_print.roc_copies or pos_print.copies,
                company_profile,
                pos_print.sumatra_path,
            )
        except Exception:
            pass
    return RedirectResponse(f"/sales/roc?success=Recibo+registrado&print_id={print_id}", status_code=303)


@router.get("/sales/roc/{recibo_id}/pdf")
def sales_roc_pdf(
    request: Request,
    recibo_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.roc")
    recibo = db.query(ReciboCaja).filter(ReciboCaja.id == recibo_id).first()
    if not recibo:
        return JSONResponse({"ok": False, "message": "Recibo no encontrado"}, status_code=404)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and recibo.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Recibo fuera de tu bodega"}, status_code=403)
    pdf_bytes = _build_roc_ticket_pdf_bytes(recibo, _company_profile_payload(db))
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=roc_{recibo.numero}.pdf"},
    )


@router.post("/sales/roc/{recibo_id}/anular")
def sales_roc_anular(
    request: Request,
    recibo_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.roc")
    recibo = db.query(ReciboCaja).filter(ReciboCaja.id == recibo_id).first()
    if not recibo:
        return RedirectResponse("/sales/roc?error=Recibo+no+encontrado", status_code=303)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and recibo.bodega_id != bodega.id:
        return RedirectResponse("/sales/roc?error=Recibo+fuera+de+tu+bodega", status_code=303)

    cierre = (
        db.query(CierreCaja)
        .filter(
            CierreCaja.fecha == recibo.fecha,
            CierreCaja.bodega_id == recibo.bodega_id,
        )
        .first()
    )
    if cierre:
        return RedirectResponse(
            "/sales/roc?error=No+se+puede+anular+porque+la+caja+ya+esta+cerrada+en+esa+fecha",
            status_code=303,
        )

    if recibo.afecta_caja:
        caja = (
            db.query(CajaDiaria)
            .filter(
                CajaDiaria.branch_id == recibo.branch_id,
                CajaDiaria.bodega_id == recibo.bodega_id,
                CajaDiaria.fecha == recibo.fecha,
            )
            .first()
        )
        if caja:
            monto_usd = Decimal(str(recibo.monto_usd or 0))
            monto_cs = Decimal(str(recibo.monto_cs or 0))
            if recibo.tipo == "INGRESO":
                caja.saldo_usd = Decimal(str(caja.saldo_usd or 0)) - monto_usd
                caja.saldo_cs = Decimal(str(caja.saldo_cs or 0)) - monto_cs
            else:
                caja.saldo_usd = Decimal(str(caja.saldo_usd or 0)) + monto_usd
                caja.saldo_cs = Decimal(str(caja.saldo_cs or 0)) + monto_cs

    db.delete(recibo)
    db.commit()
    return RedirectResponse("/sales/roc?success=Recibo+anulado", status_code=303)


@router.get("/sales/depositos")
def sales_depositos(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.depositos")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    vendedor_q = request.query_params.get("vendedor_id")
    banco_q = request.query_params.get("banco_id")
    moneda_q = request.query_params.get("moneda")
    today_value = local_today()
    start_date = today_value
    end_date = today_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today_value
            end_date = today_value

    branch, bodega = _resolve_branch_bodega(db, user)
    vendedores = _vendedores_for_bodega(db, bodega)
    bancos = db.query(Banco).order_by(Banco.nombre).all()
    cuentas = db.query(CuentaBancaria).order_by(CuentaBancaria.banco_id).all()

    depositos_query = db.query(DepositoCliente)
    if bodega:
        depositos_query = depositos_query.filter(DepositoCliente.bodega_id == bodega.id)
    depositos_query = depositos_query.filter(DepositoCliente.fecha.between(start_date, end_date))
    if vendedor_q:
        depositos_query = depositos_query.filter(DepositoCliente.vendedor_id == int(vendedor_q))
    if banco_q:
        depositos_query = depositos_query.filter(DepositoCliente.banco_id == int(banco_q))
    if moneda_q:
        depositos_query = depositos_query.filter(DepositoCliente.moneda == moneda_q.upper())
    depositos = depositos_query.order_by(DepositoCliente.fecha.desc(), DepositoCliente.id.desc()).all()

    summary = {}
    total_cs = Decimal("0")
    total_usd = Decimal("0")
    for dep in depositos:
        key = (dep.banco_id, dep.moneda)
        if key not in summary:
            summary[key] = {
                "banco": dep.banco.nombre if dep.banco else "-",
                "moneda": dep.moneda,
                "count": 0,
                "total": Decimal("0"),
            }
        summary[key]["count"] += 1
        monto_cs = Decimal(str(dep.monto_cs or 0))
        monto_usd = Decimal(str(dep.monto_usd or 0))
        if dep.moneda == "USD":
            summary[key]["total"] += monto_usd
            total_usd += monto_usd
        else:
            summary[key]["total"] += monto_cs
            total_cs += monto_cs

    summary_rows = sorted(summary.values(), key=lambda row: (row["banco"], row["moneda"]))
    summary_grouped = {}
    for row in summary_rows:
        summary_grouped.setdefault(row["banco"], []).append(row)
    summary_grouped_rows = [
        {"banco": banco, "rows": rows} for banco, rows in summary_grouped.items()
    ]

    return request.app.state.templates.TemplateResponse(
        "sales_depositos.html",
        {
            "request": request,
            "user": user,
            "branch": branch,
            "bodega": bodega,
            "vendedores": vendedores,
            "bancos": bancos,
            "cuentas": cuentas,
            "depositos": depositos,
            "summary_rows": summary_rows,
            "summary_grouped_rows": summary_grouped_rows,
            "total_cs": float(total_cs),
            "total_usd": float(total_usd),
            "today": today_value.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "vendedor_q": vendedor_q or "",
            "banco_q": banco_q or "",
            "moneda_q": moneda_q or "",
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/sales/ventas-caliente")
def sales_ventas_caliente(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.caliente")
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    today = local_today()
    first_day = date(today.year, today.month, 1)
    next_month = (first_day.replace(day=28) + timedelta(days=4)).replace(day=1)
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= today)
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    tasa_default = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

    def to_usd(factura: VentaFactura) -> Decimal:
        moneda = factura.moneda or "CS"
        if moneda == "USD":
            return Decimal(str(factura.total_usd or 0))
        tasa = Decimal(str(factura.tasa_cambio or 0)) or tasa_default
        if not tasa:
            return Decimal("0")
        return Decimal(str(factura.total_cs or 0)) / tasa

    branches = (
        _scoped_branches_query(db)
        .filter(Branch.id.in_(scoped_branch_ids))
        .order_by(Branch.name)
        .all()
    )

    base_query = (
        db.query(VentaFactura)
        .join(Bodega, VentaFactura.bodega_id == Bodega.id)
        .join(Branch, Bodega.branch_id == Branch.id)
        .filter(VentaFactura.estado != "ANULADA")
    )

    def total_branch(branch: Optional[Branch]) -> Decimal:
        if not branch:
            return Decimal("0")
        rows = (
            base_query.filter(Branch.id == branch.id)
            .filter(VentaFactura.fecha >= first_day, VentaFactura.fecha < next_month)
            .all()
        )
        total = sum(to_usd(row) for row in rows)
        return Decimal(str(total))

    def total_branch_day(branch: Optional[Branch]) -> Decimal:
        if not branch:
            return Decimal("0")
        rows = (
            base_query.filter(Branch.id == branch.id)
            .filter(func.date(VentaFactura.fecha) == today)
            .all()
        )
        total = sum(to_usd(row) for row in rows)
        return Decimal(str(total))

    branch_totals: list[dict[str, float | str | int]] = []
    total_all = Decimal("0")
    total_day_all = Decimal("0")
    badge_styles = [
        ("bg-primary-subtle text-primary", "primary"),
        ("bg-success-subtle text-success", "success"),
        ("bg-info-subtle text-info", "info"),
        ("bg-warning-subtle text-warning", "warning"),
    ]
    for idx, branch in enumerate(branches):
        month_total = total_branch(branch)
        day_total = total_branch_day(branch)
        total_all += month_total
        total_day_all += day_total
        badge_class, badge_tone = badge_styles[idx % len(badge_styles)]
        branch_totals.append(
            {
                "id": branch.id,
                "name": branch.name,
                "total_month": float(month_total),
                "total_day": float(day_total),
                "badge_class": badge_class,
                "badge_tone": badge_tone,
            }
        )

    monthly_rows = (
        base_query.filter(VentaFactura.fecha >= first_day, VentaFactura.fecha < next_month).all()
    )
    totals_by_day = {}
    for row in monthly_rows:
        day = row.fecha.date() if isinstance(row.fecha, datetime) else row.fecha
        totals_by_day[day] = totals_by_day.get(day, Decimal("0")) + to_usd(row)

    days = []
    cursor = first_day
    while cursor < next_month:
        days.append(cursor)
        cursor += timedelta(days=1)
    chart_points = [
        {
            "label": d.strftime("%d/%m"),
            "value": float(totals_by_day.get(d, Decimal("0"))),
        }
        for d in days
    ]
    month_total = sum(Decimal(str(p["value"])) for p in chart_points)

    return request.app.state.templates.TemplateResponse(
        "sales_caliente.html",
        {
            "request": request,
            "user": user,
            "branch_totals": branch_totals,
            "total_all": float(total_all),
            "total_day_all": float(total_day_all),
            "month_total": float(month_total),
            "chart_points": chart_points,
            "month_label": first_day.strftime("%B %Y").capitalize(),
            "tasa": float(tasa_default) if tasa_default else 0,
            "today_label": today.strftime("%d/%m/%Y"),
            "version": settings.UI_VERSION,
        },
    )


@router.get("/reports")
def reports_index(
    request: Request,
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    return request.app.state.templates.TemplateResponse(
        "reports_index.html",
        {
            "request": request,
            "user": user,
            "version": settings.UI_VERSION,
        },
    )


def _sales_report_filters(request: Request):
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    branch_id = request.query_params.get("branch_id")
    vendedor_id = request.query_params.get("vendedor_id")
    producto_q = (request.query_params.get("producto") or "").strip()

    today = local_today()
    start_date = today
    end_date = today
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today
            end_date = today

    if not branch_id:
        branch_id = "all"

    return start_date, end_date, branch_id, vendedor_id, producto_q


def _sales_products_report_filters(request: Request):
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    branch_id = request.query_params.get("branch_id")
    vendedor_id = request.query_params.get("vendedor_id")
    producto_id = request.query_params.get("producto_id")
    producto_q = (request.query_params.get("producto") or "").strip()

    today = local_today()
    start_date = today
    end_date = today
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today
            end_date = today

    if not branch_id:
        branch_id = "all"

    return start_date, end_date, branch_id, vendedor_id, producto_id, producto_q


def _build_sales_products_report(
    db: Session,
    user: User,
    start_date: date,
    end_date: date,
    branch_id: str | None,
    vendedor_id: str | None,
    producto_id: str | None,
    producto_q: str,
):
    allowed_codes = _allowed_branch_codes(db)
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    base_query = (
        db.query(VentaFactura, VentaItem, Producto, Branch)
        .join(VentaItem, VentaItem.factura_id == VentaFactura.id)
        .join(Producto, Producto.id == VentaItem.producto_id)
        .join(Bodega, Bodega.id == VentaFactura.bodega_id, isouter=True)
        .join(Branch, Branch.id == Bodega.branch_id, isouter=True)
        .filter(VentaFactura.fecha >= start_dt, VentaFactura.fecha < end_dt)
        .filter(VentaFactura.estado != "ANULADA")
        .filter(func.lower(Branch.code).in_(allowed_codes))
        .filter(Branch.id.in_(scoped_branch_ids))
    )
    if branch_id and branch_id != "all":
        try:
            branch_id_int = int(branch_id)
            if branch_id_int not in scoped_branch_ids:
                base_query = base_query.filter(Branch.id == -1)
            else:
                base_query = base_query.filter(Branch.id == branch_id_int)
        except ValueError:
            pass
    if vendedor_id:
        try:
            base_query = base_query.filter(VentaFactura.vendedor_id == int(vendedor_id))
        except ValueError:
            pass
    if producto_id:
        try:
            base_query = base_query.filter(VentaItem.producto_id == int(producto_id))
        except ValueError:
            pass
    if producto_q:
        like = f"%{producto_q.lower()}%"
        base_query = base_query.filter(
            or_(
                func.lower(Producto.cod_producto).like(like),
                func.lower(Producto.descripcion).like(like),
            )
        )

    rows = base_query.all()
    report_map: dict[int, dict] = {}
    facturas_set = set()
    detail_rows = []

    for factura, item, producto, branch in rows:
        moneda = factura.moneda or "CS"
        tasa_factura = Decimal(str(factura.tasa_cambio or 0))
        if moneda == "CS" and not tasa_factura:
            rate_today = (
                db.query(ExchangeRate)
                .filter(ExchangeRate.effective_date <= factura.fecha)
                .order_by(ExchangeRate.effective_date.desc())
                .first()
            )
            tasa_factura = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

        cantidad = Decimal(str(item.cantidad or 0))
        subtotal_usd = Decimal(str(item.subtotal_usd or 0))
        subtotal_cs = Decimal(str(item.subtotal_cs or 0))
        venta_usd = subtotal_usd if moneda == "USD" else (subtotal_cs / tasa_factura if tasa_factura else Decimal("0"))
        venta_cs = subtotal_cs if moneda == "CS" else (subtotal_usd * tasa_factura if tasa_factura else Decimal("0"))

        costo_cs_unit = Decimal(str(producto.costo_producto or 0))
        costo_cs = costo_cs_unit * cantidad
        tasa_producto = Decimal(str(producto.tasa_cambio or 0))
        if not tasa_producto:
            tasa_producto = tasa_factura
        if not tasa_producto:
            rate_today = (
                db.query(ExchangeRate)
                .filter(ExchangeRate.effective_date <= factura.fecha)
                .order_by(ExchangeRate.effective_date.desc())
                .first()
            )
            tasa_producto = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
        costo_usd = costo_cs / tasa_producto if tasa_producto else Decimal("0")

        if producto.id not in report_map:
            report_map[producto.id] = {
                "codigo": producto.cod_producto,
                "producto": producto.descripcion,
                "cantidad": Decimal("0"),
                "venta_usd": Decimal("0"),
                "venta_cs": Decimal("0"),
                "costo_usd": Decimal("0"),
                "costo_cs": Decimal("0"),
            }
        report_row = report_map[producto.id]
        report_row["cantidad"] += cantidad
        report_row["venta_usd"] += venta_usd
        report_row["venta_cs"] += venta_cs
        report_row["costo_usd"] += costo_usd
        report_row["costo_cs"] += costo_cs
        facturas_set.add(factura.id)
        detail_rows.append(
            {
                "fecha": factura.fecha.strftime("%d/%m/%Y") if factura.fecha else "",
                "factura": factura.numero or f"FAC-{factura.id}",
                "cliente": factura.cliente.nombre if factura.cliente else "Consumidor final",
                "vendedor": factura.vendedor.nombre if factura.vendedor else "-",
                "sucursal": branch.name if branch else "-",
                "codigo": producto.cod_producto,
                "producto": producto.descripcion,
                "cantidad": float(cantidad),
                "precio_usd": float(item.precio_unitario_usd or 0),
                "precio_cs": float(item.precio_unitario_cs or 0),
                "subtotal_usd": float(subtotal_usd),
                "subtotal_cs": float(subtotal_cs),
                "costo_usd": float(costo_usd),
                "costo_cs": float(costo_cs),
                "margen_usd": float(venta_usd - costo_usd),
            }
        )

    report_rows = []
    total_qty = Decimal("0")
    total_usd = Decimal("0")
    total_cs = Decimal("0")
    total_cost_usd = Decimal("0")
    total_cost_cs = Decimal("0")
    for row in report_map.values():
        total_qty += row["cantidad"]
        total_usd += row["venta_usd"]
        total_cs += row["venta_cs"]
        total_cost_usd += row["costo_usd"]
        total_cost_cs += row["costo_cs"]
        report_rows.append(
            {
                "codigo": row["codigo"],
                "producto": row["producto"],
                "cantidad": float(row["cantidad"]),
                "venta_usd": float(row["venta_usd"]),
                "venta_cs": float(row["venta_cs"]),
                "costo_usd": float(row["costo_usd"]),
                "costo_cs": float(row["costo_cs"]),
                "margen_usd": float(row["venta_usd"] - row["costo_usd"]),
            }
        )

    report_rows.sort(key=lambda r: r["venta_usd"], reverse=True)
    detail_rows.sort(key=lambda r: (r["fecha"], r["factura"], r["producto"]))
    total_facturas = len(facturas_set)
    return (
        report_rows,
        detail_rows,
        total_qty,
        total_usd,
        total_cs,
        total_cost_usd,
        total_cost_cs,
        total_facturas,
    )


def _depositos_report_filters(request: Request):
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    branch_id = request.query_params.get("branch_id")

    today = local_today()
    start_date = today
    end_date = today
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today
            end_date = today

    if not branch_id:
        branch_id = "all"

    return start_date, end_date, branch_id


def _depositos_report_query(
    db: Session,
    start_date: date,
    end_date: date,
    branch_id: str | None,
):
    allowed_codes = _allowed_branch_codes(db)
    query = (
        db.query(DepositoCliente)
        .join(Bodega, DepositoCliente.bodega_id == Bodega.id, isouter=True)
        .join(Branch, Bodega.branch_id == Branch.id, isouter=True)
        .filter(DepositoCliente.fecha.between(start_date, end_date))
        .filter(func.lower(Branch.code).in_(allowed_codes))
    )
    if branch_id and branch_id != "all":
        try:
            query = query.filter(Branch.id == int(branch_id))
        except ValueError:
            pass
    return query


def _kardex_report_filters(request: Request):
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    branch_id = request.query_params.get("branch_id")
    producto_q = (request.query_params.get("producto") or "").strip()

    today = local_today()
    start_date = today
    end_date = today
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today
            end_date = today

    if not branch_id:
        branch_id = "all"

    return start_date, end_date, branch_id, producto_q


def _inventory_consolidated_filters(request: Request) -> str:
    branch_id = request.query_params.get("branch_id")
    return branch_id or "all"


def _inventory_rotation_filters(request: Request):
    branch_id = request.query_params.get("branch_id") or "all"
    bodega_id = request.query_params.get("bodega_id") or "all"
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    top_n_raw = request.query_params.get("top_n")
    slow_days_raw = request.query_params.get("slow_days")
    categoria_id = request.query_params.get("categoria_id") or "all"
    trend_granularity = (request.query_params.get("trend") or "monthly").strip().lower()
    sort_sales = (request.query_params.get("sort_sales") or "venta").strip().lower()
    min_stock_days_raw = request.query_params.get("min_stock_days")
    lead_days_raw = request.query_params.get("lead_days")

    today = local_today()
    start_date = today - timedelta(days=180)
    end_date = today

    if start_raw:
        try:
            start_date = date.fromisoformat(start_raw)
        except ValueError:
            pass
    if end_raw:
        try:
            end_date = date.fromisoformat(end_raw)
        except ValueError:
            pass
    if end_date < start_date:
        end_date = start_date

    try:
        top_n = int(top_n_raw or 20)
    except ValueError:
        top_n = 20
    top_n = min(max(top_n, 5), 200)

    try:
        slow_days = int(slow_days_raw or 45)
    except ValueError:
        slow_days = 45
    slow_days = min(max(slow_days, 7), 365)

    try:
        min_stock_days = int(min_stock_days_raw or 15)
    except ValueError:
        min_stock_days = 15
    min_stock_days = min(max(min_stock_days, 5), 120)

    try:
        lead_days = int(lead_days_raw or 7)
    except ValueError:
        lead_days = 7
    lead_days = min(max(lead_days, 1), 90)

    if trend_granularity not in {"daily", "weekly", "monthly"}:
        trend_granularity = "monthly"
    if sort_sales not in {"venta", "cantidad"}:
        sort_sales = "venta"

    return (
        start_date,
        end_date,
        branch_id,
        bodega_id,
        top_n,
        slow_days,
        categoria_id,
        trend_granularity,
        sort_sales,
        min_stock_days,
        lead_days,
    )


def _build_inventory_rotation_data(
    db: Session,
    start_date: date,
    end_date: date,
    branch_id: str,
    bodega_id: str,
    top_n: int,
    slow_days: int,
    categoria_id: str,
    trend_granularity: str,
    sort_sales: str,
    min_stock_days: int,
    lead_days: int,
):
    # The financial block must obey the top date filters.
    period_start = start_date
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    lineas = db.query(Linea).filter(Linea.activo.is_(True)).order_by(Linea.linea.asc()).all()
    selected_branch = None
    if branch_id and branch_id != "all":
        try:
            selected_branch = next((b for b in branches if b.id == int(branch_id)), None)
        except ValueError:
            selected_branch = None

    selected_linea = None
    if categoria_id and categoria_id != "all":
        try:
            selected_linea = next((l for l in lineas if l.id == int(categoria_id)), None)
        except ValueError:
            selected_linea = None

    bodegas_q = _scoped_bodegas_query(db)
    if selected_branch:
        bodegas_q = bodegas_q.filter(Bodega.branch_id == selected_branch.id)
    if bodega_id and bodega_id != "all":
        try:
            bodegas_q = bodegas_q.filter(Bodega.id == int(bodega_id))
        except ValueError:
            pass
    bodegas = bodegas_q.order_by(Bodega.id.asc()).all()
    bodega_ids = [b.id for b in bodegas]
    selected_bodega = None
    if bodega_id and bodega_id != "all":
        try:
            selected_bodega = next((b for b in bodegas if b.id == int(bodega_id)), None)
        except ValueError:
            selected_bodega = None

    if not bodega_ids:
        return {
            "branches": branches,
            "lineas": lineas,
            "bodegas": bodegas,
            "selected_branch": selected_branch,
            "selected_linea": selected_linea,
            "selected_bodega": selected_bodega,
            "month_start": period_start,
            "month_end": end_date,
            "kpis": {
                "productos_stock": 0,
                "productos_sin_venta": 0,
                "inversion_actual_cs": 0.0,
                "inversion_ingresada_cs": 0.0,
                "capital_recuperado_cs": 0.0,
                "recuperacion_pct": 0.0,
                "ventas_periodo_cs": 0.0,
            },
            "unsold_rows": [],
            "slow_rows": [],
            "expensive_rows": [],
            "rotation_rows": [],
            "sales_products_rows": [],
            "balance_rows": [],
            "reorder_rows": [],
            "abc_rows": [],
            "coverage_rows": [],
            "trend_rows": [],
        }

    productos_q = db.query(Producto).filter(Producto.activo.is_(True))
    if selected_linea:
        productos_q = productos_q.filter(Producto.linea_id == selected_linea.id)
    productos = productos_q.all()
    product_map = {p.id: p for p in productos}
    product_ids_list = [p.id for p in productos]

    if not product_ids_list:
        return {
            "branches": branches,
            "lineas": lineas,
            "bodegas": bodegas,
            "selected_branch": selected_branch,
            "selected_linea": selected_linea,
            "selected_bodega": selected_bodega,
            "month_start": period_start,
            "month_end": end_date,
            "kpis": {
                "productos_stock": 0,
                "productos_sin_venta": 0,
                "inversion_actual_cs": 0.0,
                "inversion_ingresada_cs": 0.0,
                "capital_recuperado_cs": 0.0,
                "recuperacion_pct": 0.0,
                "ventas_periodo_cs": 0.0,
            },
            "unsold_rows": [],
            "slow_rows": [],
            "expensive_rows": [],
            "rotation_rows": [],
            "sales_products_rows": [],
            "balance_rows": [],
            "reorder_rows": [],
            "abc_rows": [],
            "coverage_rows": [],
            "trend_rows": [],
        }

    balances = _balances_by_bodega(db, bodega_ids, product_ids_list)
    saldo_map: dict[int, Decimal] = {}
    for producto_id in product_ids_list:
        total = Decimal("0")
        for bodega_id in bodega_ids:
            total += balances.get((producto_id, bodega_id), Decimal("0"))
        saldo_map[producto_id] = total

    ingresos_rows = (
        db.query(
            IngresoItem.producto_id,
            func.sum(IngresoItem.cantidad),
            func.sum(IngresoItem.cantidad * IngresoItem.costo_unitario_cs),
            func.min(IngresoInventario.fecha),
        )
        .join(IngresoInventario, IngresoInventario.id == IngresoItem.ingreso_id)
        .filter(IngresoInventario.bodega_id.in_(bodega_ids))
        .filter(IngresoItem.producto_id.in_(product_ids_list))
        .filter(IngresoInventario.fecha >= start_date, IngresoInventario.fecha <= end_date)
        .group_by(IngresoItem.producto_id)
        .all()
    )
    ingreso_qty_map: dict[int, Decimal] = {}
    ingreso_cost_map: dict[int, Decimal] = {}
    first_ingreso_map: dict[int, date] = {}
    for pid, qty, cost_cs, first_date in ingresos_rows:
        ingreso_qty_map[int(pid)] = Decimal(str(qty or 0))
        ingreso_cost_map[int(pid)] = Decimal(str(cost_cs or 0))
        first_ingreso_map[int(pid)] = first_date

    ventas_rows = (
        db.query(
            VentaItem.producto_id,
            func.sum(VentaItem.cantidad),
            func.sum(VentaItem.subtotal_cs),
            func.max(VentaFactura.fecha),
        )
        .join(VentaFactura, VentaFactura.id == VentaItem.factura_id)
        .filter(VentaFactura.bodega_id.in_(bodega_ids))
        .filter(VentaItem.producto_id.in_(product_ids_list))
        .filter(VentaFactura.estado != "ANULADA")
        .filter(VentaFactura.fecha >= datetime.combine(start_date, datetime.min.time()))
        .filter(VentaFactura.fecha < datetime.combine(end_date + timedelta(days=1), datetime.min.time()))
        .group_by(VentaItem.producto_id)
        .all()
    )
    sold_qty_map: dict[int, Decimal] = {}
    sold_cs_map: dict[int, Decimal] = {}
    last_sale_map: dict[int, datetime] = {}
    for pid, qty, subtotal_cs, last_dt in ventas_rows:
        sold_qty_map[int(pid)] = Decimal(str(qty or 0))
        sold_cs_map[int(pid)] = Decimal(str(subtotal_cs or 0))
        last_sale_map[int(pid)] = last_dt

    ingresos_mes_rows = (
        db.query(
            IngresoItem.producto_id,
            func.sum(IngresoItem.cantidad),
            func.sum(IngresoItem.subtotal_cs),
        )
        .join(IngresoInventario, IngresoInventario.id == IngresoItem.ingreso_id)
        .filter(IngresoInventario.bodega_id.in_(bodega_ids))
        .filter(IngresoItem.producto_id.in_(product_ids_list))
        .filter(IngresoInventario.fecha >= period_start, IngresoInventario.fecha <= end_date)
        .group_by(IngresoItem.producto_id)
        .all()
    )
    ingreso_mes_qty_map = {int(pid): Decimal(str(qty or 0)) for pid, qty, _ in ingresos_mes_rows}
    ingreso_mes_cs_map = {int(pid): Decimal(str(cost_cs or 0)) for pid, _, cost_cs in ingresos_mes_rows}

    egresos_mes_rows = (
        db.query(
            EgresoItem.producto_id,
            func.sum(EgresoItem.cantidad),
            func.sum(EgresoItem.subtotal_cs),
        )
        .join(EgresoInventario, EgresoInventario.id == EgresoItem.egreso_id)
        .filter(EgresoInventario.bodega_id.in_(bodega_ids))
        .filter(EgresoItem.producto_id.in_(product_ids_list))
        .filter(EgresoInventario.fecha >= period_start, EgresoInventario.fecha <= end_date)
        .group_by(EgresoItem.producto_id)
        .all()
    )
    egreso_mes_qty_map = {int(pid): Decimal(str(qty or 0)) for pid, qty, _ in egresos_mes_rows}
    egreso_mes_cs_map = {int(pid): Decimal(str(cost_cs or 0)) for pid, _, cost_cs in egresos_mes_rows}

    ventas_mes_rows = (
        db.query(VentaItem.producto_id, func.sum(VentaItem.cantidad))
        .join(VentaFactura, VentaFactura.id == VentaItem.factura_id)
        .filter(VentaFactura.bodega_id.in_(bodega_ids))
        .filter(VentaItem.producto_id.in_(product_ids_list))
        .filter(VentaFactura.estado != "ANULADA")
        .filter(VentaFactura.fecha >= datetime.combine(period_start, datetime.min.time()))
        .filter(VentaFactura.fecha < datetime.combine(end_date + timedelta(days=1), datetime.min.time()))
        .group_by(VentaItem.producto_id)
        .all()
    )
    ventas_mes_qty_map = {int(pid): Decimal(str(qty or 0)) for pid, qty in ventas_mes_rows}

    last_90_start = end_date - timedelta(days=89)
    ventas_90_rows = (
        db.query(VentaItem.producto_id, func.sum(VentaItem.cantidad))
        .join(VentaFactura, VentaFactura.id == VentaItem.factura_id)
        .filter(VentaFactura.bodega_id.in_(bodega_ids))
        .filter(VentaItem.producto_id.in_(product_ids_list))
        .filter(VentaFactura.estado != "ANULADA")
        .filter(VentaFactura.fecha >= datetime.combine(last_90_start, datetime.min.time()))
        .filter(VentaFactura.fecha < datetime.combine(end_date + timedelta(days=1), datetime.min.time()))
        .group_by(VentaItem.producto_id)
        .all()
    )
    sold_90_qty_map = {int(pid): Decimal(str(qty or 0)) for pid, qty in ventas_90_rows}

    latest_provider_rows = (
        db.query(IngresoItem.producto_id, IngresoInventario.fecha, Proveedor.nombre)
        .join(IngresoInventario, IngresoInventario.id == IngresoItem.ingreso_id)
        .join(Proveedor, Proveedor.id == IngresoInventario.proveedor_id, isouter=True)
        .filter(IngresoInventario.bodega_id.in_(bodega_ids))
        .filter(IngresoItem.producto_id.in_(product_ids_list))
        .order_by(IngresoInventario.fecha.desc(), IngresoInventario.id.desc())
        .all()
    )
    latest_provider_map: dict[int, str] = {}
    for pid, _fecha, prov_name in latest_provider_rows:
        key = int(pid)
        if key not in latest_provider_map:
            latest_provider_map[key] = prov_name or "-"

    rate_today_row = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= end_date)
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    rate_today = Decimal(str(rate_today_row.rate)) if rate_today_row else Decimal("0")

    trend_raw_rows = (
        db.query(
            VentaItem.producto_id,
            func.date(VentaFactura.fecha),
            func.sum(VentaItem.cantidad),
            func.sum(VentaItem.subtotal_cs),
        )
        .join(VentaFactura, VentaFactura.id == VentaItem.factura_id)
        .filter(VentaFactura.bodega_id.in_(bodega_ids))
        .filter(VentaItem.producto_id.in_(product_ids_list))
        .filter(VentaFactura.estado != "ANULADA")
        .filter(VentaFactura.fecha >= datetime.combine(start_date, datetime.min.time()))
        .filter(VentaFactura.fecha < datetime.combine(end_date + timedelta(days=1), datetime.min.time()))
        .group_by(VentaItem.producto_id, func.date(VentaFactura.fecha))
        .all()
    )
    trend_buckets: dict[tuple[date, int], dict] = {}
    for pid, raw_day, qty, subtotal_cs in trend_raw_rows:
        if isinstance(raw_day, datetime):
            day = raw_day.date()
        elif isinstance(raw_day, date):
            day = raw_day
        elif isinstance(raw_day, str):
            day = date.fromisoformat(raw_day)
        else:
            continue
        if trend_granularity == "weekly":
            period_date = day - timedelta(days=day.weekday())
        elif trend_granularity == "monthly":
            period_date = day.replace(day=1)
        else:
            period_date = day
        key = (period_date, int(pid))
        if key not in trend_buckets:
            trend_buckets[key] = {"cantidad": Decimal("0"), "venta_cs": Decimal("0")}
        trend_buckets[key]["cantidad"] += Decimal(str(qty or 0))
        trend_buckets[key]["venta_cs"] += Decimal(str(subtotal_cs or 0))

    product_ids = set(saldo_map.keys()) | set(ingreso_qty_map.keys()) | set(sold_qty_map.keys())
    rows = []
    total_inversion_actual = Decimal("0")
    total_inversion_disponible = Decimal("0")
    total_inversion_ingresada = Decimal("0")
    total_recuperado = Decimal("0")
    total_ventas_periodo_cs = Decimal("0")
    total_costo_vendido_periodo_cs = Decimal("0")
    productos_sin_venta = 0
    productos_stock = 0
    total_inversion_inicial_mes_cs = Decimal("0")
    total_ingresos_mes_cs = Decimal("0")
    total_egresos_mes_cs = Decimal("0")

    for pid in product_ids:
        producto = product_map.get(pid)
        if not producto:
            continue
        ingreso_qty = Decimal(str(ingreso_qty_map.get(pid, Decimal("0"))))
        vendido_qty = Decimal(str(sold_qty_map.get(pid, Decimal("0"))))
        saldo_qty = Decimal(str(saldo_map.get(pid, Decimal("0"))))
        costo_unit_cs = Decimal(str(producto.costo_producto or 0))
        inversion_actual_cs = saldo_qty * costo_unit_cs
        inversion_ingresada_cs = Decimal(str(ingreso_cost_map.get(pid, Decimal("0"))))
        capital_recuperado_cs = vendido_qty * costo_unit_cs
        costo_vendido_periodo_cs = vendido_qty * costo_unit_cs
        ingreso_mes_qty = Decimal(str(ingreso_mes_qty_map.get(pid, Decimal("0"))))
        egreso_mes_qty = Decimal(str(egreso_mes_qty_map.get(pid, Decimal("0"))))
        venta_mes_qty = Decimal(str(ventas_mes_qty_map.get(pid, Decimal("0"))))
        saldo_inicio_mes_qty = saldo_qty - ingreso_mes_qty + egreso_mes_qty + venta_mes_qty
        inversion_inicial_mes_cs = saldo_inicio_mes_qty * costo_unit_cs
        sell_through_pct = (
            (vendido_qty / ingreso_qty) * Decimal("100")
            if ingreso_qty > 0
            else Decimal("0")
        )
        avg_daily_90 = Decimal(str(sold_90_qty_map.get(pid, Decimal("0")))) / Decimal("90")
        min_qty_recomendado = avg_daily_90 * Decimal(str(min_stock_days))
        reorder_qty = avg_daily_90 * Decimal(str(lead_days))
        cobertura_dias = (saldo_qty / avg_daily_90) if avg_daily_90 > 0 else None
        days_to_liquidate = (
            (saldo_qty / avg_daily_90)
            if avg_daily_90 > 0 and saldo_qty > 0
            else None
        )
        last_sale_dt = last_sale_map.get(pid)
        last_sale_date = (
            last_sale_dt.date() if isinstance(last_sale_dt, datetime) else last_sale_dt
        )
        first_ingreso = first_ingreso_map.get(pid)
        days_without_sale = (
            (end_date - last_sale_date).days
            if last_sale_date
            else ((end_date - first_ingreso).days if first_ingreso else None)
        )

        if saldo_qty != 0:
            total_inversion_actual += inversion_actual_cs
        if saldo_qty > 0:
            productos_stock += 1
            total_inversion_disponible += inversion_actual_cs
        total_inversion_ingresada += inversion_ingresada_cs
        total_recuperado += capital_recuperado_cs
        total_ventas_periodo_cs += Decimal(str(sold_cs_map.get(pid, Decimal("0"))))
        total_costo_vendido_periodo_cs += costo_vendido_periodo_cs
        total_inversion_inicial_mes_cs += inversion_inicial_mes_cs
        total_ingresos_mes_cs += Decimal(str(ingreso_mes_cs_map.get(pid, Decimal("0"))))
        total_egresos_mes_cs += Decimal(str(egreso_mes_cs_map.get(pid, Decimal("0"))))
        if vendido_qty <= 0 and saldo_qty > 0:
            productos_sin_venta += 1

        if days_to_liquidate is None:
            rotacion = "Sin movimiento"
        elif days_to_liquidate <= Decimal("30"):
            rotacion = "Rapida"
        elif days_to_liquidate <= Decimal("90"):
            rotacion = "Media"
        else:
            rotacion = "Lenta"

        rows.append(
            {
                "producto_id": pid,
                "codigo": producto.cod_producto or "-",
                "descripcion": producto.descripcion or "-",
                "ingreso_qty": float(ingreso_qty),
                "vendido_qty": float(vendido_qty),
                "saldo_qty": float(saldo_qty),
                "costo_unit_cs": float(costo_unit_cs),
                "inversion_actual_cs": float(inversion_actual_cs),
                "inversion_ingresada_cs": float(inversion_ingresada_cs),
                "capital_recuperado_cs": float(capital_recuperado_cs),
                "sell_through_pct": float(sell_through_pct),
                "last_sale": last_sale_date,
                "first_ingreso": first_ingreso,
                "days_without_sale": days_without_sale if days_without_sale is not None else None,
                "avg_daily_90": float(avg_daily_90),
                "min_qty_recomendado": float(min_qty_recomendado),
                "reorder_qty": float(reorder_qty),
                "cobertura_dias": float(cobertura_dias) if cobertura_dias is not None else None,
                "days_to_liquidate": float(days_to_liquidate) if days_to_liquidate is not None else None,
                "rotacion": rotacion,
                "linea": producto.linea.linea if producto.linea else "-",
                "proveedor": latest_provider_map.get(pid, "-"),
            }
        )

    unsold_rows = [
        r for r in rows if (r["saldo_qty"] > 0 and r["sell_through_pct"] < 100)
    ]
    unsold_rows.sort(
        key=lambda r: (
            -(r["days_without_sale"] if r["days_without_sale"] is not None else 999999),
            -r["inversion_actual_cs"],
        )
    )

    slow_rows = [
        r
        for r in rows
        if r["saldo_qty"] > 0 and (r["days_without_sale"] or 0) >= slow_days
    ]
    slow_rows.sort(
        key=lambda r: (
            -(r["days_without_sale"] or 0),
            r["avg_daily_90"],
            -r["inversion_actual_cs"],
        )
    )

    expensive_rows = [
        r
        for r in rows
        if r["saldo_qty"] > 0 and (r["days_without_sale"] or 0) >= slow_days
    ]
    expensive_rows.sort(key=lambda r: -r["inversion_actual_cs"])

    rotation_rows = [r for r in rows if r["saldo_qty"] > 0]
    rotation_rows.sort(
        key=lambda r: (
            999999 if r["days_to_liquidate"] is None else r["days_to_liquidate"],
            -r["inversion_actual_cs"],
        )
    )

    recuperacion_pct = (
        (total_recuperado / total_inversion_ingresada) * Decimal("100")
        if total_inversion_ingresada > 0
        else Decimal("0")
    )
    ajuste_negativos_cs = total_inversion_disponible - total_inversion_actual
    inversion_total_fecha_cs = total_inversion_inicial_mes_cs + total_ingresos_mes_cs - total_egresos_mes_cs
    costo_sobre_venta_pct = (
        (total_costo_vendido_periodo_cs / total_ventas_periodo_cs) * Decimal("100")
        if total_ventas_periodo_cs > 0
        else Decimal("0")
    )
    recuperacion_costos_pct = (
        (total_costo_vendido_periodo_cs / inversion_total_fecha_cs) * Decimal("100")
        if inversion_total_fecha_cs > 0
        else Decimal("0")
    )

    sales_products_rows = [
        {
            "codigo": r["codigo"],
            "descripcion": r["descripcion"],
            "fecha_venta": r["last_sale"].strftime("%d/%m/%Y") if r["last_sale"] else "-",
            "cantidad": r["vendido_qty"],
            "ingreso_cs": float(sold_cs_map.get(r["producto_id"], Decimal("0"))),
        }
        for r in rows
        if r["vendido_qty"] > 0
    ]
    if sort_sales == "cantidad":
        sales_products_rows.sort(key=lambda r: (-float(r["cantidad"]), -float(r["ingreso_cs"])))
    else:
        sales_products_rows.sort(key=lambda r: (-float(r["ingreso_cs"]), -float(r["cantidad"])))

    balance_rows = []
    reorder_rows = []
    coverage_rows = []
    for r in rows:
        if r["saldo_qty"] <= 0:
            continue
        reorder_qty = Decimal(str(r["reorder_qty"] or 0))
        saldo_qty = Decimal(str(r["saldo_qty"] or 0))
        if reorder_qty > 0 and saldo_qty <= (reorder_qty * Decimal("0.5")):
            estado = "CRITICO"
        elif reorder_qty > 0 and saldo_qty <= reorder_qty:
            estado = "BAJO"
        else:
            estado = "NORMAL"

        balance_row = {
            "codigo": r["codigo"],
            "descripcion": r["descripcion"],
            "linea": r["linea"],
            "stock_actual": r["saldo_qty"],
            "stock_min": r["min_qty_recomendado"],
            "estado": estado,
        }
        balance_rows.append(balance_row)

        if reorder_qty > 0 and saldo_qty <= reorder_qty:
            reorder_rows.append(
                {
                    "codigo": r["codigo"],
                    "descripcion": r["descripcion"],
                    "stock_actual": r["saldo_qty"],
                    "punto_reorden": r["reorder_qty"],
                    "proveedor": r["proveedor"],
                    "lead_days": lead_days,
                    "accion": "Reordenar ahora",
                }
            )

        if r["cobertura_dias"] is not None:
            coverage_rows.append(
                {
                    "codigo": r["codigo"],
                    "descripcion": r["descripcion"],
                    "stock_actual": r["saldo_qty"],
                    "ventas_diarias": r["avg_daily_90"],
                    "cobertura_dias": r["cobertura_dias"],
                }
            )

    balance_rows.sort(key=lambda r: (r["estado"] != "CRITICO", r["estado"] != "BAJO", r["descripcion"]))
    reorder_rows.sort(key=lambda r: (float(r["stock_actual"] - r["punto_reorden"]), r["descripcion"]))
    coverage_rows.sort(key=lambda r: float(r["cobertura_dias"]))

    abc_base = [r for r in rows if r["saldo_qty"] > 0 or r["vendido_qty"] > 0]
    abc_base.sort(
        key=lambda r: -float(
            sold_cs_map.get(r["producto_id"], Decimal("0"))
            if sold_cs_map.get(r["producto_id"], Decimal("0")) > 0
            else Decimal(str(r["inversion_actual_cs"] or 0))
        )
    )
    abc_total = Decimal("0")
    for r in abc_base:
        metric = sold_cs_map.get(r["producto_id"], Decimal("0"))
        if metric <= 0:
            metric = Decimal(str(r["inversion_actual_cs"] or 0))
        abc_total += metric

    abc_rows = []
    acumulado = Decimal("0")
    for r in abc_base:
        valor = sold_cs_map.get(r["producto_id"], Decimal("0"))
        if valor <= 0:
            valor = Decimal(str(r["inversion_actual_cs"] or 0))
        pct = (valor / abc_total * Decimal("100")) if abc_total > 0 else Decimal("0")
        acumulado += pct
        if acumulado <= Decimal("80"):
            categoria = "A"
        elif acumulado <= Decimal("95"):
            categoria = "B"
        else:
            categoria = "C"
        abc_rows.append(
            {
                "codigo": r["codigo"],
                "descripcion": r["descripcion"],
                "valor": float(valor),
                "pct_acumulado": float(acumulado),
                "categoria": categoria,
            }
        )

    trend_rows = []
    for (period_date, pid), agg in sorted(
        trend_buckets.items(),
        key=lambda x: (x[0][0], -x[1]["venta_cs"]),
        reverse=True,
    ):
        prod = product_map.get(pid)
        if not prod:
            continue
        trend_rows.append(
            {
                "periodo": period_date.strftime("%d/%m/%Y"),
                "codigo": prod.cod_producto or "-",
                "descripcion": prod.descripcion or "-",
                "cantidad": float(agg["cantidad"]),
                "ingreso_cs": float(agg["venta_cs"]),
                "ingreso_usd": float((agg["venta_cs"] / rate_today) if rate_today else Decimal("0")),
            }
        )

    return {
        "branches": branches,
        "lineas": lineas,
        "bodegas": bodegas,
        "selected_branch": selected_branch,
        "selected_linea": selected_linea,
        "selected_bodega": selected_bodega,
        "kpis": {
            "productos_stock": int(productos_stock),
            "productos_sin_venta": int(productos_sin_venta),
            "inversion_actual_cs": float(total_inversion_actual),
            "inversion_disponible_cs": float(total_inversion_disponible),
            "ajuste_negativos_cs": float(ajuste_negativos_cs),
            "inversion_ingresada_cs": float(total_inversion_ingresada),
            "capital_recuperado_cs": float(total_recuperado),
            "recuperacion_pct": float(recuperacion_pct),
            "ventas_periodo_cs": float(total_ventas_periodo_cs),
            "inversion_inicial_mes_cs": float(total_inversion_inicial_mes_cs),
            "ingresos_mes_cs": float(total_ingresos_mes_cs),
            "egresos_mes_cs": float(total_egresos_mes_cs),
            "inversion_total_fecha_cs": float(inversion_total_fecha_cs),
            "costo_vendido_periodo_cs": float(total_costo_vendido_periodo_cs),
            "costo_sobre_venta_pct": float(costo_sobre_venta_pct),
            "recuperacion_costos_pct": float(recuperacion_costos_pct),
        },
        "month_start": period_start,
        "month_end": end_date,
        "unsold_rows": unsold_rows[:top_n],
        "slow_rows": slow_rows[:top_n],
        "expensive_rows": expensive_rows[:top_n],
        "rotation_rows": rotation_rows[:top_n],
        "sales_products_rows": sales_products_rows[: max(top_n * 2, 30)],
        "balance_rows": balance_rows[: max(top_n * 2, 30)],
        "reorder_rows": reorder_rows[: max(top_n * 2, 30)],
        "abc_rows": abc_rows[: max(top_n * 2, 30)],
        "coverage_rows": coverage_rows[: max(top_n * 2, 30)],
        "trend_rows": trend_rows[: max(top_n * 4, 60)],
    }


def _inventory_consolidated_data(
    db: Session,
    branch_id: str,
) -> tuple[list[dict], Decimal, Decimal, list[Branch], Optional[Branch], list[Bodega]]:
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .order_by(Producto.descripcion)
        .all()
    )
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    selected_branch = None
    if branch_id and branch_id != "all":
        try:
            selected_branch = next((b for b in branches if b.id == int(branch_id)), None)
        except ValueError:
            selected_branch = None

    bodegas_query = _scoped_bodegas_query(db)
    if selected_branch:
        bodegas_query = bodegas_query.filter(Bodega.branch_id == selected_branch.id)
    bodegas = bodegas_query.order_by(Bodega.id).all()
    bodega_ids = [b.id for b in bodegas]
    product_ids = [p.id for p in productos]
    balances = _balances_by_bodega(db, bodega_ids, product_ids)

    rows: list[dict] = []
    total_qty = Decimal("0")
    total_cost = Decimal("0")
    for producto in productos:
        qty = Decimal("0")
        for bodega_id in bodega_ids:
            qty += balances.get((producto.id, bodega_id), Decimal("0"))
        if qty == 0:
            continue
        costo_unit = Decimal(str(producto.costo_producto or 0))
        costo_total = costo_unit * qty
        rows.append(
            {
                "codigo": producto.cod_producto,
                "descripcion": producto.descripcion,
                "cantidad": qty,
                "costo_unitario": costo_unit,
                "costo_total": costo_total,
            }
        )
        total_qty += qty
        total_cost += costo_total
    return rows, total_qty, total_cost, branches, selected_branch, bodegas


def _kardex_cost_unit_usd(
    db: Session,
    producto: Producto,
    tasa_fallback: Decimal,
) -> Decimal:
    costo_cs = Decimal(str(producto.costo_producto or 0))
    tasa = Decimal(str(producto.tasa_cambio or 0)) or tasa_fallback
    if not tasa:
        rate_today = (
            db.query(ExchangeRate)
            .filter(ExchangeRate.effective_date <= local_today())
            .order_by(ExchangeRate.effective_date.desc())
            .first()
        )
        tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
    return costo_cs / tasa if tasa else Decimal("0")


def _build_kardex_movements(
    db: Session,
    start_date: date,
    end_date: date,
    branch_id: str | None,
    producto_q: str,
):
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    producto_filter = None
    if producto_q:
        like = f"%{producto_q.lower()}%"
        producto_filter = or_(
            func.lower(Producto.cod_producto).like(like),
            func.lower(Producto.descripcion).like(like),
        )

    branch_filter = None
    if branch_id and branch_id != "all":
        try:
            branch_filter = int(branch_id)
        except ValueError:
            branch_filter = None

    movimientos = []

    ingresos_q = (
        db.query(IngresoInventario, IngresoItem, Producto, Bodega, Branch, IngresoTipo)
        .join(IngresoItem, IngresoItem.ingreso_id == IngresoInventario.id)
        .join(Producto, Producto.id == IngresoItem.producto_id)
        .join(Bodega, Bodega.id == IngresoInventario.bodega_id)
        .join(Branch, Branch.id == Bodega.branch_id)
        .join(IngresoTipo, IngresoTipo.id == IngresoInventario.tipo_id, isouter=True)
        .filter(IngresoInventario.fecha >= start_date, IngresoInventario.fecha <= end_date)
    )
    if branch_filter:
        ingresos_q = ingresos_q.filter(Branch.id == branch_filter)
    if producto_filter is not None:
        ingresos_q = ingresos_q.filter(producto_filter)

    for ingreso, item, producto, bodega, branch, ingreso_tipo in ingresos_q.all():
        cantidad = Decimal(str(item.cantidad or 0))
        costo_unit_cs = Decimal(str(item.costo_unitario_cs or 0))
        costo_unit_usd = Decimal(str(item.costo_unitario_usd or 0))
        concepto = (ingreso_tipo.nombre or "").strip() if ingreso_tipo else ""
        tipo_label = f"Ingreso - {concepto}" if concepto else "Ingreso"
        fecha_text = ingreso.fecha.isoformat() if ingreso.fecha else ""
        source_url = (
            f"/inventory/ingresos?start_date={fecha_text}&end_date={fecha_text}"
            f"&focus_ingreso_id={ingreso.id}&focus_producto_id={producto.id}"
        )
        movimientos.append(
            {
                "fecha": ingreso.fecha,
                "tipo": tipo_label,
                "tipo_base": "Ingreso",
                "concepto": concepto or "-",
                "source_kind": "ingreso",
                "source_id": ingreso.id,
                "source_url": source_url,
                "branch": branch.name if branch else "-",
                "bodega": bodega.name if bodega else "-",
                "producto_id": producto.id,
                "codigo": producto.cod_producto,
                "descripcion": producto.descripcion,
                "vendedor": "-",
                "cantidad": cantidad,
                "costo_unit_cs": costo_unit_cs,
                "costo_unit_usd": costo_unit_usd,
            }
        )

    egresos_q = (
        db.query(EgresoInventario, EgresoItem, Producto, Bodega, Branch, EgresoTipo)
        .join(EgresoItem, EgresoItem.egreso_id == EgresoInventario.id)
        .join(Producto, Producto.id == EgresoItem.producto_id)
        .join(Bodega, Bodega.id == EgresoInventario.bodega_id)
        .join(Branch, Branch.id == Bodega.branch_id)
        .join(EgresoTipo, EgresoTipo.id == EgresoInventario.tipo_id, isouter=True)
        .filter(EgresoInventario.fecha >= start_date, EgresoInventario.fecha <= end_date)
    )
    if branch_filter:
        egresos_q = egresos_q.filter(Branch.id == branch_filter)
    if producto_filter is not None:
        egresos_q = egresos_q.filter(producto_filter)

    for egreso, item, producto, bodega, branch, egreso_tipo in egresos_q.all():
        cantidad = Decimal(str(item.cantidad or 0)) * Decimal("-1")
        costo_unit_cs = Decimal(str(item.costo_unitario_cs or 0))
        costo_unit_usd = Decimal(str(item.costo_unitario_usd or 0))
        concepto = (egreso_tipo.nombre or "").strip() if egreso_tipo else ""
        tipo_label = f"Egreso - {concepto}" if concepto else "Egreso"
        fecha_text = egreso.fecha.isoformat() if egreso.fecha else ""
        source_url = (
            f"/inventory/egresos?start_date={fecha_text}&end_date={fecha_text}"
            f"&focus_egreso_id={egreso.id}&focus_producto_id={producto.id}"
        )
        movimientos.append(
            {
                "fecha": egreso.fecha,
                "tipo": tipo_label,
                "tipo_base": "Egreso",
                "concepto": concepto or "-",
                "source_kind": "egreso",
                "source_id": egreso.id,
                "source_url": source_url,
                "branch": branch.name if branch else "-",
                "bodega": bodega.name if bodega else "-",
                "producto_id": producto.id,
                "codigo": producto.cod_producto,
                "descripcion": producto.descripcion,
                "vendedor": "-",
                "cantidad": cantidad,
                "costo_unit_cs": costo_unit_cs,
                "costo_unit_usd": costo_unit_usd,
            }
        )

    ventas_q = (
        db.query(VentaFactura, VentaItem, Producto, Bodega, Branch, Vendedor)
        .join(VentaItem, VentaItem.factura_id == VentaFactura.id)
        .join(Producto, Producto.id == VentaItem.producto_id)
        .join(Bodega, Bodega.id == VentaFactura.bodega_id, isouter=True)
        .join(Branch, Branch.id == Bodega.branch_id, isouter=True)
        .join(Vendedor, Vendedor.id == VentaFactura.vendedor_id, isouter=True)
        .filter(VentaFactura.fecha >= start_dt, VentaFactura.fecha < end_dt)
        .filter(VentaFactura.estado != "ANULADA")
    )
    if branch_filter:
        ventas_q = ventas_q.filter(Branch.id == branch_filter)
    if producto_filter is not None:
        ventas_q = ventas_q.filter(producto_filter)

    for factura, item, producto, bodega, branch, vendedor in ventas_q.all():
        cantidad = Decimal(str(item.cantidad or 0)) * Decimal("-1")
        tasa_factura = Decimal(str(factura.tasa_cambio or 0))
        costo_unit_usd = _kardex_cost_unit_usd(db, producto, tasa_factura)
        costo_unit_cs = Decimal(str(producto.costo_producto or 0))
        fecha_mov = factura.fecha.date() if isinstance(factura.fecha, datetime) else factura.fecha
        fecha_text = fecha_mov.isoformat() if fecha_mov else ""
        source_url = (
            f"/sales/utilitario?start_date={fecha_text}&end_date={fecha_text}"
            f"&focus_sale_id={factura.id}&focus_producto_id={producto.id}"
        )
        movimientos.append(
            {
                "fecha": fecha_mov,
                "tipo": "Venta",
                "tipo_base": "Venta",
                "concepto": "Venta",
                "source_kind": "venta",
                "source_id": factura.id,
                "source_url": source_url,
                "branch": branch.name if branch else "-",
                "bodega": bodega.name if bodega else "-",
                "producto_id": producto.id,
                "codigo": producto.cod_producto,
                "descripcion": producto.descripcion,
                "vendedor": vendedor.nombre if vendedor else "-",
                "cantidad": cantidad,
                "costo_unit_cs": costo_unit_cs,
                "costo_unit_usd": costo_unit_usd,
            }
        )

    movimientos.sort(key=lambda row: (row["fecha"], row["tipo"]))

    saldos = {}
    rows = []
    for mov in movimientos:
        key = (mov["producto_id"], mov["bodega"])
        saldo = saldos.get(key, Decimal("0"))
        saldo += mov["cantidad"]
        saldos[key] = saldo
        costo_unit_cs = mov["costo_unit_cs"]
        costo_unit_usd = mov["costo_unit_usd"]
        rows.append(
            {
                **mov,
                "saldo": saldo,
                "costo_total_cs": saldo * costo_unit_cs,
                "costo_total_usd": saldo * costo_unit_usd,
            }
        )

    resumen = {
        "movimientos": len(rows),
        "dias": (end_date - start_date).days + 1,
        "total_ingresos": sum((r["cantidad"] for r in rows if r.get("tipo_base") == "Ingreso"), Decimal("0")),
        "total_egresos": sum((abs(r["cantidad"]) for r in rows if r.get("tipo_base") == "Egreso"), Decimal("0")),
        "total_ventas": sum((abs(r["cantidad"]) for r in rows if r.get("tipo_base") == "Venta"), Decimal("0")),
    }

    return rows, resumen


def _build_sales_report_rows(
    db: Session,
    user: User,
    start_date: date,
    end_date: date,
    branch_id: str | None,
    vendedor_id: str | None,
    producto_q: str,
):
    allowed_codes = _allowed_branch_codes(db)
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    _, bodega_user = _resolve_branch_bodega(db, user)
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    query = (
        db.query(VentaFactura, VentaItem, Producto, Cliente, Vendedor, Branch)
        .join(VentaItem, VentaItem.factura_id == VentaFactura.id)
        .join(Producto, Producto.id == VentaItem.producto_id)
        .join(Bodega, Bodega.id == VentaFactura.bodega_id, isouter=True)
        .join(Branch, Branch.id == Bodega.branch_id, isouter=True)
        .join(Cliente, Cliente.id == VentaFactura.cliente_id, isouter=True)
        .join(Vendedor, Vendedor.id == VentaFactura.vendedor_id, isouter=True)
        .filter(VentaFactura.fecha >= start_dt, VentaFactura.fecha < end_dt)
        .filter(func.lower(Branch.code).in_(allowed_codes))
        .filter(Branch.id.in_(scoped_branch_ids))
    )
    if branch_id and branch_id != "all":
        try:
            branch_id_int = int(branch_id)
            if branch_id_int not in scoped_branch_ids:
                query = query.filter(Branch.id == -1)
            else:
                query = query.filter(Branch.id == branch_id_int)
        except ValueError:
            pass
    if vendedor_id:
        try:
            query = query.filter(VentaFactura.vendedor_id == int(vendedor_id))
        except ValueError:
            pass
    if producto_q:
        query = query.filter(
            or_(
                func.lower(Producto.cod_producto).like(f"%{producto_q.lower()}%"),
                func.lower(Producto.descripcion).like(f"%{producto_q.lower()}%"),
            )
        )

    rows = query.order_by(VentaFactura.secuencia.asc(), VentaFactura.id.asc()).all()
    report_rows = []
    total_usd = Decimal("0")
    total_cs = Decimal("0")
    facturas_set = set()
    total_items = Decimal("0")
    vendedor_totals: dict[str, Decimal] = {}
    for factura, item, producto, cliente, vendedor, branch in rows:
        moneda = factura.moneda or "CS"
        tasa = Decimal(str(factura.tasa_cambio or 0))
        if moneda == "CS" and not tasa:
            rate_today = (
                db.query(ExchangeRate)
                .filter(ExchangeRate.effective_date <= factura.fecha)
                .order_by(ExchangeRate.effective_date.desc())
                .first()
            )
            tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
        subtotal = Decimal(str(item.subtotal_usd or 0)) if moneda == "USD" else Decimal(str(item.subtotal_cs or 0))
        subtotal_usd = subtotal if moneda == "USD" else (subtotal / tasa if tasa else Decimal("0"))
        subtotal_cs = subtotal if moneda == "CS" else (subtotal * tasa if tasa else Decimal("0"))
        precio_unit = (
            Decimal(str(item.precio_unitario_usd or 0))
            if moneda == "USD"
            else Decimal(str(item.precio_unitario_cs or 0))
        )
        total_factura = (
            Decimal(str(factura.total_usd or 0)) if moneda == "USD" else Decimal(str(factura.total_cs or 0))
        )
        total_factura_usd = total_factura if moneda == "USD" else (total_factura / tasa if tasa else Decimal("0"))
        total_factura_cs = total_factura if moneda == "CS" else (total_factura * tasa if tasa else Decimal("0"))
        is_anulada = factura.estado == "ANULADA"
        if not is_anulada:
            total_usd += subtotal_usd
            total_cs += subtotal_cs
            facturas_set.add(factura.id)
            total_items += Decimal(str(item.cantidad or 0))
            vendedor_name = vendedor.nombre if vendedor else "Sin asignar"
            vendedor_totals[vendedor_name] = vendedor_totals.get(vendedor_name, Decimal("0")) + subtotal_usd
        report_rows.append(
            {
                "fecha": factura.fecha.strftime("%d/%m/%Y") if factura.fecha else "",
                "factura": factura.numero,
                "cliente": cliente.nombre if cliente else "Consumidor final",
                "vendedor": vendedor.nombre if vendedor else "-",
                "sucursal": branch.name if branch else "-",
                "codigo": producto.cod_producto if producto else "",
                "producto": producto.descripcion if producto else "",
                "cantidad": float(item.cantidad or 0),
                "moneda": moneda,
                "precio_usd": float(precio_unit if moneda == "USD" else (precio_unit / tasa if tasa else Decimal("0"))),
                "precio_cs": float(precio_unit if moneda == "CS" else (precio_unit * tasa if tasa else Decimal("0"))),
                "subtotal_usd": float(subtotal_usd),
                "subtotal_cs": float(subtotal_cs),
                "total_factura_usd": float(total_factura_usd),
                "total_factura_cs": float(total_factura_cs),
                "anulada": is_anulada,
                "factura_id": factura.id,
            }
        )
    vendor_summary = [
        {"vendedor": name, "total_usd": float(total)}
        for name, total in sorted(vendedor_totals.items(), key=lambda item: item[1], reverse=True)
    ]
    return report_rows, total_usd, total_cs, len(facturas_set), float(total_items), vendor_summary


@router.get("/reports/ventas")
def report_sales_detailed(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    start_date, end_date, branch_id, vendedor_id, producto_q = _sales_report_filters(request)
    report_rows, total_usd, total_cs, total_facturas, total_items, vendor_summary = _build_sales_report_rows(
        db,
        user,
        start_date,
        end_date,
        branch_id,
        vendedor_id,
        producto_q,
    )

    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    _, bodega = _resolve_branch_bodega(db, user)
    vendedores = _vendedores_for_bodega(db, bodega)

    return request.app.state.templates.TemplateResponse(
        "report_sales_detailed.html",
        {
            "request": request,
            "user": user,
            "rows": report_rows,
            "vendor_summary": vendor_summary,
            "branches": branches,
            "vendedores": vendedores,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "selected_branch": branch_id or "",
            "selected_vendedor": vendedor_id or "",
            "producto_q": producto_q,
            "total_usd": float(total_usd),
            "total_cs": float(total_cs),
            "total_facturas": total_facturas,
            "total_items": total_items,
            "version": settings.UI_VERSION,
        },
      )


@router.get("/reports/depositos")
def report_depositos(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    start_date, end_date, branch_id = _depositos_report_filters(request)
    depositos = (
        _depositos_report_query(db, start_date, end_date, branch_id)
        .order_by(DepositoCliente.fecha.desc(), DepositoCliente.id.desc())
        .all()
    )
    branches = _scoped_branches_query(db).order_by(Branch.name).all()

    total_cs = Decimal("0")
    total_usd = Decimal("0")
    for dep in depositos:
        if dep.moneda == "USD":
            total_usd += Decimal(str(dep.monto_usd or 0))
        else:
            total_cs += Decimal(str(dep.monto_cs or 0))

    return request.app.state.templates.TemplateResponse(
        "report_depositos.html",
        {
            "request": request,
            "user": user,
            "depositos": depositos,
            "branches": branches,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "selected_branch": branch_id or "",
            "total_cs": float(total_cs),
            "total_usd": float(total_usd),
            "version": settings.UI_VERSION,
        },
    )


@router.get("/reports/depositos/export")
def report_depositos_export(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    company_profile = _company_profile_payload(db)
    start_date, end_date, branch_id = _depositos_report_filters(request)
    depositos = (
        _depositos_report_query(db, start_date, end_date, branch_id)
        .order_by(DepositoCliente.banco_id, DepositoCliente.fecha)
        .all()
    )
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    selected_branch = None
    if branch_id and branch_id != "all":
        try:
            selected_branch = next((b for b in branches if b.id == int(branch_id)), None)
        except ValueError:
            selected_branch = None

    grouped, total_cs, total_usd = _depositos_grouped(depositos)
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    rate = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
    total_usd_equiv = total_usd + (total_cs / rate if rate else Decimal("0"))

    buffer = io.BytesIO()
    width = 470
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import portrait
    from reportlab.lib.units import mm

    c = canvas.Canvas(buffer, pagesize=portrait((width, 600)))
    y = 560
    logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
    if logo_path.exists():
        c.drawImage(str(logo_path), 24, y - 40, width=90, height=40, mask="auto")
    c.setFont("Times-Bold", 11)
    c.drawString(120, y - 8, "Informe de Depositos, Transferencias y")
    c.drawString(120, y - 24, f"Tarjetas {company_profile.get('trade_name', 'Empresa')}")
    y -= 50
    c.setFont("Times-Roman", 9)
    c.setFillColor(colors.HexColor("#4b5563"))
    if selected_branch:
        c.drawString(24, y, f"Sucursal: {selected_branch.name}")
        y -= 14
    c.drawString(24, y, f"Rango: {start_date} a {end_date}")
    y -= 14
    c.drawString(24, y, f"Tasa: {rate_today.rate if rate_today else 'N/D'}")
    y -= 14
    c.setFillColor(colors.black)
    c.line(24, y, width - 24, y)
    y -= 12

    grouped_map = {}
    for dep in depositos:
        banco_name = dep.banco.nombre if dep.banco else "-"
        if len(banco_name) > 12:
            banco_name = banco_name[:12] + "…"
        grouped_map.setdefault(dep.banco_id, {"banco": banco_name, "rows": []})
        grouped_map[dep.banco_id]["rows"].append(dep)
    grouped_list = sorted(grouped_map.values(), key=lambda row: row["banco"])

    total_count = 0
    for group in grouped_list:
        if y < 90:
            c.showPage()
            y = 560
        c.setFillColor(colors.HexColor("#1e3a8a"))
        c.roundRect(24, y - 6, width - 48, 16, 4, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Times-Bold", 9)
        c.drawString(30, y - 2, group["banco"])
        c.setFillColor(colors.black)
        y -= 20

        c.setFont("Times-Bold", 8)
        c.drawString(24, y, "Fecha")
        c.drawString(95, y, "Banco")
        c.drawRightString(210, y, "Monto Cordobas")
        c.drawRightString(300, y, "Monto Dolares")
        c.drawString(310, y, "Vendedor")
        y -= 12
        c.setFont("Times-Roman", 8)

        subtotal_cs = Decimal("0")
        subtotal_usd = Decimal("0")
        for dep in group["rows"]:
            if y < 70:
                c.showPage()
                y = 560
            total_count += 1
            fecha_text = dep.fecha.strftime("%d/%m/%Y") if dep.fecha else ""
            vendedor_text = dep.vendedor.nombre if dep.vendedor else "-"
            banco_text = dep.banco.nombre if dep.banco else "-"
            c.drawString(24, y, fecha_text)
            c.drawString(95, y, banco_text[:10])
            if dep.moneda == "USD":
                monto_usd = Decimal(str(dep.monto_usd or 0))
                subtotal_usd += monto_usd
                c.setFillColor(colors.HexColor("#16a34a"))
                c.drawRightString(210, y, "C$ 0.00")
                c.drawRightString(300, y, f"$ {monto_usd:,.2f}")
                c.setFillColor(colors.black)
            else:
                monto_cs = Decimal(str(dep.monto_cs or 0))
                subtotal_cs += monto_cs
                c.setFillColor(colors.HexColor("#1d4ed8"))
                c.drawRightString(210, y, f"C$ {monto_cs:,.2f}")
                c.drawRightString(300, y, "$ 0.00")
                c.setFillColor(colors.black)
            c.drawString(310, y, vendedor_text[:18])
            y -= 12

        y -= 6
        c.setFont("Times-Bold", 9)
        c.drawString(30, y, "Total depositos :")
        c.setFillColor(colors.HexColor("#1d4ed8"))
        c.drawString(140, y, f"C$ {subtotal_cs:,.2f}")
        c.setFillColor(colors.HexColor("#16a34a"))
        c.drawRightString(width - 24, y, f"$ {subtotal_usd:,.2f}")
        c.setFillColor(colors.black)
        y -= 18
        c.line(24, y, width - 24, y)
        y -= 16

    c.setFont("Times-Bold", 10)
    c.drawString(24, y, "Totales depositos :")
    c.setFillColor(colors.HexColor("#1d4ed8"))
    c.drawRightString(width - 120, y, f"C$ {total_cs:,.2f}")
    c.setFillColor(colors.HexColor("#16a34a"))
    c.drawRightString(width - 24, y, f"$ {total_usd:,.2f}")
    c.setFillColor(colors.black)
    y -= 18
    c.setFont("Times-Bold", 10)
    c.drawString(24, y, "Totales depositos Dolarizado")
    c.setFillColor(colors.HexColor("#16a34a"))
    c.drawRightString(width - 24, y, f"$ {total_usd_equiv:,.2f}")
    c.setFillColor(colors.black)
    y -= 14
    c.setFont("Times-Roman", 9)
    c.drawRightString(width - 24, y, f"Cantidad de DP: {total_count}")

    c.showPage()
    c.save()
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=depositos_reporte.pdf"},
    )


@router.get("/reports/ventas-productos")
def report_sales_products(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    start_date, end_date, branch_id, vendedor_id, producto_id, producto_q = _sales_products_report_filters(request)
    (
        report_rows,
        detail_rows,
        total_qty,
        total_usd,
        total_cs,
        total_cost_usd,
        total_cost_cs,
        total_facturas,
    ) = _build_sales_products_report(db, user, start_date, end_date, branch_id, vendedor_id, producto_id, producto_q)

    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    branches = (
        _scoped_branches_query(db)
        .filter(Branch.id.in_(scoped_branch_ids))
        .order_by(Branch.name)
        .all()
    )
    _, bodega = _resolve_branch_bodega(db, user)
    vendedores = _vendedores_for_bodega(db, bodega)
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .order_by(Producto.descripcion)
        .all()
    )

    return request.app.state.templates.TemplateResponse(
        "report_sales_products.html",
        {
            "request": request,
            "user": user,
            "rows": report_rows,
            "detail_rows": detail_rows,
            "branches": branches,
            "vendedores": vendedores,
            "productos": productos,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "selected_branch": branch_id or "",
            "selected_vendedor": vendedor_id or "",
            "selected_producto": producto_id or "",
            "producto_q": producto_q,
            "total_qty": float(total_qty),
            "total_usd": float(total_usd),
            "total_cs": float(total_cs),
            "total_cost_usd": float(total_cost_usd),
            "total_cost_cs": float(total_cost_cs),
            "total_facturas": total_facturas,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/reports/ventas-productos/export")
def report_sales_products_export(
    request: Request,
    format: str = "xlsx",
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    company_profile = _company_profile_payload(db)
    start_date, end_date, branch_id, vendedor_id, producto_id, producto_q = _sales_products_report_filters(request)
    (
        _report_rows,
        detail_rows,
        total_qty,
        total_usd,
        total_cs,
        _total_cost_usd,
        _total_cost_cs,
        total_facturas,
    ) = _build_sales_products_report(db, user, start_date, end_date, branch_id, vendedor_id, producto_id, producto_q)

    if format.lower() == "pdf":
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4, landscape

        buffer = io.BytesIO()
        width, height = landscape(A4)
        c = canvas.Canvas(buffer, pagesize=landscape(A4))
        margin = 20
        y = height - 28

        logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
        if logo_path.exists():
            c.drawImage(str(logo_path), margin, y - 30, width=70, height=28, mask="auto")

        c.setFont("Times-Bold", 10)
        c.drawString(margin + 80, y - 8, "Reporte Ventas por Producto - Detalle por Factura")
        c.setFont("Times-Roman", 8)
        c.drawString(margin + 80, y - 20, f"Rango: {start_date} a {end_date}")
        c.drawString(margin + 80, y - 32, f"Facturas: {total_facturas}  |  Items: {float(total_qty):,.2f}")
        c.drawString(margin + 80, y - 44, f"Total USD: {float(total_usd):,.2f}  |  Total C$: {float(total_cs):,.2f}")
        y -= 58

        headers = [
            ("Fecha", 40),
            ("Factura", 62),
            ("Cliente", 120),
            ("Vendedor", 80),
            ("Suc", 52),
            ("Codigo", 52),
            ("Producto", 130),
            ("Cant", 42),
            ("P.USD", 52),
            ("P.C$", 52),
            ("V.USD", 56),
            ("V.C$", 56),
        ]
        total_width = sum(w for _, w in headers)
        if total_width > (width - (margin * 2)):
            scale = (width - (margin * 2)) / total_width
            headers = [(label, int(w * scale)) for label, w in headers]

        def draw_header(curr_y: float):
            c.setFillColor(colors.HexColor("#0f172a"))
            c.rect(margin, curr_y - 12, width - (margin * 2), 14, fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Times-Bold", 7)
            x = margin + 2
            for label, col_w in headers:
                c.drawString(x, curr_y - 8, label)
                x += col_w
            c.setFillColor(colors.black)
            c.setFont("Times-Roman", 7)
            return curr_y - 16

        y = draw_header(y)
        for row in detail_rows:
            if y < 35:
                c.showPage()
                y = height - 28
                y = draw_header(y)
            vals = [
                row.get("fecha", ""),
                row.get("factura", ""),
                row.get("cliente", ""),
                row.get("vendedor", ""),
                row.get("sucursal", ""),
                row.get("codigo", ""),
                row.get("producto", ""),
                f"{float(row.get('cantidad') or 0):,.2f}",
                f"{float(row.get('precio_usd') or 0):,.2f}",
                f"{float(row.get('precio_cs') or 0):,.2f}",
                f"{float(row.get('subtotal_usd') or 0):,.2f}",
                f"{float(row.get('subtotal_cs') or 0):,.2f}",
            ]
            x = margin + 2
            for idx, value in enumerate(vals):
                col_w = headers[idx][1]
                c.drawString(x, y, str(value)[: max(1, col_w // 4)])
                x += col_w
            y -= 11

        c.showPage()
        c.save()
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=ventas_productos_detalle.pdf"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"
    ws.append(["Reporte", "Ventas por Producto - Detalle por Factura"])
    ws.append(["Rango", f"{start_date} a {end_date}"])
    ws.append(["Facturas", int(total_facturas or 0)])
    ws.append(["Items", float(total_qty or 0)])
    ws.append(["Total USD", float(total_usd or 0)])
    ws.append(["Total C$", float(total_cs or 0)])
    ws.append([])
    headers = [
        "Fecha",
        "Factura",
        "Cliente",
        "Vendedor",
        "Sucursal",
        "Codigo",
        "Producto",
        "Cantidad",
        "Precio USD",
        "Precio C$",
        "Venta USD",
        "Venta C$",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=8, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in detail_rows:
        ws.append(
            [
                row.get("fecha", ""),
                row.get("factura", ""),
                row.get("cliente", ""),
                row.get("vendedor", ""),
                row.get("sucursal", ""),
                row.get("codigo", ""),
                row.get("producto", ""),
                float(row.get("cantidad") or 0),
                float(row.get("precio_usd") or 0),
                float(row.get("precio_cs") or 0),
                float(row.get("subtotal_usd") or 0),
                float(row.get("subtotal_cs") or 0),
            ]
        )
    widths = [12, 16, 26, 20, 16, 14, 32, 12, 12, 12, 12, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ventas_productos_detalle.xlsx"},
    )


@router.get("/reports/inventario-consolidado")
def report_inventory_consolidated(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    branch_id = _inventory_consolidated_filters(request)
    rows, total_qty, total_cost, branches, selected_branch, _ = _inventory_consolidated_data(db, branch_id)
    branch_label = selected_branch.name if selected_branch else "Ambas"
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    rate = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
    total_usd = (total_cost / rate) if rate else None
    return request.app.state.templates.TemplateResponse(
        "report_inventory_consolidated.html",
        {
            "request": request,
            "user": user,
            "rows": rows,
            "total_qty": float(total_qty),
            "total_cost": float(total_cost),
            "total_usd": float(total_usd) if total_usd is not None else None,
            "rate_value": float(rate) if rate else None,
            "total_items": len(rows),
            "branches": branches,
            "selected_branch": branch_id or "all",
            "branch_label": branch_label,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/reports/inventario-rotacion")
def report_inventory_rotation(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    (
        start_date,
        end_date,
        branch_id,
        bodega_id,
        top_n,
        slow_days,
        categoria_id,
        trend_granularity,
        sort_sales,
        min_stock_days,
        lead_days,
    ) = _inventory_rotation_filters(request)
    data = _build_inventory_rotation_data(
        db,
        start_date,
        end_date,
        branch_id,
        bodega_id,
        top_n,
        slow_days,
        categoria_id,
        trend_granularity,
        sort_sales,
        min_stock_days,
        lead_days,
    )
    selected_branch = data["selected_branch"]
    branch_label = selected_branch.name if selected_branch else "Todas las sucursales"
    selected_bodega = data.get("selected_bodega")
    bodega_label = selected_bodega.name if selected_bodega else "Todas las bodegas"

    return request.app.state.templates.TemplateResponse(
        "report_inventory_rotation.html",
        {
            "request": request,
            "user": user,
            "branches": data["branches"],
            "lineas": data["lineas"],
            "bodegas": data["bodegas"],
            "selected_branch": branch_id or "all",
            "selected_linea": categoria_id or "all",
            "selected_bodega": bodega_id or "all",
            "branch_label": branch_label,
            "bodega_label": bodega_label,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "top_n": top_n,
            "slow_days": slow_days,
            "trend_granularity": trend_granularity,
            "sort_sales": sort_sales,
            "min_stock_days": min_stock_days,
            "lead_days": lead_days,
            "month_start": data["month_start"],
            "month_end": data["month_end"],
            "kpis": data["kpis"],
            "sales_products_rows": data["sales_products_rows"],
            "balance_rows": data["balance_rows"],
            "unsold_rows": data["unsold_rows"],
            "slow_rows": data["slow_rows"],
            "expensive_rows": data["expensive_rows"],
            "rotation_rows": data["rotation_rows"],
            "reorder_rows": data["reorder_rows"],
            "abc_rows": data["abc_rows"],
            "coverage_rows": data["coverage_rows"],
            "trend_rows": data["trend_rows"],
            "version": settings.UI_VERSION,
        },
    )


@router.get("/reports/inventario-consolidado/pdf")
def report_inventory_consolidated_pdf(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    company_profile = _company_profile_payload(db)
    branch_id = _inventory_consolidated_filters(request)
    rows, total_qty, total_cost, _, selected_branch, _ = _inventory_consolidated_data(db, branch_id)
    branch_label = selected_branch.name if selected_branch else "Ambas"
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    rate = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
    total_usd = (total_cost / rate) if rate else None

    buffer = io.BytesIO()
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    width, height = A4
    c = canvas.Canvas(buffer, pagesize=A4)
    margin = 30
    y = height - 36

    logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
    if logo_path.exists():
        c.drawImage(str(logo_path), margin, y - 40, width=90, height=40, mask="auto")

    profile_phone = company_profile.get("phone", "").strip()
    profile_address = company_profile.get("address", "").strip()
    header_text = f"Reporte de inventario consolidado : {company_profile.get('trade_name', 'Empresa')}"
    if profile_address:
        header_text += f" - {profile_address}"
    if profile_phone:
        header_text += f" - {profile_phone}"
    header_x = margin + 110
    header_width = width - margin - header_x

    def wrap_pdf(text: str, max_width: float, font_name: str, font_size: int) -> list[str]:
        words = text.split()
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if c.stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = word
        if current:
            lines.append(current)
        return lines

    c.setFont("Times-Bold", 11)
    header_lines = wrap_pdf(header_text, header_width, "Times-Bold", 11)
    header_line_height = 14
    header_top = y - 8
    for idx, line in enumerate(header_lines):
        c.drawString(header_x, header_top - idx * header_line_height, line)

    date_y = header_top - len(header_lines) * header_line_height - 2
    today = local_today()
    months = [
        "enero",
        "febrero",
        "marzo",
        "abril",
        "mayo",
        "junio",
        "julio",
        "agosto",
        "septiembre",
        "octubre",
        "noviembre",
        "diciembre",
    ]
    fecha_label = f"{today.day:02d} de {months[today.month - 1]} {today.year}"
    c.setFont("Times-Roman", 10)
    c.setFillColor(colors.HexColor("#475569"))
    c.drawString(header_x, date_y, f"Fecha: {fecha_label}")
    c.drawString(header_x, date_y - 14, f"Sucursal: {branch_label}")
    c.drawString(header_x, date_y - 28, "Expresado en moneda nacional cordobas.")
    c.setFillColor(colors.black)
    y = date_y - 46

    row_height = 15

    def draw_header():
        nonlocal y
        c.setFillColor(colors.HexColor("#1e3a8a"))
        c.rect(margin, y - 12, width - margin * 2, 18, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Times-Bold", 9)
        c.drawString(margin + 4, y - 10, "Codigo")
        c.drawString(margin + 82, y - 10, "Descripcion")
        c.drawRightString(margin + 330, y - 10, "Cantidad")
        c.drawRightString(margin + 430, y - 10, "Costo Unit")
        c.drawRightString(width - margin, y - 10, "Costo Total")
        c.setFillColor(colors.black)
        c.setFont("Times-Roman", 9)
        y -= 22

    def trunc(text: str, limit: int) -> str:
        if text is None:
            return ""
        return text if len(text) <= limit else text[: limit - 3] + "..."

    draw_header()
    for row in rows:
        if y < 80:
            c.showPage()
            y = height - 36
            draw_header()
        c.drawString(margin + 4, y, trunc(row.get("codigo") or "", 10))
        c.drawString(margin + 82, y, trunc(row.get("descripcion") or "", 52))
        c.drawRightString(margin + 330, y, f"{float(row.get('cantidad') or 0):,.2f}")
        c.drawRightString(margin + 430, y, f"C$ {float(row.get('costo_unitario') or 0):,.2f}")
        c.drawRightString(width - margin, y, f"C$ {float(row.get('costo_total') or 0):,.2f}")
        y -= row_height

    if y < 60:
        c.showPage()
        y = height - 36
        draw_header()

    y -= 4
    c.line(margin, y, width - margin, y)
    y -= 18
    c.setFont("Times-Bold", 11)
    c.drawString(margin + 4, y, "Totales")
    c.drawRightString(margin + 360, y, f"{float(total_qty or 0):,.2f}")
    c.drawRightString(width - margin, y, f"C$ {float(total_cost or 0):,.2f}")
    if total_usd is not None:
        y -= 14
        c.setFont("Times-Roman", 10)
        c.setFillColor(colors.HexColor("#16a34a"))
        c.drawRightString(width - margin, y, f"Equivalencia USD: $ {float(total_usd):,.2f}")
        c.setFillColor(colors.black)

    c.showPage()
    c.save()
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=inventario_consolidado.pdf"},
    )


@router.get("/reports/kardex")
def report_kardex(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    start_date, end_date, branch_id, producto_q = _kardex_report_filters(request)
    rows, resumen = _build_kardex_movements(db, start_date, end_date, branch_id, producto_q)
    branches = _scoped_branches_query(db).order_by(Branch.name).all()

    return request.app.state.templates.TemplateResponse(
        "report_kardex.html",
        {
            "request": request,
            "user": user,
            "rows": rows,
            "resumen": resumen,
            "branches": branches,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "selected_branch": branch_id or "",
            "producto_q": producto_q,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/reports/kardex/export")
def report_kardex_export(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    company_profile = _company_profile_payload(db)
    start_date, end_date, branch_id, producto_q = _kardex_report_filters(request)
    rows, resumen = _build_kardex_movements(db, start_date, end_date, branch_id, producto_q)
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    selected_branch = None
    if branch_id and branch_id != "all":
        try:
            selected_branch = next((b for b in branches if b.id == int(branch_id)), None)
        except ValueError:
            selected_branch = None

    buffer = io.BytesIO()
    width = 380
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import portrait

    c = canvas.Canvas(buffer, pagesize=portrait((width, 700)))
    y = 660
    logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
    if logo_path.exists():
        c.drawImage(str(logo_path), 24, y - 40, width=90, height=40, mask="auto")
    c.setFont("Times-Bold", 11)
    c.drawString(120, y - 8, "Reporte Kardex por producto")
    c.drawString(120, y - 24, "Movimientos de inventario")
    y -= 50
    c.setFont("Times-Roman", 9)
    c.setFillColor(colors.HexColor("#4b5563"))
    if selected_branch:
        c.drawString(24, y, f"Sucursal: {selected_branch.name}")
        y -= 14
    c.drawString(24, y, f"Rango: {start_date} a {end_date}")
    y -= 14
    c.drawString(24, y, f"Producto: {producto_q or 'Todos'}")
    y -= 14
    c.drawString(24, y, f"Dias: {resumen['dias']}")
    y -= 14
    c.drawString(24, y, f"Ingresos: {resumen['total_ingresos']}")
    y -= 14
    c.drawString(24, y, f"Egresos: {resumen['total_egresos']}")
    y -= 14
    c.drawString(24, y, f"Ventas: {resumen['total_ventas']}")
    y -= 14
    c.setFillColor(colors.black)
    c.line(24, y, width - 24, y)
    y -= 12

    c.setFont("Times-Bold", 8)
    c.drawString(24, y, "Fecha")
    c.drawString(70, y, "Tipo")
    c.drawString(116, y, "Sucursal/Bodega")
    c.drawString(196, y, "Producto")
    c.drawRightString(252, y, "Cant")
    c.drawRightString(284, y, "Saldo")
    c.drawRightString(336, y, "C.Unit")
    c.drawRightString(388, y, "C.Total")
    c.drawString(392, y, "Vendedor")
    y -= 12
    c.setFont("Times-Roman", 8)

    for row in rows:
        if y < 70:
            c.showPage()
            y = 660
        fecha_text = row["fecha"].strftime("%d/%m/%Y") if row["fecha"] else ""
        tipo_text = row.get("concepto") or row.get("tipo") or "-"
        sucursal_text = f"{row['branch']} / {row['bodega']}"
        prod_text = f"{row['codigo']} {row['descripcion'][:10]}"
        c.drawString(24, y, fecha_text)
        c.drawString(70, y, tipo_text[:18])
        c.drawString(116, y, sucursal_text[:16])
        c.drawString(196, y, prod_text)
        c.drawRightString(252, y, f"{row['cantidad']:.2f}")
        c.drawRightString(284, y, f"{row['saldo']:.2f}")
        c.drawRightString(336, y, f"{row['costo_unit_cs']:.2f}")
        c.drawRightString(388, y, f"{row['costo_total_cs']:.2f}")
        c.drawString(392, y, (row.get("vendedor") or "-")[:10])
        y -= 12

    c.showPage()
    c.save()
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=kardex.pdf"},
    )


@router.get("/reports/ventas/export")
def report_sales_export(
    request: Request,
    format: str = "xlsx",
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.reports")
    scoped_branch_ids = _user_scoped_branch_ids(db, user)
    company_profile = _company_profile_payload(db)
    start_date, end_date, branch_id, vendedor_id, producto_q = _sales_report_filters(request)
    report_rows, total_usd, total_cs, total_facturas, total_items, vendor_summary = _build_sales_report_rows(
        db,
        user,
        start_date,
        end_date,
        branch_id,
        vendedor_id,
        producto_q,
    )

    if format.lower() == "pdf":
        buffer = io.BytesIO()
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors

        width, height = A4
        c = canvas.Canvas(buffer, pagesize=A4)
        margin = 24
        y = height - 36

        logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
        if logo_path.exists():
            c.drawImage(str(logo_path), margin, y - 42, width=90, height=40, mask="auto")

        c.setFont("Times-Bold", 10)
        c.drawString(margin + 110, y - 10, "Reporte Detallado de Ventas")
        c.setFont("Times-Roman", 8)
        c.setFillColor(colors.HexColor("#475569"))
        c.drawString(margin + 110, y - 24, f"Rango: {start_date} a {end_date}")

        selected_branch = None
        if branch_id and branch_id != "all":
            try:
                branch_id_int = int(branch_id)
                if branch_id_int in scoped_branch_ids:
                    selected_branch = db.query(Branch).filter(Branch.id == branch_id_int).first()
            except Exception:
                selected_branch = None
        branch_label = selected_branch.name if selected_branch else "Todas"
        c.drawString(margin + 110, y - 40, f"Sucursal: {branch_label}")
        c.setFillColor(colors.black)
        y -= 64

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

        rate_cache: dict[date, Decimal] = {}
        def rate_for_date(value_date: date | datetime | None) -> Decimal:
            if not value_date:
                return Decimal('0')
            key = value_date if isinstance(value_date, date) and not isinstance(value_date, datetime) else value_date.date()
            if key not in rate_cache:
                rate_row = (
                    db.query(ExchangeRate)
                    .filter(ExchangeRate.effective_date <= key)
                    .order_by(ExchangeRate.effective_date.desc())
                    .first()
                )
                rate_cache[key] = Decimal(str(rate_row.rate)) if rate_row else Decimal('0')
            return rate_cache[key]

        def to_usd(moneda: str, monto_cs: Decimal, monto_usd: Decimal, tasa: Decimal, value_date) -> Decimal:
            if moneda == 'USD':
                return Decimal(str(monto_usd or 0))
            rate = tasa if tasa else rate_for_date(value_date)
            return (Decimal(str(monto_cs or 0)) / rate) if rate else Decimal('0')

        bodega_ids = [row[0] for row in db.query(Bodega.id).filter(Bodega.branch_id.in_(scoped_branch_ids)).all()]
        if branch_id and branch_id != 'all':
            try:
                branch_int = int(branch_id)
                if branch_int in scoped_branch_ids:
                    bodega_ids = [row[0] for row in db.query(Bodega.id).filter(Bodega.branch_id == branch_int).all()]
                else:
                    bodega_ids = []
            except Exception:
                bodega_ids = []

        total_ventas_usd = Decimal('0')
        ventas_query = db.query(VentaFactura).filter(
            VentaFactura.fecha >= start_dt,
            VentaFactura.fecha < end_dt,
            VentaFactura.estado != 'ANULADA',
        )
        if bodega_ids is not None:
            ventas_query = ventas_query.filter(VentaFactura.bodega_id.in_(bodega_ids))
        if vendedor_id:
            try:
                ventas_query = ventas_query.filter(VentaFactura.vendedor_id == int(vendedor_id))
            except Exception:
                pass
        for factura in ventas_query.all():
            moneda = factura.moneda or 'CS'
            tasa = Decimal(str(factura.tasa_cambio or 0))
            total_ventas_usd += to_usd(moneda, Decimal(str(factura.total_cs or 0)), Decimal(str(factura.total_usd or 0)), tasa, factura.fecha)

        total_egresos_usd = Decimal('0')
        recibos_query = db.query(ReciboCaja).filter(
            ReciboCaja.fecha.between(start_date, end_date),
            ReciboCaja.afecta_caja.is_(True),
            ReciboCaja.tipo == 'EGRESO',
        )
        if bodega_ids is not None:
            recibos_query = recibos_query.filter(ReciboCaja.bodega_id.in_(bodega_ids))
        for recibo in recibos_query.all():
            moneda = recibo.moneda or 'CS'
            tasa = Decimal(str(recibo.tasa_cambio or 0))
            total_egresos_usd += to_usd(moneda, Decimal(str(recibo.monto_cs or 0)), Decimal(str(recibo.monto_usd or 0)), tasa, recibo.fecha)

        total_depositos_usd = Decimal('0')
        depositos_query = db.query(DepositoCliente).filter(
            DepositoCliente.fecha.between(start_date, end_date),
        )
        if bodega_ids is not None:
            depositos_query = depositos_query.filter(DepositoCliente.bodega_id.in_(bodega_ids))
        if vendedor_id:
            try:
                depositos_query = depositos_query.filter(DepositoCliente.vendedor_id == int(vendedor_id))
            except Exception:
                pass
        for dep in depositos_query.all():
            moneda = dep.moneda or 'CS'
            tasa = Decimal(str(dep.tasa_cambio or 0))
            total_depositos_usd += to_usd(moneda, Decimal(str(dep.monto_cs or 0)), Decimal(str(dep.monto_usd or 0)), tasa, dep.fecha)

        total_creditos_usd = Decimal('0')
        creditos_query = db.query(VentaFactura).filter(
            VentaFactura.fecha >= start_dt,
            VentaFactura.fecha < end_dt,
            VentaFactura.estado != 'ANULADA',
            VentaFactura.estado_cobranza == 'PENDIENTE',
        )
        if bodega_ids is not None:
            creditos_query = creditos_query.filter(VentaFactura.bodega_id.in_(bodega_ids))
        if vendedor_id:
            try:
                creditos_query = creditos_query.filter(VentaFactura.vendedor_id == int(vendedor_id))
            except Exception:
                pass
        for factura in creditos_query.all():
            moneda = factura.moneda or 'CS'
            tasa = Decimal(str(factura.tasa_cambio or 0))
            if moneda == 'USD':
                paid_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos)
                due_usd = Decimal(str(factura.total_usd or 0))
                saldo_usd = max(due_usd - paid_usd, Decimal('0'))
                total_creditos_usd += saldo_usd
            else:
                paid_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos)
                due_cs = Decimal(str(factura.total_cs or 0))
                saldo_cs = max(due_cs - paid_cs, Decimal('0'))
                total_creditos_usd += to_usd('CS', saldo_cs, Decimal('0'), tasa, factura.fecha)

        total_residuo_usd = (
            Decimal(str(total_ventas_usd))
            - Decimal(str(total_depositos_usd))
            - Decimal(str(total_egresos_usd))
            - Decimal(str(total_creditos_usd))
        )

        cobranza_rows = []
        total_creditos_pendientes_usd = Decimal("0")
        cobranza_query = db.query(VentaFactura).filter(
            VentaFactura.fecha >= start_dt,
            VentaFactura.fecha < end_dt,
            VentaFactura.estado != "ANULADA",
            VentaFactura.estado_cobranza == "PENDIENTE",
        )
        if bodega_ids is not None:
            cobranza_query = cobranza_query.filter(VentaFactura.bodega_id.in_(bodega_ids))
        if vendedor_id:
            try:
                cobranza_query = cobranza_query.filter(VentaFactura.vendedor_id == int(vendedor_id))
            except Exception:
                pass

        for factura in cobranza_query.order_by(VentaFactura.fecha, VentaFactura.numero).all():
            moneda = factura.moneda or "CS"
            tasa = Decimal(str(factura.tasa_cambio or 0))
            total_usd = to_usd(
                moneda,
                Decimal(str(factura.total_cs or 0)),
                Decimal(str(factura.total_usd or 0)),
                tasa,
                factura.fecha,
            )
            abonos_usd = Decimal("0")
            for abono in factura.abonos:
                abono_moneda = abono.moneda or "CS"
                abono_tasa = Decimal(str(abono.tasa_cambio or 0))
                abonos_usd += to_usd(
                    abono_moneda,
                    Decimal(str(abono.monto_cs or 0)),
                    Decimal(str(abono.monto_usd or 0)),
                    abono_tasa,
                    abono.fecha,
                )
            saldo_usd = max(total_usd - abonos_usd, Decimal("0"))
            if saldo_usd <= 0:
                continue
            total_creditos_pendientes_usd += saldo_usd
            cobranza_rows.append(
                {
                    "factura": factura.numero,
                    "cliente": factura.cliente.nombre if factura.cliente else "Consumidor final",
                    "vendedor": factura.vendedor.nombre if factura.vendedor else "-",
                    "total_usd": total_usd,
                    "abono_usd": abonos_usd,
                    "saldo_usd": saldo_usd,
                }
            )

        content_width = width - (margin * 2)
        factura_x = margin
        cliente_x = margin + 60
        producto_x = margin + 170
        vendedor_x = margin + 320
        qty_right = margin + content_width - 140
        price_right = margin + content_width - 70
        subtotal_right = margin + content_width

        def max_chars_for_width(width_pts: float, font_size: int = 8) -> int:
            return max(8, int(width_pts / (font_size * 0.55)))

        def draw_header():
            nonlocal y
            c.setFont("Times-Bold", 8)
            c.setFillColor(colors.HexColor("#1e293b"))
            c.drawString(factura_x, y, "Factura")
            c.drawString(cliente_x, y, "Cliente")
            c.drawString(producto_x, y, "Producto")
            c.drawString(vendedor_x, y, "Vendedor")
            c.drawRightString(qty_right, y, "Cant")
            c.drawRightString(price_right, y, "Precio")
            c.drawRightString(subtotal_right, y, "Subtotal")
            c.setFillColor(colors.HexColor("#e2e8f0"))
            c.line(margin, y - 6, width - margin, y - 6)
            c.setFillColor(colors.black)
            c.setFont("Times-Roman", 8)
            y -= 16

        def trunc(text: str, limit: int) -> str:
            if text is None:
                return ""
            return text if len(text) <= limit else text[: limit - 3] + "..."

        def wrap_lines(text: str, limit: int, max_lines: int = 2) -> list[str]:
            if not text:
                return [""]
            words = text.split()
            lines: list[str] = []
            current = ""
            for word in words:
                candidate = f"{current} {word}".strip()
                if len(candidate) <= limit:
                    current = candidate
                else:
                    lines.append(current)
                    current = word
                    if len(lines) >= max_lines:
                        break
            if len(lines) < max_lines and current:
                lines.append(current)
            if len(lines) > max_lines:
                lines = lines[:max_lines]
            if len(lines) == max_lines and len(words) > 0:
                full = " ".join(lines)
                if len(full) < len(text):
                    lines[-1] = trunc(lines[-1], max(3, limit - 1))
            return lines or [trunc(text, limit)]

        draw_header()
        line_height = 12
        for row in report_rows:
            moneda = row.get("moneda") or "CS"
            label = "$" if moneda == "USD" else "C$"
            precio = row["precio_usd"] if moneda == "USD" else row["precio_cs"]
            subtotal = row["subtotal_usd"] if moneda == "USD" else row["subtotal_cs"]
            product_text = f"{row.get('producto') or ''}".strip()
            product_limit = max_chars_for_width(vendedor_x - producto_x - 8, 8)
            producto_lines = wrap_lines(product_text, product_limit, 2)
            row_height = (line_height * max(1, len(producto_lines))) + 4

            if y - row_height < 70:
                c.showPage()
                y = height - 36
                draw_header()

            row_y = y
            c.drawString(factura_x, row_y, trunc(str(row["factura"] or ""), 16))
            c.drawString(cliente_x, row_y, trunc(str(row.get("cliente") or ""), 22))
            if row.get("anulada"):
                c.setFont("Times-Bold", 8)
                c.drawString(producto_x, row_y, "ANULADO")
                c.setFont("Times-Roman", 8)
            else:
                c.drawString(producto_x, row_y, producto_lines[0] if producto_lines else "")
                c.drawString(vendedor_x, row_y, trunc(str(row.get("vendedor") or ""), 16))
                c.drawRightString(qty_right, row_y, f"{row.get('cantidad', 0):,.2f}")
                c.drawRightString(price_right, row_y, f"{label} {float(precio or 0):,.2f}")
                c.drawRightString(subtotal_right, row_y, f"{label} {float(subtotal or 0):,.2f}")

                if len(producto_lines) > 1:
                    c.drawString(producto_x, row_y - line_height, producto_lines[1])

            y -= row_height

        rate_row = (
            db.query(ExchangeRate)
            .filter(ExchangeRate.effective_date <= end_date)
            .order_by(ExchangeRate.effective_date.desc())
            .first()
        )
        rate = Decimal(str(rate_row.rate)) if rate_row else Decimal("0")
        total_cs_decimal = Decimal(str(total_cs or 0))
        total_usd_conv = (total_cs_decimal / rate) if rate > 0 else Decimal("0")

        y -= 10
        c.setFont("Times-Bold", 10)
        c.setFillColor(colors.HexColor("#1e3a8a"))
        c.drawString(margin, y, "Totales")
        c.setFillColor(colors.black)
        y -= 16
        c.setFont("Times-Roman", 9)
        c.drawString(margin, y, f"Total final (C$): {total_cs:,.2f}")
        c.drawString(margin + 220, y, f"Bultos vendidos: {float(total_items or 0):,.2f}")
        y -= 14
        c.drawString(margin, y, f"Total C$: {total_cs:,.2f}")
        if rate > 0:
            c.drawString(margin + 220, y, f"Total USD conversión: {float(total_usd_conv):,.2f}")
            c.drawString(margin + 420, y, f"Tasa: {float(rate):,.4f}")
        else:
            c.drawString(margin + 220, y, "Total USD conversión: -")
        y -= 24

        if y < 120:
            c.showPage()
            y = height - 50

        c.setFont("Times-Bold", 10)
        c.setFillColor(colors.HexColor("#1e3a8a"))
        c.drawString(margin, y, "Resumen arqueo (USD)")
        c.setFillColor(colors.black)
        y -= 16
        c.setFont("Times-Roman", 9)
        c.drawString(margin, y, f"Total ventas: $ {float(total_ventas_usd):,.2f}")
        y -= 12
        c.drawString(margin, y, f"- Depositos: $ {float(total_depositos_usd):,.2f}")
        y -= 12
        c.drawString(margin, y, f"- Gastos recibos de caja: $ {float(total_egresos_usd):,.2f}")
        y -= 12
        c.drawString(margin, y, f"- Pendientes deudas: $ {float(total_creditos_usd):,.2f}")
        y -= 12
        c.setFont("Times-Bold", 9)
        c.drawString(margin, y, f"Total residuo esperado (efectivo): $ {float(total_residuo_usd):,.2f}")
        y -= 22

        c.setFont("Times-Bold", 10)
        c.setFillColor(colors.HexColor("#1e3a8a"))
        if y < 90:
            c.showPage()
            y = height - 50
        c.drawString(margin, y, "Resumen por vendedor (USD)")
        c.setFillColor(colors.black)
        y -= 16
        c.setFont("Times-Roman", 9)
        for row in vendor_summary:
            if y < 50:
                c.showPage()
                y = height - 60
            c.drawString(margin, y, trunc(row["vendedor"], 25))
            c.drawRightString(margin + 260, y, f"$ {float(row['total_usd'] or 0):,.2f}")
            y -= 14
        c.showPage()
        y = height - 50
        c.setFont("Times-Bold", 10)
        c.setFillColor(colors.HexColor("#1e3a8a"))
        c.drawString(margin, y, "Reporte de Cuentas por Cobrar - Anexo.")
        c.setFillColor(colors.black)
        y -= 16

        annex_content_width = width - (margin * 2)
        factura_x = margin
        cliente_x = margin + 70
        vendedor_x = margin + 250
        monto_right = margin + annex_content_width - 120
        abono_right = margin + annex_content_width - 60
        saldo_right = margin + annex_content_width

        def draw_annex_header():
            nonlocal y
            c.setFont("Times-Bold", 8)
            c.setFillColor(colors.HexColor("#1e293b"))
            c.drawString(factura_x, y, "Factura")
            c.drawString(cliente_x, y, "Cliente")
            c.drawString(vendedor_x, y, "Vendedor")
            c.drawRightString(monto_right, y, "Monto USD")
            c.drawRightString(abono_right, y, "Abono USD")
            c.drawRightString(saldo_right, y, "Saldo USD")
            c.setFillColor(colors.HexColor("#e2e8f0"))
            c.line(margin, y - 6, width - margin, y - 6)
            c.setFillColor(colors.black)
            c.setFont("Times-Roman", 8)
            y -= 16

        draw_annex_header()
        factura_limit = max_chars_for_width(cliente_x - factura_x - 6, 8)
        cliente_limit = max_chars_for_width(vendedor_x - cliente_x - 6, 8)
        vendedor_limit = max_chars_for_width(monto_right - vendedor_x - 6, 8)

        for row in cobranza_rows:
            if y < 70:
                c.showPage()
                y = height - 50
                c.setFont("Times-Bold", 10)
                c.setFillColor(colors.HexColor("#1e3a8a"))
                c.drawString(margin, y, "Reporte de Cuentas por Cobrar - Anexo.")
                c.setFillColor(colors.black)
                y -= 16
                draw_annex_header()

            c.drawString(factura_x, y, trunc(str(row["factura"] or ""), factura_limit))
            c.drawString(cliente_x, y, trunc(row["cliente"] or "-", cliente_limit))
            c.drawString(vendedor_x, y, trunc(row["vendedor"] or "-", vendedor_limit))
            c.drawRightString(monto_right, y, f"$ {float(row['total_usd'] or 0):,.2f}")
            c.drawRightString(abono_right, y, f"$ {float(row['abono_usd'] or 0):,.2f}")
            c.drawRightString(saldo_right, y, f"$ {float(row['saldo_usd'] or 0):,.2f}")
            y -= 12

        if y < 90:
            c.showPage()
            y = height - 50
            c.setFont("Times-Bold", 10)
            c.setFillColor(colors.HexColor("#1e3a8a"))
            c.drawString(margin, y, "Reporte de Cuentas por Cobrar - Anexo.")
            c.setFillColor(colors.black)
            y -= 16
        c.setFont("Times-Bold", 9)
        c.drawRightString(
            saldo_right,
            y,
            f"Total creditos pendientes: $ {float(total_creditos_pendientes_usd):,.2f}",
        )
        y -= 14

        c.save()
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "inline; filename=ventas_detallado.pdf"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas Detallado"
    headers = [
        "Fecha",
        "Factura",
        "Cliente",
        "Vendedor",
        "Sucursal",
        "Codigo",
        "Producto",
        "Cantidad",
        "Moneda",
        "Precio",
        "Subtotal",
        "Total Factura",
        "Estado",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in report_rows:
        estado = "ANULADA" if row["anulada"] else "ACTIVA"
        ws.append(
            [
                row["fecha"],
                row["factura"],
                row["cliente"],
                row["vendedor"],
                row["sucursal"],
                row.get("codigo", ""),
                row["producto"],
                row["cantidad"],
                row["moneda"],
                row["precio_usd"] if row["moneda"] == "USD" else row["precio_cs"],
                row["subtotal_usd"] if row["moneda"] == "USD" else row["subtotal_cs"],
                row["total_factura_usd"] if row["moneda"] == "USD" else row["total_factura_cs"],
                estado,
            ]
        )

    ws.append([])
    ws.append(["Total USD", float(total_usd)])
    ws.append(["Total C$", float(total_cs)])
    ws.append(["Facturas", total_facturas])
    ws.append(["Bultos", total_items])
    ws.append([])
    ws.append(["Resumen vendedores (USD)"])
    for row in vendor_summary:
        ws.append([row["vendedor"], float(row["total_usd"])])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": "attachment; filename=ventas_detallado.xlsx"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _depositos_filters(request: Request):
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    vendedor_q = request.query_params.get("vendedor_id")
    banco_q = request.query_params.get("banco_id")
    moneda_q = request.query_params.get("moneda")
    today_value = local_today()
    start_date = today_value
    end_date = today_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today_value
            end_date = today_value
    return start_date, end_date, vendedor_q, banco_q, moneda_q


def _depositos_grouped(depositos):
    summary = {}
    total_cs = Decimal("0")
    total_usd = Decimal("0")
    for dep in depositos:
        key = dep.banco_id
        if key not in summary:
            summary[key] = {
                "banco": dep.banco.nombre if dep.banco else "-",
                "total_cs": Decimal("0"),
                "total_usd": Decimal("0"),
            }
        monto_cs = Decimal(str(dep.monto_cs or 0))
        monto_usd = Decimal(str(dep.monto_usd or 0))
        if dep.moneda == "USD":
            summary[key]["total_usd"] += monto_usd
            total_usd += monto_usd
        else:
            summary[key]["total_cs"] += monto_cs
            total_cs += monto_cs
    grouped = sorted(summary.values(), key=lambda row: row["banco"])
    return grouped, total_cs, total_usd


@router.get("/sales/depositos/export")
def sales_depositos_export(
    request: Request,
    format: str = "xlsx",
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.depositos")
    company_profile = _company_profile_payload(db)
    start_date, end_date, vendedor_q, banco_q, moneda_q = _depositos_filters(request)
    branch, bodega = _resolve_branch_bodega(db, user)
    vendedores = _vendedores_for_bodega(db, bodega)
    depositos_query = db.query(DepositoCliente)
    if bodega:
        depositos_query = depositos_query.filter(DepositoCliente.bodega_id == bodega.id)
    depositos_query = depositos_query.filter(DepositoCliente.fecha.between(start_date, end_date))
    if vendedor_q:
        depositos_query = depositos_query.filter(DepositoCliente.vendedor_id == int(vendedor_q))
    if banco_q:
        depositos_query = depositos_query.filter(DepositoCliente.banco_id == int(banco_q))
    if moneda_q:
        depositos_query = depositos_query.filter(DepositoCliente.moneda == moneda_q.upper())
    depositos = depositos_query.order_by(DepositoCliente.banco_id, DepositoCliente.fecha).all()

    grouped, total_cs, total_usd = _depositos_grouped(depositos)
    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    rate = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
    total_usd_equiv = total_usd + (total_cs / rate if rate else Decimal("0"))

    if format.lower() == "pdf":
        buffer = io.BytesIO()
        width = 380
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import portrait
        from reportlab.lib.units import mm

        c = canvas.Canvas(buffer, pagesize=portrait((width, 600)))
        y = 560
        logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
        if logo_path.exists():
            c.drawImage(str(logo_path), 24, y - 40, width=90, height=40, mask="auto")
        c.setFont("Times-Bold", 11)
        c.drawString(120, y - 8, "Informe de Depositos, Transferencias y")
        c.drawString(120, y - 24, f"Tarjetas {company_profile.get('trade_name', 'Empresa')}")
        y -= 50
        c.setFont("Times-Roman", 9)
        c.setFillColor(colors.HexColor("#4b5563"))
        if branch:
            c.drawString(24, y, f"Sucursal: {branch.name}")
            y -= 14
        c.drawString(24, y, f"Rango: {start_date} a {end_date}")
        y -= 14
        c.drawString(24, y, f"Tasa: {rate_today.rate if rate_today else 'N/D'}")
        y -= 14
        c.setFillColor(colors.black)
        c.line(24, y, width - 24, y)
        y -= 12

        grouped_map = {}
        for dep in depositos:
            banco_name = dep.banco.nombre if dep.banco else "-"
            if len(banco_name) > 12:
                banco_name = banco_name[:12] + "…"
            grouped_map.setdefault(dep.banco_id, {"banco": banco_name, "rows": []})
            grouped_map[dep.banco_id]["rows"].append(dep)
        grouped_list = sorted(grouped_map.values(), key=lambda row: row["banco"])

        total_count = 0
        for group in grouped_list:
            if y < 90:
                c.showPage()
                y = 560
            c.setFillColor(colors.HexColor("#1e3a8a"))
            c.roundRect(24, y - 6, width - 48, 16, 4, fill=1, stroke=0)
            c.setFillColor(colors.white)
            c.setFont("Times-Bold", 9)
            c.drawString(30, y - 2, group["banco"])
            c.setFillColor(colors.black)
            y -= 20

            c.setFont("Times-Bold", 8)
            c.setFillColor(colors.HexColor("#475569"))
            c.drawString(30, y, "Fecha")
            c.drawString(78, y, "Banco")
            c.drawString(140, y, "Monto Cordobas")
            c.drawString(230, y, "Monto Dolares")
            c.drawString(305, y, "Vendedor")
            y -= 8
            c.setFillColor(colors.black)
            c.line(24, y, width - 24, y)
            y -= 10

            subtotal_cs = Decimal("0")
            subtotal_usd = Decimal("0")
            for dep in group["rows"]:
                if y < 90:
                    c.showPage()
                    y = 560
                monto_cs = Decimal(str(dep.monto_cs or 0))
                monto_usd = Decimal(str(dep.monto_usd or 0))
                total_count += 1
                c.setFont("Times-Roman", 7)
                c.setFillColor(colors.black)
                c.drawString(30, y, str(dep.fecha))
                banco_row = dep.banco.nombre if dep.banco else "-"
                if len(banco_row) > 12:
                    banco_row = banco_row[:12] + "…"
                c.drawString(78, y, banco_row)
                c.setFillColor(colors.HexColor("#1d4ed8"))
                display_cs = monto_cs if dep.moneda == "CS" else Decimal("0")
                display_usd = monto_usd if dep.moneda == "USD" else Decimal("0")
                c.drawString(140, y, f"C$ {display_cs:,.2f}")
                c.setFillColor(colors.HexColor("#16a34a"))
                c.drawString(230, y, f"$ {display_usd:,.2f}")
                c.setFillColor(colors.black)
                c.drawString(305, y, dep.vendedor.nombre if dep.vendedor else "-")
                y -= 12
                subtotal_cs += display_cs
                subtotal_usd += display_usd

            y -= 4
            c.setFont("Times-Bold", 8)
            c.setFillColor(colors.HexColor("#475569"))
            c.drawString(30, y, "Total depositos :")
            c.setFillColor(colors.HexColor("#1d4ed8"))
            c.drawString(140, y, f"C$ {subtotal_cs:,.2f}")
            c.setFillColor(colors.HexColor("#16a34a"))
            c.drawString(230, y, f"$ {subtotal_usd:,.2f}")
            c.setFillColor(colors.black)
            y -= 16

        y -= 6
        c.setFont("Times-Bold", 9)
        c.line(24, y, width - 24, y)
        y -= 14
        c.drawString(24, y, "Totales depositos :")
        c.setFillColor(colors.HexColor("#1d4ed8"))
        c.drawRightString(width - 120, y, f"C$ {total_cs:,.2f}")
        c.setFillColor(colors.HexColor("#16a34a"))
        c.drawRightString(width - 24, y, f"$ {total_usd:,.2f}")
        c.setFillColor(colors.black)
        y -= 14
        c.drawString(24, y, "Totales depositos Dolarizado")
        c.setFillColor(colors.HexColor("#16a34a"))
        c.drawRightString(width - 24, y, f"$ {total_usd_equiv:,.2f}")
        c.setFillColor(colors.black)
        y -= 14
        c.drawString(24, y, f"Cantidad de DP: {total_count}")
        c.showPage()
        c.save()
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=depositos.pdf"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Depositos"
    ws["A1"] = "Informe de Depositos"
    ws["A2"] = f"Rango: {start_date} a {end_date}"
    ws["A3"] = f"Tasa: {rate_today.rate if rate_today else 'N/D'}"
    ws["A5"] = "Banco"
    ws["B5"] = "Total C$"
    ws["C5"] = "Total USD"
    for cell in ("A1", "A5", "B5", "C5"):
        ws[cell].font = Font(bold=True)
    row_idx = 6
    for row in grouped:
        ws.cell(row=row_idx, column=1, value=row["banco"])
        ws.cell(row=row_idx, column=2, value=float(row["total_cs"]))
        ws.cell(row=row_idx, column=3, value=float(row["total_usd"]))
        row_idx += 1
    ws.cell(row=row_idx, column=1, value="Totales").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=float(total_cs)).font = Font(bold=True)
    ws.cell(row=row_idx, column=3, value=float(total_usd)).font = Font(bold=True)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Total USD (convertido)").font = Font(bold=True)
    ws.cell(row=row_idx, column=3, value=float(total_usd_equiv)).font = Font(bold=True)
    for col in range(1, 4):
        ws.column_dimensions[chr(64 + col)].width = 20
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=depositos.xlsx"},
    )


@router.post("/sales/depositos")
async def sales_depositos_create(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.depositos")
    form = await request.form()
    deposito_id = form.get("deposito_id")
    vendedor_id = form.get("vendedor_id")
    banco_id = form.get("banco_id")
    cuenta_id = form.get("cuenta_id") or None
    fecha_raw = form.get("fecha")
    moneda = (form.get("moneda") or "CS").upper()
    monto_raw = form.get("monto")
    observacion = (form.get("observacion") or "").strip()

    if not vendedor_id or not banco_id or not monto_raw:
        return RedirectResponse("/sales/depositos?error=Datos+incompletos", status_code=303)
    if moneda not in {"CS", "USD"}:
        return RedirectResponse("/sales/depositos?error=Moneda+no+valida", status_code=303)

    def parse_decimal(value: str) -> Decimal:
        raw = re.sub(r"[^0-9.,-]", "", str(value or "0"))
        if "," in raw and "." in raw:
            if raw.rfind(",") > raw.rfind("."):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "," in raw and "." not in raw:
            parts = raw.split(",")
            if len(parts) == 2 and len(parts[1]) == 2:
                raw = raw.replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "." in raw and "," not in raw:
            parts = raw.split(".")
            if len(parts) == 2 and len(parts[1]) == 2:
                raw = raw
            else:
                raw = raw.replace(".", "")
        try:
            return Decimal(raw)
        except Exception:
            return Decimal("0")

    monto = parse_decimal(monto_raw)
    if monto <= 0:
        return RedirectResponse("/sales/depositos?error=Monto+no+valido", status_code=303)

    if fecha_raw:
        try:
            fecha_value = date.fromisoformat(str(fecha_raw).split("T")[0])
        except ValueError:
            fecha_value = local_today()
    else:
        fecha_value = local_today()

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if moneda == "USD" and not rate_today:
        return RedirectResponse("/sales/depositos?error=Tasa+de+cambio+no+configurada", status_code=303)
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

    branch, bodega = _resolve_branch_bodega(db, user)
    if not branch:
        return RedirectResponse("/sales/depositos?error=Usuario+sin+sucursal+asignada", status_code=303)
    if not bodega:
        return RedirectResponse("/sales/depositos?error=Bodega+no+configurada+para+la+sucursal", status_code=303)

    vendedor = db.query(Vendedor).filter(Vendedor.id == int(vendedor_id)).first()
    banco = db.query(Banco).filter(Banco.id == int(banco_id)).first()
    if not vendedor or not banco:
        return RedirectResponse("/sales/depositos?error=Vendedor+o+banco+no+valido", status_code=303)
    cuenta = None
    if cuenta_id:
        cuenta = db.query(CuentaBancaria).filter(CuentaBancaria.id == int(cuenta_id)).first()

    if moneda == "USD":
        monto_usd = monto
        monto_cs = monto * tasa
    else:
        monto_cs = monto
        monto_usd = monto / tasa if tasa else Decimal("0")

    if deposito_id:
        deposito = (
            db.query(DepositoCliente)
            .filter(DepositoCliente.id == int(deposito_id), DepositoCliente.bodega_id == bodega.id)
            .first()
        )
        if not deposito:
            return RedirectResponse("/sales/depositos?error=Deposito+no+encontrado", status_code=303)
        deposito.vendedor_id = vendedor.id
        deposito.banco_id = banco.id
        deposito.cuenta_id = cuenta.id if cuenta else None
        deposito.fecha = fecha_value
        deposito.moneda = moneda
        deposito.tasa_cambio = tasa if tasa else None
        deposito.monto_usd = monto_usd
        deposito.monto_cs = monto_cs
        deposito.observacion = observacion
        deposito.usuario_registro = user.full_name
    else:
        deposito = DepositoCliente(
            branch_id=branch.id,
            bodega_id=bodega.id,
            vendedor_id=vendedor.id,
            banco_id=banco.id,
            cuenta_id=cuenta.id if cuenta else None,
            fecha=fecha_value,
            moneda=moneda,
            tasa_cambio=tasa if tasa else None,
            monto_usd=monto_usd,
            monto_cs=monto_cs,
            observacion=observacion,
            usuario_registro=user.full_name,
        )
        db.add(deposito)
    db.commit()

    msg = "Deposito+actualizado" if deposito_id else "Deposito+registrado"
    return RedirectResponse(f"/sales/depositos?success={msg}", status_code=303)


@router.post("/sales/depositos/{deposito_id}/delete")
def sales_depositos_delete(
    request: Request,
    deposito_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.depositos")
    _, bodega = _resolve_branch_bodega(db, user)
    query = db.query(DepositoCliente).filter(DepositoCliente.id == deposito_id)
    if bodega:
        query = query.filter(DepositoCliente.bodega_id == bodega.id)
    deposito = query.first()
    if not deposito:
        return RedirectResponse("/sales/depositos?error=Deposito+no+encontrado", status_code=303)
    db.delete(deposito)
    db.commit()
    return RedirectResponse("/sales/depositos?success=Deposito+eliminado", status_code=303)


@router.get("/sales/{venta_id}/detail")
def sales_detail(
    request: Request,
    venta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.cobranza")
    factura = (
        db.query(VentaFactura)
        .filter(VentaFactura.id == venta_id)
        .first()
    )
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)

    items = []
    for item in factura.items:
        items.append(
            {
                "producto_id": item.producto_id,
                "codigo": item.producto.cod_producto if item.producto else "",
                "descripcion": item.producto.descripcion if item.producto else "",
                "cantidad": float(item.cantidad or 0),
                "precio_usd": float(item.precio_unitario_usd or 0),
                "precio_cs": float(item.precio_unitario_cs or 0),
                "subtotal_usd": float(item.subtotal_usd or 0),
                "subtotal_cs": float(item.subtotal_cs or 0),
            }
        )

    return JSONResponse(
        {
            "ok": True,
            "factura": {
                "id": factura.id,
                "numero": factura.numero,
                "fecha": factura.fecha.isoformat() if factura.fecha else "",
                "hora": factura.created_at.strftime("%H:%M") if factura.created_at else "",
                "cliente": factura.cliente.nombre if factura.cliente else "Consumidor final",
                "vendedor": factura.vendedor.nombre if factura.vendedor else "-",
                "sucursal": factura.bodega.branch.name if factura.bodega and factura.bodega.branch else "-",
                "bodega": factura.bodega.name if factura.bodega else "-",
                "moneda": factura.moneda,
                "total_usd": float(factura.total_usd or 0),
                "total_cs": float(factura.total_cs or 0),
                "items": items,
            },
        }
    )


@router.post("/sales/{venta_id}/reprint")
def sales_reprint(
    venta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    factura = (
        db.query(VentaFactura)
        .filter(VentaFactura.id == venta_id)
        .first()
    )
    if not factura or not factura.bodega or not factura.bodega.branch:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    return JSONResponse(
        {
            "ok": True,
            "print_url": f"/sales/{factura.id}/ticket/print?copies=2",
        }
    )


@router.post("/sales/{venta_id}/reversion/request")
async def sales_reversion_request(
    venta_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.reversion")
    form = await request.form()
    motivo = (form.get("motivo") or "").strip()
    if not motivo:
        return JSONResponse({"ok": False, "message": "Motivo requerido"}, status_code=400)

    factura = db.query(VentaFactura).filter(VentaFactura.id == venta_id).first()
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    if factura.estado == "ANULADA":
        return JSONResponse({"ok": False, "message": "Factura ya anulada"}, status_code=400)
    if factura.abonos and len(factura.abonos) > 0:
        return JSONResponse({"ok": False, "message": "No se puede anular con abonos aplicados"}, status_code=400)

    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and factura.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Factura fuera de tu bodega"}, status_code=403)

    config = db.query(EmailConfig).first()
    if not config or not config.active:
        return JSONResponse({"ok": False, "message": "Configura el correo emisor"}, status_code=400)
    recipients = (
        db.query(NotificationRecipient)
        .filter(NotificationRecipient.active.is_(True))
        .all()
    )
    recipient_emails = [r.email for r in recipients]
    if not recipient_emails:
        return JSONResponse({"ok": False, "message": "No hay destinatarios activos"}, status_code=400)

    token = _generate_token()
    expires_at = datetime.utcnow() + timedelta(minutes=10)
    db.add(
        ReversionToken(
            factura_id=factura.id,
            token=token,
            motivo=motivo,
            solicitado_por=user.full_name,
            expires_at=expires_at,
        )
    )
    db.commit()

    branch = factura.bodega.branch if factura.bodega else None
    sucursal = branch.name if branch else "-"
    bodega_name = factura.bodega.name if factura.bodega else "-"
    cliente = factura.cliente.nombre if factura.cliente else "Consumidor final"
    vendedor = factura.vendedor.nombre if factura.vendedor else "-"
    fecha_base = factura.created_at or factura.fecha
    fecha_str = ""
    hora_str = ""
    if fecha_base:
        try:
            fecha_str = fecha_base.strftime("%d/%m/%Y")
            hora_str = fecha_base.strftime("%H:%M")
        except AttributeError:
            fecha_str = str(fecha_base)
    moneda = factura.moneda or "CS"
    total_amount = float(factura.total_cs or 0) if moneda == "CS" else float(factura.total_usd or 0)
    currency_label = "C$" if moneda == "CS" else "$"

    items_html = "".join(
        f"""
        <tr>
          <td style="padding:6px 8px;border-bottom:1px solid #eee;">{item.producto.cod_producto if item.producto else ''}</td>
          <td style="padding:6px 8px;border-bottom:1px solid #eee;">{item.producto.descripcion if item.producto else ''}</td>
          <td style="padding:6px 8px;border-bottom:1px solid #eee;text-align:right;">{float(item.cantidad or 0):.2f}</td>
          <td style="padding:6px 8px;border-bottom:1px solid #eee;text-align:right;">{currency_label} {float(item.subtotal_cs if moneda == 'CS' else item.subtotal_usd or 0):,.2f}</td>
        </tr>
        """
        for item in factura.items
    )

    html_body = f"""
    <div style="font-family:Arial,sans-serif;background:#f7f4fb;padding:24px;">
      <div style="max-width:780px;margin:0 auto;background:#ffffff;border-radius:18px;padding:28px;border:1px solid #eadff2;">
        <div style="margin-bottom:16px;">
          <div style="font-size:12px;letter-spacing:2px;text-transform:uppercase;color:#8b7aa8;">Solicitud de anulacion</div>
          <h2 style="margin:6px 0 0;color:#5b2a86;">Sucursal: {sucursal}</h2>
          <div style="margin-top:6px;color:#6a5b86;font-size:14px;">Serie/Factura: <strong>{factura.numero}</strong> | Bodega: {bodega_name}</div>
        </div>

        <div style="background:#f3eefd;border:1px solid #e2d7f5;border-radius:14px;padding:14px;text-align:center;margin-bottom:18px;">
          <div style="font-size:12px;text-transform:uppercase;color:#7b6a98;">Codigo de autorizacion</div>
          <div style="font-size:26px;font-weight:700;color:#5b2a86;letter-spacing:3px;margin-top:6px;">{token}</div>
        </div>

        <div style="background:#faf7ff;border:1px solid #eee3fb;border-radius:12px;padding:12px;margin-bottom:18px;">
          <div style="font-size:12px;text-transform:uppercase;color:#7b6a98;margin-bottom:6px;">Motivo</div>
          <div style="font-size:14px;color:#3b2f52;font-weight:600;">{motivo}</div>
        </div>

        <table style="width:100%;font-size:14px;color:#333;margin-bottom:16px;">
          <tr><td>Cliente:</td><td><strong>{cliente}</strong></td></tr>
          <tr><td>Vendedor:</td><td>{vendedor}</td></tr>
          <tr><td>Usuario:</td><td>{user.full_name}</td></tr>
          <tr><td>Fecha/Hora:</td><td>{fecha_str} {hora_str}</td></tr>
          <tr><td>Total:</td><td><strong>{currency_label} {total_amount:,.2f}</strong></td></tr>
        </table>

        <h3 style="margin:12px 0;color:#5b2a86;">Detalle de items</h3>
        <table style="width:100%;border-collapse:collapse;font-size:13px;">
          <thead>
            <tr style="background:#f1e8f8;color:#5b2a86;text-align:left;">
              <th style="padding:6px 8px;">Codigo</th>
              <th style="padding:6px 8px;">Descripcion</th>
              <th style="padding:6px 8px;text-align:right;">Cantidad</th>
              <th style="padding:6px 8px;text-align:right;">Subtotal</th>
            </tr>
          </thead>
          <tbody>
            {items_html}
          </tbody>
        </table>
      </div>
    </div>
    """

    send_error = _send_reversion_email(
        subject=f"Reversion {factura.numero}",
        html_body=html_body,
        recipients=recipient_emails,
        sender_email=config.sender_email if config else None,
        sender_name=config.sender_name if config else None,
    )
    if send_error:
        return JSONResponse({"ok": False, "message": send_error}, status_code=500)

    return JSONResponse({"ok": True, "message": "Codigo enviado"})


@router.post("/sales/{venta_id}/reversion/confirm")
async def sales_reversion_confirm(
    venta_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.reversion")
    form = await request.form()
    token = (form.get("token") or "").strip()
    if not token:
        return JSONResponse({"ok": False, "message": "Codigo requerido"}, status_code=400)

    factura = db.query(VentaFactura).filter(VentaFactura.id == venta_id).first()
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    if factura.estado == "ANULADA":
        return JSONResponse({"ok": False, "message": "Factura ya anulada"}, status_code=400)
    if factura.abonos and len(factura.abonos) > 0:
        return JSONResponse({"ok": False, "message": "No se puede anular con abonos aplicados"}, status_code=400)

    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and factura.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Factura fuera de tu bodega"}, status_code=403)

    token_row = (
        db.query(ReversionToken)
        .filter(
            ReversionToken.factura_id == factura.id,
            ReversionToken.token == token,
            ReversionToken.used_at.is_(None),
        )
        .order_by(ReversionToken.created_at.desc())
        .first()
    )
    if not token_row:
        return JSONResponse({"ok": False, "message": "Codigo invalido"}, status_code=400)
    if token_row.expires_at < datetime.utcnow():
        return JSONResponse({"ok": False, "message": "Codigo expirado"}, status_code=400)

    def to_decimal(value: Optional[float]) -> Decimal:
        return Decimal(str(value or 0))

    for item in factura.items:
        if item.producto and item.producto.saldo:
            existencia_actual = to_decimal(item.producto.saldo.existencia)
            item.producto.saldo.existencia = existencia_actual + to_decimal(item.cantidad)

    factura.estado = "ANULADA"
    factura.reversion_motivo = token_row.motivo
    factura.revertida_por = user.full_name
    factura.revertida_at = local_now_naive()
    token_row.used_at = local_now_naive()
    db.commit()
    return JSONResponse({"ok": True, "message": "Factura anulada"})

@router.get("/data")
def data_home(
    request: Request,
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data")
    return request.app.state.templates.TemplateResponse(
        "data.html",
        {
            "request": request,
            "user": user,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/data/entornos")
def data_entornos(
    request: Request,
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    profiles = get_company_profiles()
    active_key = get_active_company_key()
    current_database_url = get_current_database_url()
    return request.app.state.templates.TemplateResponse(
        "data_entornos.html",
        {
            "request": request,
            "user": user,
            "profiles": profiles,
            "active_key": active_key,
            "current_database_url": current_database_url,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/entornos")
def data_entornos_create(
    request: Request,
    company_key: str = Form(...),
    company_name: str = Form(...),
    database_url: str = Form(...),
    activate: Optional[str] = Form(None),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    normalized_key = _normalize_company_key(company_key)
    if not normalized_key:
        return RedirectResponse("/data/entornos?error=Clave+de+empresa+invalida", status_code=303)
    database_url = (database_url or "").strip()
    if not database_url:
        return RedirectResponse("/data/entornos?error=DATABASE_URL+requerida", status_code=303)

    connect_error = _validate_database_url(database_url)
    if connect_error:
        return RedirectResponse(f"/data/entornos?{urlencode({'error': connect_error})}", status_code=303)

    activate_now = activate == "on"
    try:
        upsert_company_profile(
            key=normalized_key,
            name=company_name,
            database_url=database_url,
            activate=activate_now,
        )
        if activate_now or normalized_key == get_active_company_key():
            refresh_engine(force=True)
            init_db()
    except ValueError as exc:
        return RedirectResponse(f"/data/entornos?{urlencode({'error': str(exc)})}", status_code=303)

    success_message = "Perfil de empresa guardado"
    if activate_now:
        success_message += " y activado"
    return RedirectResponse(f"/data/entornos?{urlencode({'success': success_message})}", status_code=303)


@router.post("/data/entornos/{company_key}/update")
def data_entornos_update(
    request: Request,
    company_key: str,
    company_name: str = Form(...),
    database_url: str = Form(...),
    activate: Optional[str] = Form(None),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    normalized_key = _normalize_company_key(company_key)
    if not normalized_key:
        return RedirectResponse("/data/entornos?error=Clave+de+empresa+invalida", status_code=303)

    connect_error = _validate_database_url(database_url)
    if connect_error:
        return RedirectResponse(f"/data/entornos?{urlencode({'error': connect_error})}", status_code=303)

    activate_now = activate == "on"
    try:
        upsert_company_profile(
            key=normalized_key,
            name=company_name,
            database_url=database_url,
            activate=activate_now,
        )
        if activate_now or normalized_key == get_active_company_key():
            refresh_engine(force=True)
            init_db()
    except ValueError as exc:
        return RedirectResponse(f"/data/entornos?{urlencode({'error': str(exc)})}", status_code=303)

    return RedirectResponse("/data/entornos?success=Perfil+actualizado", status_code=303)


@router.post("/data/entornos/{company_key}/activar")
def data_entornos_activate(
    request: Request,
    company_key: str,
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    normalized_key = _normalize_company_key(company_key)
    if not normalized_key:
        return RedirectResponse("/data/entornos?error=Empresa+invalida", status_code=303)

    profiles = get_company_profiles()
    profile = next((item for item in profiles if item["key"] == normalized_key), None)
    if not profile:
        return RedirectResponse("/data/entornos?error=Empresa+no+registrada", status_code=303)

    connect_error = _validate_database_url(profile["database_url"])
    if connect_error:
        return RedirectResponse(f"/data/entornos?{urlencode({'error': connect_error})}", status_code=303)

    set_active_company(normalized_key)
    refresh_engine(force=True)
    init_db()
    return RedirectResponse("/data/entornos?success=Empresa+activa+actualizada", status_code=303)


@router.get("/data/interfaz-ventas")
def data_interfaz_ventas(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    setting = _get_sales_interface_setting(db)
    return request.app.state.templates.TemplateResponse(
        "data_interfaz_ventas.html",
        {
            "request": request,
            "user": user,
            "setting": setting,
            "options": SALES_INTERFACE_OPTIONS,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/interfaz-ventas")
def data_interfaz_ventas_update(
    request: Request,
    interface_code: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    selected = (interface_code or "").strip().lower()
    if selected not in _allowed_sales_interface_codes():
        return RedirectResponse("/data/interfaz-ventas?error=Interfaz+no+valida", status_code=303)

    setting = _get_sales_interface_setting(db)
    setting.interface_code = selected
    setting.updated_by = user.full_name
    db.commit()
    return RedirectResponse("/data/interfaz-ventas?success=Interfaz+de+ventas+actualizada", status_code=303)


@router.get("/data/empresa")
def data_empresa(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    profile = _get_company_profile_setting(db)
    return request.app.state.templates.TemplateResponse(
        "data_empresa.html",
        {
            "request": request,
            "user": user,
            "profile": profile,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/empresa")
def data_empresa_update(
    request: Request,
    legal_name: str = Form(...),
    trade_name: str = Form(...),
    app_title: str = Form(...),
    sidebar_subtitle: str = Form(...),
    website: Optional[str] = Form(None),
    ruc: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    logo_url: Optional[str] = Form(None),
    pos_logo_url: Optional[str] = Form(None),
    favicon_url: Optional[str] = Form(None),
    inventory_cs_only: Optional[str] = Form(None),
    multi_branch_enabled: Optional[str] = Form(None),
    price_auto_from_cost_enabled: Optional[str] = Form(None),
    price_margin_percent: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    profile = _get_company_profile_setting(db)
    profile.legal_name = (legal_name or "").strip() or "Empresa"
    profile.trade_name = (trade_name or "").strip() or profile.legal_name
    profile.app_title = (app_title or "").strip() or f"ERP {profile.trade_name}"
    profile.sidebar_subtitle = (sidebar_subtitle or "").strip() or "ERP"
    profile.website = (website or "").strip()
    profile.ruc = (ruc or "").strip()
    profile.phone = (phone or "").strip()
    profile.address = (address or "").strip()
    profile.email = (email or "").strip()
    profile.logo_url = (logo_url or "").strip() or "/static/logo_hollywood.png"
    profile.pos_logo_url = (pos_logo_url or "").strip() or profile.logo_url
    profile.favicon_url = (favicon_url or "").strip() or "/static/favicon.ico"
    profile.inventory_cs_only = inventory_cs_only == "on"
    profile.multi_branch_enabled = multi_branch_enabled == "on"
    raw_margin = (price_margin_percent or "").strip()
    margin_value = 0
    if raw_margin:
        if not raw_margin.isdigit():
            return RedirectResponse("/data/empresa?error=Porcentaje+de+ganancia+invalido+(solo+enteros)", status_code=303)
        margin_value = int(raw_margin)
    auto_margin_enabled = price_auto_from_cost_enabled == "on"
    profile.price_auto_from_cost_enabled = auto_margin_enabled
    profile.price_margin_percent = margin_value
    profile.updated_by = user.full_name
    db.commit()
    return RedirectResponse("/data/empresa?success=Perfil+empresarial+actualizado", status_code=303)


@router.get("/data/pos-print")
def data_pos_print(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(PosPrintSetting).filter(PosPrintSetting.id == int(edit_id)).first()
    items = db.query(PosPrintSetting).order_by(PosPrintSetting.branch_id).all()
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    return request.app.state.templates.TemplateResponse(
        "data_pos_print.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "branches": branches,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.get("/data/notificaciones")
def data_notificaciones(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    config = db.query(EmailConfig).first()
    recipients = db.query(NotificationRecipient).order_by(NotificationRecipient.email).all()
    return request.app.state.templates.TemplateResponse(
        "data_notificaciones.html",
        {
            "request": request,
            "user": user,
            "config": config,
            "recipients": recipients,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/notificaciones/config")
def data_notificaciones_config(
    request: Request,
    sender_email: str = Form(...),
    sender_name: Optional[str] = Form(None),
    active: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    sender_email = sender_email.strip()
    if not sender_email:
        return RedirectResponse("/data/notificaciones?error=Correo+requerido", status_code=303)
    config = db.query(EmailConfig).first()
    if not config:
        config = EmailConfig(sender_email=sender_email)
        db.add(config)
    config.sender_email = sender_email
    config.sender_name = sender_name.strip() if sender_name else None
    config.active = active == "on"
    db.commit()
    return RedirectResponse("/data/notificaciones?success=Configuracion+guardada", status_code=303)


@router.post("/data/notificaciones/recipients")
def data_notificaciones_add_recipient(
    request: Request,
    email: str = Form(...),
    name: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    email = email.strip().lower()
    if not email:
        return RedirectResponse("/data/notificaciones?error=Correo+requerido", status_code=303)
    exists = db.query(NotificationRecipient).filter(NotificationRecipient.email == email).first()
    if exists:
        return RedirectResponse("/data/notificaciones?error=Correo+ya+existe", status_code=303)
    db.add(NotificationRecipient(email=email, name=name, active=True))
    db.commit()
    return RedirectResponse("/data/notificaciones?success=Destinatario+agregado", status_code=303)


@router.post("/data/notificaciones/recipients/{recipient_id}/update")
def data_notificaciones_update_recipient(
    request: Request,
    recipient_id: int,
    email: str = Form(...),
    name: Optional[str] = Form(None),
    active: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    recipient = db.query(NotificationRecipient).filter(NotificationRecipient.id == recipient_id).first()
    if not recipient:
        return RedirectResponse("/data/notificaciones?error=Destinatario+no+existe", status_code=303)
    recipient.email = email.strip().lower()
    recipient.name = name
    recipient.active = active == "on"
    db.commit()
    return RedirectResponse("/data/notificaciones?success=Destinatario+actualizado", status_code=303)


@router.post("/data/pos-print")
def data_pos_print_save(
    request: Request,
    branch_id: int = Form(...),
    printer_name: str = Form(...),
    copies: int = Form(2),
    auto_print: Optional[str] = Form(None),
    roc_printer_name: Optional[str] = Form(None),
    roc_copies: Optional[int] = Form(None),
    roc_auto_print: Optional[str] = Form(None),
    cierre_printer_name: Optional[str] = Form(None),
    cierre_copies: Optional[int] = Form(None),
    cierre_auto_print: Optional[str] = Form(None),
    sumatra_path: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    printer_name = printer_name.strip()
    if not printer_name or not branch_id:
        return RedirectResponse("/data/pos-print?error=Datos+incompletos", status_code=303)
    setting = db.query(PosPrintSetting).filter(PosPrintSetting.branch_id == branch_id).first()
    if setting:
        setting.printer_name = printer_name
        setting.copies = max(copies, 2)
        setting.auto_print = auto_print == "on"
        setting.roc_printer_name = roc_printer_name.strip() if roc_printer_name else None
        setting.roc_copies = max(roc_copies or 1, 1) if roc_copies is not None else None
        setting.roc_auto_print = roc_auto_print == "on"
        setting.cierre_printer_name = cierre_printer_name.strip() if cierre_printer_name else None
        setting.cierre_copies = max(cierre_copies or 1, 1) if cierre_copies is not None else None
        setting.cierre_auto_print = cierre_auto_print == "on"
        setting.sumatra_path = sumatra_path.strip() if sumatra_path else None
    else:
        setting = PosPrintSetting(
            branch_id=branch_id,
            printer_name=printer_name,
            copies=max(copies, 2),
            auto_print=auto_print == "on",
            roc_printer_name=roc_printer_name.strip() if roc_printer_name else None,
            roc_copies=max(roc_copies or 1, 1) if roc_copies is not None else None,
            roc_auto_print=roc_auto_print == "on",
            cierre_printer_name=cierre_printer_name.strip() if cierre_printer_name else None,
            cierre_copies=max(cierre_copies or 1, 1) if cierre_copies is not None else None,
            cierre_auto_print=cierre_auto_print == "on",
            sumatra_path=sumatra_path.strip() if sumatra_path else None,
        )
        db.add(setting)
    db.commit()
    return RedirectResponse("/data/pos-print?success=Configuracion+guardada", status_code=303)


@router.post("/data/pos-print/{setting_id}/update")
def data_pos_print_update(
    request: Request,
    setting_id: int,
    branch_id: int = Form(...),
    printer_name: str = Form(...),
    copies: int = Form(2),
    auto_print: Optional[str] = Form(None),
    roc_printer_name: Optional[str] = Form(None),
    roc_copies: Optional[int] = Form(None),
    roc_auto_print: Optional[str] = Form(None),
    cierre_printer_name: Optional[str] = Form(None),
    cierre_copies: Optional[int] = Form(None),
    cierre_auto_print: Optional[str] = Form(None),
    sumatra_path: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    setting = db.query(PosPrintSetting).filter(PosPrintSetting.id == setting_id).first()
    if not setting:
        return RedirectResponse("/data/pos-print?error=Configuracion+no+existe", status_code=303)
    setting.branch_id = branch_id
    setting.printer_name = printer_name.strip()
    setting.copies = max(copies, 2)
    setting.auto_print = auto_print == "on"
    setting.roc_printer_name = roc_printer_name.strip() if roc_printer_name else None
    setting.roc_copies = max(roc_copies or 1, 1) if roc_copies is not None else None
    setting.roc_auto_print = roc_auto_print == "on"
    setting.cierre_printer_name = cierre_printer_name.strip() if cierre_printer_name else None
    setting.cierre_copies = max(cierre_copies or 1, 1) if cierre_copies is not None else None
    setting.cierre_auto_print = cierre_auto_print == "on"
    setting.sumatra_path = sumatra_path.strip() if sumatra_path else None
    db.commit()
    return RedirectResponse("/data/pos-print?success=Configuracion+actualizada", status_code=303)


@router.get("/data/vendedores")
def data_vendedores(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(Vendedor).filter(Vendedor.id == int(edit_id)).first()
    items = db.query(Vendedor).order_by(Vendedor.nombre).all()
    bodegas = _scoped_bodegas_query(db).order_by(Bodega.name).all()
    edit_bodega_ids = []
    edit_default_bodega_id = None
    if edit_item:
        edit_bodega_ids = [asig.bodega_id for asig in (edit_item.assignments or [])]
        default_assignment = next((asig for asig in (edit_item.assignments or []) if asig.is_default), None)
        edit_default_bodega_id = default_assignment.bodega_id if default_assignment else None
    return request.app.state.templates.TemplateResponse(
        "data_vendedores.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "edit_item": edit_item,
            "bodegas": bodegas,
            "edit_bodega_ids": edit_bodega_ids,
            "edit_default_bodega_id": edit_default_bodega_id,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/vendedores")
def data_create_vendedor(
    request: Request,
    nombre: str = Form(...),
    telefono: Optional[str] = Form(None),
    bodega_ids: Optional[list[str]] = Form(None),
    default_bodega_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    nombre = nombre.strip()
    if not nombre:
        return RedirectResponse("/data/vendedores?error=Nombre+requerido", status_code=303)
    exists = db.query(Vendedor).filter(func.lower(Vendedor.nombre) == nombre.lower()).first()
    if exists:
        return RedirectResponse("/data/vendedores?error=Vendedor+ya+existe", status_code=303)
    vendedor = Vendedor(nombre=nombre, telefono=telefono, activo=True)
    db.add(vendedor)
    db.flush()
    selected_ids = {int(b) for b in (bodega_ids or []) if str(b).strip()}
    if default_bodega_id:
        selected_ids.add(int(default_bodega_id))
    for bodega_id in sorted(selected_ids):
        db.add(
            VendedorBodega(
                vendedor_id=vendedor.id,
                bodega_id=bodega_id,
                is_default=default_bodega_id and int(default_bodega_id) == bodega_id,
            )
        )
    db.commit()
    return RedirectResponse("/data/vendedores?success=Vendedor+creado", status_code=303)


@router.post("/data/vendedores/{item_id}/update")
def data_update_vendedor(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    telefono: Optional[str] = Form(None),
    activo: Optional[str] = Form(None),
    bodega_ids: Optional[list[str]] = Form(None),
    default_bodega_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    vendedor = db.query(Vendedor).filter(Vendedor.id == item_id).first()
    if not vendedor:
        return RedirectResponse("/data/vendedores?error=Vendedor+no+existe", status_code=303)
    nombre = nombre.strip()
    if not nombre:
        return RedirectResponse("/data/vendedores?error=Nombre+requerido", status_code=303)
    exists = (
        db.query(Vendedor)
        .filter(func.lower(Vendedor.nombre) == nombre.lower(), Vendedor.id != item_id)
        .first()
    )
    if exists:
        return RedirectResponse("/data/vendedores?error=Ya+existe+otro+vendedor+con+ese+nombre", status_code=303)
    vendedor.nombre = nombre
    vendedor.telefono = telefono
    vendedor.activo = activo == "on"
    selected_ids = {int(b) for b in (bodega_ids or []) if str(b).strip()}
    if default_bodega_id:
        selected_ids.add(int(default_bodega_id))
    vendedor.assignments.clear()
    for bodega_id in sorted(selected_ids):
        vendedor.assignments.append(
            VendedorBodega(
                vendedor_id=vendedor.id,
                bodega_id=bodega_id,
                is_default=default_bodega_id and int(default_bodega_id) == bodega_id,
            )
        )
    db.commit()
    return RedirectResponse("/data/vendedores?success=Vendedor+actualizado", status_code=303)


@router.post("/data/vendedores/{item_id}/toggle")
def data_toggle_vendedor(
    request: Request,
    item_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    vendedor = db.query(Vendedor).filter(Vendedor.id == item_id).first()
    if not vendedor:
        return RedirectResponse("/data/vendedores?error=Vendedor+no+existe", status_code=303)
    vendedor.activo = not bool(vendedor.activo)
    db.commit()
    if vendedor.activo:
        return RedirectResponse("/data/vendedores?success=Vendedor+activado", status_code=303)
    return RedirectResponse("/data/vendedores?success=Vendedor+desactivado", status_code=303)


@router.get("/data/bancos")
def data_bancos(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(Banco).filter(Banco.id == int(edit_id)).first()
    items = db.query(Banco).order_by(Banco.nombre).all()
    return request.app.state.templates.TemplateResponse(
        "data_bancos.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/bancos")
def data_create_banco(
    request: Request,
    nombre: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    nombre = nombre.strip()
    if not nombre:
        return RedirectResponse("/data/bancos?error=Nombre+requerido", status_code=303)
    exists = db.query(Banco).filter(Banco.nombre == nombre).first()
    if exists:
        return RedirectResponse("/data/bancos?error=Banco+ya+existe", status_code=303)
    banco = Banco(nombre=nombre)
    db.add(banco)
    db.commit()
    return RedirectResponse("/data/bancos?success=Banco+creado", status_code=303)


@router.post("/data/bancos/{item_id}/update")
def data_update_banco(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    banco = db.query(Banco).filter(Banco.id == item_id).first()
    if not banco:
        return RedirectResponse("/data/bancos?error=Banco+no+existe", status_code=303)
    banco.nombre = nombre.strip()
    db.commit()
    return RedirectResponse("/data/bancos?success=Banco+actualizado", status_code=303)


@router.get("/data/sucursales")
def data_sucursales(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(Branch).filter(Branch.id == int(edit_id)).first()
    items = db.query(Branch).order_by(Branch.name).all()
    return request.app.state.templates.TemplateResponse(
        "data_sucursales.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/sucursales")
def data_create_sucursal(
    request: Request,
    code: str = Form(...),
    name: str = Form(...),
    company_name: str = Form(...),
    ruc: str = Form(...),
    telefono: str = Form(...),
    direccion: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    code = code.strip().lower()
    name = name.strip()
    company_name = company_name.strip()
    ruc = ruc.strip()
    telefono = telefono.strip()
    direccion = direccion.strip()
    if not code or not name or not company_name or not ruc or not telefono or not direccion:
        return RedirectResponse("/data/sucursales?error=Datos+incompletos", status_code=303)
    exists = (
        db.query(Branch)
        .filter((func.lower(Branch.code) == code) | (func.lower(Branch.name) == name.lower()))
        .first()
    )
    if exists:
        return RedirectResponse("/data/sucursales?error=Sucursal+ya+existe", status_code=303)
    db.add(
        Branch(
            code=code,
            name=name,
            company_name=company_name,
            ruc=ruc,
            telefono=telefono,
            direccion=direccion,
        )
    )
    db.commit()
    return RedirectResponse("/data/sucursales?success=Sucursal+creada", status_code=303)


@router.post("/data/sucursales/{item_id}/update")
def data_update_sucursal(
    request: Request,
    item_id: int,
    code: str = Form(...),
    name: str = Form(...),
    company_name: str = Form(...),
    ruc: str = Form(...),
    telefono: str = Form(...),
    direccion: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    branch = db.query(Branch).filter(Branch.id == item_id).first()
    if not branch:
        return RedirectResponse("/data/sucursales?error=Sucursal+no+existe", status_code=303)
    code = code.strip().lower()
    name = name.strip()
    company_name = company_name.strip()
    ruc = ruc.strip()
    telefono = telefono.strip()
    direccion = direccion.strip()
    if not code or not name or not company_name or not ruc or not telefono or not direccion:
        return RedirectResponse("/data/sucursales?error=Datos+incompletos", status_code=303)
    exists = (
        db.query(Branch)
        .filter(Branch.id != item_id)
        .filter((func.lower(Branch.code) == code) | (func.lower(Branch.name) == name.lower()))
        .first()
    )
    if exists:
        return RedirectResponse("/data/sucursales?error=Sucursal+ya+existe", status_code=303)
    branch.code = code
    branch.name = name
    branch.company_name = company_name
    branch.ruc = ruc
    branch.telefono = telefono
    branch.direccion = direccion
    db.commit()
    return RedirectResponse("/data/sucursales?success=Sucursal+actualizada", status_code=303)


@router.get("/data/bodegas")
def data_bodegas(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(Bodega).filter(Bodega.id == int(edit_id)).first()
    items = _scoped_bodegas_query(db).order_by(Bodega.name).all()
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    return request.app.state.templates.TemplateResponse(
        "data_bodegas.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "branches": branches,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/bodegas")
def data_create_bodega(
    request: Request,
    code: str = Form(...),
    name: str = Form(...),
    branch_id: int = Form(...),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    code = code.strip().lower()
    name = name.strip()
    if not code or not name:
        return RedirectResponse("/data/bodegas?error=Datos+incompletos", status_code=303)
    branch = db.query(Branch).filter(Branch.id == branch_id).first()
    if not branch:
        return RedirectResponse("/data/bodegas?error=Sucursal+no+valida", status_code=303)
    exists = (
        db.query(Bodega)
        .filter(func.lower(Bodega.code) == code)
        .first()
    )
    if exists:
        return RedirectResponse("/data/bodegas?error=Bodega+ya+existe", status_code=303)
    db.add(Bodega(code=code, name=name, branch_id=branch.id, activo=activo == "on"))
    db.commit()
    return RedirectResponse("/data/bodegas?success=Bodega+creada", status_code=303)


@router.post("/data/bodegas/{item_id}/update")
def data_update_bodega(
    request: Request,
    item_id: int,
    code: str = Form(...),
    name: str = Form(...),
    branch_id: int = Form(...),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    bodega = db.query(Bodega).filter(Bodega.id == item_id).first()
    if not bodega:
        return RedirectResponse("/data/bodegas?error=Bodega+no+existe", status_code=303)
    code = code.strip().lower()
    name = name.strip()
    if not code or not name:
        return RedirectResponse("/data/bodegas?error=Datos+incompletos", status_code=303)
    branch = db.query(Branch).filter(Branch.id == branch_id).first()
    if not branch:
        return RedirectResponse("/data/bodegas?error=Sucursal+no+valida", status_code=303)
    exists = (
        db.query(Bodega)
        .filter(Bodega.id != item_id)
        .filter(func.lower(Bodega.code) == code)
        .first()
    )
    if exists:
        return RedirectResponse("/data/bodegas?error=Bodega+ya+existe", status_code=303)
    bodega.code = code
    bodega.name = name
    bodega.branch_id = branch.id
    bodega.activo = activo == "on"
    db.commit()
    return RedirectResponse("/data/bodegas?success=Bodega+actualizada", status_code=303)


@router.get("/data/formas-pago")
def data_formas_pago(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(FormaPago).filter(FormaPago.id == int(edit_id)).first()
    items = db.query(FormaPago).order_by(FormaPago.nombre).all()
    return request.app.state.templates.TemplateResponse(
        "data_formas_pago.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/formas-pago")
def data_create_forma_pago(
    request: Request,
    nombre: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    nombre = nombre.strip()
    if not nombre:
        return RedirectResponse("/data/formas-pago?error=Nombre+requerido", status_code=303)
    exists = db.query(FormaPago).filter(FormaPago.nombre == nombre).first()
    if exists:
        return RedirectResponse("/data/formas-pago?error=Forma+ya+existe", status_code=303)
    forma = FormaPago(nombre=nombre)
    db.add(forma)
    db.commit()
    return RedirectResponse("/data/formas-pago?success=Forma+creada", status_code=303)


@router.post("/data/formas-pago/{item_id}/update")
def data_update_forma_pago(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    forma = db.query(FormaPago).filter(FormaPago.id == item_id).first()
    if not forma:
        return RedirectResponse("/data/formas-pago?error=Forma+no+existe", status_code=303)
    forma.nombre = nombre.strip()
    db.commit()
    return RedirectResponse("/data/formas-pago?success=Forma+actualizada", status_code=303)


@router.get("/data/recibos-rubros")
def data_recibos_rubros(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(ReciboRubro).filter(ReciboRubro.id == int(edit_id)).first()
    items = db.query(ReciboRubro).order_by(ReciboRubro.nombre).all()
    cuentas = db.query(CuentaContable).filter(CuentaContable.activo.is_(True)).order_by(CuentaContable.codigo).all()
    return request.app.state.templates.TemplateResponse(
        "data_recibos_rubros.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "cuentas": cuentas,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/recibos-rubros")
def data_recibos_rubros_create(
    request: Request,
    nombre: str = Form(...),
    cuenta_id: Optional[int] = Form(None),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    nombre = nombre.strip()
    if not nombre:
        return RedirectResponse("/data/recibos-rubros?error=Nombre+requerido", status_code=303)
    exists = db.query(ReciboRubro).filter(func.lower(ReciboRubro.nombre) == nombre.lower()).first()
    if exists:
        return RedirectResponse("/data/recibos-rubros?error=Rubro+ya+existe", status_code=303)
    db.add(ReciboRubro(nombre=nombre, activo=activo == "on", cuenta_id=cuenta_id))
    db.commit()
    return RedirectResponse("/data/recibos-rubros?success=Rubro+creado", status_code=303)


@router.post("/data/recibos-rubros/{item_id}/update")
def data_recibos_rubros_update(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    cuenta_id: Optional[int] = Form(None),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    item = db.query(ReciboRubro).filter(ReciboRubro.id == item_id).first()
    if not item:
        return RedirectResponse("/data/recibos-rubros?error=Rubro+no+existe", status_code=303)
    item.nombre = nombre.strip()
    item.activo = activo == "on"
    item.cuenta_id = cuenta_id
    db.commit()
    return RedirectResponse("/data/recibos-rubros?success=Rubro+actualizado", status_code=303)


@router.get("/data/recibos-motivos")
def data_recibos_motivos(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(ReciboMotivo).filter(ReciboMotivo.id == int(edit_id)).first()
    items = db.query(ReciboMotivo).order_by(ReciboMotivo.tipo, ReciboMotivo.nombre).all()
    return request.app.state.templates.TemplateResponse(
        "data_recibos_motivos.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/recibos-motivos")
def data_recibos_motivos_create(
    request: Request,
    nombre: str = Form(...),
    tipo: str = Form(...),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    nombre = nombre.strip()
    tipo = tipo.strip().upper()
    if tipo not in {"INGRESO", "EGRESO"}:
        return RedirectResponse("/data/recibos-motivos?error=Tipo+no+valido", status_code=303)
    if not nombre:
        return RedirectResponse("/data/recibos-motivos?error=Nombre+requerido", status_code=303)
    exists = db.query(ReciboMotivo).filter(func.lower(ReciboMotivo.nombre) == nombre.lower()).first()
    if exists:
        return RedirectResponse("/data/recibos-motivos?error=Motivo+ya+existe", status_code=303)
    db.add(ReciboMotivo(nombre=nombre, tipo=tipo, activo=activo == "on"))
    db.commit()
    return RedirectResponse("/data/recibos-motivos?success=Motivo+creado", status_code=303)


@router.post("/data/recibos-motivos/{item_id}/update")
def data_recibos_motivos_update(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    tipo: str = Form(...),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    item = db.query(ReciboMotivo).filter(ReciboMotivo.id == item_id).first()
    if not item:
        return RedirectResponse("/data/recibos-motivos?error=Motivo+no+existe", status_code=303)
    item.nombre = nombre.strip()
    item.tipo = tipo.strip().upper()
    item.activo = activo == "on"
    db.commit()
    return RedirectResponse("/data/recibos-motivos?success=Motivo+actualizado", status_code=303)


@router.get("/data/cuentas-contables")
def data_cuentas_contables(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(CuentaContable).filter(CuentaContable.id == int(edit_id)).first()
    items = db.query(CuentaContable).order_by(CuentaContable.codigo).all()
    cuentas = db.query(CuentaContable).order_by(CuentaContable.codigo).all()
    return request.app.state.templates.TemplateResponse(
        "data_cuentas_contables.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "cuentas": cuentas,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/cuentas-contables")
def data_cuentas_contables_create(
    request: Request,
    codigo: str = Form(...),
    nombre: str = Form(...),
    tipo: str = Form(...),
    naturaleza: str = Form(...),
    parent_id: Optional[int] = Form(None),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    codigo = codigo.strip()
    nombre = nombre.strip()
    tipo = tipo.strip().upper()
    naturaleza = naturaleza.strip().upper()
    if not codigo or not nombre:
        return RedirectResponse("/data/cuentas-contables?error=Datos+incompletos", status_code=303)
    if tipo not in {"BALANCE", "RESULTADO"}:
        return RedirectResponse("/data/cuentas-contables?error=Tipo+no+valido", status_code=303)
    if naturaleza not in {"DEBE", "HABER"}:
        return RedirectResponse("/data/cuentas-contables?error=Naturaleza+no+valida", status_code=303)
    exists = db.query(CuentaContable).filter(func.lower(CuentaContable.codigo) == codigo.lower()).first()
    if exists:
        return RedirectResponse("/data/cuentas-contables?error=Codigo+ya+existe", status_code=303)
    nivel = 1
    if parent_id:
        parent = db.query(CuentaContable).filter(CuentaContable.id == parent_id).first()
        if not parent:
            return RedirectResponse("/data/cuentas-contables?error=Cuenta+padre+no+valida", status_code=303)
        nivel = (parent.nivel or 1) + 1
        if nivel > 4:
            return RedirectResponse("/data/cuentas-contables?error=Maximo+4+niveles", status_code=303)
    db.add(
        CuentaContable(
            codigo=codigo,
            nombre=nombre,
            tipo=tipo,
            naturaleza=naturaleza,
            parent_id=parent_id,
            nivel=nivel,
            activo=activo == "on",
        )
    )
    db.commit()
    return RedirectResponse("/data/cuentas-contables?success=Cuenta+creada", status_code=303)


@router.post("/data/cuentas-contables/{item_id}/update")
def data_cuentas_contables_update(
    request: Request,
    item_id: int,
    codigo: str = Form(...),
    nombre: str = Form(...),
    tipo: str = Form(...),
    naturaleza: str = Form(...),
    parent_id: Optional[int] = Form(None),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    item = db.query(CuentaContable).filter(CuentaContable.id == item_id).first()
    if not item:
        return RedirectResponse("/data/cuentas-contables?error=Cuenta+no+existe", status_code=303)
    codigo = codigo.strip()
    nombre = nombre.strip()
    tipo = tipo.strip().upper()
    naturaleza = naturaleza.strip().upper()
    exists = (
        db.query(CuentaContable)
        .filter(CuentaContable.id != item_id)
        .filter(func.lower(CuentaContable.codigo) == codigo.lower())
        .first()
    )
    if exists:
        return RedirectResponse("/data/cuentas-contables?error=Codigo+ya+existe", status_code=303)
    nivel = 1
    if parent_id:
        parent = db.query(CuentaContable).filter(CuentaContable.id == parent_id).first()
        if not parent:
            return RedirectResponse("/data/cuentas-contables?error=Cuenta+padre+no+valida", status_code=303)
        nivel = (parent.nivel or 1) + 1
        if nivel > 4:
            return RedirectResponse("/data/cuentas-contables?error=Maximo+4+niveles", status_code=303)
    item.codigo = codigo
    item.nombre = nombre
    item.tipo = tipo
    item.naturaleza = naturaleza
    item.parent_id = parent_id
    item.nivel = nivel
    item.activo = activo == "on"
    db.commit()
    return RedirectResponse("/data/cuentas-contables?success=Cuenta+actualizada", status_code=303)


@router.get("/data/roles")
def data_roles(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.roles")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(Role).filter(Role.id == int(edit_id)).first()
    items = db.query(Role).order_by(Role.name).all()
    return request.app.state.templates.TemplateResponse(
        "data_roles.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/roles")
def data_create_role(
    request: Request,
    nombre: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.roles")
    nombre = nombre.strip().lower()
    if not nombre:
        return RedirectResponse("/data/roles?error=Nombre+requerido", status_code=303)
    exists = db.query(Role).filter(Role.name == nombre).first()
    if exists:
        return RedirectResponse("/data/roles?error=Rol+ya+existe", status_code=303)
    role = Role(name=nombre)
    db.add(role)
    db.commit()
    return RedirectResponse("/data/roles?success=Rol+creado", status_code=303)


@router.post("/data/roles/{item_id}/update")
def data_update_role(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.roles")
    role = db.query(Role).filter(Role.id == item_id).first()
    if not role:
        return RedirectResponse("/data/roles?error=Rol+no+existe", status_code=303)
    role.name = nombre.strip().lower()
    db.commit()
    return RedirectResponse("/data/roles?success=Rol+actualizado", status_code=303)


@router.get("/data/permisos")
def data_permisos(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.permissions")
    _ensure_permission_catalog_in_db(db)
    role_id = request.query_params.get("role_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    roles = db.query(Role).order_by(Role.name).all()
    selected_role = None
    if role_id:
        selected_role = db.query(Role).filter(Role.id == int(role_id)).first()
    if not selected_role and roles:
        selected_role = roles[0]
    selected_permissions = set()
    if selected_role:
        if selected_role.name == "administrador":
            selected_permissions = set(_permission_catalog_names())
        else:
            selected_permissions = {perm.name for perm in (selected_role.permissions or [])}
    return request.app.state.templates.TemplateResponse(
        "data_permisos.html",
        {
            "request": request,
            "user": user,
            "roles": roles,
            "selected_role": selected_role,
            "selected_permissions": selected_permissions,
            "permission_groups": PERMISSION_GROUPS,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/permisos")
def data_permisos_update(
    request: Request,
    role_id: int = Form(...),
    perms: Optional[list[str]] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.permissions")
    _ensure_permission_catalog_in_db(db)
    role = db.query(Role).filter(Role.id == role_id).first()
    if not role:
        return RedirectResponse("/data/permisos?error=Rol+no+existe", status_code=303)
    perm_names = _permission_catalog_names()
    selected = set(perms or [])
    selected = {name for name in selected if name in perm_names}
    permissions = []
    if role.name == "administrador":
        permissions = db.query(Permission).filter(Permission.name.in_(perm_names)).all()
    else:
        if selected:
            permissions = db.query(Permission).filter(Permission.name.in_(selected)).all()
    role.permissions = permissions
    db.commit()
    return RedirectResponse(
        f"/data/permisos?role_id={role_id}&success=Permisos+actualizados",
        status_code=303,
    )


@router.get("/data/usuarios")
def data_usuarios(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.users")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(User).filter(User.id == int(edit_id)).first()
    items = db.query(User).order_by(User.full_name).all()
    roles = db.query(Role).order_by(Role.name).all()
    branches = _scoped_branches_query(db).order_by(Branch.name).all()
    bodegas = _scoped_bodegas_query(db).order_by(Bodega.name).all()
    return request.app.state.templates.TemplateResponse(
        "data_usuarios.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "roles": roles,
            "branches": branches,
            "bodegas": bodegas,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/usuarios")
def data_create_usuario(
    request: Request,
    full_name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    role_ids: Optional[list[int]] = Form(None),
    branch_id: Optional[int] = Form(None),
    bodega_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.users")
    full_name = full_name.strip()
    email = email.strip().lower()
    if not full_name or not email or not password:
        return RedirectResponse("/data/usuarios?error=Datos+incompletos", status_code=303)
    exists = db.query(User).filter(func.lower(User.email) == email).first()
    if exists:
        return RedirectResponse("/data/usuarios?error=Email+ya+existe", status_code=303)
    roles = []
    if role_ids:
        roles = db.query(Role).filter(Role.id.in_(role_ids)).all()
    if not branch_id or not bodega_id:
        return RedirectResponse("/data/usuarios?error=Sucursal+y+bodega+requeridas", status_code=303)
    branch = db.query(Branch).filter(Branch.id == branch_id).first()
    if not branch:
        return RedirectResponse("/data/usuarios?error=Sucursal+no+valida", status_code=303)
    bodega = db.query(Bodega).filter(Bodega.id == bodega_id).first()
    if not bodega or bodega.branch_id != branch.id:
        return RedirectResponse("/data/usuarios?error=Bodega+no+corresponde+a+la+sucursal", status_code=303)
    new_user = User(
        full_name=full_name,
        email=email,
        hashed_password=hash_password(password),
        is_active=True,
        roles=roles,
        branches=[branch],
        default_branch_id=branch.id,
        default_bodega_id=bodega.id,
    )
    db.add(new_user)
    db.commit()
    return RedirectResponse("/data/usuarios?success=Usuario+creado", status_code=303)


@router.post("/data/usuarios/{item_id}/update")
def data_update_usuario(
    request: Request,
    item_id: int,
    full_name: str = Form(...),
    email: str = Form(...),
    new_password: Optional[str] = Form(None),
    role_ids: Optional[list[int]] = Form(None),
    is_active: Optional[str] = Form(None),
    branch_id: Optional[int] = Form(None),
    bodega_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.users")
    edit_user = db.query(User).filter(User.id == item_id).first()
    if not edit_user:
        return RedirectResponse("/data/usuarios?error=Usuario+no+existe", status_code=303)
    edit_user.full_name = full_name.strip()
    edit_user.email = email.strip().lower()
    if new_password:
        edit_user.hashed_password = hash_password(new_password)
    edit_user.is_active = is_active == "on"
    roles = []
    if role_ids:
        roles = db.query(Role).filter(Role.id.in_(role_ids)).all()
    edit_user.roles = roles
    if not branch_id or not bodega_id:
        return RedirectResponse("/data/usuarios?error=Sucursal+y+bodega+requeridas", status_code=303)
    branch = db.query(Branch).filter(Branch.id == branch_id).first()
    if not branch:
        return RedirectResponse("/data/usuarios?error=Sucursal+no+valida", status_code=303)
    bodega = db.query(Bodega).filter(Bodega.id == bodega_id).first()
    if not bodega or bodega.branch_id != branch.id:
        return RedirectResponse("/data/usuarios?error=Bodega+no+corresponde+a+la+sucursal", status_code=303)
    edit_user.branches = [branch]
    edit_user.default_branch_id = branch.id
    edit_user.default_bodega_id = bodega.id
    db.commit()
    return RedirectResponse("/data/usuarios?success=Usuario+actualizado", status_code=303)


@router.get("/data/clientes")
def data_clientes(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    edit_id = request.query_params.get("edit_id")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    edit_item = None
    if edit_id:
        edit_item = db.query(Cliente).filter(Cliente.id == int(edit_id)).first()
    items = db.query(Cliente).order_by(Cliente.nombre).all()
    return request.app.state.templates.TemplateResponse(
        "data_clientes.html",
        {
            "request": request,
            "user": user,
            "items": items,
            "edit_item": edit_item,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/data/clientes")
def data_create_cliente(
    request: Request,
    nombre: str = Form(...),
    identificacion: Optional[str] = Form(None),
    telefono: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    direccion: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    nombre = nombre.strip()
    if not nombre:
        return RedirectResponse("/data/clientes?error=Nombre+requerido", status_code=303)
    exists = db.query(Cliente).filter(func.lower(Cliente.nombre) == nombre.lower()).first()
    if exists:
        return RedirectResponse("/data/clientes?error=Cliente+ya+existe", status_code=303)
    cliente = Cliente(
        nombre=nombre,
        identificacion=identificacion.strip() if identificacion else None,
        telefono=telefono,
        email=email,
        direccion=direccion,
        activo=True,
    )
    db.add(cliente)
    db.commit()
    return RedirectResponse("/data/clientes?success=Cliente+creado", status_code=303)


@router.post("/data/clientes/{item_id}/update")
def data_update_cliente(
    request: Request,
    item_id: int,
    nombre: str = Form(...),
    identificacion: Optional[str] = Form(None),
    telefono: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    direccion: Optional[str] = Form(None),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.data.catalogs")
    cliente = db.query(Cliente).filter(Cliente.id == item_id).first()
    if not cliente:
        return RedirectResponse("/data/clientes?error=Cliente+no+existe", status_code=303)
    cliente.nombre = nombre.strip()
    cliente.identificacion = identificacion.strip() if identificacion else None
    cliente.telefono = telefono
    cliente.email = email
    cliente.direccion = direccion
    cliente.activo = activo == "on"
    db.commit()
    return RedirectResponse("/data/clientes?success=Cliente+actualizado", status_code=303)

@router.get("/sales/products/search")
def sales_products_search(
    q: str = "",
    bodega_id: Optional[int] = None,
    vendedor_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    query = q.strip()
    if len(query) < 2:
        return JSONResponse({"ok": True, "items": []})
    like = f"%{query.lower()}%"
    productos = (
        db.query(Producto)
        .filter(Producto.activo.is_(True))
        .filter(
            or_(
                func.lower(Producto.cod_producto).like(like),
                func.lower(Producto.descripcion).like(like),
            )
        )
        .order_by(Producto.descripcion)
        .limit(100)
        .all()
    )
    _, resolved_bodega = _resolve_branch_bodega(db, user)
    bodega = resolved_bodega
    if bodega_id:
        allowed_branch_ids = {b.id for b in (user.branches or [])}
        requested = (
            db.query(Bodega)
            .filter(Bodega.id == bodega_id, Bodega.activo.is_(True))
            .first()
        )
        if requested and (not allowed_branch_ids or requested.branch_id in allowed_branch_ids):
            bodega = requested
    balances: dict[tuple[int, int], Decimal] = {}
    if bodega and productos:
        product_ids = [p.id for p in productos]
        balances = _balances_by_bodega(db, [bodega.id], product_ids)
    reserved_totals: dict[int, Decimal] = {}
    reserved_details: dict[int, list[dict[str, object]]] = {}
    if bodega and productos:
        reserved_totals, reserved_details = _preventa_reserved_bulk_by_others(
            db,
            bodega_id=bodega.id,
            producto_ids=[p.id for p in productos],
            vendedor_id=vendedor_id,
            include_same_vendedor=True,
        )

    items = []
    for producto in productos:
        existencia = 0.0
        if bodega and balances:
            existencia = float(balances.get((producto.id, bodega.id), Decimal("0")) or 0)
        reserved_qty = float(reserved_totals.get(producto.id, Decimal("0")) or 0)
        free_qty = max(0.0, existencia - reserved_qty)
        items.append(
            {
                "id": producto.id,
                "cod_producto": producto.cod_producto,
                "descripcion": producto.descripcion,
                "precio_venta1_usd": float(producto.precio_venta1_usd or 0),
                "precio_venta1": float(producto.precio_venta1 or 0),
                "existencia": existencia,
                "reserved_qty": reserved_qty,
                "free_qty": free_qty,
                "reserved_details": reserved_details.get(producto.id, []),
                "combo_count": len(producto.combo_children or []),
            }
        )
    return JSONResponse({"ok": True, "items": items})


@router.get("/sales/combo/{parent_id}/items")
def sales_combo_items(
    request: Request,
    parent_id: int,
    bodega_id: Optional[int] = None,
    vendedor_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.registrar")
    parent = db.query(Producto).filter(Producto.id == parent_id, Producto.activo.is_(True)).first()
    if not parent:
        return JSONResponse({"ok": False, "message": "Producto no encontrado"}, status_code=404)
    _, resolved_bodega = _resolve_branch_bodega(db, user)
    bodega = resolved_bodega
    if bodega_id:
        allowed_branch_ids = {b.id for b in (user.branches or [])}
        requested = (
            db.query(Bodega)
            .filter(Bodega.id == bodega_id, Bodega.activo.is_(True))
            .first()
        )
        if requested and (not allowed_branch_ids or requested.branch_id in allowed_branch_ids):
            bodega = requested
    rows = (
        db.query(ProductoCombo, Producto)
        .join(Producto, Producto.id == ProductoCombo.child_producto_id)
        .filter(ProductoCombo.parent_producto_id == parent_id)
        .order_by(ProductoCombo.id.asc())
        .all()
    )
    children = [producto for _, producto in rows]
    balances: dict[tuple[int, int], Decimal] = {}
    reserved_totals: dict[int, Decimal] = {}
    reserved_details: dict[int, list[dict[str, object]]] = {}
    if bodega and children:
        product_ids = [p.id for p in children]
        balances = _balances_by_bodega(db, [bodega.id], product_ids)
        reserved_totals, reserved_details = _preventa_reserved_bulk_by_others(
            db,
            bodega_id=bodega.id,
            producto_ids=product_ids,
            vendedor_id=vendedor_id,
            include_same_vendedor=True,
        )
    items: list[dict[str, object]] = []
    for combo, producto in rows:
        existencia = float(balances.get((producto.id, bodega.id), Decimal("0")) or 0) if bodega else 0.0
        reserved_qty = float(reserved_totals.get(producto.id, Decimal("0")) or 0)
        free_qty = max(0.0, existencia - reserved_qty)
        items.append(
            {
                "id": int(combo.id),
                "child_id": int(producto.id),
                "cod_producto": producto.cod_producto,
                "descripcion": producto.descripcion,
                "cantidad": float(combo.cantidad or 1),
                "precio_venta1_usd": float(producto.precio_venta1_usd or 0),
                "precio_venta1": float(producto.precio_venta1 or 0),
                "existencia": existencia,
                "reserved_qty": reserved_qty,
                "free_qty": free_qty,
                "reserved_details": reserved_details.get(producto.id, []),
            }
        )
    return JSONResponse({"ok": True, "items": items})


@router.get("/inventory/ingresos/{ingreso_id}/pdf")
def inventory_ingreso_pdf(
    request: Request,
    ingreso_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.ingresos")
    company_profile = _company_profile_payload(db)
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="ReportLab no esta instalado") from exc

    ingreso = (
        db.query(IngresoInventario)
        .filter(IngresoInventario.id == ingreso_id)
        .first()
    )
    if not ingreso:
        raise HTTPException(status_code=404, detail="Ingreso no encontrado")

    total_items = len(ingreso.items or [])
    total_bultos = sum(float(item.cantidad or 0) for item in (ingreso.items or []))

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 36

    logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
    if logo_path.exists():
        pdf.drawImage(
            str(logo_path),
            margin,
            height - 78,
            width=90,
            height=60,
            preserveAspectRatio=True,
            mask="auto",
        )

    info_x = margin + 110
    info_y = height - 44
    branch = ingreso.bodega.branch if ingreso.bodega else None
    identity = _company_identity(branch, company_profile)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(info_x, info_y, identity["company_name"])
    pdf.setFont("Helvetica", 9)
    pdf.drawString(info_x, info_y - 14, f"Telf. {identity['telefono']}")
    pdf.drawString(info_x, info_y - 28, f"Direccion: {identity['direccion']}")

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(margin, height - 120, "Informe de ingreso de mercaderias")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(margin, height - 136, f"Reimpresion ingreso #{ingreso.id}")
    pdf.setStrokeColorRGB(0.75, 0.75, 0.75)
    pdf.setLineWidth(0.6)
    pdf.line(margin, height - 146, width - margin, height - 146)

    pdf.setFont("Helvetica", 9)
    pdf.drawString(
        margin,
        height - 160,
        f"Fecha: {ingreso.fecha.isoformat()}",
    )
    pdf.drawString(
        margin + 200,
        height - 160,
        f"Tipo: {ingreso.tipo.nombre if ingreso.tipo else '-'}",
    )
    pdf.drawString(
        margin,
        height - 174,
        f"Bodega: {ingreso.bodega.name if ingreso.bodega else '-'}",
    )
    pdf.drawString(
        margin + 200,
        height - 174,
        f"Proveedor: {ingreso.proveedor.nombre if ingreso.proveedor else '-'}",
    )
    pdf.drawString(
        margin,
        height - 188,
        f"Moneda: {ingreso.moneda}",
    )
    pdf.drawString(
        margin + 200,
        height - 188,
        f"Tasa: {('C$ %.4f' % float(ingreso.tasa_cambio)) if ingreso.tasa_cambio else '-'}",
    )
    observacion = ingreso.observacion or "-"
    if len(observacion) > 120:
        observacion = f"{observacion[:117]}..."
    pdf.drawString(
        margin,
        height - 202,
        f"Descripcion: {observacion}",
    )

    y = height - 224
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(margin, y, "Codigo")
    pdf.drawString(margin + 80, y, "Descripcion")
    pdf.drawRightString(margin + 340, y, "Cant.")
    pdf.drawRightString(margin + 420, y, "Costo USD")
    pdf.drawRightString(margin + 500, y, "Subtotal USD")
    y -= 12

    pdf.setFont("Helvetica", 8)
    for item in ingreso.items:
        if y < margin + 60:
            pdf.setFont("Helvetica", 8)
            pdf.drawRightString(
                width - margin,
                margin - 18,
                f"Pagina {pdf.getPageNumber()}",
            )
            pdf.showPage()
            y = height - margin
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(margin, y, "Codigo")
            pdf.drawString(margin + 80, y, "Descripcion")
            pdf.drawRightString(margin + 340, y, "Cant.")
            pdf.drawRightString(margin + 420, y, "Costo USD")
            pdf.drawRightString(margin + 500, y, "Subtotal USD")
            y -= 12
            pdf.setFont("Helvetica", 8)

        codigo = item.producto.cod_producto if item.producto else ""
        descripcion = item.producto.descripcion if item.producto else ""
        if len(descripcion) > 48:
            descripcion = f"{descripcion[:45]}..."
        pdf.drawString(margin, y, codigo)
        pdf.drawString(margin + 80, y, descripcion)
        pdf.drawRightString(margin + 340, y, f"{float(item.cantidad or 0):.2f}")
        pdf.drawRightString(margin + 420, y, f"{float(item.costo_unitario_usd or 0):.2f}")
        pdf.drawRightString(margin + 500, y, f"{float(item.subtotal_usd or 0):.2f}")
        y -= 12

    y -= 8
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(margin + 420, y, "Total bultos:")
    pdf.drawRightString(margin + 500, y, f"{float(total_bultos or 0):.2f}")
    y -= 12
    pdf.drawRightString(margin + 420, y, "Total items:")
    pdf.drawRightString(margin + 500, y, f"{int(total_items or 0)}")
    y -= 12
    pdf.drawRightString(margin + 420, y, "Total USD:")
    pdf.drawRightString(margin + 500, y, f"{float(ingreso.total_usd or 0):.2f}")
    y -= 12
    pdf.drawRightString(margin + 420, y, "Total C$:")
    pdf.drawRightString(margin + 500, y, f"{float(ingreso.total_cs or 0):.2f}")

    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(
        width - margin,
        margin - 18,
        f"Pagina {pdf.getPageNumber()}",
    )

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    headers = {
        "Content-Disposition": f"inline; filename=ingreso_{ingreso.id}.pdf"
    }
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


@router.get("/inventory/egresos/{egreso_id}/pdf")
def inventory_egreso_pdf(
    request: Request,
    egreso_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.egresos")
    company_profile = _company_profile_payload(db)
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="ReportLab no esta instalado") from exc

    egreso = (
        db.query(EgresoInventario)
        .filter(EgresoInventario.id == egreso_id)
        .first()
    )
    if not egreso:
        raise HTTPException(status_code=404, detail="Egreso no encontrado")

    total_items = len(egreso.items or [])
    total_bultos = sum(float(item.cantidad or 0) for item in (egreso.items or []))

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 36

    logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
    if logo_path.exists():
        pdf.drawImage(
            str(logo_path),
            margin,
            height - 78,
            width=90,
            height=60,
            preserveAspectRatio=True,
            mask="auto",
        )

    info_x = margin + 110
    info_y = height - 44
    branch = egreso.bodega.branch if egreso.bodega else None
    identity = _company_identity(branch, company_profile)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(info_x, info_y, identity["company_name"])
    pdf.setFont("Helvetica", 9)
    pdf.drawString(info_x, info_y - 14, f"Telf. {identity['telefono']}")
    pdf.drawString(info_x, info_y - 28, f"Direccion: {identity['direccion']}")

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(margin, height - 120, "Informe de egreso de mercaderias")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(margin, height - 136, f"Reimpresion egreso #{egreso.id}")
    pdf.setStrokeColorRGB(0.75, 0.75, 0.75)
    pdf.setLineWidth(0.6)
    pdf.line(margin, height - 146, width - margin, height - 146)

    pdf.setFont("Helvetica", 9)
    pdf.drawString(
        margin,
        height - 160,
        f"Fecha: {egreso.fecha.isoformat()}",
    )
    pdf.drawString(
        margin + 200,
        height - 160,
        f"Tipo: {egreso.tipo.nombre if egreso.tipo else '-'}",
    )
    pdf.drawString(
        margin,
        height - 174,
        f"Bodega origen: {egreso.bodega.name if egreso.bodega else '-'}",
    )
    pdf.drawString(
        margin + 260,
        height - 174,
        f"Bodega destino: {egreso.bodega_destino.name if egreso.bodega_destino else '-'}",
    )
    pdf.drawString(
        margin,
        height - 188,
        f"Moneda: {egreso.moneda}",
    )
    pdf.drawString(
        margin + 200,
        height - 188,
        f"Tasa: {('C$ %.4f' % float(egreso.tasa_cambio)) if egreso.tasa_cambio else '-'}",
    )
    observacion = egreso.observacion or "-"
    if len(observacion) > 120:
        observacion = f"{observacion[:117]}..."
    pdf.drawString(
        margin,
        height - 202,
        f"Motivo: {observacion}",
    )

    y = height - 224
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(margin, y, "Codigo")
    pdf.drawString(margin + 80, y, "Descripcion")
    pdf.drawRightString(margin + 340, y, "Cant.")
    pdf.drawRightString(margin + 420, y, "Costo USD")
    pdf.drawRightString(margin + 500, y, "Subtotal USD")
    y -= 12

    pdf.setFont("Helvetica", 8)
    for item in egreso.items:
        if y < margin + 60:
            pdf.setFont("Helvetica", 8)
            pdf.drawRightString(
                width - margin,
                margin - 18,
                f"Pagina {pdf.getPageNumber()}",
            )
            pdf.showPage()
            y = height - margin
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(margin, y, "Codigo")
            pdf.drawString(margin + 80, y, "Descripcion")
            pdf.drawRightString(margin + 340, y, "Cant.")
            pdf.drawRightString(margin + 420, y, "Costo USD")
            pdf.drawRightString(margin + 500, y, "Subtotal USD")
            y -= 12
            pdf.setFont("Helvetica", 8)

        codigo = item.producto.cod_producto if item.producto else ""
        descripcion = item.producto.descripcion if item.producto else ""
        if len(descripcion) > 48:
            descripcion = f"{descripcion[:45]}..."
        pdf.drawString(margin, y, codigo)
        pdf.drawString(margin + 80, y, descripcion)
        pdf.drawRightString(margin + 340, y, f"{float(item.cantidad or 0):.2f}")
        pdf.drawRightString(margin + 420, y, f"{float(item.costo_unitario_usd or 0):.2f}")
        pdf.drawRightString(margin + 500, y, f"{float(item.subtotal_usd or 0):.2f}")
        y -= 12

    y -= 8
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(margin + 420, y, "Total bultos:")
    pdf.drawRightString(margin + 500, y, f"{float(total_bultos or 0):.2f}")
    y -= 12
    pdf.drawRightString(margin + 420, y, "Total items:")
    pdf.drawRightString(margin + 500, y, f"{int(total_items or 0)}")
    y -= 12
    pdf.drawRightString(margin + 420, y, "Total USD:")
    pdf.drawRightString(margin + 500, y, f"{float(egreso.total_usd or 0):.2f}")
    y -= 12
    pdf.drawRightString(margin + 420, y, "Total C$:")
    pdf.drawRightString(margin + 500, y, f"{float(egreso.total_cs or 0):.2f}")

    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(
        width - margin,
        margin - 18,
        f"Pagina {pdf.getPageNumber()}",
    )

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    headers = {
        "Content-Disposition": f"inline; filename=egreso_{egreso.id}.pdf"
    }
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


@router.get("/sales/{venta_id}/pdf")
def sales_invoice_pdf(
    venta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="ReportLab no esta instalado") from exc

    factura = (
        db.query(VentaFactura)
        .filter(VentaFactura.id == venta_id)
        .first()
    )
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    company_profile = _company_profile_payload(db)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    margin = 36

    logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
    if logo_path.exists():
        pdf.drawImage(
            str(logo_path),
            margin,
            height - 78,
            width=90,
            height=60,
            preserveAspectRatio=True,
            mask="auto",
        )

    info_x = margin + 110
    info_y = height - 44
    branch = factura.bodega.branch if factura.bodega else None
    identity = _company_identity(branch, company_profile)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(info_x, info_y, identity["company_name"])
    pdf.setFont("Helvetica", 9)
    pdf.drawString(info_x, info_y - 14, f"Telf. {identity['telefono']}")
    pdf.drawString(info_x, info_y - 28, f"Direccion: {identity['direccion']}")

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(margin, height - 120, "Factura de venta")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(margin, height - 136, f"{factura.numero}")
    pdf.setStrokeColorRGB(0.75, 0.75, 0.75)
    pdf.setLineWidth(0.6)
    pdf.line(margin, height - 146, width - margin, height - 146)

    pdf.setFont("Helvetica", 9)
    pdf.drawString(
        margin,
        height - 160,
        f"Fecha: {factura.fecha.date().isoformat()}",
    )
    pdf.drawString(
        margin + 200,
        height - 160,
        f"Vendedor: {factura.vendedor.nombre if factura.vendedor else '-'}",
    )
    pdf.drawString(
        margin,
        height - 174,
        f"Cliente: {factura.cliente.nombre if factura.cliente else 'Consumidor final'}",
    )
    pdf.drawString(
        margin + 200,
        height - 174,
        f"Moneda: {factura.moneda}",
    )
    pdf.drawString(
        margin,
        height - 188,
        f"Bodega: {factura.bodega.name if factura.bodega else '-'}",
    )
    pdf.drawString(
        margin + 200,
        height - 188,
        f"Tasa: {('C$ %.4f' % float(factura.tasa_cambio)) if factura.tasa_cambio else '-'}",
    )

    y = height - 210
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(margin, y, "Codigo")
    pdf.drawString(margin + 80, y, "Descripcion")
    pdf.drawRightString(margin + 340, y, "Cant.")
    pdf.drawRightString(margin + 420, y, "Precio USD")
    pdf.drawRightString(margin + 500, y, "Subtotal USD")
    y -= 12

    pdf.setFont("Helvetica", 8)
    for item in factura.items:
        if y < margin + 60:
            pdf.setFont("Helvetica", 8)
            pdf.drawRightString(
                width - margin,
                margin - 18,
                f"Pagina {pdf.getPageNumber()}",
            )
            pdf.showPage()
            y = height - margin
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(margin, y, "Codigo")
            pdf.drawString(margin + 80, y, "Descripcion")
            pdf.drawRightString(margin + 340, y, "Cant.")
            pdf.drawRightString(margin + 420, y, "Precio USD")
            pdf.drawRightString(margin + 500, y, "Subtotal USD")
            y -= 12
            pdf.setFont("Helvetica", 8)

        codigo = item.producto.cod_producto if item.producto else ""
        descripcion = item.producto.descripcion if item.producto else ""
        if len(descripcion) > 48:
            descripcion = f"{descripcion[:45]}..."
        pdf.drawString(margin, y, codigo)
        pdf.drawString(margin + 80, y, descripcion)
        pdf.drawRightString(margin + 340, y, f"{float(item.cantidad or 0):.2f}")
        pdf.drawRightString(margin + 420, y, f"{float(item.precio_unitario_usd or 0):.2f}")
        pdf.drawRightString(margin + 500, y, f"{float(item.subtotal_usd or 0):.2f}")
        y -= 12

    y -= 8
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(margin + 420, y, "Total USD:")
    pdf.drawRightString(margin + 500, y, f"{float(factura.total_usd or 0):.2f}")
    y -= 12
    pdf.drawRightString(margin + 420, y, "Total C$:")
    pdf.drawRightString(margin + 500, y, f"{float(factura.total_cs or 0):.2f}")

    pago = factura.pagos[0] if factura.pagos else None
    if pago:
        y -= 18
        pdf.setFont("Helvetica", 9)
        pdf.drawString(
            margin,
            y,
            f"Forma de pago: {pago.forma_pago.nombre if pago.forma_pago else '-'}",
        )
        banco = pago.banco.nombre if pago.banco else "-"
        pdf.drawString(margin + 200, y, f"Banco: {banco}")

    pdf.setFont("Helvetica", 8)
    pdf.drawRightString(
        width - margin,
        margin - 18,
        f"Pagina {pdf.getPageNumber()}",
    )

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    headers = {"Content-Disposition": f"inline; filename={factura.numero}.pdf"}
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


@router.get("/sales/{venta_id}/ticket")
def sales_ticket_pos(
    request: Request,
    venta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales")
    try:
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="ReportLab no esta instalado") from exc

    factura = (
        db.query(VentaFactura)
        .filter(VentaFactura.id == venta_id)
        .first()
    )
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    pdf_bytes = _build_pos_ticket_pdf_bytes(factura, _company_profile_payload(db))
    buffer = io.BytesIO(pdf_bytes)
    headers = {"Content-Disposition": f"inline; filename={factura.numero}_ticket.pdf"}
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


@router.get("/sales/{venta_id}/ticket/print")
def sales_ticket_print(
    request: Request,
    venta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales")
    copies_value = request.query_params.get("copies", "2")
    try:
        copies = max(int(copies_value), 1)
    except ValueError:
        copies = 2
    factura = (
        db.query(VentaFactura)
        .filter(VentaFactura.id == venta_id)
        .first()
    )
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    def wrap_text(text: str, max_chars: int) -> list[str]:
        if not text:
            return [""]
        words = text.split()
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if len(candidate) > max_chars:
                if current:
                    lines.append(current)
                current = word
            else:
                current = candidate
        if current:
            lines.append(current)
        return lines or [text]

    def extract_weight_lbs(text: str) -> float:
        if not text:
            return 0.0
        match = re.search(r"\b(\d+(?:\.\d+)?)\s*(lbs)\b", text.lower())
        if not match:
            return 0.0
        try:
            return float(match.group(1))
        except ValueError:
            return 0.0

    profile = _company_profile_payload(db)
    branch = factura.bodega.branch if factura.bodega else None
    identity = _company_identity(branch, profile)
    company_name = identity["company_name"]
    ruc = identity["ruc"]
    telefono = identity["telefono"]
    direccion = identity["direccion"]
    direccion_lines = wrap_text(direccion, 32)[:2]
    sucursal = identity["sucursal"]

    cliente = factura.cliente.nombre if factura.cliente else "Consumidor final"
    cliente_id = factura.cliente.identificacion if factura.cliente and factura.cliente.identificacion else "-"
    vendedor = factura.vendedor.nombre if factura.vendedor else "-"

    fecha_base = factura.created_at or factura.fecha
    fecha_str = ""
    hora_str = ""
    if fecha_base:
        try:
            fecha_str = fecha_base.strftime("%d/%m/%Y")
            hora_str = fecha_base.strftime("%H:%M")
        except AttributeError:
            fecha_str = str(fecha_base)

    moneda = factura.moneda or "CS"
    currency_label = "C$" if moneda == "CS" else "$"
    total_amount = float(factura.total_cs or 0) if moneda == "CS" else float(factura.total_usd or 0)
    subtotal_amount = total_amount

    pagos = factura.pagos or []
    total_paid = sum(
        float(pago.monto_cs or 0) if moneda == "CS" else float(pago.monto_usd or 0)
        for pago in pagos
    )
    saldo = total_paid - total_amount

    def format_amount(value: float) -> str:
        return f"{value:,.2f}"

    items = []
    total_bultos = 0.0
    total_lbs = 0.0
    line_count = 0
    line_count += 1  # company
    line_count += 1  # ruc
    line_count += 1  # telefono
    line_count += len(direccion_lines)
    line_count += 1  # sucursal
    line_count += 1  # divider
    line_count += 5  # factura/fecha/cliente/id/vendedor
    line_count += 1  # divider
    for item in factura.items:
        desc_lines = wrap_text(item.producto.descripcion if item.producto else "-", 32)
        line_count += 1  # codigo
        line_count += len(desc_lines)
        line_count += 2  # qty/price + desc/subtotal
        line_count += 1  # divider
        qty = float(item.cantidad or 0)
        total_bultos += qty
        desc_text = item.producto.descripcion if item.producto else ""
        lbs_per_unit = extract_weight_lbs(desc_text)
        if lbs_per_unit:
            total_lbs += lbs_per_unit * qty
        price = (
            float(item.precio_unitario_cs or 0)
            if moneda == "CS"
            else float(item.precio_unitario_usd or 0)
        )
        subtotal = (
            float(item.subtotal_cs or 0)
            if moneda == "CS"
            else float(item.subtotal_usd or 0)
        )
        items.append(
            {
                "codigo": item.producto.cod_producto if item.producto else "-",
                "descripcion": item.producto.descripcion if item.producto else "-",
                "cantidad": qty,
                "precio": price,
                "subtotal": subtotal,
            }
        )

    pagos_render = []
    line_count += 5  # total bultos/lbs + subtotal/desc/total
    for pago in pagos:
        forma = pago.forma_pago.nombre if pago.forma_pago else "Pago"
        banco = pago.banco.nombre if pago.banco else ""
        label = f"{forma} {banco}".strip()
        monto = (
            float(pago.monto_cs or 0)
            if moneda == "CS"
            else float(pago.monto_usd or 0)
        )
        pagos_render.append({"label": label, "monto": monto})
    if pagos_render:
        line_count += 2  # divider + title
        line_count += len(pagos_render)
    line_count += 1  # vuelto/saldo
    line_count += 4  # footer

    # Keep page height tight to content to avoid trailing white space in POS print
    line_height_mm = 3.55
    page_height_mm = max(96.0, 8.0 + line_count * line_height_mm + 8.0)

    return request.app.state.templates.TemplateResponse(
        "sales_ticket_print.html",
        {
            "request": request,
            "factura": factura,
            "company_name": company_name,
            "ruc": ruc,
            "telefono": telefono,
            "direccion": direccion,
            "direccion_lines": direccion_lines,
            "sucursal": sucursal,
            "cliente": cliente,
            "cliente_id": cliente_id,
            "vendedor": vendedor,
            "fecha_str": fecha_str,
            "hora_str": hora_str,
            "moneda": moneda,
            "currency_label": currency_label,
            "total_amount": total_amount,
            "subtotal_amount": subtotal_amount,
            "saldo": saldo,
            "total_bultos": total_bultos,
            "total_lbs": total_lbs,
            "items": items,
            "pagos": pagos_render,
            "format_amount": format_amount,
            "copies": copies,
            "page_height_mm": page_height_mm,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/inventory/product")
def inventory_create_product(
    request: Request,
    cod_producto: str = Form(...),
    descripcion: str = Form(...),
    linea_id: Optional[str] = Form(None),
    segmento_id: Optional[str] = Form(None),
    marca: Optional[str] = Form(None),
    referencia_producto: Optional[str] = Form(None),
    precio_venta1_usd: float = Form(0),
    precio_venta2_usd: float = Form(0),
    precio_venta3_usd: float = Form(0),
    costo_producto_usd: float = Form(0),
    existencia: float = Form(0),
    activo: Optional[str] = Form(None),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    if not (
        _has_permission(user, "access.inventory.productos")
        or _has_permission(user, "access.inventory.ingresos")
    ):
        _enforce_permission(request, user, "access.inventory.productos")
    accept_header = request.headers.get("accept", "")
    is_fetch = (
        request.headers.get("x-requested-with") == "fetch"
        or "application/json" in accept_header
        or request.headers.get("hx-request") == "true"
    )

    def _error(message: str):
        if is_fetch:
            return JSONResponse({"ok": False, "message": message}, status_code=400)
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error={message.replace(' ', '+')}", status_code=303)

    cod_producto = cod_producto.strip()
    descripcion = descripcion.strip()
    marca = (marca or "").strip() or "Sin Marca"
    if not cod_producto or not descripcion:
        return _error("Faltan datos obligatorios")
    if float(precio_venta1_usd or 0) <= 0:
        return _error("Precio de venta obligatorio")
    if float(costo_producto_usd or 0) <= 0:
        return _error("Costo obligatorio")

    def _to_int(value: Optional[str]) -> Optional[int]:
        if not value:
            return None
        return int(value) if value.isdigit() else None

    exists = db.query(Producto).filter(Producto.cod_producto == cod_producto).first()
    if exists:
        return _error("Codigo existente")

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if not rate_today:
        return _error("Tasa de cambio no configurada")

    tasa = float(rate_today.rate)
    inventory_cs_only = _inventory_cs_only_mode(db)
    if inventory_cs_only:
        precio_venta1_cs = float(precio_venta1_usd or 0)
        precio_venta2_cs = float(precio_venta2_usd or 0)
        precio_venta3_cs = float(precio_venta3_usd or 0)
        costo_producto_cs = float(costo_producto_usd or 0)
        precio_venta1_usd = (precio_venta1_cs / tasa) if tasa else 0
        precio_venta2_usd = (precio_venta2_cs / tasa) if tasa else 0
        precio_venta3_usd = (precio_venta3_cs / tasa) if tasa else 0
        costo_producto_usd = (costo_producto_cs / tasa) if tasa else 0
    else:
        precio_venta1_cs = precio_venta1_usd * tasa
        precio_venta2_cs = precio_venta2_usd * tasa
        precio_venta3_cs = precio_venta3_usd * tasa
        costo_producto_cs = costo_producto_usd * tasa
    active_flag = True if activo is None else activo == "on"
    producto = Producto(
        cod_producto=cod_producto,
        descripcion=descripcion,
        linea_id=_to_int(linea_id),
        segmento_id=_to_int(segmento_id),
        marca=marca,
        referencia_producto=referencia_producto,
        precio_venta1=precio_venta1_cs,
        precio_venta2=precio_venta2_cs,
        precio_venta3=precio_venta3_cs,
        precio_venta1_usd=precio_venta1_usd,
        precio_venta2_usd=precio_venta2_usd,
        precio_venta3_usd=precio_venta3_usd,
        tasa_cambio=tasa,
        costo_producto=costo_producto_cs,
        activo=active_flag,
)
    db.add(producto)
    db.flush()
    db.add(SaldoProducto(producto_id=producto.id, existencia=existencia))
    db.commit()
    if is_fetch:
        return JSONResponse(
            {
                "ok": True,
                "message": "Producto registrado exitosamente",
                "id": producto.id,
                "cod_producto": producto.cod_producto,
                "descripcion": producto.descripcion,
                "precio_venta1_usd": float(producto.precio_venta1_usd or 0),
                "costo_usd": float(costo_producto_usd or 0),
                "precio_venta1": float(producto.precio_venta1 or 0),
                "costo_cs": float(producto.costo_producto or 0),
                "activo": bool(producto.activo),
            }
        )
    target = redirect_to or "/inventory"
    return RedirectResponse(f"{target}?success=Producto+registrado", status_code=303)


@router.post("/inventory/product/{product_id}/update")
def inventory_update_product(
    request: Request,
    product_id: int,
    descripcion: str = Form(...),
    linea_id: Optional[str] = Form(None),
    segmento_id: Optional[str] = Form(None),
    precio_venta1_usd: float = Form(0),
    precio_venta2_usd: float = Form(0),
    precio_venta3_usd: float = Form(0),
    costo_producto_usd: float = Form(0),
    existencia: Optional[float] = Form(None),
    activo: Optional[str] = Form(None),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    accept_header = request.headers.get("accept", "")
    is_fetch = (
        request.headers.get("x-requested-with") == "fetch"
        or "application/json" in accept_header
        or request.headers.get("hx-request") == "true"
    )
    descripcion = descripcion.strip()
    if not descripcion:
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Faltan+datos+obligatorios", status_code=303)

    def _to_int(value: Optional[str]) -> Optional[int]:
        if not value:
            return None
        return int(value) if value.isdigit() else None

    producto = db.query(Producto).filter(Producto.id == product_id).first()
    if not producto:
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Producto+no+encontrado", status_code=303)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if not rate_today:
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Tasa+de+cambio+no+configurada", status_code=303)

    tasa = float(rate_today.rate)
    inventory_cs_only = _inventory_cs_only_mode(db)
    if inventory_cs_only:
        precio_venta1_cs = float(precio_venta1_usd or 0)
        precio_venta2_cs = float(precio_venta2_usd or 0)
        precio_venta3_cs = float(precio_venta3_usd or 0)
        costo_producto_cs = float(costo_producto_usd or 0)
        precio_venta1_usd = (precio_venta1_cs / tasa) if tasa else 0
        precio_venta2_usd = (precio_venta2_cs / tasa) if tasa else 0
        precio_venta3_usd = (precio_venta3_cs / tasa) if tasa else 0
        costo_producto_usd = (costo_producto_cs / tasa) if tasa else 0
    else:
        precio_venta1_cs = precio_venta1_usd * tasa
        precio_venta2_cs = precio_venta2_usd * tasa
        precio_venta3_cs = precio_venta3_usd * tasa
        costo_producto_cs = costo_producto_usd * tasa
    producto.descripcion = descripcion
    producto.linea_id = _to_int(linea_id)
    producto.segmento_id = _to_int(segmento_id)
    producto.precio_venta1_usd = precio_venta1_usd
    producto.precio_venta2_usd = precio_venta2_usd
    producto.precio_venta3_usd = precio_venta3_usd
    producto.precio_venta1 = precio_venta1_cs
    producto.precio_venta2 = precio_venta2_cs
    producto.precio_venta3 = precio_venta3_cs
    producto.costo_producto = costo_producto_cs
    producto.tasa_cambio = tasa
    producto.activo = activo == "on"

    if existencia is not None:
        if producto.saldo:
            producto.saldo.existencia = existencia
        else:
            db.add(SaldoProducto(producto_id=producto.id, existencia=existencia))

    db.commit()
    if is_fetch:
        return JSONResponse(
            {
                "ok": True,
                "message": "Producto actualizado exitosamente",
                "id": producto.id,
                "cod_producto": producto.cod_producto,
                "descripcion": producto.descripcion,
                "precio_venta1_usd": float(producto.precio_venta1_usd or 0),
                "precio_venta2_usd": float(producto.precio_venta2_usd or 0),
                "precio_venta3_usd": float(producto.precio_venta3_usd or 0),
                "costo_usd": float(costo_producto_usd or 0),
                "activo": bool(producto.activo),
            }
        )
    return RedirectResponse(redirect_to or "/inventory", status_code=303)


@router.get("/inventory/product/{product_id}/json")
def inventory_product_json(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    producto = db.query(Producto).filter(Producto.id == product_id).first()
    if not producto:
        return JSONResponse({"ok": False, "message": "Producto no encontrado"}, status_code=404)
    tasa = float(producto.tasa_cambio or 0)
    costo_usd = float(producto.costo_producto or 0) / tasa if tasa else 0.0
    return JSONResponse(
        {
            "ok": True,
            "id": producto.id,
            "cod_producto": producto.cod_producto,
            "descripcion": producto.descripcion,
            "linea_id": producto.linea_id,
            "segmento_id": producto.segmento_id,
            "precio_venta1_usd": float(producto.precio_venta1_usd or 0),
            "precio_venta2_usd": float(producto.precio_venta2_usd or 0),
            "precio_venta3_usd": float(producto.precio_venta3_usd or 0),
            "costo_usd": costo_usd,
            "activo": bool(producto.activo),
        }
    )


@router.get("/inventory/product/{product_id}/combo")
def inventory_product_combo_list(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    producto = db.query(Producto).filter(Producto.id == product_id).first()
    if not producto:
        return JSONResponse({"ok": False, "message": "Producto no encontrado"}, status_code=404)
    combos = (
        db.query(ProductoCombo)
        .filter(ProductoCombo.parent_producto_id == product_id)
        .order_by(ProductoCombo.id)
        .all()
    )
    items = []
    for combo in combos:
        child = combo.child
        items.append(
            {
                "id": combo.id,
                "child_id": child.id if child else None,
                "cod_producto": child.cod_producto if child else "",
                "descripcion": child.descripcion if child else "",
                "cantidad": float(combo.cantidad or 0),
                "precio_venta1_usd": float(child.precio_venta1_usd or 0) if child else 0.0,
                "precio_venta1": float(child.precio_venta1 or 0) if child else 0.0,
                "existencia": float(child.saldo.existencia or 0) if child and child.saldo else 0.0,
            }
        )
    return JSONResponse({"ok": True, "items": items})


@router.post("/inventory/product/{product_id}/combo")
def inventory_product_combo_add(
    request: Request,
    product_id: int,
    child_id: int = Form(...),
    cantidad: float = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    producto = db.query(Producto).filter(Producto.id == product_id).first()
    if not producto:
        return JSONResponse({"ok": False, "message": "Producto no encontrado"}, status_code=404)
    child = db.query(Producto).filter(Producto.id == child_id).first()
    if not child:
        return JSONResponse({"ok": False, "message": "Regalia no encontrada"}, status_code=404)
    if cantidad <= 0:
        return JSONResponse({"ok": False, "message": "Cantidad invalida"}, status_code=400)
    existing = (
        db.query(ProductoCombo)
        .filter(
            ProductoCombo.parent_producto_id == product_id,
            ProductoCombo.child_producto_id == child_id,
        )
        .first()
    )
    if existing:
        existing.cantidad = cantidad
        existing.activo = True
    else:
        db.add(
            ProductoCombo(
                parent_producto_id=product_id,
                child_producto_id=child_id,
                cantidad=cantidad,
                activo=True,
            )
        )
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/inventory/product/{product_id}/combo/{combo_id}/delete")
def inventory_product_combo_delete(
    request: Request,
    product_id: int,
    combo_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    combo = (
        db.query(ProductoCombo)
        .filter(
            ProductoCombo.id == combo_id,
            ProductoCombo.parent_producto_id == product_id,
        )
        .first()
    )
    if not combo:
        return JSONResponse({"ok": False, "message": "Regalia no encontrada"}, status_code=404)
    db.delete(combo)
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/inventory/linea")
def inventory_create_linea(
    request: Request,
    cod_linea: Optional[str] = Form(None),
    linea: str = Form(...),
    activo: Optional[str] = Form(None),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    if not (
        _has_permission(user, "access.inventory.productos")
        or _has_permission(user, "access.inventory.ingresos")
    ):
        _enforce_permission(request, user, "access.inventory.productos")
    is_fetch = request.headers.get("x-requested-with") == "fetch" or "application/json" in (
        request.headers.get("accept", "")
    )
    linea = (linea or "").strip()
    if not linea:
        if is_fetch:
            return JSONResponse({"ok": False, "message": "Nombre de linea requerido"}, status_code=400)
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Nombre+de+linea+requerido", status_code=303)
    exists = db.query(Linea).filter(func.lower(Linea.linea) == linea.lower()).first()
    if exists:
        if is_fetch:
            return JSONResponse(
                {
                    "ok": False,
                    "message": "Linea ya existe",
                    "id": exists.id,
                    "cod_linea": exists.cod_linea,
                    "linea": exists.linea,
                    "activo": bool(exists.activo),
                },
                status_code=409,
            )
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Linea+ya+existe", status_code=303)
    # Codigo de linea 100% automatico, independiente del input del usuario.
    base_code = re.sub(r"[^A-Za-z0-9]+", "_", linea.upper()).strip("_")
    if not base_code:
        base_code = "LINEA"
    base_code = base_code[:40]
    generated_code = base_code
    seq = 2
    while (
        db.query(Linea)
        .filter(func.lower(Linea.cod_linea) == generated_code.lower())
        .first()
        is not None
    ):
        suffix = f"_{seq}"
        generated_code = f"{base_code[: max(1, 50 - len(suffix))]}{suffix}"
        seq += 1

    nueva = Linea(cod_linea=generated_code, linea=linea, activo=activo == "on")
    db.add(nueva)
    db.commit()
    if is_fetch:
        return JSONResponse(
            {
                "ok": True,
                "message": "Linea creada",
                "id": nueva.id,
                "cod_linea": nueva.cod_linea,
                "linea": nueva.linea,
                "activo": bool(nueva.activo),
            }
        )
    return RedirectResponse(redirect_to or "/inventory", status_code=303)


@router.post("/inventory/linea/{linea_id}/update")
def inventory_update_linea(
    request: Request,
    linea_id: int,
    linea: str = Form(...),
    activo: Optional[str] = Form(None),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    linea_obj = db.query(Linea).filter(Linea.id == linea_id).first()
    if linea_obj:
        linea_obj.linea = linea.strip()
        linea_obj.activo = activo == "on"
        db.commit()
    return RedirectResponse(redirect_to or "/inventory", status_code=303)


@router.post("/inventory/segmento")
def inventory_create_segmento(
    request: Request,
    segmento: str = Form(...),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    if not (
        _has_permission(user, "access.inventory.productos")
        or _has_permission(user, "access.inventory.ingresos")
    ):
        _enforce_permission(request, user, "access.inventory.productos")
    is_fetch = request.headers.get("x-requested-with") == "fetch" or "application/json" in (
        request.headers.get("accept", "")
    )
    segmento = (segmento or "").strip()
    if not segmento:
        if is_fetch:
            return JSONResponse({"ok": False, "message": "Nombre de segmento requerido"}, status_code=400)
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Nombre+de+segmento+requerido", status_code=303)
    exists = db.query(Segmento).filter(func.lower(Segmento.segmento) == segmento.lower()).first()
    if exists:
        if is_fetch:
            return JSONResponse(
                {
                    "ok": False,
                    "message": "Segmento ya existe",
                    "id": exists.id,
                    "segmento": exists.segmento,
                },
                status_code=409,
            )
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Segmento+ya+existe", status_code=303)
    nuevo = Segmento(segmento=segmento)
    db.add(nuevo)
    db.commit()
    if is_fetch:
        return JSONResponse(
            {
                "ok": True,
                "message": "Segmento creado",
                "id": nuevo.id,
                "segmento": nuevo.segmento,
            }
        )
    return RedirectResponse(redirect_to or "/inventory", status_code=303)


@router.post("/inventory/segmento/{segmento_id}/update")
def inventory_update_segmento(
    request: Request,
    segmento_id: int,
    segmento: str = Form(...),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    segmento_obj = db.query(Segmento).filter(Segmento.id == segmento_id).first()
    if segmento_obj:
        segmento_obj.segmento = segmento.strip()
        db.commit()
    return RedirectResponse(redirect_to or "/inventory", status_code=303)


@router.post("/inventory/marca")
def inventory_create_marca(
    request: Request,
    nombre: str = Form(...),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    if not (
        _has_permission(user, "access.inventory.productos")
        or _has_permission(user, "access.inventory.ingresos")
    ):
        _enforce_permission(request, user, "access.inventory.productos")
    is_fetch = request.headers.get("x-requested-with") == "fetch" or "application/json" in (
        request.headers.get("accept", "")
    )
    nombre = (nombre or "").strip()
    if not nombre:
        if is_fetch:
            return JSONResponse({"ok": False, "message": "Nombre de marca requerido"}, status_code=400)
        target = redirect_to or "/inventory/ingresos"
        return RedirectResponse(f"{target}?error=Nombre+de+marca+requerido", status_code=303)
    exists = db.query(Marca).filter(func.lower(Marca.nombre) == nombre.lower()).first()
    if exists:
        if is_fetch:
            return JSONResponse(
                {"ok": True, "message": "Marca existente", "id": exists.id, "nombre": exists.nombre},
                status_code=200,
            )
        return RedirectResponse(redirect_to or "/inventory/ingresos", status_code=303)
    marca = Marca(nombre=nombre, activo=True)
    db.add(marca)
    db.commit()
    if is_fetch:
        return JSONResponse(
            {"ok": True, "message": "Marca creada", "id": marca.id, "nombre": marca.nombre},
            status_code=200,
        )
    return RedirectResponse(redirect_to or "/inventory/ingresos", status_code=303)


@router.post("/inventory/proveedor")
def inventory_create_proveedor(
    request: Request,
    nombre: str = Form(...),
    tipo: Optional[str] = Form(None),
    activo: Optional[str] = Form(None),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: User = Depends(_require_admin_web),
):
    nombre = nombre.strip()
    if not nombre:
        if request.headers.get("X-Requested-With") == "fetch":
            return JSONResponse({"ok": False, "message": "Nombre requerido"}, status_code=400)
        return RedirectResponse(redirect_to or "/inventory/ingresos", status_code=303)
    exists = db.query(Proveedor).filter(func.lower(Proveedor.nombre) == nombre.lower()).first()
    if not exists:
        proveedor = Proveedor(nombre=nombre, tipo=tipo, activo=activo == "on")
        db.add(proveedor)
        db.commit()
        if request.headers.get("X-Requested-With") == "fetch":
            return JSONResponse(
                {
                    "ok": True,
                    "id": proveedor.id,
                    "nombre": proveedor.nombre,
                    "tipo": proveedor.tipo,
                    "activo": proveedor.activo,
                }
            )
    if request.headers.get("X-Requested-With") == "fetch":
        return JSONResponse({"ok": False, "message": "Proveedor ya existe"}, status_code=409)
    return RedirectResponse(redirect_to or "/inventory/ingresos", status_code=303)


@router.post("/inventory/proveedor/{proveedor_id}/update")
def inventory_update_proveedor(
    request: Request,
    proveedor_id: int,
    nombre: str = Form(...),
    tipo: Optional[str] = Form(None),
    activo: Optional[str] = Form(None),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: User = Depends(_require_admin_web),
):
    proveedor_obj = db.query(Proveedor).filter(Proveedor.id == proveedor_id).first()
    if proveedor_obj:
        proveedor_obj.nombre = nombre.strip()
        proveedor_obj.tipo = tipo
        proveedor_obj.activo = activo == "on"
        db.commit()
        if request.headers.get("X-Requested-With") == "fetch":
            return JSONResponse(
                {
                    "ok": True,
                    "id": proveedor_obj.id,
                    "nombre": proveedor_obj.nombre,
                    "tipo": proveedor_obj.tipo,
                    "activo": proveedor_obj.activo,
                }
            )
    if request.headers.get("X-Requested-With") == "fetch":
        return JSONResponse({"ok": False, "message": "Proveedor no encontrado"}, status_code=404)
    return RedirectResponse(redirect_to or "/inventory/ingresos", status_code=303)


@router.post("/inventory/ingresos")
async def inventory_create_ingreso(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.ingresos")
    form = await request.form()
    tipo_id = form.get("tipo_id")
    bodega_id = form.get("bodega_id")
    proveedor_id = form.get("proveedor_id") or None
    fecha = form.get("fecha")
    moneda = form.get("moneda")
    observacion = form.get("observacion") or None
    item_ids = form.getlist("item_producto_id")
    item_qtys = form.getlist("item_cantidad")
    item_costs = form.getlist("item_costo")
    item_prices = form.getlist("item_precio")

    inventory_cs_only = _inventory_cs_only_mode(db)
    if inventory_cs_only:
        moneda = "CS"
    if not tipo_id or not bodega_id or not fecha or not moneda:
        return RedirectResponse("/inventory/ingresos?error=Faltan+datos+obligatorios", status_code=303)
    if not item_ids:
        return RedirectResponse("/inventory/ingresos?error=Agrega+productos+al+ingreso", status_code=303)

    tipo = db.query(IngresoTipo).filter(IngresoTipo.id == int(tipo_id)).first()
    if not tipo:
        return RedirectResponse("/inventory/ingresos?error=Tipo+no+valido", status_code=303)
    if "traslado" in (tipo.nombre or "").lower():
        return RedirectResponse(
            "/inventory/ingresos?error=El+tipo+Traslado+se+genera+desde+egresos",
            status_code=303,
        )
    if tipo.requiere_proveedor and not proveedor_id:
        return RedirectResponse("/inventory/ingresos?error=Proveedor+requerido", status_code=303)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if moneda == "USD" and not rate_today:
        return RedirectResponse("/inventory/ingresos?error=Tasa+de+cambio+no+configurada", status_code=303)

    def to_float(value: Optional[str]) -> float:
        if not value:
            return 0.0
        try:
            return float(value)
        except ValueError:
            return 0.0

    def to_decimal(value: Optional[float]) -> Decimal:
        return Decimal(str(value or 0))

    tasa = float(rate_today.rate) if rate_today else 0
    fecha_value = date.fromisoformat(str(fecha).split("T")[0])
    ingreso = IngresoInventario(
        tipo_id=int(tipo_id),
        bodega_id=int(bodega_id),
        proveedor_id=int(proveedor_id) if proveedor_id else None,
        fecha=fecha_value,
        moneda=moneda,
        tasa_cambio=tasa if moneda == "USD" else None,
        observacion=observacion,
        usuario_registro=user.full_name,
    )
    db.add(ingreso)
    db.flush()

    total_usd = 0.0
    total_cs = 0.0
    for index, product_id in enumerate(item_ids):
        qty = to_float(item_qtys[index] if index < len(item_qtys) else 0)
        cost = to_float(item_costs[index] if index < len(item_costs) else 0)
        price = to_float(item_prices[index] if index < len(item_prices) else 0)
        if qty <= 0:
            continue
        qty_dec = to_decimal(qty)

        if moneda == "USD":
            costo_usd = cost
            costo_cs = cost * tasa
            precio_usd = price
            precio_cs = price * tasa
        else:
            costo_cs = cost
            costo_usd = cost / tasa if tasa else 0
            precio_cs = price
            precio_usd = price / tasa if tasa else 0

        subtotal_usd = costo_usd * qty
        subtotal_cs = costo_cs * qty
        total_usd += subtotal_usd
        total_cs += subtotal_cs

        item = IngresoItem(
            ingreso_id=ingreso.id,
            producto_id=int(product_id),
            cantidad=qty,
            costo_unitario_usd=costo_usd,
            costo_unitario_cs=costo_cs,
            subtotal_usd=subtotal_usd,
            subtotal_cs=subtotal_cs,
        )
        db.add(item)

        producto = db.query(Producto).filter(Producto.id == int(product_id)).first()
        if producto:
            if producto.saldo:
                current = Decimal(str(producto.saldo.existencia or 0))
                producto.saldo.existencia = current + qty_dec
            else:
                db.add(SaldoProducto(producto_id=producto.id, existencia=qty_dec))
            if cost > 0:
                producto.costo_producto = costo_cs
            if price > 0:
                producto.precio_venta1 = precio_cs
                if precio_usd > 0:
                    producto.precio_venta1_usd = precio_usd

    ingreso.total_usd = total_usd
    ingreso.total_cs = total_cs
    bodega_obj = db.query(Bodega).filter(Bodega.id == int(bodega_id)).first()
    auto_amount = to_decimal(total_usd if moneda == "USD" else total_cs)
    auto_entry = _build_auto_accounting_entry(
        db,
        event_code="INV_IN",
        branch_id=bodega_obj.branch_id if bodega_obj else None,
        entry_date=fecha_value,
        amount=auto_amount,
        reference=f"AUTO-ING-{ingreso.id}",
        description=f"Asiento automatico por ingreso inventario #{ingreso.id}",
    )
    if auto_entry:
        db.add(auto_entry)
    db.commit()
    return RedirectResponse(
        f"/inventory/ingresos?success=Ingreso+registrado&print_id={ingreso.id}",
        status_code=303,
    )


@router.post("/inventory/egresos")
async def inventory_create_egreso(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.egresos")
    form = await request.form()
    tipo_id = form.get("tipo_id")
    bodega_id = form.get("bodega_id")
    bodega_destino_id = form.get("bodega_destino_id") or None
    fecha = form.get("fecha")
    moneda = form.get("moneda")
    observacion = form.get("observacion") or None
    item_ids = form.getlist("item_producto_id")
    item_qtys = form.getlist("item_cantidad")
    item_costs = form.getlist("item_costo")
    item_prices = form.getlist("item_precio")

    inventory_cs_only = _inventory_cs_only_mode(db)
    if inventory_cs_only:
        moneda = "CS"
    if not tipo_id or not bodega_id or not fecha or not moneda:
        return RedirectResponse("/inventory/egresos?error=Faltan+datos+obligatorios", status_code=303)
    if not item_ids:
        return RedirectResponse("/inventory/egresos?error=Agrega+productos+al+egreso", status_code=303)

    tipo = db.query(EgresoTipo).filter(EgresoTipo.id == int(tipo_id)).first()
    if not tipo:
        return RedirectResponse("/inventory/egresos?error=Tipo+no+valido", status_code=303)
    es_traslado = "traslado" in (tipo.nombre or "").lower()
    bodega_destino_obj = None
    if es_traslado:
        if not bodega_destino_id:
            return RedirectResponse("/inventory/egresos?error=Selecciona+bodega+destino+para+traslado", status_code=303)
        if int(bodega_destino_id) == int(bodega_id):
            return RedirectResponse("/inventory/egresos?error=La+bodega+destino+debe+ser+distinta+al+origen", status_code=303)
        bodega_destino_obj = (
            db.query(Bodega)
            .filter(Bodega.id == int(bodega_destino_id), Bodega.activo.is_(True))
            .first()
        )
        if not bodega_destino_obj:
            return RedirectResponse("/inventory/egresos?error=Bodega+destino+no+valida", status_code=303)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if moneda == "USD" and not rate_today:
        return RedirectResponse("/inventory/egresos?error=Tasa+de+cambio+no+configurada", status_code=303)

    def to_float(value: Optional[str]) -> float:
        if not value:
            return 0.0
        try:
            return float(value)
        except ValueError:
            return 0.0

    tasa = float(rate_today.rate) if rate_today else 0
    fecha_value = date.fromisoformat(str(fecha).split("T")[0])
    egreso = EgresoInventario(
        tipo_id=int(tipo_id),
        bodega_id=int(bodega_id),
        bodega_destino_id=int(bodega_destino_id) if bodega_destino_id else None,
        fecha=fecha_value,
        moneda=moneda,
        tasa_cambio=tasa if moneda == "USD" else None,
        observacion=observacion,
        usuario_registro=user.full_name,
    )
    db.add(egreso)
    db.flush()

    total_usd = 0.0
    total_cs = 0.0
    traslado_items: list[dict[str, float | int]] = []
    product_ids = [int(pid) for pid in item_ids if str(pid).isdigit()]
    balances = _balances_by_bodega(db, [int(bodega_id)], list(set(product_ids))) if product_ids else {}
    for index, product_id in enumerate(item_ids):
        qty = to_float(item_qtys[index] if index < len(item_qtys) else 0)
        cost = to_float(item_costs[index] if index < len(item_costs) else 0)
        if qty <= 0:
            continue
        producto = db.query(Producto).filter(Producto.id == int(product_id)).first()
        if not producto:
            db.rollback()
            return RedirectResponse("/inventory/egresos?error=Producto+no+encontrado", status_code=303)

        existencia = float(balances.get((producto.id, int(bodega_id)), Decimal("0")) or 0)
        if existencia < qty:
            db.rollback()
            mensaje = f"Stock+insuficiente+para+{producto.cod_producto}"
            return RedirectResponse(f"/inventory/egresos?error={mensaje}", status_code=303)
        balances[(producto.id, int(bodega_id))] = Decimal(str(existencia)) - to_decimal(qty)

        if moneda == "USD":
            costo_usd = cost
            costo_cs = cost * tasa
        else:
            costo_cs = cost
            costo_usd = cost / tasa if tasa else 0

        subtotal_usd = costo_usd * qty
        subtotal_cs = costo_cs * qty
        total_usd += subtotal_usd
        total_cs += subtotal_cs

        item = EgresoItem(
            egreso_id=egreso.id,
            producto_id=int(product_id),
            cantidad=qty,
            costo_unitario_usd=costo_usd,
            costo_unitario_cs=costo_cs,
            subtotal_usd=subtotal_usd,
            subtotal_cs=subtotal_cs,
        )
        db.add(item)

        if producto.saldo:
            existencia_actual = to_decimal(producto.saldo.existencia)
            producto.saldo.existencia = existencia_actual - to_decimal(qty)
        traslado_items.append(
            {
                "producto_id": int(product_id),
                "cantidad": qty,
                "costo_unitario_usd": costo_usd,
                "costo_unitario_cs": costo_cs,
                "subtotal_usd": subtotal_usd,
                "subtotal_cs": subtotal_cs,
            }
        )

    if es_traslado and bodega_destino_obj and traslado_items:
        ingreso_tipo = (
            db.query(IngresoTipo)
            .filter(func.lower(IngresoTipo.nombre) == "traslado entre bodegas")
            .first()
        )
        if not ingreso_tipo:
            ingreso_tipo = IngresoTipo(nombre="Traslado entre bodegas", requiere_proveedor=False)
            db.add(ingreso_tipo)
            db.flush()
        bodega_origen_obj = db.query(Bodega).filter(Bodega.id == int(bodega_id)).first()
        traslado_obs = (
            f"Traslado desde {bodega_origen_obj.name if bodega_origen_obj else 'origen'} "
            f"hacia {bodega_destino_obj.name}. Egreso #{egreso.id}"
        )
        if observacion:
            traslado_obs = f"{traslado_obs} | {observacion}"
        ingreso = IngresoInventario(
            tipo_id=ingreso_tipo.id,
            bodega_id=bodega_destino_obj.id,
            proveedor_id=None,
            fecha=fecha_value,
            moneda=moneda,
            tasa_cambio=tasa if moneda == "USD" else None,
            total_usd=total_usd,
            total_cs=total_cs,
            observacion=traslado_obs[:300],
            usuario_registro=user.full_name,
        )
        db.add(ingreso)
        db.flush()
        for row in traslado_items:
            db.add(
                IngresoItem(
                    ingreso_id=ingreso.id,
                    producto_id=int(row["producto_id"]),
                    cantidad=float(row["cantidad"]),
                    costo_unitario_usd=float(row["costo_unitario_usd"]),
                    costo_unitario_cs=float(row["costo_unitario_cs"]),
                    subtotal_usd=float(row["subtotal_usd"]),
                    subtotal_cs=float(row["subtotal_cs"]),
                )
            )
            producto = db.query(Producto).filter(Producto.id == int(row["producto_id"])).first()
            if producto and producto.saldo:
                producto.saldo.existencia = to_decimal(producto.saldo.existencia) + to_decimal(float(row["cantidad"]))
            elif producto:
                db.add(SaldoProducto(producto_id=producto.id, existencia=to_decimal(float(row["cantidad"]))))

    egreso.total_usd = total_usd
    egreso.total_cs = total_cs
    bodega_obj = db.query(Bodega).filter(Bodega.id == int(bodega_id)).first()
    auto_amount = to_decimal(total_usd if moneda == "USD" else total_cs)
    auto_entry = _build_auto_accounting_entry(
        db,
        event_code="INV_OUT",
        branch_id=bodega_obj.branch_id if bodega_obj else None,
        entry_date=fecha_value,
        amount=auto_amount,
        reference=f"AUTO-EGR-{egreso.id}",
        description=f"Asiento automatico por egreso inventario #{egreso.id}",
    )
    if auto_entry:
        db.add(auto_entry)
    db.commit()
    return RedirectResponse(
        f"/inventory/egresos?success=Egreso+registrado&print_id={egreso.id}",
        status_code=303,
    )


@router.post("/sales/cliente")
def sales_create_cliente(
    request: Request,
    nombre: str = Form(...),
    identificacion: Optional[str] = Form(None),
    telefono: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    direccion: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales")
    nombre = nombre.strip()
    if not nombre:
        return JSONResponse({"ok": False, "message": "Nombre requerido"}, status_code=400)
    exists = db.query(Cliente).filter(func.lower(Cliente.nombre) == nombre.lower()).first()
    if exists:
        return JSONResponse({"ok": False, "message": "Cliente ya existe"}, status_code=409)
    cliente = Cliente(
        nombre=nombre,
        identificacion=identificacion.strip() if identificacion else None,
        telefono=telefono,
        email=email,
        direccion=direccion,
        activo=True,
    )
    db.add(cliente)
    db.commit()
    return JSONResponse({"ok": True, "id": cliente.id, "nombre": cliente.nombre})


@router.get("/sales/clientes/search")
def sales_clientes_search(
    request: Request,
    q: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales")
    term = (q or "").strip()
    if len(term) < 1:
        return JSONResponse({"ok": True, "items": []})
    clientes = (
        db.query(Cliente)
        .filter(func.lower(Cliente.nombre).like(f"%{term.lower()}%"))
        .order_by(Cliente.nombre)
        .limit(25)
        .all()
    )
    return JSONResponse({"ok": True, "items": [{"id": c.id, "nombre": c.nombre} for c in clientes]})


@router.post("/sales/cliente/{cliente_id}/update")
def sales_update_cliente(
    cliente_id: int,
    nombre: str = Form(...),
    identificacion: Optional[str] = Form(None),
    telefono: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    direccion: Optional[str] = Form(None),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: User = Depends(_require_admin_web),
):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        return JSONResponse({"ok": False, "message": "Cliente no encontrado"}, status_code=404)
    cliente.nombre = nombre.strip()
    cliente.identificacion = identificacion.strip() if identificacion else None
    cliente.telefono = telefono
    cliente.email = email
    cliente.direccion = direccion
    cliente.activo = activo == "on"
    db.commit()
    return JSONResponse({"ok": True, "id": cliente.id, "nombre": cliente.nombre})


@router.post("/sales/vendedor")
def sales_create_vendedor(
    nombre: str = Form(...),
    telefono: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: User = Depends(_require_admin_web),
):
    nombre = nombre.strip()
    if not nombre:
        return JSONResponse({"ok": False, "message": "Nombre requerido"}, status_code=400)
    exists = db.query(Vendedor).filter(func.lower(Vendedor.nombre) == nombre.lower()).first()
    if exists:
        return JSONResponse({"ok": False, "message": "Vendedor ya existe"}, status_code=409)
    vendedor = Vendedor(nombre=nombre, telefono=telefono, activo=True)
    db.add(vendedor)
    db.commit()
    return JSONResponse({"ok": True, "id": vendedor.id, "nombre": vendedor.nombre})


@router.post("/sales/vendedor/{vendedor_id}/update")
def sales_update_vendedor(
    vendedor_id: int,
    nombre: str = Form(...),
    telefono: Optional[str] = Form(None),
    activo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    _: User = Depends(_require_admin_web),
):
    vendedor = db.query(Vendedor).filter(Vendedor.id == vendedor_id).first()
    if not vendedor:
        return JSONResponse({"ok": False, "message": "Vendedor no encontrado"}, status_code=404)
    vendedor.nombre = nombre.strip()
    vendedor.telefono = telefono
    vendedor.activo = activo == "on"
    db.commit()
    return JSONResponse({"ok": True, "id": vendedor.id, "nombre": vendedor.nombre})


@router.post("/sales")
async def sales_create_invoice(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.sales.registrar")
    form = await request.form()
    cliente_id = form.get("cliente_id") or None
    vendedor_id = form.get("vendedor_id") or None
    fecha = form.get("fecha")
    moneda = form.get("moneda")
    forma_pago_id = form.get("forma_pago_id") or None
    banco_id = form.get("banco_id") or None
    cuenta_id = form.get("cuenta_id") or None
    pago_monto = form.get("pago_monto") or None
    pago_forma_ids = form.getlist("pago_forma_id")
    pago_monedas = form.getlist("pago_moneda")
    pago_montos = form.getlist("pago_monto")
    pago_banco_ids = form.getlist("pago_banco_id")
    pago_cuenta_ids = form.getlist("pago_cuenta_id")
    item_ids = form.getlist("item_producto_id")
    item_qtys = form.getlist("item_cantidad")
    item_prices = form.getlist("item_precio")
    item_roles = form.getlist("item_role")
    item_combo_groups = form.getlist("item_combo_group")
    preventa_id_raw = str(form.get("preventa_id") or "").strip()

    if not fecha:
        fecha = local_today().isoformat()
    if not moneda:
        moneda = "CS"
    if not vendedor_id:
        vendedor = db.query(Vendedor).filter(Vendedor.activo.is_(True)).order_by(Vendedor.id).first()
        if vendedor:
            vendedor_id = str(vendedor.id)
    if not vendedor_id or not fecha or not moneda:
        return RedirectResponse("/sales?error=Faltan+datos+obligatorios", status_code=303)
    if not item_ids and not preventa_id_raw:
        return RedirectResponse("/sales?error=Agrega+productos+a+la+venta", status_code=303)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if moneda == "USD" and not rate_today:
        return RedirectResponse("/sales?error=Tasa+de+cambio+no+configurada", status_code=303)

    def to_float(value: Optional[str]) -> float:
        if not value:
            return 0.0
        try:
            return float(value)
        except ValueError:
            return 0.0

    def to_decimal(value: Optional[float]) -> Decimal:
        return Decimal(str(value or 0))

    tasa = float(rate_today.rate) if rate_today else 0
    branch, bodega = _resolve_branch_bodega(db, user)
    if not branch:
        return RedirectResponse("/sales?error=Usuario+sin+sucursal+asignada", status_code=303)
    if not bodega:
        return RedirectResponse("/sales?error=Bodega+no+configurada+para+la+sucursal", status_code=303)
    last_factura = (
        db.query(VentaFactura)
        .filter(VentaFactura.bodega_id == bodega.id)
        .order_by(VentaFactura.secuencia.desc())
        .first()
    )
    next_seq = (last_factura.secuencia if last_factura else 0) + 1
    branch_code = (branch.code or "").lower()
    prefix = "C" if branch_code == "central" else "E" if branch_code == "esteli" else branch_code[:1].upper()
    width = 6
    numero = f"{prefix}-{next_seq:0{width}d}"

    now_local = local_now()
    try:
        fecha_value = date.fromisoformat(str(fecha).split("T")[0])
    except (TypeError, ValueError):
        fecha_value = local_today()
    fecha_dt = datetime.combine(fecha_value, now_local.time()).replace(tzinfo=None)
    factura = VentaFactura(
        secuencia=next_seq,
        numero=numero,
        bodega_id=bodega.id,
        cliente_id=int(cliente_id) if cliente_id else None,
        vendedor_id=int(vendedor_id) if vendedor_id else None,
        fecha=fecha_dt,
        moneda=moneda,
        tasa_cambio=tasa if moneda == "USD" else None,
        usuario_registro=user.full_name,
        created_at=local_now_naive(),
    )
    db.add(factura)
    db.flush()
    preventa: Optional[Preventa] = None
    source_items: list[dict[str, object]] = []
    if preventa_id_raw:
        if not preventa_id_raw.isdigit():
            db.rollback()
            return RedirectResponse("/sales?error=Preventa+invalida", status_code=303)
        preventa = (
            db.query(Preventa)
            .filter(Preventa.id == int(preventa_id_raw))
            .first()
        )
        if not preventa:
            db.rollback()
            return RedirectResponse("/sales?error=Preventa+no+encontrada", status_code=303)
        if preventa.estado not in {"PENDIENTE", "REVISION"}:
            db.rollback()
            return RedirectResponse("/sales?error=Preventa+no+disponible+para+facturar", status_code=303)
        if int(preventa.bodega_id or 0) != int(bodega.id):
            db.rollback()
            return RedirectResponse("/sales?error=Preventa+de+otra+bodega", status_code=303)
        p_items = (
            db.query(PreventaItem)
            .filter(PreventaItem.preventa_id == preventa.id)
            .order_by(PreventaItem.id.asc())
            .all()
        )
        if not p_items:
            db.rollback()
            return RedirectResponse("/sales?error=Preventa+sin+items", status_code=303)
        for p_item in p_items:
            source_items.append(
                {
                    "product_id": int(p_item.producto_id),
                    "qty": to_float(str(p_item.cantidad or 0)),
                    "price_usd": to_float(str(p_item.precio_unitario_usd or 0)),
                    "price_cs": to_float(str(p_item.precio_unitario_cs or 0)),
                    "role": p_item.combo_role or None,
                    "combo_group": p_item.combo_group or None,
                }
            )
    else:
        for index, product_id in enumerate(item_ids):
            if not str(product_id).isdigit():
                continue
            source_items.append(
                {
                    "product_id": int(product_id),
                    "qty": to_float(item_qtys[index] if index < len(item_qtys) else 0),
                    "price_usd": None,
                    "price_cs": None,
                    "price_input": to_float(item_prices[index] if index < len(item_prices) else 0),
                    "role": item_roles[index] if index < len(item_roles) else None,
                    "combo_group": item_combo_groups[index] if index < len(item_combo_groups) else None,
                }
            )
    if not source_items:
        db.rollback()
        return RedirectResponse("/sales?error=No+hay+items+validos", status_code=303)

    total_usd = 0.0
    total_cs = 0.0
    total_items = 0.0
    product_ids = [int(it["product_id"]) for it in source_items if int(it["product_id"]) > 0]
    balances = _balances_by_bodega(db, [bodega.id], list(set(product_ids))) if product_ids else {}
    for src in source_items:
        product_id = int(src["product_id"])
        qty = to_float(str(src.get("qty") or 0))
        if qty <= 0:
            continue

        producto = db.query(Producto).filter(Producto.id == int(product_id)).first()
        if not producto:
            db.rollback()
            return RedirectResponse("/sales?error=Producto+no+encontrado", status_code=303)

        existencia = float(balances.get((producto.id, bodega.id), Decimal("0")) or 0)
        if existencia < qty:
            db.rollback()
            mensaje = f"Stock+insuficiente+para+{producto.cod_producto}"
            return RedirectResponse(f"/sales?error={mensaje}", status_code=303)
        balances[(producto.id, bodega.id)] = Decimal(str(existencia)) - to_decimal(qty)

        if preventa:
            precio_usd = to_float(str(src.get("price_usd") or 0))
            precio_cs = to_float(str(src.get("price_cs") or 0))
        else:
            price = to_float(str(src.get("price_input") or 0))
            if moneda == "USD":
                precio_usd = price
                precio_cs = price * tasa
            else:
                precio_cs = price
                precio_usd = price / tasa if tasa else 0

        subtotal_usd = precio_usd * qty
        subtotal_cs = precio_cs * qty
        total_usd += subtotal_usd
        total_cs += subtotal_cs
        total_items += qty

        combo_role = src.get("role")
        combo_group = src.get("combo_group")
        combo_role = combo_role or None
        combo_group = combo_group or None
        item = VentaItem(
            factura_id=factura.id,
            producto_id=int(product_id),
            cantidad=qty,
            precio_unitario_usd=precio_usd,
            precio_unitario_cs=precio_cs,
            subtotal_usd=subtotal_usd,
            subtotal_cs=subtotal_cs,
            combo_role=combo_role,
            combo_group=combo_group,
        )
        db.add(item)

        # No usar saldo global; el stock se calcula por movimientos/bodega.

    factura.total_usd = total_usd
    factura.total_cs = total_cs
    factura.total_items = total_items

    pagos = []
    if pago_forma_ids:
        for index, forma_id in enumerate(pago_forma_ids):
            moneda_pago = pago_monedas[index] if index < len(pago_monedas) else moneda
            monto_pago = to_float(pago_montos[index] if index < len(pago_montos) else 0)
            banco_pago = pago_banco_ids[index] if index < len(pago_banco_ids) else None
            cuenta_pago = pago_cuenta_ids[index] if index < len(pago_cuenta_ids) else None
            if monto_pago <= 0:
                continue
            if moneda_pago != moneda and not tasa:
                db.rollback()
                return RedirectResponse("/sales?error=Tasa+de+cambio+no+configurada", status_code=303)
            pago_usd = monto_pago if moneda_pago == "USD" else monto_pago / tasa if tasa else 0
            pago_cs = monto_pago if moneda_pago == "CS" else monto_pago * tasa
            pagos.append(
                VentaPago(
                    factura_id=factura.id,
                    forma_pago_id=int(forma_id),
                    banco_id=int(banco_pago) if banco_pago else None,
                    cuenta_id=int(cuenta_pago) if cuenta_pago else None,
                    monto_usd=pago_usd,
                    monto_cs=pago_cs,
                )
            )
    elif forma_pago_id:
        monto_pago = to_float(pago_monto) if pago_monto is not None else (total_usd if moneda == "USD" else total_cs)
        pago_usd = monto_pago if moneda == "USD" else monto_pago / tasa if tasa else 0
        pago_cs = monto_pago if moneda == "CS" else monto_pago * tasa
        pagos.append(
            VentaPago(
                factura_id=factura.id,
                forma_pago_id=int(forma_pago_id),
                banco_id=int(banco_id) if banco_id else None,
                cuenta_id=int(cuenta_id) if cuenta_id else None,
                monto_usd=pago_usd,
                monto_cs=pago_cs,
            )
        )

    if not pagos:
        db.rollback()
        return RedirectResponse("/sales?error=Agrega+pagos+para+registrar", status_code=303)

    total_paid = sum(pago.monto_usd for pago in pagos) if moneda == "USD" else sum(pago.monto_cs for pago in pagos)
    due_total = total_usd if moneda == "USD" else total_cs
    if float(total_paid) < float(due_total):
        db.rollback()
        return RedirectResponse("/sales?error=Pago+incompleto", status_code=303)

    factura.estado_cobranza = "PAGADA"
    if preventa:
        preventa.estado = "FACTURADA"
        preventa.facturada_at = local_now_naive()
        preventa.venta_factura_id = factura.id
    for pago in pagos:
        db.add(pago)

    auto_amount = to_decimal(total_usd if moneda == "USD" else total_cs)
    auto_entry = _build_auto_accounting_entry(
        db,
        event_code="SALE",
        branch_id=bodega.branch_id if bodega else None,
        entry_date=fecha_value,
        amount=auto_amount,
        reference=f"AUTO-VTA-{factura.id}",
        description=f"Asiento automatico por venta {factura.numero}",
    )
    if auto_entry:
        db.add(auto_entry)

    db.commit()
    pos_print = (
        db.query(PosPrintSetting)
        .filter(PosPrintSetting.branch_id == branch.id)
        .first()
    )
    if pos_print and pos_print.auto_print:
        try:
            company_profile = _company_profile_payload(db)
            copies_to_print = max(int(pos_print.copies or 0), 2)
            _print_pos_ticket(
                factura,
                pos_print.printer_name,
                copies_to_print,
                company_profile,
                pos_print.sumatra_path,
            )
        except Exception:
            pass
    return RedirectResponse(
        f"/sales?success=Venta+registrada&print_id={factura.id}",
        status_code=303,
    )


@router.get("/sales/cobranza")
def sales_cobranza(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.cobranza")
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    producto_q = (request.query_params.get("producto") or "").strip()
    vendedor_q = (request.query_params.get("vendedor_id") or "").strip()
    today_value = local_today()
    start_date = today_value
    end_date = today_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today_value
            end_date = today_value

    _, bodega = _resolve_branch_bodega(db, user)
    ventas_query = db.query(VentaFactura)
    if bodega:
        ventas_query = ventas_query.filter(VentaFactura.bodega_id == bodega.id)
    ventas_query = ventas_query.filter(VentaFactura.fecha >= start_date, VentaFactura.fecha <= end_date)
    if vendedor_q:
        try:
            ventas_query = ventas_query.filter(VentaFactura.vendedor_id == int(vendedor_q))
        except ValueError:
            pass
    if producto_q:
        ventas_query = ventas_query.join(VentaItem).join(Producto).filter(
            or_(
                func.lower(Producto.cod_producto).like(f"%{producto_q.lower()}%"),
                func.lower(Producto.descripcion).like(f"%{producto_q.lower()}%"),
            )
        )
    ventas = ventas_query.order_by(VentaFactura.fecha.desc(), VentaFactura.id.desc()).all()

    results = []
    total_saldo_cs = Decimal("0")
    total_saldo_usd = Decimal("0")
    total_pacas = Decimal("0")
    total_vendido_cs = Decimal("0")
    for factura in ventas:
        if factura.estado == "ANULADA":
            total_abono_usd = Decimal("0")
            total_abono_cs = Decimal("0")
            saldo_usd = Decimal("0")
            saldo_cs = Decimal("0")
        else:
            total_pacas += Decimal(str(factura.total_items or 0))
            total_vendido_cs += Decimal(str(factura.total_cs or 0))
            total_abono_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos)
            total_abono_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos)
            total_due_usd = Decimal(str(factura.total_usd or 0))
            total_due_cs = Decimal(str(factura.total_cs or 0))

            if factura.estado_cobranza == "PENDIENTE":
                total_paid_usd = total_abono_usd
                total_paid_cs = total_abono_cs
            else:
                total_paid_usd = sum(Decimal(str(p.monto_usd or 0)) for p in factura.pagos) + total_abono_usd
                total_paid_cs = sum(Decimal(str(p.monto_cs or 0)) for p in factura.pagos) + total_abono_cs

            saldo_usd = max(total_due_usd - total_paid_usd, Decimal("0"))
            saldo_cs = max(total_due_cs - total_paid_cs, Decimal("0"))
            total_saldo_cs += saldo_cs
            total_saldo_usd += saldo_usd
        results.append(
            {
                "factura": factura,
                "abonos_usd": total_abono_usd,
                "abonos_cs": total_abono_cs,
                "saldo_usd": saldo_usd,
                "saldo_cs": saldo_cs,
            }
        )

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")
    vendedores = _vendedores_for_bodega(db, bodega)

    return request.app.state.templates.TemplateResponse(
        "sales_cobranza.html",
        {
            "request": request,
            "user": user,
            "ventas": results,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "producto_q": producto_q,
            "vendedor_q": vendedor_q,
            "vendedores": vendedores,
            "default_vendedor_id": _default_vendedor_id(db, bodega),
            "total_saldo_cs": float(total_saldo_cs),
            "total_saldo_usd": float(total_saldo_usd),
            "total_pacas": float(total_pacas),
            "total_vendido_cs": float(total_vendido_cs),
            "version": settings.UI_VERSION,
        },
    )


@router.get("/sales/cobranza/export")
def sales_cobranza_export(
    request: Request,
    format: str = "pdf",
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.cobranza")
    company_profile = _company_profile_payload(db)
    start_raw = request.query_params.get("start_date")
    end_raw = request.query_params.get("end_date")
    producto_q = (request.query_params.get("producto") or "").strip()
    vendedor_q = (request.query_params.get("vendedor_id") or "").strip()
    today_value = local_today()
    start_date = today_value
    end_date = today_value
    if start_raw or end_raw:
        try:
            if start_raw:
                start_date = date.fromisoformat(start_raw)
            if end_raw:
                end_date = date.fromisoformat(end_raw)
        except ValueError:
            start_date = today_value
            end_date = today_value

    _, bodega = _resolve_branch_bodega(db, user)
    ventas_query = db.query(VentaFactura)
    if bodega:
        ventas_query = ventas_query.filter(VentaFactura.bodega_id == bodega.id)
    ventas_query = ventas_query.filter(VentaFactura.fecha >= start_date, VentaFactura.fecha <= end_date)
    if vendedor_q:
        try:
            ventas_query = ventas_query.filter(VentaFactura.vendedor_id == int(vendedor_q))
        except ValueError:
            pass
    if producto_q:
        ventas_query = ventas_query.join(VentaItem).join(Producto).filter(
            or_(
                func.lower(Producto.cod_producto).like(f"%{producto_q.lower()}%"),
                func.lower(Producto.descripcion).like(f"%{producto_q.lower()}%"),
            )
        )
    ventas = ventas_query.order_by(VentaFactura.fecha.desc(), VentaFactura.id.desc()).all()

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

    rows = []
    resumen = {}
    total_saldo_cs = Decimal("0")
    total_saldo_usd = Decimal("0")
    for factura in ventas:
        if factura.estado == "ANULADA":
            total_abono_usd = Decimal("0")
            total_abono_cs = Decimal("0")
            saldo_usd = Decimal("0")
            saldo_cs = Decimal("0")
        else:
            total_abono_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos)
            total_abono_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos)
            total_due_usd = Decimal(str(factura.total_usd or 0))
            total_due_cs = Decimal(str(factura.total_cs or 0))
            if factura.estado_cobranza == "PENDIENTE":
                total_paid_usd = total_abono_usd
                total_paid_cs = total_abono_cs
            else:
                total_paid_usd = sum(Decimal(str(p.monto_usd or 0)) for p in factura.pagos) + total_abono_usd
                total_paid_cs = sum(Decimal(str(p.monto_cs or 0)) for p in factura.pagos) + total_abono_cs
            saldo_usd = max(total_due_usd - total_paid_usd, Decimal("0"))
            saldo_cs = max(total_due_cs - total_paid_cs, Decimal("0"))
            total_saldo_cs += saldo_cs
            total_saldo_usd += saldo_usd
        vendedor_name = factura.vendedor.nombre if factura.vendedor else "-"
        resumen.setdefault(vendedor_name, {"cs": Decimal("0"), "usd": Decimal("0")})
        resumen[vendedor_name]["cs"] += saldo_cs
        resumen[vendedor_name]["usd"] += saldo_usd
        rows.append(
            {
                "numero": factura.numero,
                "fecha": factura.fecha,
                "cliente": factura.cliente.nombre if factura.cliente else "Consumidor final",
                "vendedor": vendedor_name,
                "saldo_usd": saldo_usd,
                "saldo_cs": saldo_cs,
                "estado": factura.estado_cobranza,
            }
        )

    if format.lower() != "pdf":
        return JSONResponse({"ok": False, "message": "Formato no soportado"}, status_code=400)

    buffer = io.BytesIO()
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 40
    logo_path = _resolve_logo_path(company_profile.get("logo_url", ""))
    if logo_path.exists():
        c.drawImage(str(logo_path), 40, y - 30, width=90, height=30, mask="auto")
    c.setFont("Times-Bold", 14)
    c.drawString(150, y - 8, "Informe de Cuentas por Cobrar")
    y -= 35
    c.setFont("Times-Roman", 10)
    branch_name = "-"
    if bodega and bodega.branch:
        branch_name = bodega.branch.name
    c.drawString(40, y, f"Sucursal: {branch_name}")
    y -= 14
    c.drawString(40, y, f"Rango: {start_date} a {end_date}")
    y -= 14
    if vendedor_q:
        vendedor_name = next((v.nombre for v in db.query(Vendedor).all() if str(v.id) == vendedor_q), "")
        c.drawString(40, y, f"Vendedor: {vendedor_name}")
        y -= 14
    c.drawString(40, y, f"Tasa: {rate_today.rate if rate_today else 'N/D'}")
    y -= 16
    c.line(40, y, width - 40, y)
    y -= 16

    c.setFont("Times-Bold", 11)
    c.drawString(40, y, "Resumen por vendedor")
    y -= 12
    c.setFont("Times-Bold", 9)
    c.drawString(50, y, "Vendedor")
    c.drawString(320, y, "Saldo C$")
    c.drawString(430, y, "Saldo $")
    y -= 8
    c.line(40, y, width - 40, y)
    y -= 10
    c.setFont("Times-Roman", 9)
    resumen_total_cs = Decimal("0")
    resumen_total_usd = Decimal("0")
    for vend, totals in resumen.items():
        if y < 120:
            c.showPage()
            y = height - 60
        c.drawString(50, y, vend)
        c.setFillColor(colors.HexColor("#1d4ed8"))
        c.drawRightString(400, y, f"C$ {totals['cs']:,.2f}")
        c.setFillColor(colors.HexColor("#16a34a"))
        c.drawRightString(500, y, f"$ {totals['usd']:,.2f}")
        c.setFillColor(colors.black)
        resumen_total_cs += totals["cs"]
        resumen_total_usd += totals["usd"]
        y -= 12

    y -= 4
    c.setFillColor(colors.HexColor("#1e3a8a"))
    c.roundRect(40, y - 6, width - 80, 12, 4, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Times-Bold", 9)
    c.drawString(50, y - 2, "Total resumen")
    c.drawRightString(400, y - 2, f"C$ {resumen_total_cs:,.2f}")
    c.drawRightString(500, y - 2, f"$ {resumen_total_usd:,.2f}")
    c.setFillColor(colors.black)
    y -= 18

    c.setFillColor(colors.HexColor("#1e3a8a"))
    c.roundRect(40, y - 6, width - 80, 12, 4, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Times-Bold", 9)
    c.drawString(50, y - 2, "Detalle de creditos")
    c.setFillColor(colors.black)
    y -= 18

    c.setFillColor(colors.black)
    c.setFont("Times-Bold", 11)
    c.setFont("Times-Bold", 9)
    c.drawString(40, y, "Factura")
    c.drawString(110, y, "Cliente")
    c.drawString(260, y, "Vendedor")
    c.drawString(370, y, "Estado")
    c.drawString(430, y, "Saldo C$")
    c.drawString(510, y, "Saldo $")
    y -= 8
    c.line(40, y, width - 40, y)
    y -= 12
    c.setFont("Times-Roman", 9)
    for row in rows:
        if y < 80:
            c.showPage()
            y = height - 60
        c.drawString(40, y, row["numero"])
        c.drawString(110, y, (row["cliente"][:20] + "…") if len(row["cliente"]) > 20 else row["cliente"])
        c.drawString(260, y, (row["vendedor"][:14] + "…") if len(row["vendedor"]) > 14 else row["vendedor"])
        c.drawString(370, y, row["estado"])
        c.setFillColor(colors.HexColor("#1d4ed8"))
        c.drawRightString(495, y, f"C$ {row['saldo_cs']:,.2f}")
        c.setFillColor(colors.HexColor("#16a34a"))
        c.drawRightString(570, y, f"$ {row['saldo_usd']:,.2f}")
        c.setFillColor(colors.black)
        y -= 12

    y -= 4
    c.line(40, y, width - 40, y)
    y -= 14
    c.setFont("Times-Bold", 10)
    c.drawString(40, y, "Totales pendientes:")
    c.setFillColor(colors.HexColor("#1d4ed8"))
    c.drawRightString(495, y, f"C$ {total_saldo_cs:,.2f}")
    c.setFillColor(colors.HexColor("#16a34a"))
    c.drawRightString(570, y, f"$ {total_saldo_usd:,.2f}")
    c.setFillColor(colors.black)
    y -= 14
    c.showPage()
    c.save()
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=cobranza.pdf"},
    )


@router.get("/sales/cobranza/{venta_id}/abonos")
def sales_cobranza_abonos(
    request: Request,
    venta_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.cobranza")
    factura = db.query(VentaFactura).filter(VentaFactura.id == venta_id).first()
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and factura.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Factura fuera de tu bodega"}, status_code=403)
    abonos = (
        db.query(CobranzaAbono)
        .filter(CobranzaAbono.factura_id == factura.id)
        .order_by(CobranzaAbono.fecha.desc(), CobranzaAbono.id.desc())
        .all()
    )
    items = []
    for abono in abonos:
        items.append(
            {
                "id": abono.id,
                "numero": abono.numero,
                "fecha": abono.fecha.isoformat() if abono.fecha else "",
                "moneda": abono.moneda,
                "monto": float(abono.monto_usd if abono.moneda == "USD" else abono.monto_cs),
                "observacion": abono.observacion or "",
            }
        )
    return JSONResponse({"ok": True, "items": items})


@router.post("/sales/cobranza/{venta_id}/abono")
async def sales_cobranza_abono(
    venta_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.pagos")
    form = await request.form()
    moneda = (form.get("moneda") or "CS").upper()
    monto_raw = form.get("monto")
    observacion = (form.get("observacion") or "").strip()
    if not monto_raw:
        return JSONResponse({"ok": False, "message": "Monto requerido"}, status_code=400)
    if moneda not in {"CS", "USD"}:
        return JSONResponse({"ok": False, "message": "Moneda invalida"}, status_code=400)

    def parse_decimal(value: str) -> Decimal:
        raw = re.sub(r"[^0-9.,-]", "", str(value or "0"))
        if "," in raw and "." in raw:
            if raw.rfind(",") > raw.rfind("."):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "," in raw and "." not in raw:
            parts = raw.split(",")
            if len(parts) == 2 and len(parts[1]) == 2:
                raw = raw.replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "." in raw and "," not in raw:
            parts = raw.split(".")
            if len(parts) == 2 and len(parts[1]) == 2:
                raw = raw
            else:
                raw = raw.replace(".", "")
        try:
            return Decimal(raw)
        except Exception:
            return Decimal("0")

    monto = parse_decimal(monto_raw)
    if monto <= 0:
        return JSONResponse({"ok": False, "message": "Monto invalido"}, status_code=400)

    factura = db.query(VentaFactura).filter(VentaFactura.id == venta_id).first()
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    if factura.estado == "ANULADA":
        return JSONResponse({"ok": False, "message": "Factura anulada"}, status_code=400)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and factura.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Factura fuera de tu bodega"}, status_code=403)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if moneda == "USD" and not rate_today:
        return JSONResponse({"ok": False, "message": "Tasa no configurada"}, status_code=400)
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

    if moneda == "USD":
        monto_usd = monto
        monto_cs = monto * tasa
    else:
        monto_cs = monto
        monto_usd = monto / tasa if tasa else Decimal("0")

    if factura.estado_cobranza == "PAGADA":
        return JSONResponse({"ok": False, "message": "Factura ya pagada"}, status_code=400)

    last_abono = (
        db.query(CobranzaAbono)
        .filter(CobranzaAbono.bodega_id == factura.bodega_id)
        .order_by(CobranzaAbono.secuencia.desc())
        .first()
    )
    next_seq = (last_abono.secuencia if last_abono else 0) + 1
    branch_code = (factura.bodega.branch.code if factura.bodega and factura.bodega.branch else "").lower()
    prefix = "C" if branch_code == "central" else "E" if branch_code == "esteli" else (branch_code[:1].upper() or "X")
    numero = f"RE-{prefix}-{next_seq:06d}"

    abono = CobranzaAbono(
        factura_id=factura.id,
        branch_id=factura.bodega.branch_id if factura.bodega else user.default_branch_id,
        bodega_id=factura.bodega_id or user.default_bodega_id,
        secuencia=next_seq,
        numero=numero,
        fecha=local_today(),
        moneda=moneda,
        tasa_cambio=tasa if tasa else None,
        monto_usd=monto_usd,
        monto_cs=monto_cs,
        observacion=observacion,
        usuario_registro=user.full_name,
    )
    db.add(abono)

    total_paid_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos) + monto_usd
    total_paid_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos) + monto_cs
    due_usd = Decimal(str(factura.total_usd or 0))
    due_cs = Decimal(str(factura.total_cs or 0))
    if (factura.moneda or "CS") == "USD":
        factura.estado_cobranza = "PAGADA" if total_paid_usd >= due_usd else "PENDIENTE"
    else:
        factura.estado_cobranza = "PAGADA" if total_paid_cs >= due_cs else "PENDIENTE"

    db.commit()
    return JSONResponse({"ok": True, "message": "Abono registrado"})


@router.post("/sales/cobranza/{venta_id}/abono/{abono_id}")
async def sales_cobranza_abono_update(
    venta_id: int,
    abono_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.pagos")
    form = await request.form()
    moneda = (form.get("moneda") or "CS").upper()
    monto_raw = form.get("monto")
    observacion = (form.get("observacion") or "").strip()
    if not monto_raw:
        return JSONResponse({"ok": False, "message": "Monto requerido"}, status_code=400)
    if moneda not in {"CS", "USD"}:
        return JSONResponse({"ok": False, "message": "Moneda invalida"}, status_code=400)

    def parse_decimal(value: str) -> Decimal:
        raw = re.sub(r"[^0-9.,-]", "", str(value or "0"))
        if "," in raw and "." in raw:
            if raw.rfind(",") > raw.rfind("."):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "," in raw and "." not in raw:
            parts = raw.split(",")
            if len(parts) == 2 and len(parts[1]) == 2:
                raw = raw.replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "." in raw and "," not in raw:
            parts = raw.split(".")
            if len(parts) == 2 and len(parts[1]) == 2:
                raw = raw
            else:
                raw = raw.replace(".", "")
        try:
            return Decimal(raw)
        except Exception:
            return Decimal("0")

    monto = parse_decimal(monto_raw)
    if monto <= 0:
        return JSONResponse({"ok": False, "message": "Monto invalido"}, status_code=400)

    factura = db.query(VentaFactura).filter(VentaFactura.id == venta_id).first()
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and factura.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Factura fuera de tu bodega"}, status_code=403)

    abono = db.query(CobranzaAbono).filter(CobranzaAbono.id == abono_id, CobranzaAbono.factura_id == factura.id).first()
    if not abono:
        return JSONResponse({"ok": False, "message": "Abono no encontrado"}, status_code=404)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if moneda == "USD" and not rate_today:
        return JSONResponse({"ok": False, "message": "Tasa no configurada"}, status_code=400)
    tasa = Decimal(str(rate_today.rate)) if rate_today else Decimal("0")

    if moneda == "USD":
        monto_usd = monto
        monto_cs = monto * tasa
    else:
        monto_cs = monto
        monto_usd = monto / tasa if tasa else Decimal("0")

    abono.moneda = moneda
    abono.tasa_cambio = tasa if tasa else None
    abono.monto_usd = monto_usd
    abono.monto_cs = monto_cs
    abono.observacion = observacion
    db.commit()

    total_abono_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos)
    total_abono_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos)
    due_usd = Decimal(str(factura.total_usd or 0))
    due_cs = Decimal(str(factura.total_cs or 0))
    if (factura.moneda or "CS") == "USD":
        factura.estado_cobranza = "PAGADA" if total_abono_usd >= due_usd else "PENDIENTE"
    else:
        factura.estado_cobranza = "PAGADA" if total_abono_cs >= due_cs else "PENDIENTE"
    db.commit()

    return JSONResponse({"ok": True, "message": "Abono actualizado"})


@router.post("/sales/cobranza/{venta_id}/abono/{abono_id}/delete")
def sales_cobranza_abono_delete(
    request: Request,
    venta_id: int,
    abono_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.pagos")
    factura = db.query(VentaFactura).filter(VentaFactura.id == venta_id).first()
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    if factura.estado == "ANULADA":
        return JSONResponse({"ok": False, "message": "Factura anulada"}, status_code=400)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and factura.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Factura fuera de tu bodega"}, status_code=403)

    abono = (
        db.query(CobranzaAbono)
        .filter(CobranzaAbono.id == abono_id, CobranzaAbono.factura_id == factura.id)
        .first()
    )
    if not abono:
        return JSONResponse({"ok": False, "message": "Abono no encontrado"}, status_code=404)

    db.delete(abono)

    total_abono_usd = sum(Decimal(str(a.monto_usd or 0)) for a in factura.abonos if a.id != abono_id)
    total_abono_cs = sum(Decimal(str(a.monto_cs or 0)) for a in factura.abonos if a.id != abono_id)
    due_usd = Decimal(str(factura.total_usd or 0))
    due_cs = Decimal(str(factura.total_cs or 0))
    if (factura.moneda or "CS") == "USD":
        factura.estado_cobranza = "PAGADA" if total_abono_usd >= due_usd else "PENDIENTE"
    else:
        factura.estado_cobranza = "PAGADA" if total_abono_cs >= due_cs else "PENDIENTE"
    db.commit()

    return JSONResponse({"ok": True, "message": "Abono eliminado"})


@router.post("/sales/cobranza/{venta_id}/estado")
async def sales_cobranza_estado(
    venta_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_user_web),
):
    _enforce_permission(request, user, "access.sales.cobranza")
    form = await request.form()
    estado = (form.get("estado") or "PENDIENTE").upper()
    if estado not in {"PENDIENTE", "PAGADA"}:
        return JSONResponse({"ok": False, "message": "Estado invalido"}, status_code=400)
    factura = db.query(VentaFactura).filter(VentaFactura.id == venta_id).first()
    if not factura:
        return JSONResponse({"ok": False, "message": "Factura no encontrada"}, status_code=404)
    _, bodega = _resolve_branch_bodega(db, user)
    if bodega and factura.bodega_id != bodega.id:
        return JSONResponse({"ok": False, "message": "Factura fuera de tu bodega"}, status_code=403)
    factura.estado_cobranza = estado
    db.commit()
    return JSONResponse({"ok": True, "message": "Estado actualizado"})


@router.post("/inventory/import")
def inventory_import_products(
    request: Request,
    file: UploadFile = File(...),
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Archivo+Excel+(.xlsx)+requerido", status_code=303)

    rate_today = (
        db.query(ExchangeRate)
        .filter(ExchangeRate.effective_date <= local_today())
        .order_by(ExchangeRate.effective_date.desc())
        .first()
    )
    if not rate_today:
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Tasa+de+cambio+no+configurada", status_code=303)

    content = file.file.read()
    if not content:
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Archivo+vacio", status_code=303)

    def to_decimal(value) -> Decimal:
        if value is None:
            return Decimal("0")
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        cleaned = str(value).strip().replace(",", "")
        if not cleaned:
            return Decimal("0")
        try:
            return Decimal(cleaned)
        except Exception:
            return Decimal("0")

    tasa = Decimal(str(rate_today.rate))
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    header_map = {name.lower(): idx for idx, name in enumerate(header_row)}

    def col_idx(*names: str) -> Optional[int]:
        for name in names:
            idx = header_map.get(name.lower())
            if idx is not None:
                return idx
        return None

    idx_codigo = col_idx("codigo", "cod_producto", "cod")
    idx_desc = col_idx("descripcion", "descripción", "nombre")
    idx_linea = col_idx("linea")
    idx_segmento = col_idx("segmento")
    idx_costo_usd = col_idx("costo_usd", "costo_producto_usd", "costo")
    idx_precio_usd = col_idx("precio_usd", "precio_venta1_usd")
    idx_precio_cs = col_idx("precio_cs", "precio_cordobas", "precio_venta1")
    idx_saldo_central = col_idx("saldo_central", "existencia_central")
    idx_saldo_esteli = col_idx("saldo_esteli", "existencia_esteli")

    if idx_codigo is None or idx_desc is None:
        target = redirect_to or "/inventory"
        return RedirectResponse(f"{target}?error=Columnas+requeridas:+codigo+y+descripcion", status_code=303)

    bodegas = _scoped_bodegas_query(db).all()
    bodega_central = next(
        (b for b in bodegas if (b.code or "").lower() == "central"),
        None,
    )
    if not bodega_central:
        bodega_central = next((b for b in bodegas if "central" in (b.name or "").lower()), None)
    bodega_esteli = next(
        (b for b in bodegas if (b.code or "").lower() == "esteli"),
        None,
    )
    if not bodega_esteli:
        bodega_esteli = next((b for b in bodegas if "esteli" in (b.name or "").lower()), None)

    ingreso_tipo = (
        db.query(IngresoTipo)
        .filter(func.lower(IngresoTipo.nombre).like("%ajuste%"))
        .first()
    )
    if not ingreso_tipo:
        ingreso_tipo = (
            db.query(IngresoTipo)
            .filter(func.lower(IngresoTipo.nombre).like("%apertura%"))
            .first()
        )
    if not ingreso_tipo:
        ingreso_tipo = db.query(IngresoTipo).order_by(IngresoTipo.id).first()

    central_items: list[tuple[Producto, Decimal]] = []
    esteli_items: list[tuple[Producto, Decimal]] = []
    total_rows = 0
    skipped_rows = 0
    created_count = 0
    updated_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        cod = str(row[idx_codigo]).strip() if row[idx_codigo] is not None else ""
        descripcion = str(row[idx_desc]).strip() if row[idx_desc] is not None else ""
        if not cod or not descripcion:
            skipped_rows += 1
            continue

        linea = None
        segmento = None
        linea_name = str(row[idx_linea]).strip() if idx_linea is not None and row[idx_linea] is not None else ""
        segmento_name = str(row[idx_segmento]).strip() if idx_segmento is not None and row[idx_segmento] is not None else ""
        if linea_name:
            linea = (
                db.query(Linea)
                .filter(func.lower(Linea.linea) == linea_name.lower())
                .first()
            )
            if not linea:
                linea = Linea(cod_linea=linea_name[:50], linea=linea_name, activo=True)
                db.add(linea)
                db.flush()
        if segmento_name:
            segmento = (
                db.query(Segmento)
                .filter(func.lower(Segmento.segmento) == segmento_name.lower())
                .first()
            )
            if not segmento:
                segmento = Segmento(segmento=segmento_name)
                db.add(segmento)
                db.flush()

        costo_usd = to_decimal(row[idx_costo_usd]) if idx_costo_usd is not None else Decimal("0")
        precio_usd = to_decimal(row[idx_precio_usd]) if idx_precio_usd is not None else Decimal("0")
        precio_cs = to_decimal(row[idx_precio_cs]) if idx_precio_cs is not None else Decimal("0")
        if precio_usd == 0 and precio_cs > 0:
            precio_usd = (precio_cs / tasa) if tasa else Decimal("0")
        if precio_cs == 0 and precio_usd > 0:
            precio_cs = precio_usd * tasa

        producto = db.query(Producto).filter(Producto.cod_producto == cod).first()
        if not producto:
            producto = Producto(
                cod_producto=cod,
                descripcion=descripcion,
                linea_id=linea.id if linea else None,
                segmento_id=segmento.id if segmento else None,
                precio_venta1=precio_cs,
                precio_venta2=Decimal("0"),
                precio_venta3=Decimal("0"),
                precio_venta1_usd=precio_usd,
                precio_venta2_usd=Decimal("0"),
                precio_venta3_usd=Decimal("0"),
                tasa_cambio=tasa,
                costo_producto=costo_usd * tasa,
                activo=True,
            )
            db.add(producto)
            db.flush()
            created_count += 1
        else:
            producto.descripcion = descripcion
            if linea:
                producto.linea_id = linea.id
            if segmento:
                producto.segmento_id = segmento.id
            if precio_usd > 0:
                producto.precio_venta1_usd = precio_usd
                producto.precio_venta1 = precio_cs
            if costo_usd > 0:
                producto.costo_producto = costo_usd * tasa
            producto.tasa_cambio = tasa
            updated_count += 1

        saldo_central = to_decimal(row[idx_saldo_central]) if idx_saldo_central is not None else Decimal("0")
        saldo_esteli = to_decimal(row[idx_saldo_esteli]) if idx_saldo_esteli is not None else Decimal("0")

        if bodega_central and saldo_central > 0:
            central_items.append((producto, saldo_central))
        if bodega_esteli and saldo_esteli > 0:
            esteli_items.append((producto, saldo_esteli))

        # No mezclar bodegas en el saldo global: se mantiene en cero y
        # el saldo por bodega se calcula desde los movimientos.
        saldo_row = db.query(SaldoProducto).filter(SaldoProducto.producto_id == producto.id).first()
        if not saldo_row:
            db.add(SaldoProducto(producto_id=producto.id, existencia=Decimal("0")))
        else:
            saldo_row.existencia = Decimal("0")

    def create_ingreso(bodega: Bodega, items: list[tuple[Producto, Decimal]]) -> None:
        if not items:
            return
        ingreso = IngresoInventario(
            tipo_id=ingreso_tipo.id if ingreso_tipo else 1,
            bodega_id=bodega.id,
            proveedor_id=None,
            fecha=local_today(),
            moneda="USD",
            tasa_cambio=tasa,
            total_usd=Decimal("0"),
            total_cs=Decimal("0"),
            observacion="Carga inicial por importacion",
            usuario_registro="Importador",
        )
        db.add(ingreso)
        db.flush()
        total_usd = Decimal("0")
        total_cs = Decimal("0")
        for producto, qty in items:
            costo_unit_usd = Decimal(str(producto.costo_producto or 0)) / tasa if tasa else Decimal("0")
            subtotal_usd = costo_unit_usd * qty
            subtotal_cs = subtotal_usd * tasa
            total_usd += subtotal_usd
            total_cs += subtotal_cs
            db.add(
                IngresoItem(
                    ingreso_id=ingreso.id,
                    producto_id=producto.id,
                    cantidad=qty,
                    costo_unitario_usd=costo_unit_usd,
                    costo_unitario_cs=costo_unit_usd * tasa,
                    subtotal_usd=subtotal_usd,
                    subtotal_cs=subtotal_cs,
                )
            )
        ingreso.total_usd = total_usd
        ingreso.total_cs = total_cs

    if bodega_central:
        create_ingreso(bodega_central, central_items)
    if bodega_esteli:
        create_ingreso(bodega_esteli, esteli_items)

    db.commit()
    target = redirect_to or "/inventory"
    msg = (
        f"Importacion completa. Filas: {total_rows}. "
        f"Creados: {created_count}. Actualizados: {updated_count}. Omitidos: {skipped_rows}."
    )
    return RedirectResponse(f"{target}?success={msg}", status_code=303)


@router.get("/inventory/import/template")
def inventory_import_template(
    request: Request,
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    headers = [
        "codigo",
        "descripcion",
        "linea",
        "segmento",
        "costo_usd",
        "precio_usd",
        "precio_cs",
        "saldo_central",
        "saldo_esteli",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos_template.xlsx"},
    )


@router.post("/inventory/import/preview")
def inventory_import_preview(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return HTMLResponse(
            "<div class='alert alert-warning py-2 px-3'>Archivo Excel (.xlsx) requerido.</div>"
        )
    content = file.file.read()
    if not content:
        return HTMLResponse(
            "<div class='alert alert-warning py-2 px-3'>Archivo vacio.</div>"
        )
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows = []
    total_rows = 0
    skipped_rows = 0
    idx_codigo = None
    idx_desc = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        total_rows += 1
        rows.append([cell if cell is not None else "" for cell in row])
    if headers:
        header_map = {name.lower(): idx for idx, name in enumerate(headers)}
        idx_codigo = header_map.get("codigo")
        idx_desc = header_map.get("descripcion")
    if idx_codigo is not None or idx_desc is not None:
        for row in rows:
            cod = str(row[idx_codigo]).strip() if idx_codigo is not None and row[idx_codigo] is not None else ""
            desc = str(row[idx_desc]).strip() if idx_desc is not None and row[idx_desc] is not None else ""
            if not cod or not desc:
                skipped_rows += 1
    if not headers:
        return HTMLResponse(
            "<div class='alert alert-warning py-2 px-3'>No se detectaron columnas.</div>"
        )
    table = [
        "<div class='mt-3'>",
        "<div class='fw-semibold mb-2'>Vista previa (todas las filas)</div>",
        f"<div class='small text-muted mb-2'>Filas detectadas: {total_rows}. Omitidas por codigo/descripcion vacios: {skipped_rows}.</div>",
        "<div class='table-responsive'>",
        "<table class='table table-sm table-striped align-middle'>",
        "<thead><tr>",
    ]
    for h in headers:
        table.append(f"<th class='text-nowrap'>{h}</th>")
    table.append("</tr></thead><tbody>")
    for row in rows:
        table.append("<tr>")
        for cell in row:
            table.append(f"<td class='text-nowrap'>{cell}</td>")
        table.append("</tr>")
    table.append("</tbody></table></div></div>")
    return HTMLResponse("".join(table))


@router.post("/inventory/import/reset")
def inventory_import_reset(
    request: Request,
    redirect_to: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    # Guard rail: avoid wiping products if there are sales linked.
    ventas_count = db.query(VentaItem).count()
    if ventas_count > 0:
        target = redirect_to or "/inventory"
        return RedirectResponse(
            f"{target}?error=No+se+puede+limpiar+inventario+con+ventas+registradas",
            status_code=303,
        )

    db.query(IngresoItem).delete()
    db.query(IngresoInventario).delete()
    db.query(EgresoItem).delete()
    db.query(EgresoInventario).delete()
    db.query(SaldoProducto).delete()
    db.query(ProductoCombo).delete()
    db.query(Producto).delete()
    db.commit()

    target = redirect_to or "/inventory"
    return RedirectResponse(f"{target}?success=Inventario+limpio", status_code=303)


@router.get("/finance/rates")
def finance_rates(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.finance.rates")
    error = request.query_params.get("error")
    success = request.query_params.get("success")
    rates = db.query(ExchangeRate).order_by(ExchangeRate.effective_date.desc()).all()
    return request.app.state.templates.TemplateResponse(
        "finance_rates.html",
        {
            "request": request,
            "user": user,
            "rates": rates,
            "error": error,
            "success": success,
            "version": settings.UI_VERSION,
        },
    )


@router.post("/finance/rates")
def finance_rates_create(
    request: Request,
    effective_date: date = Form(...),
    period: str = Form(...),
    rate: float = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.finance.rates")
    if rate <= 0:
        return RedirectResponse("/finance/rates?error=Tasa+no+valida", status_code=303)
    exists = (
        db.query(ExchangeRate)
        .filter(
            ExchangeRate.effective_date == effective_date,
            ExchangeRate.period == period,
        )
        .first()
    )
    if not exists:
        db.add(ExchangeRate(effective_date=effective_date, period=period, rate=rate))
        db.commit()
        return RedirectResponse("/finance/rates?success=Tasa+creada", status_code=303)
    return RedirectResponse("/finance/rates?error=Ya+existe+una+tasa+con+esa+fecha+y+periodo", status_code=303)


@router.post("/finance/rates/{rate_id}/update")
def finance_rates_update(
    request: Request,
    rate_id: int,
    effective_date: date = Form(...),
    period: str = Form(...),
    rate: float = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.finance.rates")
    if rate <= 0:
        return RedirectResponse("/finance/rates?error=Tasa+no+valida", status_code=303)
    row = db.query(ExchangeRate).filter(ExchangeRate.id == rate_id).first()
    if not row:
        return RedirectResponse("/finance/rates?error=Registro+no+encontrado", status_code=303)
    exists = (
        db.query(ExchangeRate)
        .filter(
            ExchangeRate.effective_date == effective_date,
            ExchangeRate.period == period,
            ExchangeRate.id != rate_id,
        )
        .first()
    )
    if exists:
        return RedirectResponse("/finance/rates?error=Ya+existe+otra+tasa+con+esa+fecha+y+periodo", status_code=303)
    row.effective_date = effective_date
    row.period = period
    row.rate = rate
    db.commit()
    return RedirectResponse("/finance/rates?success=Tasa+actualizada", status_code=303)


@router.post("/finance/rates/{rate_id}/delete")
def finance_rates_delete(
    request: Request,
    rate_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.finance.rates")
    row = db.query(ExchangeRate).filter(ExchangeRate.id == rate_id).first()
    if not row:
        return RedirectResponse("/finance/rates?error=Registro+no+encontrado", status_code=303)
    db.delete(row)
    db.commit()
    return RedirectResponse("/finance/rates?success=Tasa+eliminada", status_code=303)


@router.post("/inventory/product/{product_id}/deactivate")
def inventory_deactivate_product(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    producto = db.query(Producto).filter(Producto.id == product_id).first()
    if producto:
        existencia = float(producto.saldo.existencia) if producto.saldo else 0
        if existencia <= 0:
            producto.activo = False
            db.commit()
            return RedirectResponse("/inventory", status_code=303)
        return RedirectResponse("/inventory?error=No+se+puede+desactivar+con+existencia", status_code=303)
    return RedirectResponse("/inventory?error=Producto+no+encontrado", status_code=303)


@router.post("/inventory/product/{product_id}/activate")
def inventory_activate_product(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_require_admin_web),
):
    _enforce_permission(request, user, "access.inventory.productos")
    producto = db.query(Producto).filter(Producto.id == product_id).first()
    if producto:
        producto.activo = True
        db.commit()
        return RedirectResponse("/inventory", status_code=303)
    return RedirectResponse("/inventory?error=Producto+no+encontrado", status_code=303)

